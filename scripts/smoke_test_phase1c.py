import sys
import os
import shutil
import subprocess
import glob
import openpyxl
import pandas as pd

# Define paths
SITE_DATA_SRC = 'Info/input/site_pr_po_view.xlsx'
TEST_OUTPUT_DIR = 'output_ti_test'
TEMP_SITE_DATA = os.path.join(TEST_OUTPUT_DIR, 'site_pr_po_view_smoke_temp.xlsx')

print("=" * 100)
print("PHASE 1C REGRESSION / SMOKE VALIDATION SUITE")
print("=" * 100)

# Step 1: Initialize isolated test environment
print(f"\n[STEP 1] Preparing isolated environment under {TEST_OUTPUT_DIR}/...")
if os.path.exists(TEST_OUTPUT_DIR):
    # Keep the directory itself but clear old contents
    for item in os.listdir(TEST_OUTPUT_DIR):
        item_path = os.path.join(TEST_OUTPUT_DIR, item)
        if os.path.isdir(item_path):
            shutil.rmtree(item_path)
        else:
            os.remove(item_path)
else:
    os.makedirs(TEST_OUTPUT_DIR)
print(f"  ✓ Cleaned isolated environment directory")

# Step 2: Copy and patch the site data file
print(f"\n[STEP 2] Creating temporary site view file: {TEMP_SITE_DATA}")
shutil.copy2(SITE_DATA_SRC, TEMP_SITE_DATA)

print("  Preparing Excel sheet to clear duplicate flags ('Subcon PR - TI' col)...")
wb = openpyxl.load_workbook(TEMP_SITE_DATA)
# Try loading sheet named 'data', fall back to active sheet if not found
ws = wb['data'] if 'data' in wb.sheetnames else wb.active

# Find columns dynamically by matching row 4 headers
pr_col_idx = None
site_code_col_idx = None

# Scan row 4 for headers
for c in range(1, ws.max_column + 1):
    val = ws.cell(4, c).value
    if val == 'Subcon PR - TI':
        pr_col_idx = c
    elif val == 'customer site code':
        site_code_col_idx = c

if pr_col_idx is None:
    # Fallback to column 32 (AF) if not found dynamically
    pr_col_idx = 32
    print(f"  ⚠ 'Subcon PR - TI' header not found in row 4. Using fallback col {pr_col_idx}")
else:
    print(f"  ✓ Found 'Subcon PR - TI' header dynamically at column {pr_col_idx}")

if site_code_col_idx is None:
    # Fallback to column 1 (A) if not found dynamically
    site_code_col_idx = 1
    print(f"  ⚠ 'customer site code' header not found in row 4. Using fallback col {site_code_col_idx}")
else:
    print(f"  ✓ Found 'customer site code' header dynamically at column {site_code_col_idx}")

# Target sites we need to test
test_sites = {'1106L_HU', '1007D_HU', '9743C_AD', 'A01073_AD', '9313A_AD'}
cleared_count = 0

for r in range(5, ws.max_row + 1):
    site_val = str(ws.cell(r, site_code_col_idx).value).strip()
    if site_val in test_sites:
        existing_val = ws.cell(r, pr_col_idx).value
        if existing_val is not None:
            ws.cell(r, pr_col_idx).value = None
            cleared_count += 1
            print(f"    - Cleared 'Subcon PR - TI' for {site_val} at row {r} (was: '{existing_val}')")

wb.save(TEMP_SITE_DATA)
wb.close()
print(f"  ✓ Saved temporary patched site data. Cleared duplicate flags on {cleared_count} rows.")

print("  Adding synthetic TI review scenarios...")
wb = openpyxl.load_workbook(TEMP_SITE_DATA)
ws = wb['data'] if 'data' in wb.sheetnames else wb.active
header_map = {str(ws.cell(4, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(4, c).value is not None}

def find_row(site_code):
    col = header_map['customer site code']
    for row_idx in range(5, ws.max_row + 1):
        if str(ws.cell(row_idx, col).value).strip() == site_code:
            return row_idx
    raise RuntimeError(f"Source site row not found: {site_code}")

def clone_site(source_site, target_site, overrides):
    source_row = find_row(source_site)
    target_row = ws.max_row + 1
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(target_row, col_idx).value = ws.cell(source_row, col_idx).value
    ws.cell(target_row, header_map['customer site code']).value = target_site
    for header, value in overrides.items():
        if header in header_map:
            ws.cell(target_row, header_map[header]).value = value
    print(f"    - Added synthetic row {target_site} from {source_site}")

clone_site('9313A_AD', 'QA_RE_NO_MODEL', {
    'Subcon PR - TI': None,
    'Tx SOW': 'MW Re-eng Unsupported',
    'Province/State': 'Selangor',
    'region': 'Central'
})

clone_site('1106L_HU', 'QA_SABAH_COORD', {
    'Subcon PR - TI': None,
    'SubCon - TI Team': 'Allstar',
    'Tx SOW': 'MW Swap',
    'region': 'Sabah',
    'Province/State': 'Sabah',
    'Latitude (North Plus South Minus)': 5.9804,
    'Longitude (East Plus West Minus)': 116.0735,
    'MW Config Antenna Size NE': 0.6,
    'MW Config Antenna Size FE': 0.6
})

wb.save(TEMP_SITE_DATA)
wb.close()
print("  ✓ Synthetic rows added.")

# Step 3: Run the generator for each site under isolated outputs
print(f"\n[STEP 3] Running generator script on target validation sites...")
test_results = {}

sites_to_test = [
    # (site_code, expected_ecc_rows, expect_in_review, expected_reason_substring)
    ("1106L_HU", 3, False, None),
    ("1007D_HU", 0, True, "No matching antenna group item"),
    ("9743C_AD", 1, True, "MW Reroute decom antenna size ambiguous"),
    ("A01073_AD", 0, True, "Missing TI antenna size - review required"),
    ("9313A_AD", 1, False, None),
    ("QA_RE_NO_MODEL", 0, True, "NO_MATCHING_TI_PR_MODEL_ITEM"),
    ("QA_SABAH_COORD", 0, True, "COORDINATE_RESOLUTION_UNSUPPORTED")
]

for site, exp_ecc, exp_rev, exp_reason in sites_to_test:
    site_out_dir = os.path.join(TEST_OUTPUT_DIR, site)
    os.makedirs(site_out_dir, exist_ok=True)
    
    print(f"  * Generating TI PR ECC for site {site} -> {site_out_dir}/...")
    cmd = [
        "python", "scripts/generate_tss_pr_ecc.py",
        "--site-data", TEMP_SITE_DATA,
        "--site-code", site,
        "--scope", "TI",
        "--output", site_out_dir
    ]
    
    # Run generator via subprocess
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        print(f"    ✗ Generator failed with exit code {proc.returncode}")
        print(proc.stderr)
        test_results[site] = {
            "status": "FAIL",
            "error": f"Generator execution failed with exit code {proc.returncode}",
            "ecc_count": 0,
            "in_review": False,
            "reason": None
        }
        continue
        
    # Analyze outputs for this site
    # 1. Count actual ECC rows from generated Excel files
    excel_files = glob.glob(os.path.join(site_out_dir, "*.xls")) + glob.glob(os.path.join(site_out_dir, "*.xlsx"))
    total_ecc_rows = 0
    cutover_found_in_any = False
    
    for xls_path in excel_files:
        try:
            # openpyxl engine is robust for the .xls file output since it is actually openxml structured
            df_xls = pd.read_excel(xls_path, sheet_name=0)
            
            # Count rows that belong to this site
            if 'Site ID*' in df_xls.columns:
                site_rows = df_xls[df_xls['Site ID*'] == site]
                total_ecc_rows += len(site_rows)
            else:
                total_ecc_rows += len(df_xls)
                
            # Verify Phase 1B: MW Hardware Cutover exclusion in SOW* column
            sow_col = 'SOW*' if 'SOW*' in df_xls.columns else ('SOW' if 'SOW' in df_xls.columns else None)
            if sow_col:
                for val in df_xls[sow_col].dropna().astype(str):
                    if "mw hardware cutover" in val.lower():
                        cutover_found_in_any = True
                        print(f"    ✗ CRITICAL: 'MW Hardware Cutover' was found in generated ECC row: '{val}'")
        except Exception as e:
            print(f"    ⚠ Error reading generated Excel file {xls_path}: {e}")
            
    # 2. Check REVIEW_REQUIRED CSV files
    review_files = glob.glob(os.path.join(site_out_dir, "REVIEW_REQUIRED_TI_*.csv"))
    is_in_review = False
    actual_reason = None
    
    if review_files:
        latest_csv = review_files[0]
        try:
            df_rev = pd.read_csv(latest_csv)
            site_rev = df_rev[df_rev['Site_ID'] == site]
            if len(site_rev) > 0:
                is_in_review = True
                reason_parts = []
                for column in ['Reason_Code', 'Reason_Description', 'Required_Action', 'Technical_Detail', 'Review_Reason']:
                    if column in site_rev.columns:
                        reason_parts.append(str(site_rev.iloc[0][column]).strip())
                actual_reason = ' | '.join([part for part in reason_parts if part])
        except Exception as e:
            print(f"    ⚠ Error reading generated Review CSV {latest_csv}: {e}")
            
    # Perform strict assertions
    case_failures = []
    
    # Assert ECC counts
    if total_ecc_rows != exp_ecc:
        case_failures.append(f"ECC row count mismatch (got {total_ecc_rows}, expected {exp_ecc})")
        
    # Assert review status
    if is_in_review != exp_rev:
        case_failures.append(f"Review required mismatch (got {is_in_review}, expected {exp_rev})")
        
    # Assert review reasons
    if exp_rev and exp_reason:
        if not actual_reason or exp_reason.lower() not in actual_reason.lower():
            case_failures.append(f"Review reason mismatch (got '{actual_reason}', expected containing '{exp_reason}')")
            
    # Assert Phase 1B
    if cutover_found_in_any:
        case_failures.append("MW Hardware Cutover was not excluded from generated ECC rows")
        
    # Save results
    if not case_failures:
        test_results[site] = {
            "status": "PASS",
            "ecc_count": total_ecc_rows,
            "in_review": is_in_review,
            "reason": actual_reason
        }
        print(f"    ✓ Site {site} check completed successfully [PASS]")
    else:
        test_results[site] = {
            "status": "FAIL",
            "errors": case_failures,
            "ecc_count": total_ecc_rows,
            "in_review": is_in_review,
            "reason": actual_reason
        }
        print(f"    ✗ Site {site} check failed [FAIL]: {', '.join(case_failures)}")

# Step 4: Summary and Clean-up
print(f"\n[STEP 4] Cleaning up temporary patched files...")
if os.path.exists(TEMP_SITE_DATA):
    os.remove(TEMP_SITE_DATA)
    print("  ✓ Removed temporary patched site data file")

print("\n" + "=" * 100)
print("TEST SUMMARY DASHBOARD")
print("=" * 100)

suite_failed = False
for site, res in test_results.items():
    status = res["status"]
    if status == "FAIL":
        suite_failed = True
        
    in_rev_str = "YES" if res["in_review"] else "NO"
    reason_str = f" ({res['reason']})" if res["reason"] else ""
    
    print(f"Site: {site:<12} | Status: {status:<5} | ECC Rows: {res['ecc_count']:<2} | Review Req: {in_rev_str:<3}{reason_str}")
    if status == "FAIL":
        for err in res.get("errors", [res.get("error", "Unknown error")]):
            print(f"  -> ERROR: {err}")

print("=" * 100)
if suite_failed:
    print("❌ SMOKE TEST SUITE FAILED! Regressions or safety violations detected.")
    print("=" * 100 + "\n")
    sys.exit(1)
else:
    print("🎉 ALL SMOKE TEST CASES PASSED SUCCESSFULLY!")
    print("=" * 100 + "\n")
    sys.exit(0)
