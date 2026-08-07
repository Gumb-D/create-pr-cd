#!/usr/bin/env python3
"""Read-only profiler for original four-header iEPMS exports.

This module deliberately has no dependency on the ECC generator. It inventories
source headers, calculates deterministic raw and structural header hashes, and
produces only UNVERIFIED mapping suggestions. It never generates PR or ECC output.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HEADER_ROW_COUNT = 4
SCHEMA_VERSION = "1.0"
FINGERPRINT_FIELDS = ("field_code", "wbs_stage", "task_name", "display_header")
SITE_IDENTITY_PATTERN = re.compile(
    r"^site\|fix00012\|(?P<du_model_id>\d+)\|(?P<view_id>\d+)$"
)
STRUCTURAL_VIEW_TOKEN = "<VIEW_ID>"

# This is discovery-only matching. Suggested candidates are always marked
# UNVERIFIED and must not be consumed by a production DU profile.
CANONICAL_FIELD_KEYWORDS: Dict[str, Sequence[Sequence[str]]] = {
    "site_code": (("customer", "site", "code"), ("site", "code"), ("site", "id")),
    "site_name": (("customer", "site", "name"), ("site", "name")),
    "du_key": (("du", "code"), ("du", "key")),
    "tx_sow_raw": (("tx", "sow"), ("scope", "of", "work")),
    "tx_upgrade_scope_raw": (("tx", "upgrade", "scope"),),
    "region": (("region",),),
    "state": (("province", "state"), ("state",)),
    "subcontractor_ti": (("subcon", "ti"), ("subcontractor", "ti")),
    "subcontractor_planning": (("subcon", "planning"), ("subcontractor", "planning")),
    "existing_tss_pr_status": (("tss", "pr"),),
    "existing_ti_pr_status": (("ti", "pr"),),
    "latitude": (("latitude",),),
    "longitude": (("longitude",),),
    "antenna_size_ne": (("antenna", "size", "ne"),),
    "antenna_size_fe": (("antenna", "size", "fe"),),
    "boq_configuration": (("boq", "configuration"),),
    "tx_sow_details": (("tx", "sow", "details"),),
    "ne_sow_details": (("ne", "sow", "details"),),
    "fe_sow_details": (("fe", "sow", "details"),),
}

PR_CRITICAL_FIELDS = (
    "site_code",
    "tx_sow_raw",
    "region",
    "subcontractor_ti",
    "existing_tss_pr_status",
    "existing_ti_pr_status",
)


def normalize_header_value(value: Any) -> str:
    """Normalize only for deterministic comparison; retain raw values separately."""
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split())


def make_header_fingerprint(header_values: Sequence[Any]) -> Dict[str, str]:
    if len(header_values) != HEADER_ROW_COUNT:
        raise ValueError(f"A header fingerprint requires {HEADER_ROW_COUNT} values.")
    normalized = [normalize_header_value(value) for value in header_values]
    return {
        "field_code": normalized[0],
        "wbs_stage": normalized[1],
        "task_name": normalized[2],
        "display_header": normalized[3],
    }


def _canonical_fingerprint(fingerprint: Mapping[str, Any]) -> Dict[str, str]:
    return {
        key: normalize_header_value(fingerprint.get(key, ""))
        for key in FINGERPRINT_FIELDS
    }


def parse_site_identity_field_code(field_code: Any) -> Dict[str, str] | None:
    """Parse the one dynamic iEPMS site identity field-code form strictly."""
    match = SITE_IDENTITY_PATTERN.fullmatch(normalize_header_value(field_code))
    if match is None:
        return None
    return {
        "du_model_id": match.group("du_model_id"),
        "view_id": match.group("view_id"),
    }


def extract_du_identities(inventory: Mapping[str, Any]) -> List[Dict[str, str]]:
    """Return unique runtime model/view identities found in site identity columns."""
    identities: Dict[tuple[str, str], Dict[str, str]] = {}
    for sheet in inventory.get("sheets", []):
        for column in sheet.get("columns", []):
            parsed = parse_site_identity_field_code(
                column.get("fingerprint", {}).get("field_code", "")
            )
            if parsed is None:
                continue
            key = (parsed["du_model_id"], parsed["view_id"])
            identities[key] = {
                **parsed,
                "sheet_name": str(sheet.get("sheet_name", "")),
            }
    return list(identities.values())


def fingerprint_key(fingerprint: Mapping[str, Any]) -> str:
    """Exact normalized four-layer identity used for raw row lookup/provenance."""
    return json.dumps(
        _canonical_fingerprint(fingerprint),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def structural_fingerprint(fingerprint: Mapping[str, Any]) -> Dict[str, str]:
    """Normalize only the dynamic View suffix of the site identity field code."""
    canonical = _canonical_fingerprint(fingerprint)
    parsed = parse_site_identity_field_code(canonical["field_code"])
    if parsed is not None:
        canonical["field_code"] = (
            f"site|fix00012|{parsed['du_model_id']}|{STRUCTURAL_VIEW_TOKEN}"
        )
    return canonical


def structural_fingerprint_key(fingerprint: Mapping[str, Any]) -> str:
    """Comparison identity that ignores only the known dynamic View suffix."""
    return json.dumps(
        structural_fingerprint(fingerprint),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _columns_from_header_rows(sheet_name: str, header_rows: Sequence[Sequence[Any]]) -> List[Dict[str, Any]]:
    max_columns = max((len(row) for row in header_rows), default=0)
    columns: List[Dict[str, Any]] = []
    for index in range(max_columns):
        raw_values = [row[index] if index < len(row) else None for row in header_rows]
        if not any(normalize_header_value(value) for value in raw_values):
            continue
        fingerprint = make_header_fingerprint(raw_values)
        columns.append(
            {
                "source_position": {"excel_column": get_column_letter(index + 1), "one_based_index": index + 1},
                "raw_header_values": [_safe_json_value(value) for value in raw_values],
                "normalized_header_values": [normalize_header_value(value) for value in raw_values],
                "fingerprint": fingerprint,
                "fingerprint_key": fingerprint_key(fingerprint),
                "structural_fingerprint_key": structural_fingerprint_key(fingerprint),
            }
        )
    return columns


def _read_xlsx_inventory(path: Path) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets: List[Dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows(min_row=1, max_row=HEADER_ROW_COUNT, values_only=True):
            rows.append(list(row))
        while len(rows) < HEADER_ROW_COUNT:
            rows.append([])
        sheets.append(
            {
                "sheet_name": worksheet.title,
                "header_row_count": HEADER_ROW_COUNT,
                "columns": _columns_from_header_rows(worksheet.title, rows),
            }
        )
    workbook.close()
    return sheets


def _read_csv_inventory(path: Path) -> List[Dict[str, Any]]:
    try:
        text_handle = path.open("r", encoding="utf-8-sig", newline="")
        rows = []
        with text_handle as handle:
            reader = csv.reader(handle)
            for _ in range(HEADER_ROW_COUNT):
                try:
                    rows.append(next(reader))
                except StopIteration:
                    break
    except UnicodeDecodeError:
        with path.open("r", encoding="latin-1", newline="") as handle:
            reader = csv.reader(handle)
            rows = []
            for _ in range(HEADER_ROW_COUNT):
                try:
                    rows.append(next(reader))
                except StopIteration:
                    break
    while len(rows) < HEADER_ROW_COUNT:
        rows.append([])
    return [
        {
            "sheet_name": "CSV",
            "header_row_count": HEADER_ROW_COUNT,
            "columns": _columns_from_header_rows("CSV", rows),
        }
    ]


def build_header_inventory(input_path: Path) -> Dict[str, Any]:
    """Read candidate sheets and preserve the first four header rows in the inventory."""
    input_path = Path(input_path)
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        sheets = _read_xlsx_inventory(input_path)
    elif suffix == ".csv":
        sheets = _read_csv_inventory(input_path)
    else:
        raise ValueError("Only .xlsx, .xlsm, and .csv exports are supported by the profiler.")

    return {
        "schema_version": SCHEMA_VERSION,
        "source": {
            "file_name": input_path.name,
            "source_file_hash": sha256_file(input_path),
            "format": suffix.lstrip("."),
            "header_row_count": HEADER_ROW_COUNT,
        },
        "sheets": sheets,
    }


def _sheet_has_site_identity(sheet: Mapping[str, Any]) -> bool:
    """Return whether a sheet contains the strict iEPMS DU site-identity column."""
    return any(
        parse_site_identity_field_code(
            column.get("fingerprint", {}).get("field_code", "")
        )
        is not None
        for column in sheet.get("columns", [])
    )


def resolve_authoritative_sheets(
    inventory: Mapping[str, Any],
    profile: Mapping[str, Any] | None = None,
) -> List[Mapping[str, Any]]:
    """Select the sheet(s) that define DU record structure for header approval.

    The full workbook inventory remains untouched for audit/discovery. Header
    approval is scoped separately: an explicit profile selector wins; otherwise
    a single-sheet source remains backward compatible; multi-sheet workbooks
    must contain exactly one sheet with the strict DU site-identity field.
    """
    sheets = list(inventory.get("sheets", []))
    if not sheets:
        raise ValueError("No authoritative DU sheet can be resolved from an empty inventory.")

    selector: Any = None
    if profile is not None:
        selector = profile.get("export_structure", {}).get("sheet_selector")

    if selector not in (None, "", []):
        if isinstance(selector, str):
            selected_names = [selector]
        elif isinstance(selector, (list, tuple)) and selector and all(
            isinstance(value, str) and value for value in selector
        ):
            selected_names = list(selector)
        else:
            raise ValueError("Configured authoritative sheet_selector must be a sheet name or a non-empty list of sheet names.")

        if len(set(selected_names)) != len(selected_names):
            raise ValueError("Configured authoritative sheet_selector contains duplicate sheet names.")

        by_name = {str(sheet.get("sheet_name", "")): sheet for sheet in sheets}
        missing = [name for name in selected_names if name not in by_name]
        if missing:
            raise ValueError(
                "Configured authoritative sheet_selector did not resolve: "
                + ", ".join(missing)
            )
        selected = [by_name[name] for name in selected_names]
        if not all(_sheet_has_site_identity(sheet) for sheet in selected):
            raise ValueError("Configured authoritative sheet_selector includes a sheet without a strict DU site identity column.")
        return selected

    if len(sheets) == 1:
        return sheets

    candidates = [sheet for sheet in sheets if _sheet_has_site_identity(sheet)]
    if len(candidates) != 1:
        candidate_names = [str(sheet.get("sheet_name", "")) for sheet in candidates]
        raise ValueError(
            "Unable to resolve exactly one authoritative DU sheet from a multi-sheet export; "
            f"strict site-identity candidates={candidate_names}."
        )
    return candidates


def _header_hash(
    inventory: Mapping[str, Any],
    *,
    structural: bool = False,
    reference_view_id: str | None = None,
    profile: Mapping[str, Any] | None = None,
) -> str:
    def selected_fingerprint(column: Mapping[str, Any]) -> Dict[str, str]:
        fingerprint = _canonical_fingerprint(column.get("fingerprint", {}))
        parsed = parse_site_identity_field_code(fingerprint["field_code"])
        if parsed is not None:
            if structural:
                fingerprint["field_code"] = (
                    f"site|fix00012|{parsed['du_model_id']}|{STRUCTURAL_VIEW_TOKEN}"
                )
            elif reference_view_id is not None:
                fingerprint["field_code"] = (
                    f"site|fix00012|{parsed['du_model_id']}|{reference_view_id}"
                )
        return fingerprint

    authoritative_sheets = resolve_authoritative_sheets(inventory, profile)
    hash_payload = {
        "schema_version": inventory.get("schema_version", SCHEMA_VERSION),
        "header_row_count": HEADER_ROW_COUNT,
        "sheets": [
            {
                "sheet_name": sheet["sheet_name"],
                "columns": [
                    selected_fingerprint(column)
                    for column in sheet.get("columns", [])
                ],
            }
            for sheet in authoritative_sheets
        ],
    }
    canonical = json.dumps(
        hash_payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def calculate_header_hash(
    inventory: Mapping[str, Any],
    profile: Mapping[str, Any] | None = None,
) -> str:
    """Hash exact normalized headers from authoritative DU sheet(s) only."""
    return _header_hash(inventory, profile=profile)


def calculate_structural_header_hash(
    inventory: Mapping[str, Any],
    profile: Mapping[str, Any] | None = None,
) -> str:
    """Hash authoritative headers while normalizing only the site identity View suffix."""
    return _header_hash(inventory, structural=True, profile=profile)


def _approved_site_layout_view_ids(profile: Mapping[str, Any]) -> set[str]:
    """Return layout references from APPROVED site-code mapping evidence."""
    view_ids: set[str] = set()
    site_mapping = profile.get("field_mapping", {}).get("site_code", {})
    for candidate in site_mapping.get("source_candidates", []):
        if str(candidate.get("mapping_status", "")) != "APPROVED":
            continue
        parsed = parse_site_identity_field_code(
            candidate.get("fingerprint", {}).get("field_code", "")
        )
        if parsed is not None:
            view_ids.add(parsed["view_id"])
    return view_ids


def resolve_approved_header_structure(
    inventory: Mapping[str, Any],
    profile: Mapping[str, Any],
) -> Dict[str, str | bool]:
    """Validate exact or View-normalized structure against approved raw hashes.

    Existing approved hashes remain authoritative. For an unseen runtime View,
    the authoritative inventory is rebound only at the site identity View suffix
    represented by an APPROVED site-code mapping fingerprint. Identity
    accepted_view_ids is never consulted. A match proves the selected DU layout
    is otherwise identical; auxiliary workbook sheets remain audit evidence only.
    """
    raw_hash = calculate_header_hash(inventory, profile)
    structural_hash = calculate_structural_header_hash(inventory, profile)
    approved_hashes = {
        str(value)
        for value in profile.get("export_structure", {}).get("approved_header_hashes", [])
    }
    if raw_hash in approved_hashes:
        return {
            "approved": True,
            "raw_header_hash": raw_hash,
            "structural_header_hash": structural_hash,
            "approved_header_hash": raw_hash,
            "approval_basis": "RAW_HEADER_HASH",
        }

    layout_view_ids = _approved_site_layout_view_ids(profile)
    for view_id in sorted(layout_view_ids):
        candidate_hash = _header_hash(
            inventory,
            reference_view_id=view_id,
            profile=profile,
        )
        if candidate_hash in approved_hashes:
            return {
                "approved": True,
                "raw_header_hash": raw_hash,
                "structural_header_hash": structural_hash,
                "approved_header_hash": candidate_hash,
                "approval_basis": "VIEW_NORMALIZED_TO_APPROVED_LAYOUT",
            }

    return {
        "approved": False,
        "raw_header_hash": raw_hash,
        "structural_header_hash": structural_hash,
        "approved_header_hash": "",
        "approval_basis": "UNAPPROVED_STRUCTURE",
    }


def _header_search_text(column: Mapping[str, Any]) -> str:
    return " ".join(column.get("normalized_header_values", [])).lower()


def _candidate_matches(text: str, keyword_sets: Iterable[Sequence[str]]) -> List[List[str]]:
    return [list(keywords) for keywords in keyword_sets if all(keyword in text for keyword in keywords)]


def identify_canonical_field_candidates(inventory: Mapping[str, Any]) -> Dict[str, Any]:
    fields: Dict[str, Any] = {}
    for canonical_field, keyword_sets in CANONICAL_FIELD_KEYWORDS.items():
        candidates = []
        for sheet in inventory.get("sheets", []):
            for column in sheet.get("columns", []):
                matches = _candidate_matches(_header_search_text(column), keyword_sets)
                if matches:
                    candidates.append(
                        {
                            "sheet_name": sheet["sheet_name"],
                            "fingerprint": column["fingerprint"],
                            "fingerprint_key": column["fingerprint_key"],
                            "evidence": {
                                "matched_keyword_sets": matches,
                                "normalized_header_values": column["normalized_header_values"],
                            },
                            "mapping_status": "UNVERIFIED",
                        }
                    )
        if not candidates:
            status = "MISSING"
        elif len(candidates) == 1:
            status = "UNVERIFIED"
        else:
            status = "AMBIGUOUS"
        fields[canonical_field] = {"status": status, "candidates": candidates}
    return {
        "schema_version": SCHEMA_VERSION,
        "discovery_only": True,
        "fields": fields,
    }


def _slug(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return cleaned or "unidentified_du_model"


def build_draft_du_profile(
    inventory: Mapping[str, Any],
    candidates: Mapping[str, Any],
    *,
    project_key: str | None = None,
    du_model_name: str | None = None,
    du_model_id: str | None = None,
    view_id: str | None = None,
) -> Dict[str, Any]:
    """Create a non-runnable DRAFT profile. All suggestions remain UNVERIFIED."""
    field_mapping: Dict[str, Any] = {}
    candidate_fields = candidates.get("fields", {})
    for canonical_field in CANONICAL_FIELD_KEYWORDS:
        suggested = candidate_fields.get(canonical_field, {})
        field_mapping[canonical_field] = {
            "required": canonical_field in PR_CRITICAL_FIELDS,
            "source_candidates": [
                {
                    "fingerprint": candidate["fingerprint"],
                    "mapping_status": "UNVERIFIED",
                }
                for candidate in suggested.get("candidates", [])
            ],
            "transforms": ["trim"],
        }

    profile_name = du_model_name or "unidentified_du_model"
    return {
        "profile_id": f"draft_{_slug(profile_name)}_v1",
        "profile_version": "0.1.0",
        "mapping_version": "discovery-0.1.0",
        "status": "DRAFT",
        "identity": {
            "project_key": project_key or "UNVERIFIED",
            "accepted_du_models": [du_model_name] if du_model_name else [],
            "accepted_du_model_ids": [str(du_model_id)] if du_model_id else [],
            "accepted_view_ids": [str(view_id)] if view_id else [],
        },
        "export_structure": {
            "sheet_selector": None,
            "header_rows": [0, 1, 2, 3],
            "header_hash_policy": "strict",
            "observed_header_hash": calculate_header_hash(inventory),
            "observed_structural_header_hash": calculate_structural_header_hash(inventory),
            "approved_header_hashes": [],
        },
        "field_mapping": field_mapping,
        "validation": {
            "reject_unknown_headers": True,
            "reject_ambiguous_source_mapping": True,
            "require_evidence_for_every_canonical_field": True,
            "profile_notes": [
                "Generated by read-only profiler.",
                "No source candidate is approved.",
                "DRAFT profiles must never be used for production ECC generation.",
            ],
        },
    }


def header_inventory_markdown(inventory: Mapping[str, Any], header_hash: str) -> str:
    lines = [
        "# iEPMS Four-Header Inventory",
        "",
        f"- Source file: `{inventory['source']['file_name']}`",
        f"- Source file SHA-256: `{inventory['source']['source_file_hash']}`",
        f"- Header hash: `{header_hash}`",
        f"- Structural header hash: `{calculate_structural_header_hash(inventory)}`",
        f"- Header rows preserved: `{HEADER_ROW_COUNT}`",
        "",
        "This is a read-only profiler output. Column position is informational only; production profiles must match by four-layer fingerprint.",
    ]
    for sheet in inventory.get("sheets", []):
        lines.extend(["", f"## Sheet: {sheet['sheet_name']}", "", "| Position | Field ID / Code | WBS Stage | Task Name | Display Header |", "|---|---|---|---|---|"])
        for column in sheet.get("columns", []):
            fp = column["fingerprint"]
            lines.append(
                "| {position} | {field} | {wbs} | {task} | {display} |".format(
                    position=column["source_position"]["excel_column"],
                    field=fp["field_code"].replace("|", "\\|"),
                    wbs=fp["wbs_stage"].replace("|", "\\|"),
                    task=fp["task_name"].replace("|", "\\|"),
                    display=fp["display_header"].replace("|", "\\|"),
                )
            )
    return "\n".join(lines) + "\n"


def missing_pr_critical_fields_markdown(candidates: Mapping[str, Any]) -> str:
    lines = [
        "# PR-Critical Source Field Assessment",
        "",
        "All profiler matches are UNVERIFIED. This file is not a production approval.",
        "",
        "| Canonical field | Discovery status | Required handling |",
        "|---|---|---|",
    ]
    for field in PR_CRITICAL_FIELDS:
        status = candidates["fields"][field]["status"]
        if status == "UNVERIFIED":
            action = "Business validation required before profile promotion."
        elif status == "AMBIGUOUS":
            action = "Quarantine: select one approved four-layer fingerprint."
        else:
            action = "Incomplete: identify a source field before PR generation."
        lines.append(f"| `{field}` | `{status}` | {action} |")

    lines.extend(
        [
            "",
            "## Conditional PR fields",
            "",
            "Coordinates and antenna fields become mandatory only when the selected shared PR rule requires them. Their mappings still require profile evidence before use.",
        ]
    )
    for field in ("latitude", "longitude", "antenna_size_ne", "antenna_size_fe", "tx_upgrade_scope_raw"):
        status = candidates["fields"][field]["status"]
        lines.append(f"- `{field}`: `{status}`")
    return "\n".join(lines) + "\n"


def write_profiler_outputs(
    output_dir: Path,
    inventory: Mapping[str, Any],
    candidates: Mapping[str, Any],
    profile: Mapping[str, Any],
) -> Dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    header_hash = calculate_header_hash(inventory)
    outputs = {
        "header_inventory_json": output_dir / "header_inventory.json",
        "header_inventory_markdown": output_dir / "header_inventory.md",
        "header_hash": output_dir / "header_hash.txt",
        "canonical_field_candidates": output_dir / "canonical_field_candidates.json",
        "missing_pr_critical_fields": output_dir / "missing_pr_critical_fields.md",
        "draft_du_profile": output_dir / "draft_du_profile.yaml",
    }
    outputs["header_inventory_json"].write_text(json.dumps(inventory, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    outputs["header_inventory_markdown"].write_text(header_inventory_markdown(inventory, header_hash), encoding="utf-8")
    outputs["header_hash"].write_text(header_hash + "\n", encoding="utf-8")
    outputs["canonical_field_candidates"].write_text(json.dumps(candidates, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    outputs["missing_pr_critical_fields"].write_text(missing_pr_critical_fields_markdown(candidates), encoding="utf-8")
    outputs["draft_du_profile"].write_text(json.dumps(profile, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return outputs


def profile_export(
    input_path: Path,
    output_dir: Path,
    *,
    project_key: str | None = None,
    du_model_name: str | None = None,
    du_model_id: str | None = None,
    view_id: str | None = None,
) -> Dict[str, Any]:
    inventory = build_header_inventory(Path(input_path))
    candidates = identify_canonical_field_candidates(inventory)
    profile = build_draft_du_profile(
        inventory,
        candidates,
        project_key=project_key,
        du_model_name=du_model_name,
        du_model_id=du_model_id,
        view_id=view_id,
    )
    outputs = write_profiler_outputs(Path(output_dir), inventory, candidates, profile)
    return {"inventory": inventory, "candidates": candidates, "profile": profile, "outputs": outputs}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Read-only profiler for original four-header iEPMS exports.")
    parser.add_argument("--input", required=True, help="Original iEPMS .xlsx/.xlsm/.csv export")
    parser.add_argument("--output", required=True, help="Directory for profiler artifacts")
    parser.add_argument("--project-key")
    parser.add_argument("--du-model-name")
    parser.add_argument("--du-model-id")
    parser.add_argument("--view-id")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = profile_export(
        Path(args.input),
        Path(args.output),
        project_key=args.project_key,
        du_model_name=args.du_model_name,
        du_model_id=args.du_model_id,
        view_id=args.view_id,
    )
    print("Profiler completed. No ECC generation was attempted.")
    print(f"Header hash: {calculate_header_hash(result['inventory'])}")
    print(f"Structural header hash: {calculate_structural_header_hash(result['inventory'])}")
    for name, output in result["outputs"].items():
        print(f"{name}: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())