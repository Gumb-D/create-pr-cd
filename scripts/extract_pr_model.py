import pandas as pd
import openpyxl
from openpyxl import load_workbook

print("=" * 100)
print("EXTRACTING PR MODEL DATA")
print("=" * 100)

pr_file = 'Info/input/pr_model.xlsx'

# Load workbook
wb = load_workbook(pr_file, data_only=True)
print(f"\n✓ PR Model sheets: {wb.sheetnames}")

# Extract TX Line Item sheet
tx_sheet = wb["TX Line Item (After 21-Apr 26)"]
print(f"\n[TX LINE ITEM SHEET]")
print(f"Max row: {tx_sheet.max_row}, Max col: {tx_sheet.max_column}")

# Read raw data to find structure
print("\n\nFirst 30 rows (searching for headers and data):\n")

for row_idx in range(1, min(31, tx_sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(15, tx_sheet.max_column + 1)):
        cell = tx_sheet.cell(row_idx, col_idx)
        row_data.append(str(cell.value)[:40] if cell.value else "")
    
    non_empty = any(row_data)
    if non_empty:
        print(f"Row {row_idx}: {' | '.join(row_data)}")

# Try to read as pandas
print("\n\n" + "=" * 100)
print("READING AS DATAFRAME")
print("=" * 100)

try:
    df_tx = pd.read_excel(pr_file, sheet_name="TX Line Item (After 21-Apr 26)", header=None)
    print(f"\nDataFrame shape: {df_tx.shape}")
    print(f"\nFirst 20 rows, first 12 columns:")
    print(df_tx.iloc[:20, :12].to_string())
    
    # Find rows that contain TSS, TI, Planning
    print("\n\n[SEARCHING FOR SCOPE MARKERS]")
    for idx in range(len(df_tx)):
        row_str = ' '.join([str(val)[:50] if val else '' for val in df_tx.iloc[idx, :5]])
        if any(keyword in row_str.upper() for keyword in ['TSS', 'TECHNICAL SITE', 'MW NEW', 'MW SWAP', 'MW DISMANTLE', 'PBOM']):
            print(f"Row {idx}: {row_str[:100]}")
            
except Exception as e:
    print(f"Error: {e}")

# Also check PBOM sheet
print("\n\n" + "=" * 100)
print("PBOM SHEET ANALYSIS")
print("=" * 100)

pbom_sheet = wb["PBOM (After 21-Apr 26)"]
print(f"Max row: {pbom_sheet.max_row}, Max col: {pbom_sheet.max_column}")

try:
    df_pbom = pd.read_excel(pr_file, sheet_name="PBOM (After 21-Apr 26)", header=None)
    print(f"\nDataFrame shape: {df_pbom.shape}")
    print(f"\nFirst 30 rows, first 10 columns:")
    print(df_pbom.iloc[:30, :10].to_string())
except Exception as e:
    print(f"Error: {e}")

print("\n" + "=" * 100)
