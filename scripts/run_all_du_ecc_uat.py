#!/usr/bin/env python3
"""Generate consolidated NON_PRODUCTION_UAT ECC verification packs."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import create_pr
from du_profile_resolver import DuProfileResolutionError, resolve_du_profile
from pr_safety_controls import DISALLOWED_CONTRACT_VALUES, SafetyControlError, normalize_subcontractor


ROOT = Path(__file__).resolve().parent.parent
PROFILE_ROOT = ROOT / "config" / "du_profiles"
IDENTITY_REGISTRY = ROOT / "config" / "registries" / "mw_du_profile_identity_registry.yaml"
DEFAULT_CONTRACT_REFERENCE = ROOT / "Info" / "input" / "contract_info_reference.md"
DEFAULT_POLICY = ROOT / "config" / "subcontractor_pr_policy.json"
DEFAULT_PR_MODEL = ROOT / "Info" / "input" / "pr_model.xlsx"
DEFAULT_TEMPLATE = ROOT / "Info" / "input" / "ecc_template.xls"
UAT_MARKER = "NON_PRODUCTION_UAT"
ELIGIBLE_STATUSES = {"PR_INPUT_READY", "PRODUCTION"}
VERIFICATION_STATUSES = ("PENDING", "PASS", "FAIL", "WAIVED_WITH_ACCEPTED_RISK")

MANIFEST_COLUMNS = (
    "Run ID",
    "Pack Type",
    "Project Key",
    "DU Model",
    "DU Model ID",
    "View ID",
    "Profile ID",
    "Profile Version",
    "Mapping Version",
    "Profile Status",
    "Header Hash",
    "Source File Name",
    "Source File SHA256",
    "Scope",
    "Region",
    "Subcontractor",
    "Contract Number",
    "Site Count",
    "ECC Row Count",
    "Tx SOW Distribution",
    "SM Count",
    "Missing Contract Count",
    "Review Required Count",
    "Duplicate Count",
    "Ignored Count",
    "Output File",
    "Verification Status",
    "Reviewer Comment",
)

BLOCKED_COLUMNS = (
    "Profile ID",
    "Profile Status",
    "Reason Code",
    "Reason",
    "Source Export",
)


class BatchUatError(RuntimeError):
    def __init__(self, code: str, message: str, details: Mapping[str, Any] | None = None):
        super().__init__(message)
        self.code = code
        self.details = dict(details or {})


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _new_run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_json(path: Path, code: str, label: str) -> Any:
    candidate = Path(path)
    if not candidate.is_file():
        raise BatchUatError(code, f"{label} is missing: {candidate}", {"path": str(candidate)})
    try:
        return json.loads(candidate.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as error:
        raise BatchUatError(code, f"{label} is invalid: {candidate}", {"path": str(candidate), "error": str(error)}) from error


def load_input_manifest(path: Path) -> dict[str, dict[str, Any]]:
    payload = _load_json(path, "UAT_INPUT_MANIFEST_INVALID", "UAT input manifest")
    if not isinstance(payload, dict) or str(payload.get("schema_version", "")).strip() != "1.0":
        raise BatchUatError(
            "UAT_INPUT_MANIFEST_INVALID",
            "UAT input manifest must be an object with schema_version 1.0.",
            {"path": str(path)},
        )
    profiles = payload.get("profiles")
    if not isinstance(profiles, list):
        raise BatchUatError(
            "UAT_INPUT_MANIFEST_INVALID",
            "UAT input manifest profiles must be a list.",
            {"path": str(path)},
        )
    result: dict[str, dict[str, Any]] = {}
    for index, raw_entry in enumerate(profiles):
        if not isinstance(raw_entry, dict):
            raise BatchUatError(
                "UAT_INPUT_MANIFEST_INVALID",
                "Every UAT manifest profile entry must be an object.",
                {"path": str(path), "index": index},
            )
        profile_id = str(raw_entry.get("profile_id", "")).strip()
        if not profile_id or profile_id in result:
            raise BatchUatError(
                "UAT_INPUT_MANIFEST_INVALID",
                "UAT manifest profile_id values must be non-blank and unique.",
                {"path": str(path), "index": index, "profile_id": profile_id},
            )
        result[profile_id] = dict(raw_entry)
    return result


def load_structured_profiles(profile_root: Path = PROFILE_ROOT) -> dict[str, dict[str, Any]]:
    profiles: dict[str, dict[str, Any]] = {}
    for path in sorted(Path(profile_root).glob("*.yaml")):
        payload = _load_json(path, "DU_PROFILE_CONFIG_INVALID", "DU Profile")
        if not isinstance(payload, dict):
            raise BatchUatError("DU_PROFILE_CONFIG_INVALID", "DU Profile must be a JSON object.", {"path": str(path)})
        profile_id = str(payload.get("profile_id", "")).strip()
        if not profile_id or profile_id in profiles:
            raise BatchUatError(
                "DU_PROFILE_CONFIG_INVALID",
                "DU Profile IDs must be non-blank and unique.",
                {"path": str(path), "profile_id": profile_id},
            )
        profiles[profile_id] = payload
    if not profiles:
        raise BatchUatError("DU_PROFILE_CONFIG_INVALID", "No DU Profiles were found.", {"profile_root": str(profile_root)})
    return profiles


def resolve_source_export(raw_path: Any, manifest_path: Path) -> Path | None:
    value = os.path.expandvars(str(raw_path or "").strip())
    if not value:
        return None
    candidate = Path(value).expanduser()
    if not candidate.is_absolute():
        candidate = manifest_path.resolve().parent / candidate
    return candidate.resolve()


def deterministic_review_site_codes(
    ecc_files: Iterable[Path],
    max_combinations: int,
) -> list[str]:
    if max_combinations < 1:
        raise BatchUatError("INVALID_REVIEW_SAMPLE_CAP", "Review sample cap must be at least 1.")
    combinations: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    for path in sorted(Path(file) for file in ecc_files):
        for row in read_ecc_rows(path):
            combination = (
                str(row["source_tx_sow"] or "").strip(),
                str(row["subcontractor"] or "").strip(),
                str(row["region"] or "").strip(),
            )
            site_code = str(row["site_code"] or "").strip().upper()
            if site_code:
                combinations[combination].add(site_code)
    ordered = sorted(combinations)
    if len(ordered) <= max_combinations:
        chosen = ordered
    else:
        by_category: dict[tuple[str, str], list[tuple[str, str, str]]] = defaultdict(list)
        for combination in ordered:
            by_category[(combination[0], combination[1])].append(combination)
        mandatory = [sorted(values)[0] for _, values in sorted(by_category.items())]
        if len(mandatory) > max_combinations:
            raise BatchUatError(
                "REVIEW_SAMPLE_CAP_TOO_LOW",
                "Review sample cap cannot preserve every Tx SOW and subcontractor category.",
                {"category_count": len(mandatory), "max_combinations": max_combinations},
            )
        chosen = list(mandatory)
        chosen_set = set(chosen)
        for combination in ordered:
            if len(chosen) >= max_combinations:
                break
            if combination not in chosen_set:
                chosen.append(combination)
                chosen_set.add(combination)
        chosen.sort()

    site_codes: list[str] = []
    seen: set[str] = set()
    for combination in chosen:
        site_code = sorted(combinations[combination])[0]
        if site_code not in seen:
            seen.add(site_code)
            site_codes.append(site_code)
    return site_codes


def read_ecc_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if "details" not in workbook.sheetnames:
            raise BatchUatError(
                "ECC_DETAILS_SHEET_MISSING",
                "Generated ECC workbook does not contain the details sheet.",
                {"path": str(path)},
            )
        worksheet = workbook["details"]
        rows: list[dict[str, Any]] = []
        for row_number in range(2, worksheet.max_row + 1):
            site_code = worksheet.cell(row_number, 4).value
            if site_code in (None, ""):
                continue
            rows.append(
                {
                    "site_code": site_code,
                    "region": worksheet.cell(row_number, 3).value,
                    "contract_number": worksheet.cell(row_number, 8).value,
                    "subcontractor": worksheet.cell(row_number, 9).value,
                    "sow": worksheet.cell(row_number, 11).value,
                    "source_tx_sow": worksheet.cell(row_number, 15).value,
                    "repeated_contract_number": worksheet.cell(row_number, 16).value,
                }
            )
        return rows
    finally:
        workbook.close()


def _artifact_name(path: Path, pack_type: str, run_id: str) -> str:
    stem = path.stem
    if pack_type not in stem:
        stem = f"{stem}_{pack_type}"
    if UAT_MARKER not in stem:
        stem = f"{stem}_{UAT_MARKER}"
    if run_id not in stem:
        stem = f"{stem}_{run_id}"
    return f"{stem}{path.suffix}"


def materialize_scope_artifacts(
    engine_root: Path,
    scope_dir: Path,
    summary: dict[str, Any],
    pack_type: str,
    batch_run_id: str,
) -> tuple[dict[str, Any], list[Path]]:
    scope_dir.mkdir(parents=True, exist_ok=True)
    old_summary = Path(summary.get("summary_path", "")).resolve() if summary.get("summary_path") else None
    path_map: dict[str, str] = {}
    generated_xlsx: list[Path] = []
    for source in sorted(path for path in engine_root.rglob("*") if path.is_file()):
        source_resolved = source.resolve()
        if old_summary and source_resolved == old_summary:
            continue
        target = scope_dir / _artifact_name(source, pack_type, batch_run_id)
        if target.exists():
            raise BatchUatError(
                "UAT_ARTIFACT_COLLISION",
                "Batch UAT artefact target already exists.",
                {"source": str(source), "target": str(target)},
            )
        shutil.move(str(source), str(target))
        path_map[str(source_resolved)] = str(target.resolve())
        if target.suffix.lower() == ".xlsx":
            generated_xlsx.append(target.resolve())
    if engine_root.exists():
        shutil.rmtree(engine_root)

    adjusted = dict(summary)
    adjusted["pack_type"] = pack_type
    adjusted["batch_run_id"] = batch_run_id
    adjusted["output_root"] = str(scope_dir.resolve())
    adjusted["created_files"] = [
        path_map.get(str(Path(path).resolve()), str(Path(path).resolve()))
        for path in summary.get("created_files", [])
        if str(Path(path).resolve()) in path_map
    ]
    for key in ("review_report", "contract_mapping_review_report"):
        old_value = summary.get(key)
        adjusted[key] = path_map.get(str(Path(old_value).resolve())) if old_value else None
    summary_path = scope_dir / f"CREATE_PR_SUMMARY_{adjusted['scope']}_{pack_type}_{UAT_MARKER}_{batch_run_id}.json"
    adjusted["summary_path"] = str(summary_path.resolve())
    summary_path.write_text(json.dumps(adjusted, ensure_ascii=False, indent=2), encoding="utf-8")
    return adjusted, sorted(generated_xlsx)


def _create_pr_namespace(
    source_export: Path,
    output: Path,
    scope: str,
    site_codes: list[str] | None,
    args: argparse.Namespace,
    child_run_id: str,
) -> argparse.Namespace:
    return argparse.Namespace(
        site_data=source_export,
        output=output,
        scope=scope,
        site_code=",".join(site_codes) if site_codes else None,
        all_sites=not bool(site_codes),
        pr_model=args.pr_model,
        template=args.template,
        mapping=args.mapping,
        subcontractor_policy=args.subcontractor_policy,
        non_production_uat=True,
        uat_run_id=child_run_id,
    )


def run_scope_pack(
    source_export: Path,
    scope_dir: Path,
    scope: str,
    pack_type: str,
    site_codes: list[str] | None,
    profile_id: str,
    batch_run_id: str,
    args: argparse.Namespace,
) -> tuple[dict[str, Any], list[Path]]:
    engine_root = scope_dir / "_ENGINE"
    child_run_id = f"{batch_run_id}_{profile_id}_{pack_type}_{scope}"
    parsed = _create_pr_namespace(source_export, engine_root, scope, site_codes, args, child_run_id)
    summary = create_pr.run(parsed)
    return materialize_scope_artifacts(engine_root, scope_dir, summary, pack_type, batch_run_id)


def write_empty_review_summary(
    scope_dir: Path,
    full_summary: Mapping[str, Any],
    batch_run_id: str,
) -> dict[str, Any]:
    scope_dir.mkdir(parents=True, exist_ok=True)
    summary = {
        key: value
        for key, value in full_summary.items()
        if key not in {"created_files", "summary_path", "review_report", "contract_mapping_review_report", "output_root"}
    }
    summary.update(
        {
            "status": "NO_ELIGIBLE_CANDIDATES",
            "pack_type": "REVIEW_PACK",
            "batch_run_id": batch_run_id,
            "candidate_count": 0,
            "created_files": [],
            "review_report": None,
            "contract_mapping_review_report": None,
            "output_root": str(scope_dir.resolve()),
        }
    )
    path = scope_dir / f"CREATE_PR_SUMMARY_{summary['scope']}_REVIEW_PACK_{UAT_MARKER}_{batch_run_id}.json"
    summary["summary_path"] = str(path.resolve())
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def _manifest_row(
    path: Path,
    pack_type: str,
    scope_summary: Mapping[str, Any],
    source_name: str,
    source_sha: str,
    run_id: str,
) -> dict[str, Any]:
    rows = read_ecc_rows(path)
    regions = sorted({str(row["region"] or "").strip() for row in rows if str(row["region"] or "").strip()})
    subcontractors = sorted(
        {str(row["subcontractor"] or "").strip() for row in rows if str(row["subcontractor"] or "").strip()},
        key=str.casefold,
    )
    contracts = sorted(
        {str(row["contract_number"] or "").strip() for row in rows if str(row["contract_number"] or "").strip()}
    )
    site_count = len({str(row["site_code"] or "").strip().upper() for row in rows})
    sow_distribution = Counter(str(row["source_tx_sow"] or "").strip() for row in rows)
    sm_count = sum(1 for row in rows if normalize_subcontractor(row["subcontractor"]) == "SM")
    missing_contract_count = sum(
        1
        for row in rows
        if str(row["contract_number"] or "").strip().upper() in DISALLOWED_CONTRACT_VALUES
        or str(row["repeated_contract_number"] or "").strip().upper() in DISALLOWED_CONTRACT_VALUES
    )
    return {
        "Run ID": run_id,
        "Pack Type": pack_type,
        "Project Key": scope_summary.get("project_key", ""),
        "DU Model": scope_summary.get("du_model_name", ""),
        "DU Model ID": scope_summary.get("du_model_id", ""),
        "View ID": scope_summary.get("view_id", ""),
        "Profile ID": scope_summary.get("profile_id", ""),
        "Profile Version": scope_summary.get("profile_version", ""),
        "Mapping Version": scope_summary.get("mapping_version", ""),
        "Profile Status": scope_summary.get("profile_status", ""),
        "Header Hash": scope_summary.get("header_hash", ""),
        "Source File Name": source_name,
        "Source File SHA256": source_sha,
        "Scope": scope_summary.get("scope", ""),
        "Region": " | ".join(regions),
        "Subcontractor": " | ".join(subcontractors),
        "Contract Number": " | ".join(contracts),
        "Site Count": site_count,
        "ECC Row Count": len(rows),
        "Tx SOW Distribution": json.dumps(dict(sorted(sow_distribution.items())), ensure_ascii=False),
        "SM Count": sm_count,
        "Missing Contract Count": missing_contract_count,
        "Review Required Count": scope_summary.get("review_required_count", 0),
        "Duplicate Count": scope_summary.get("duplicate_count", 0),
        "Ignored Count": scope_summary.get("ignored_count", 0),
        "Output File": str(path.resolve()),
        "Verification Status": "PENDING",
        "Reviewer Comment": "",
    }


def write_master_manifest(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ECC Files"
    worksheet.append(list(MANIFEST_COLUMNS))
    for row in rows:
        worksheet.append([row.get(column, "") for column in MANIFEST_COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 38

    widths = {
        "A": 28, "B": 15, "C": 24, "D": 24, "E": 20, "F": 20, "G": 34, "H": 15,
        "I": 32, "J": 18, "K": 68, "L": 32, "M": 68, "N": 10, "O": 18, "P": 24,
        "Q": 28, "R": 12, "S": 14, "T": 44, "U": 10, "V": 18, "W": 18, "X": 14,
        "Y": 12, "Z": 80, "AA": 28, "AB": 48,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for row_number in range(2, worksheet.max_row + 1):
        for cell in worksheet[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    status_column = MANIFEST_COLUMNS.index("Verification Status") + 1
    if worksheet.max_row >= 2:
        validation = DataValidation(type="list", formula1='"' + ",".join(VERIFICATION_STATUSES) + '"', allow_blank=False)
        worksheet.add_data_validation(validation)
        validation.add(f"{worksheet.cell(2, status_column).column_letter}2:{worksheet.cell(worksheet.max_row, status_column).column_letter}{worksheet.max_row}")
        pending_fill = PatternFill("solid", fgColor="FFF2CC")
        worksheet.conditional_formatting.add(
            f"{worksheet.cell(2, status_column).column_letter}2:{worksheet.cell(worksheet.max_row, status_column).column_letter}{worksheet.max_row}",
            FormulaRule(formula=[f'{worksheet.cell(2, status_column).column_letter}2="PENDING"'], fill=pending_fill),
        )

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["NON_PRODUCTION_UAT Business Verification"])
    instructions.append(["Verification Status", "Meaning"])
    for status, meaning in (
        ("PENDING", "Not yet reviewed by the business owner."),
        ("PASS", "Output verified and accepted for the stated UAT evidence only."),
        ("FAIL", "Output contains a defect or unresolved business issue."),
        ("WAIVED_WITH_ACCEPTED_RISK", "Business owner accepted a documented risk; this is not automatic production promotion."),
    ):
        instructions.append([status, meaning])
    instructions["A1"].font = Font(bold=True, size=14)
    instructions.column_dimensions["A"].width = 34
    instructions.column_dimensions["B"].width = 100
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(path)


def write_blocked_profiles(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=BLOCKED_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in BLOCKED_COLUMNS})


def _blocked_row(
    profile_id: str,
    profile_status: str,
    reason_code: str,
    reason: str,
    source_export: Path | None,
) -> dict[str, Any]:
    return {
        "Profile ID": profile_id,
        "Profile Status": profile_status,
        "Reason Code": reason_code,
        "Reason": reason,
        "Source Export": str(source_export) if source_export else "",
    }


def run_batch(args: argparse.Namespace) -> dict[str, Any]:
    started_at = _utc_now()
    run_id = str(args.run_id or _new_run_id()).strip()
    if not run_id:
        raise BatchUatError("INVALID_RUN_ID", "Batch UAT run ID must not be blank.")
    output_root = Path(args.output).resolve() / UAT_MARKER / run_id
    if output_root.exists():
        raise BatchUatError(
            "UAT_RUN_ALREADY_EXISTS",
            "Batch UAT run directory already exists; prior evidence will not be overwritten.",
            {"output_root": str(output_root)},
        )
    output_root.mkdir(parents=True, exist_ok=False)

    manifest_path = Path(args.manifest).resolve()
    manifest_entries = load_input_manifest(manifest_path)
    profiles = load_structured_profiles()
    blocked: list[dict[str, Any]] = []
    profile_results: list[dict[str, Any]] = []
    manifest_rows: list[dict[str, Any]] = []
    generated_files: list[Path] = []
    full_summaries: list[dict[str, Any]] = []
    eligible_profile_count = 0
    successful_scope_runs = 0
    failed_scope_runs = 0

    for unknown_profile_id in sorted(set(manifest_entries) - set(profiles)):
        source = resolve_source_export(manifest_entries[unknown_profile_id].get("source_export"), manifest_path)
        blocked.append(
            _blocked_row(
                unknown_profile_id,
                "UNKNOWN",
                "UNKNOWN_PROFILE",
                "The manifest references a DU Profile that is not registered.",
                source,
            )
        )

    for profile_id, profile in sorted(profiles.items()):
        profile_status = str(profile.get("status", "")).strip().upper()
        entry = manifest_entries.get(profile_id)
        source_export = resolve_source_export((entry or {}).get("source_export"), manifest_path)
        if profile_status not in ELIGIBLE_STATUSES:
            blocked.append(
                _blocked_row(
                    profile_id,
                    profile_status,
                    "PROFILE_STATUS_BLOCKED",
                    f"DU Profile status {profile_status or '(blank)'} is not eligible for NON_PRODUCTION_UAT ECC generation.",
                    source_export,
                )
            )
            continue
        if entry is None:
            blocked.append(
                _blocked_row(
                    profile_id,
                    profile_status,
                    "MISSING_MANIFEST_ENTRY",
                    "Eligible DU Profile has no explicit source-export entry in the input manifest.",
                    None,
                )
            )
            continue
        if entry.get("enabled", True) is False:
            blocked.append(
                _blocked_row(
                    profile_id,
                    profile_status,
                    "MANIFEST_ENTRY_DISABLED",
                    "The manifest explicitly disabled this DU Profile.",
                    source_export,
                )
            )
            continue
        if source_export is None or not source_export.is_file():
            blocked.append(
                _blocked_row(
                    profile_id,
                    profile_status,
                    "MISSING_SOURCE_EXPORT",
                    "The current raw iEPMS source export is not available at the explicit manifest path.",
                    source_export,
                )
            )
            continue

        source_sha = _sha256(source_export)
        expected_sha = str(entry.get("expected_source_sha256", "")).strip().lower()
        if expected_sha and expected_sha != source_sha:
            blocked.append(
                _blocked_row(
                    profile_id,
                    profile_status,
                    "SOURCE_SHA256_MISMATCH",
                    "The source export SHA256 does not match the manifest-approved value.",
                    source_export,
                )
            )
            continue

        try:
            resolution = resolve_du_profile(
                source_export,
                profile_root=PROFILE_ROOT,
                identity_registry_path=IDENTITY_REGISTRY,
            )
        except DuProfileResolutionError as error:
            blocked.append(_blocked_row(profile_id, profile_status, error.code, str(error), source_export))
            continue
        resolved_profile_id = str(resolution.get("profile", {}).get("profile_id", "")).strip()
        if resolved_profile_id != profile_id:
            blocked.append(
                _blocked_row(
                    profile_id,
                    profile_status,
                    "SOURCE_PROFILE_IDENTITY_MISMATCH",
                    f"Source export resolved to {resolved_profile_id or '(unknown)'} instead of {profile_id}.",
                    source_export,
                )
            )
            continue

        eligible_profile_count += 1
        profile_result: dict[str, Any] = {
            "profile_id": profile_id,
            "profile_status": profile_status,
            "source_export": str(source_export),
            "source_file_sha256": source_sha,
            "scopes": {},
        }
        for scope in ("TSS", "TI"):
            full_scope_dir = output_root / profile_id / "FULL_PACK" / scope
            review_scope_dir = output_root / profile_id / "REVIEW_PACK" / scope
            try:
                full_summary, full_files = run_scope_pack(
                    source_export,
                    full_scope_dir,
                    scope,
                    "FULL_PACK",
                    None,
                    profile_id,
                    run_id,
                    args,
                )
                full_summaries.append(full_summary)
                generated_files.extend(full_files)
                review_site_codes = deterministic_review_site_codes(full_files, args.review_max_combinations)
                if review_site_codes:
                    review_summary, review_files = run_scope_pack(
                        source_export,
                        review_scope_dir,
                        scope,
                        "REVIEW_PACK",
                        review_site_codes,
                        profile_id,
                        run_id,
                        args,
                    )
                else:
                    review_summary = write_empty_review_summary(review_scope_dir, full_summary, run_id)
                    review_files = []
                generated_files.extend(review_files)
                profile_result["scopes"][scope] = {
                    "status": "SUCCESS",
                    "review_site_codes": review_site_codes,
                    "full_pack": full_summary,
                    "review_pack": review_summary,
                }
                successful_scope_runs += 1

                for pack_type, summary, files in (
                    ("FULL_PACK", full_summary, full_files),
                    ("REVIEW_PACK", review_summary, review_files),
                ):
                    for path in files:
                        manifest_rows.append(
                            _manifest_row(path, pack_type, summary, source_export.name, source_sha, run_id)
                        )
            except (BatchUatError, SafetyControlError, create_pr.CreatePrError, DuProfileResolutionError, OSError) as error:
                error_code = getattr(error, "code", "SCOPE_RUN_FAILED")
                profile_result["scopes"][scope] = {
                    "status": "FAILED",
                    "code": error_code,
                    "message": str(error),
                    "details": getattr(error, "details", {}),
                }
                failed_scope_runs += 1
        profile_results.append(profile_result)

    manifest_rows.sort(
        key=lambda row: (
            row["Profile ID"],
            row["Pack Type"],
            row["Scope"],
            row["Region"],
            row["Subcontractor"],
            row["Output File"],
        )
    )
    blocked.sort(key=lambda row: (row["Profile ID"], row["Reason Code"]))

    manifest_path_out = output_root / "UAT_MASTER_MANIFEST.xlsx"
    blocked_path = output_root / "UAT_BLOCKED_PROFILES.csv"
    summary_path = output_root / "UAT_MASTER_SUMMARY.json"
    write_master_manifest(manifest_path_out, manifest_rows)
    write_blocked_profiles(blocked_path, blocked)

    generated_xlsx = sorted(path.resolve() for path in generated_files if path.suffix.lower() == ".xlsx")
    manifested_paths = sorted(Path(row["Output File"]).resolve() for row in manifest_rows)
    reconciliation_ok = generated_xlsx == manifested_paths
    unsafe_rows = [
        row for row in manifest_rows if row["SM Count"] or row["Missing Contract Count"]
    ]
    if unsafe_rows:
        reconciliation_ok = False

    summary = {
        "run_id": run_id,
        "run_mode": UAT_MARKER,
        "started_at": started_at,
        "completed_at": _utc_now(),
        "input_manifest": str(manifest_path),
        "output_root": str(output_root),
        "eligible_profile_count": eligible_profile_count,
        "blocked_profile_count": len(blocked),
        "successful_scope_runs": successful_scope_runs,
        "failed_scope_runs": failed_scope_runs,
        "generated_ecc_file_count": len(manifest_rows),
        "generated_ecc_row_count": sum(int(row["ECC Row Count"]) for row in manifest_rows),
        "candidate_count": sum(int(summary.get("candidate_count", 0)) for summary in full_summaries),
        "duplicate_count": sum(int(summary.get("duplicate_count", 0)) for summary in full_summaries),
        "ignored_count": sum(int(summary.get("ignored_count", 0)) for summary in full_summaries),
        "review_required_count": sum(int(summary.get("review_required_count", 0)) for summary in full_summaries),
        "contract_mapping_missing_count": sum(
            int(summary.get("contract_mapping_missing_count", 0)) for summary in full_summaries
        ),
        "sm_excluded_count": sum(int(summary.get("sm_excluded_count", 0)) for summary in full_summaries),
        "profile_results": profile_results,
        "blocked_profiles": blocked,
        "master_manifest": str(manifest_path_out),
        "blocked_profiles_report": str(blocked_path),
        "manifest_row_count": len(manifest_rows),
        "manifest_reconciliation_ok": reconciliation_ok,
        "unsafe_manifest_row_count": len(unsafe_rows),
        "business_verification_status": "PENDING",
    }
    summary["status"] = "SUCCESS" if failed_scope_runs == 0 and reconciliation_ok else "COMPLETED_WITH_BLOCKS"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate consolidated all-DU NON_PRODUCTION_UAT ECC packs.")
    parser.add_argument("--manifest", required=True, type=Path, help="Explicit profile-to-source-export JSON manifest")
    parser.add_argument("--output", type=Path, default=ROOT / "output", help="Root directory for generated UAT evidence")
    parser.add_argument("--review-max-combinations", type=int, default=500)
    parser.add_argument("--run-id", help="Optional deterministic run ID; existing runs are never overwritten")
    parser.add_argument("--pr-model", type=Path, default=DEFAULT_PR_MODEL)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--mapping", type=Path, default=DEFAULT_CONTRACT_REFERENCE)
    parser.add_argument("--subcontractor-policy", type=Path, default=DEFAULT_POLICY)
    return parser.parse_args()


def main() -> int:
    try:
        summary = run_batch(parse_args())
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0 if summary["status"] == "SUCCESS" else 2
    except (BatchUatError, SafetyControlError, create_pr.CreatePrError, DuProfileResolutionError) as error:
        payload = {
            "status": "ERROR",
            "code": getattr(error, "code", "ALL_DU_UAT_FAILED"),
            "message": str(error),
            "details": getattr(error, "details", {}),
        }
    except Exception as error:
        payload = {"status": "ERROR", "code": "ALL_DU_UAT_FAILED", "message": str(error)}
    print(json.dumps(payload, ensure_ascii=False, indent=2), file=sys.stderr)
    return 1


if __name__ == "__main__":
    sys.exit(main())
