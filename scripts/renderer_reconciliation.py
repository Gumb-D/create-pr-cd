"""Read renderer artifacts and return one terminal disposition per renderer candidate."""
from __future__ import annotations

import csv
import hashlib
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping

from openpyxl import load_workbook


def _artifact_signature(path: Path) -> tuple[int, int, str]:
    stat = path.stat()
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return stat.st_size, stat.st_mtime_ns, digest


def snapshot_renderer_artifacts(output: Path) -> dict[str, tuple[int, int, str]]:
    """Snapshot direct output files so same-name overwrites can be detected later."""
    root = Path(output)
    if not root.exists():
        return {}
    return {
        str(path.resolve()): _artifact_signature(path)
        for path in root.glob("*")
        if path.is_file()
    }


def touched_renderer_artifacts(
    output: Path,
    before: Mapping[str, tuple[int, int, str]] | None = None,
) -> list[Path]:
    """Return files created or overwritten since the supplied snapshot."""
    baseline = dict(before or {})
    current = snapshot_renderer_artifacts(output)
    return sorted(
        Path(path)
        for path, signature in current.items()
        if baseline.get(path) != signature
    )


def _csv_site_codes(path: Path) -> tuple[set[str], dict[str, str]]:
    sites: set[str] = set()
    reasons: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            code = str(row.get("Site_ID", "") or "").strip().upper()
            if not code:
                continue
            sites.add(code)
            reasons[code] = str(row.get("Reason_Code", "") or row.get("Reason", "") or "").strip()
    return sites, reasons


def _ecc_site_codes(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "details" not in workbook.sheetnames:
            return set()
        worksheet = workbook["details"]
        header = [str(value or "").strip() for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            site_index = header.index("Site ID*")
        except ValueError:
            return set()
        return {
            str(row[site_index] or "").strip().upper()
            for row in worksheet.iter_rows(min_row=2, values_only=True)
            if site_index < len(row) and str(row[site_index] or "").strip()
        }
    finally:
        workbook.close()


def collect_renderer_reconciliation(
    output: Path,
    candidates: list[Mapping[str, Any]],
    scope: str,
    render_site_code: Callable[[Mapping[str, Any]], str],
    *,
    before: set[Path] | None = None,
    created_paths: Iterable[Path | str] | None = None,
) -> dict[str, Any]:
    """Reconcile renderer-created ECC/review artifacts back to source site codes."""
    if created_paths is not None:
        created = [Path(path).resolve() for path in created_paths]
    else:
        excluded = {Path(path).resolve() for path in (before or set())}
        created = [
            path.resolve()
            for path in Path(output).glob("*")
            if path.is_file() and path.resolve() not in excluded
        ]

    generated_rendered: set[str] = set()
    review_rendered: set[str] = set()
    duplicate_rendered: set[str] = set()
    review_reasons: dict[str, str] = {}

    for path in created:
        if not path.exists():
            continue
        upper_name = path.name.upper()
        if path.suffix.lower() == ".xlsx" and " PR " in upper_name:
            generated_rendered.update(_ecc_site_codes(path))
        elif path.suffix.lower() == ".csv" and upper_name.startswith(f"REVIEW_REQUIRED_{scope.upper()}_"):
            sites, reasons = _csv_site_codes(path)
            review_rendered.update(sites)
            review_reasons.update(reasons)
        elif path.suffix.lower() == ".csv" and upper_name.startswith(f"DUPLICATES_SKIPPED_{scope.upper()}_"):
            sites, _ = _csv_site_codes(path)
            duplicate_rendered.update(sites)

    dispositions = []
    for record in candidates:
        source_code = str(record.get("site", {}).get("site_code", "") or "").strip()
        rendered_code = str(render_site_code(record) or "").strip()
        key = rendered_code.upper()
        has_generated = key in generated_rendered
        has_review = key in review_rendered
        has_duplicate = key in duplicate_rendered

        # REVIEW_REQUIRED is a stronger terminal state than GENERATED. A renderer may
        # emit partial ECC rows while also quarantining unresolved line items.
        if has_review:
            disposition = "REVIEW_REQUIRED"
            reason_code = review_reasons.get(key) or "RENDERER_REVIEW_REQUIRED"
        elif has_generated:
            disposition = "GENERATED"
            reason_code = "ECC_GENERATED"
        elif has_duplicate:
            disposition = "DUPLICATE_BLOCKED"
            reason_code = "RENDERER_DUPLICATE_BLOCKED"
        else:
            disposition = "FAILED"
            reason_code = "RENDERER_SITE_UNACCOUNTED"
        dispositions.append(
            {
                "site_code": source_code,
                "rendered_site_code": rendered_code,
                "disposition": disposition,
                "reason_code": reason_code,
                "ecc_evidence_present": has_generated,
                "review_evidence_present": has_review,
                "duplicate_evidence_present": has_duplicate,
            }
        )

    return {"site_dispositions": dispositions}
