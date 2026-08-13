#!/usr/bin/env python3
"""Deterministic Operation Backoffice ECC renderer for Issue #94."""
from __future__ import annotations

import argparse
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backoffice_pr_runtime import load_service_registry

ROOT = Path(__file__).resolve().parent.parent
SERVICE_REGISTRY = ROOT / "config" / "backoffice_service_registry.yaml"
BACKOFFICE_PBOMS = {"350000592793", "350000592794"}
MAX_SITES_PER_FILE = 30
ECC_HEADERS = [
    "SN.", "Purchasing Area*", "Region*", "Site ID*", "Site Name*",
    "Delivery Unit Code*", "Logical Site Name", "Contract Number *",
    "Subcontractor*", "PBOM Code*", "SOW*", "Unit*", "Quantity*",
    "Remarks", "", "Contract Number",
]
REQUIRED_INPUT_COLUMNS = {
    "customer site code", "customer site name", "du code", "region",
    "Backoffice Event Code", "Backoffice Trigger Date", "Backoffice Billing Month",
    "Backoffice PBOM Code", "Backoffice Unit", "Backoffice Quantity",
    "Backoffice Subcontractor", "Backoffice Contract Number", "Backoffice Issue Type",
    "Backoffice Warnings",
}


class BackofficeRendererError(RuntimeError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


def _text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").strip().split())


def _pbom(value: object) -> str:
    text = _text(value)
    return text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text


def _parse_date(value: object) -> date:
    text = _text(value)
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError as error:
        raise BackofficeRendererError("BACKOFFICE_TRIGGER_DATE_INVALID", f"Invalid Backoffice trigger date: {text}") from error


def _service_for_date(registry: Mapping[str, Any], trigger: date) -> tuple[str, str] | None:
    matches: list[tuple[str, str]] = []
    for item in registry.get("services", []):
        if not isinstance(item, Mapping):
            continue
        try:
            start = date.fromisoformat(_text(item.get("effective_from")))
            raw_end = _text(item.get("effective_to"))
            end = date.fromisoformat(raw_end) if raw_end else None
        except ValueError:
            continue
        if trigger < start or (end is not None and trigger > end):
            continue
        subcon = _text(item.get("subcontractor"))
        contract = _text(item.get("contract_number"))
        if subcon and contract:
            matches.append((subcon, contract))
    return matches[0] if len(matches) == 1 else None


def load_backoffice_model_item(path: Path, pbom_code: str) -> dict[str, Any]:
    pbom_code = _pbom(pbom_code)
    if pbom_code not in BACKOFFICE_PBOMS:
        raise BackofficeRendererError("BACKOFFICE_PR_MODEL_PBOM_NOT_APPROVED", f"Unsupported Backoffice PBOM: {pbom_code}")
    try:
        workbook = load_workbook(Path(path), read_only=True, data_only=True)
    except Exception as error:
        raise BackofficeRendererError("BACKOFFICE_PR_MODEL_UNREADABLE", f"Cannot read PR Model: {error}") from error
    matches: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                values = list(row)
                if len(values) < 5:
                    continue
                if _text(values[0]).casefold() != "operation back office":
                    continue
                if _pbom(values[1]) != pbom_code:
                    continue
                matches.append({
                    "sheet": worksheet.title,
                    "sow": _text(values[0]),
                    "pbom_code": _pbom(values[1]),
                    "description": _text(values[2]),
                    "unit": _text(values[3]),
                    "quantity": int(float(values[4])) if _text(values[4]) else 0,
                    "rules": _text(values[5]) if len(values) > 5 else "",
                })
    finally:
        workbook.close()
    if not matches:
        raise BackofficeRendererError("BACKOFFICE_PR_MODEL_PBOM_MISSING", f"Backoffice PBOM {pbom_code} is missing from the PR Model.")
    if len(matches) != 1:
        raise BackofficeRendererError("BACKOFFICE_PR_MODEL_PBOM_AMBIGUOUS", f"Backoffice PBOM {pbom_code} occurs {len(matches)} times in the PR Model.")
    item = matches[0]
    if item["unit"].casefold() != "hop" or item["quantity"] != 1:
        raise BackofficeRendererError("BACKOFFICE_PR_MODEL_LINE_ITEM_INVALID", f"Backoffice PBOM {pbom_code} must be Unit Hop and Quantity 1.")
    return item


def _parse_mapping(path: Path) -> dict[str, str]:
    if not Path(path).exists():
        raise BackofficeRendererError("MAPPING_FILE_NOT_FOUND", f"Mapping reference not found: {path}")
    region_mapping: dict[str, str] = {}
    section = ""
    for raw in Path(path).read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if line.startswith("## Region to Purchasing Area"):
            section = "region"
            continue
        if line.startswith("## "):
            section = ""
            continue
        if section != "region" or not line.startswith("|") or "---" in line:
            continue
        parts = [part.strip() for part in line.strip("|").split("|")]
        if len(parts) >= 2 and parts[0] not in {"Region*", "Region"} and parts[0] and parts[1]:
            region_mapping[_text(parts[0]).casefold()] = _text(parts[1])
    return region_mapping


def _load_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        if "data" not in workbook.sheetnames:
            raise BackofficeRendererError("CANONICAL_INPUT_SCHEMA_ERROR", "Backoffice renderer requires a data sheet.")
        worksheet = workbook["data"]
        header_row = 0
        headers: list[str] = []
        for row_number in range(1, min(10, worksheet.max_row) + 1):
            values = [_text(v) for v in next(worksheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True))]
            if REQUIRED_INPUT_COLUMNS.issubset(set(values)):
                header_row, headers = row_number, values
                break
        if not header_row:
            raise BackofficeRendererError("CANONICAL_INPUT_SCHEMA_ERROR", "Backoffice renderer input is missing required canonical columns.")
        index = {header: pos for pos, header in enumerate(headers) if header}
        result: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            row = {header: values[pos] if pos < len(values) else None for header, pos in index.items()}
            if any(_text(v) for v in row.values()):
                result.append(row)
        return result
    finally:
        workbook.close()


def _filter_sites(rows: list[dict[str, Any]], all_sites: bool, raw_site_codes: str | None) -> list[dict[str, Any]]:
    requested = [token.strip().upper() for token in str(raw_site_codes or "").split(",") if token.strip()]
    if all_sites == bool(requested):
        raise BackofficeRendererError("INVALID_SITE_SELECTION", "Use exactly one of --all-sites or --site-code.")
    if all_sites:
        return rows
    available = {_text(row.get("customer site code")).upper() for row in rows}
    missing = [code for code in requested if code not in available]
    if missing:
        raise BackofficeRendererError("SITE_CODES_NOT_FOUND", "Requested site codes were not found: " + ", ".join(missing))
    wanted = set(requested)
    return [row for row in rows if _text(row.get("customer site code")).upper() in wanted]


def _validated_ecc_rows(rows: list[dict[str, Any]], pr_model: Path, mapping: Path) -> tuple[list[dict[str, Any]], str, str, str]:
    if not rows:
        return [], "", "", ""
    months = {_text(row.get("Backoffice Billing Month")) for row in rows}
    pboms = {_pbom(row.get("Backoffice PBOM Code")) for row in rows}
    issue_types = {_text(row.get("Backoffice Issue Type")).upper() for row in rows}
    if len(months) != 1 or len(pboms) != 1 or len(issue_types) != 1:
        raise BackofficeRendererError("BACKOFFICE_BATCH_MIXED_GOVERNANCE", "One Backoffice ECC batch must contain exactly one billing month, PBOM and issue type.")
    billing_month = next(iter(months))
    pbom = next(iter(pboms))
    issue_type = next(iter(issue_types))
    if issue_type not in {"MAIN", "SUPPLEMENTARY"}:
        raise BackofficeRendererError("BACKOFFICE_ISSUE_TYPE_INVALID", f"Invalid Backoffice issue type: {issue_type}")
    model_item = load_backoffice_model_item(pr_model, pbom)
    region_mapping = _parse_mapping(mapping)
    registry = load_service_registry(SERVICE_REGISTRY)
    rendered: list[dict[str, Any]] = []
    for row in rows:
        site = _text(row.get("customer site code"))
        site_name = _text(row.get("customer site name"))
        du = _text(row.get("du code"))
        region = _text(row.get("region"))
        if not site or not site_name or not du or not region:
            raise BackofficeRendererError("BACKOFFICE_CANDIDATE_IDENTITY_MISSING", "Site ID, Site Name, Delivery Unit Code and Region are required.")
        purchasing_area = region_mapping.get(region.casefold())
        if not purchasing_area:
            raise BackofficeRendererError("PURCHASING_AREA_NOT_FOUND", f"No Purchasing Area mapping exists for Region {region!r}.")
        trigger = _parse_date(row.get("Backoffice Trigger Date"))
        effective = _service_for_date(registry, trigger)
        if effective is None:
            raise BackofficeRendererError("BACKOFFICE_SERVICE_CONTRACT_NOT_EFFECTIVE", f"No unique Backoffice service contract is effective on {trigger.isoformat()}.")
        pipeline_subcon = _text(row.get("Backoffice Subcontractor"))
        pipeline_contract = _text(row.get("Backoffice Contract Number"))
        if (pipeline_subcon, pipeline_contract) != effective:
            raise BackofficeRendererError("BACKOFFICE_SERVICE_PIPELINE_MISMATCH", f"Backoffice provider/contract mismatch for {site}.")
        if _text(row.get("Backoffice Unit")).casefold() != model_item["unit"].casefold():
            raise BackofficeRendererError("BACKOFFICE_PR_MODEL_PIPELINE_MISMATCH", f"Backoffice unit mismatch for {site}.")
        try:
            quantity = int(float(row.get("Backoffice Quantity")))
        except (TypeError, ValueError) as error:
            raise BackofficeRendererError("BACKOFFICE_PR_MODEL_PIPELINE_MISMATCH", f"Backoffice quantity is invalid for {site}.") from error
        if quantity != model_item["quantity"]:
            raise BackofficeRendererError("BACKOFFICE_PR_MODEL_PIPELINE_MISMATCH", f"Backoffice quantity mismatch for {site}.")
        rendered.append({
            "Purchasing_Area": purchasing_area,
            "Region": region,
            "Site_ID": site,
            "Site_Name": site_name,
            "Delivery_Unit_Code": du,
            "Logical_Site_Name": site_name,
            "Contract_Number": pipeline_contract,
            "Subcontractor": pipeline_subcon,
            "PBOM_Code": pbom,
            "SOW": model_item["description"],
            "Unit": model_item["unit"],
            "Quantity": model_item["quantity"],
            "Remarks": f"{_text(row.get('Backoffice Event Code'))}; Trigger={trigger.isoformat()}; Billing={billing_month}; {issue_type}",
            "Source_Tx_SOW": _text(row.get("Backoffice Warnings")),
        })
    return rendered, billing_month, issue_type


def _safe_filename(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*]+', " ", value).strip().strip(".")


def _split_rows_by_unique_site(rows: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    site_ids = sorted({_text(row.get("Site_ID")) for row in rows if _text(row.get("Site_ID"))})
    parts: list[list[dict[str, Any]]] = []
    for offset in range(0, len(site_ids), MAX_SITES_PER_FILE):
        selected_sites = set(site_ids[offset : offset + MAX_SITES_PER_FILE])
        parts.append([row for row in rows if _text(row.get("Site_ID")) in selected_sites])
    return parts


def _allocate_output_path(output: Path, filename: str) -> Path:
    path = output / filename
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    batch = 2
    while True:
        candidate = output / f"{stem} Batch {batch}{suffix}"
        if not candidate.exists():
            return candidate
        batch += 1


def _write_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "details"
    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    font = Font(bold=True)
    for col, header in enumerate(ECC_HEADERS, 1):
        cell = worksheet.cell(1, col, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for serial, row in enumerate(rows, 1):
        values = [serial,row["Purchasing_Area"],row["Region"],row["Site_ID"],row["Site_Name"],row["Delivery_Unit_Code"],row["Logical_Site_Name"],row["Contract_Number"],row["Subcontractor"],row["PBOM_Code"],row["SOW"],row["Unit"],row["Quantity"],row["Remarks"],row["Source_Tx_SOW"],row["Contract_Number"]]
        for col, value in enumerate(values, 1):
            worksheet.cell(serial + 1, col, value)
    for col, width in enumerate([5,20,12,15,20,18,20,18,15,15,72,8,10,48,20,18],1):
        worksheet.column_dimensions[get_column_letter(col)].width = width
    workbook.save(path)
    workbook.close()


def render(args: argparse.Namespace) -> list[Path]:
    if _text(args.scope).upper() != "BACKOFFICE":
        raise BackofficeRendererError("INVALID_SCOPE", "Backoffice renderer accepts only --scope BACKOFFICE.")
    rows = _filter_sites(_load_rows(args.site_data), args.all_sites, args.site_code)
    ecc_rows, billing_month, issue_type = _validated_ecc_rows(rows, args.pr_model, args.mapping)
    if not ecc_rows:
        return []
    output = Path(args.output)
    output.mkdir(parents=True, exist_ok=True)
    date_stamp = datetime.now().strftime("%Y%m%d")
    provider_contract_groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in ecc_rows:
        provider = _text(row.get("Subcontractor"))
        contract = _text(row.get("Contract_Number"))
        if not provider or not contract:
            raise BackofficeRendererError("BACKOFFICE_CANDIDATE_IDENTITY_MISSING", "Validated provider and contract are required for Backoffice output partitioning.")
        provider_contract_groups.setdefault((provider, contract), []).append(row)
    created: list[Path] = []
    for (provider, _contract), group_rows in sorted(provider_contract_groups.items()):
        parts = _split_rows_by_unique_site(group_rows)
        for part_number, part_rows in enumerate(parts, 1):
            part_suffix = f" Part {part_number}" if len(parts) > 1 else ""
            filename = _safe_filename(
                f"TX Outsource-{provider} Backoffice {issue_type} {billing_month} PR {date_stamp}{part_suffix}.xlsx"
            )
            path = _allocate_output_path(output, filename)
            _write_workbook(path, part_rows)
            created.append(path)
    return created


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render deterministic Operation Backoffice PR ECC output.")
    parser.add_argument("--site-data", required=True, type=Path)
    parser.add_argument("--pr-model", required=True, type=Path)
    parser.add_argument("--mapping", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--scope", required=True, type=str)
    parser.add_argument("--du-model-name", required=False, type=str, default="")
    parser.add_argument("--template", type=Path)
    parser.add_argument("--all-sites", action="store_true")
    parser.add_argument("--site-code")
    return parser.parse_args()


def main() -> int:
    try:
        created = render(parse_args())
        for path in created:
            print(path)
        return 0
    except BackofficeRendererError as error:
        print(f"ERROR [{error.code}] {error}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
