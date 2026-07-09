#!/usr/bin/env python3
"""Read-only local inventory for DU reference exports under Info/reference.

The script scans for local Excel/CSV exports, summarizes metadata only, and
writes the inventory under output/. It never mutates files inside Info/reference
and never emits raw customer data rows into committed documentation.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence

from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".csv", ".xls")
HEADER_ROW_COUNT = 4
DU_NAME_PATTERN = re.compile(
    r"^A-[^-]+-(?P<du_model_name>.+)-(?P<view_label>.+)-(?P<timestamp>\d{14})\.(?:xlsx|xlsm|csv|xls)$",
    re.IGNORECASE,
)
HEADER_HINTS = (
    "site",
    "code",
    "region",
    "tx",
    "sow",
    "subcon",
    "latitude",
    "longitude",
    "pr",
)


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split()).strip()


def infer_du_model(path: Path, sheet_names: Sequence[str] | None = None) -> str | None:
    match = DU_NAME_PATTERN.match(path.name)
    if match:
        return match.group("du_model_name").replace("_", " ")
    tokens = " ".join([path.stem.replace("_", " ")] + list(sheet_names or ())).lower()
    candidates = (
        "MW EOS Swap",
        "ZTE TX MINI",
        "TX Mini Project",
        "Jendela TX Migration",
        "Celcomdigi USP",
        "CD consolidation 2023",
        "2023 TX Rollout",
        "2023 Celcomdigi BAU",
        "2024 Celcomdigi BAU",
    )
    for candidate in candidates:
        if candidate.lower() in tokens:
            return candidate
    return None


def looks_like_iepms_four_layer_headers(rows: Sequence[Sequence[Any]]) -> bool:
    if len(rows) < HEADER_ROW_COUNT:
        return False
    normalized_rows = [[_normalize(cell).lower() for cell in row] for row in rows[:HEADER_ROW_COUNT]]
    non_empty_counts = [sum(1 for cell in row if cell) for row in normalized_rows]
    if min(non_empty_counts or [0]) == 0:
        return False
    combined = " ".join(cell for row in normalized_rows for cell in row if cell)
    hint_hits = sum(1 for hint in HEADER_HINTS if hint in combined)
    return hint_hits >= 3 and non_empty_counts[3] >= min(3, max(non_empty_counts))


def _csv_rows(path: Path) -> List[List[str]]:
    encodings = ("utf-8-sig", "utf-8", "latin-1")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.reader(handle))
        except UnicodeDecodeError as error:
            last_error = error
    raise last_error or UnicodeDecodeError("utf-8", b"", 0, 1, "Unable to decode CSV")


def _sheet_summary_from_rows(sheet_name: str, rows: Sequence[Sequence[Any]]) -> Dict[str, Any]:
    row_count = len(rows)
    column_count = max((len(row) for row in rows), default=0)
    header_rows = [list(row) for row in rows[:HEADER_ROW_COUNT]]
    return {
        "sheet_name": sheet_name,
        "row_count": row_count,
        "column_count": column_count,
        "looks_like_iepms_four_layer_headers": looks_like_iepms_four_layer_headers(header_rows),
    }


def summarize_excel_file(path: Path) -> Dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    candidate_sheets: List[Dict[str, Any]] = []
    sheet_names = list(workbook.sheetnames)
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            summary = _sheet_summary_from_rows(sheet_name, rows)
            if summary["looks_like_iepms_four_layer_headers"] or summary["row_count"] > 0:
                candidate_sheets.append(summary)
    finally:
        workbook.close()
    return {
        "sheet_names": sheet_names,
        "candidate_sheets": candidate_sheets,
        "candidate_du_model": infer_du_model(path, sheet_names),
        "appears_suitable_for_du_export_profiling": any(
            sheet["looks_like_iepms_four_layer_headers"] for sheet in candidate_sheets
        ),
    }


def summarize_csv_file(path: Path) -> Dict[str, Any]:
    rows = _csv_rows(path)
    summary = _sheet_summary_from_rows("CSV", rows)
    return {
        "sheet_names": ["CSV"],
        "candidate_sheets": [summary],
        "candidate_du_model": infer_du_model(path, ("CSV",)),
        "appears_suitable_for_du_export_profiling": summary["looks_like_iepms_four_layer_headers"],
    }


def summarize_reference_file(path: Path, root: Path) -> Dict[str, Any]:
    stat = path.stat()
    result: Dict[str, Any] = {
        "full_local_path": str(path.resolve()),
        "relative_path": str(path.resolve().relative_to(root.resolve())),
        "original_file_name": path.name,
        "extension": path.suffix.lower(),
        "file_size_bytes": stat.st_size,
        "last_modified_time": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        "candidate_du_model": None,
        "sheet_names": [],
        "candidate_sheets": [],
        "looks_like_iepms_four_layer_headers": False,
        "appears_suitable_for_du_export_profiling": False,
        "read_errors": [],
    }
    try:
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            summary = summarize_excel_file(path)
        elif path.suffix.lower() == ".csv":
            summary = summarize_csv_file(path)
        else:
            raise ValueError("Legacy .xls requires an optional reader that is not installed in this repo.")
        result.update(summary)
        result["looks_like_iepms_four_layer_headers"] = any(
            sheet["looks_like_iepms_four_layer_headers"] for sheet in result["candidate_sheets"]
        )
    except Exception as error:  # pragma: no cover - safety path exercised by local files
        result["read_errors"].append(str(error))
    return result


def discover_reference_files(reference_root: Path) -> List[Dict[str, Any]]:
    root = reference_root.resolve()
    files = sorted(
        path
        for path in reference_root.rglob("*")
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    )
    return [summarize_reference_file(path, root) for path in files]


def inventory_markdown(reference_root: Path, inventory: Sequence[Mapping[str, Any]]) -> str:
    lines = [
        "# Local DU Reference Discovery Inventory",
        "",
        f"- Reference root: `{reference_root.resolve()}`",
        f"- Generated at: `{datetime.now(timezone.utc).isoformat()}`",
        f"- Files inventoried: `{len(inventory)}`",
        "",
        "This inventory is local-review metadata only. It excludes raw row data and does not approve mappings, lifecycle changes, or ECC output.",
        "",
        "| File | Ext | Size (bytes) | Candidate DU model | Four-layer header look | Suitable for profiling | Read errors |",
        "|---|---:|---:|---|---|---|---|",
    ]
    for item in inventory:
        lines.append(
            "| {name} | `{ext}` | `{size}` | {du} | `{headers}` | `{suitable}` | {errors} |".format(
                name=item["original_file_name"].replace("|", "\\|"),
                ext=item["extension"],
                size=item["file_size_bytes"],
                du=(item["candidate_du_model"] or "Unknown").replace("|", "\\|"),
                headers=item["looks_like_iepms_four_layer_headers"],
                suitable=item["appears_suitable_for_du_export_profiling"],
                errors="; ".join(item["read_errors"]).replace("|", "\\|") if item["read_errors"] else "None",
            )
        )
    return "\n".join(lines) + "\n"


def write_inventory(output_path: Path, inventory: Sequence[Mapping[str, Any]], reference_root: Path) -> Dict[str, Path]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path = output_path.with_suffix(".md")
    payload = {
        "schema_version": "1.0",
        "reference_root": str(reference_root.resolve()),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "inventory": list(inventory),
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    markdown_path.write_text(inventory_markdown(reference_root, inventory), encoding="utf-8")
    return {"json": output_path, "markdown": markdown_path}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Read-only inventory for local DU reference exports.")
    parser.add_argument(
        "--reference-root",
        default="Info/reference",
        help="Local raw DU reference root. Defaults to Info/reference.",
    )
    parser.add_argument(
        "--output",
        default="output/local_du_reference_inventory.json",
        help="JSON output path. Markdown summary is written beside it.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    reference_root = Path(args.reference_root)
    inventory = discover_reference_files(reference_root)
    outputs = write_inventory(Path(args.output), inventory, reference_root)
    print(f"Discovered {len(inventory)} local reference files.")
    for name, path in outputs.items():
        print(f"{name}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
