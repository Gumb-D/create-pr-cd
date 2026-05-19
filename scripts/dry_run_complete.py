import pandas as pd
import openpyxl
from pathlib import Path

print("=" * 100)
print("DRY RUN: 5 TSS PR GENERATION")
print("=" * 100)

# File paths
info_folder = Path("Info")
site_data_file = info_folder / "input" / "site_pr_po_view.xlsx"
pr_model_file = info_folder / "input" / "pr_model.xlsx"

# Load site data (database format - using row 2 which has the human-readable headers)
print("\n[STEP 1] Loading Site Data...")
site_df = pd.read_excel(site_data_file, sheet_name='data', header=1)  # Use row 2 as header
site_df_cols = site_df.columns.tolist()

print(f"✓ Loaded {len(site_df)} site records")
print(f"\n  Data columns ({len(site_df_cols)} total):")
for i, col in enumerate(site_df_cols[:30], 1):
    print(f"    {i}. {col}")

# Load PR Model Reference
print("\n\n[STEP 2] Loading PR Model Reference...")
pr_wb = openpyxl.load_workbook(pr_model_file)
print(f"✓ Available sheets: {pr_wb.sheetnames}")

# Get actual headers from the first data row
tx_sheet = pr_wb["TX Line Item (After 21-Apr 26)"]
pbom_sheet = pr_wb["PBOM (After 21-Apr 26)"]

# Find header row (scan for non-empty cells)
def find_header_row(sheet):
    for row_idx in range(1, min(5, sheet.max_row + 1)):
        cell_values = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
        if any(cell_values):
            return row_idx, cell_values
    return None, None

header_idx_tx, headers_tx = find_header_row(tx_sheet)
header_idx_pbom, headers_pbom = find_header_row(pbom_sheet)

print(f"\n  TX Line Item sheet - Headers at row {header_idx_tx}:")
for i, h in enumerate(headers_tx[:12], 1):
    if h:
        print(f"    {i}. {h}")

print(f"\n  PBOM sheet - Headers at row {header_idx_pbom}:")
for i, h in enumerate(headers_pbom[:12], 1):
    if h:
        print(f"    {i}. {h}")

# Load ECC Template
print("\n\n[STEP 3] Loading ECC Template...")
ecc_template = pd.read_excel(info_folder / "input" / "ecc_template.xls", sheet_name='details')
print(f"✓ ECC Template columns ({len(ecc_template.columns)}):")
for i, col in enumerate(ecc_template.columns, 1):
    print(f"    {i}. {col}")

# Load Sample Output for reference
print("\n\n[STEP 4] Analyzing Sample TSS PR Output...")
sample_df = pd.read_excel(info_folder / "sample" / "Northern-GCI TX Mini Project TSS PR 20260515.xls", sheet_name='details')
sample_contract = pd.read_excel(info_folder / "sample" / "Northern-GCI TX Mini Project TSS PR 20260515.xls", sheet_name='contract infor')

print(f"✓ Sample output has {len(sample_df)} PR lines")
print(f"\n  Sample contract infor sheet:")
print(sample_contract.head(10).to_string())

# Map site data columns to human-readable names
print("\n\n[STEP 5] Searching for Site Data Trigger Columns...")

# Try to find relevant columns by pattern
relevant_patterns = ['site|fix00012', 'region', 'SOW', 'antenna', 'TSS', 'TI', 'Planning', 'actual_end_date']
found_cols = {}

for col in site_df_cols:
    col_lower = col.lower()
    if 'site' in col_lower and 'fix00012' in col_lower:
        found_cols['Site ID'] = col
    elif 'region' in col_lower:
        found_cols['Region'] = col
    elif 'sow' in col_lower:
        found_cols['SOW'] = col
    elif 'antenna' in col_lower:
        found_cols['Antenna'] = col
    elif 'tss' in col_lower:
        found_cols['TSS'] = col
    elif 'ti' in col_lower and 'team' in col_lower:
        found_cols['TI'] = col
    elif 'planning' in col_lower:
        found_cols['Planning'] = col
    elif 'actual_end_date' in col_lower:
        found_cols['Operation'] = col

print(f"\n  Identified columns:")
for key, col in found_cols.items():
    print(f"    {key}: {col}")

# Show sample data
print("\n\n[STEP 6] Sample Site Data (first 10 rows with relevant columns):")
display_cols = list(found_cols.values()) + ['Site Name', 'DU Code']
available_display = [col for col in display_cols if col in site_df_cols]
if available_display:
    print(site_df[available_display].head(10).to_string())
else:
    print("Could not identify matching columns for display")

print("\n\n" + "=" * 100)
print("ANALYSIS SUMMARY")
print("=" * 100)

print(f"""
Key Findings:
1. Site Data: 2,557 records with 142 columns (database format)
2. PR Model Reference: Two sheets with TSS/TI/Planning models and PBOM codes
3. ECC Template: Standard format with 16 columns including mandatory fields
4. Sample Output: Shows 2 TSS PR lines for Northern-GCI group

To generate 5 TSS PR:
STEP 1: Identify sites with SubCon - TSS Team populated and no existing TSS PR
STEP 2: Extract primary SOW from Tx SOW field
STEP 3: Match SOW to TSS PR model (antenna size NOT required for TSS)
STEP 4: Retrieve mandatory line items for matched model
STEP 5: Get contract number and purchasing area from contract infor
STEP 6: Generate 5 ECC output rows
STEP 7: Group by Region + Subcontractor and save to file

Expected Output:
- File name: <Region>-<Subcon> TX Mini Project TSS PR 20260515.xls
- Format: Excel with 'details' sheet (PR lines) + 'contract infor' sheet
- Each PR line has: Purchasing Area, Region, Site ID/Name, Contract No, Subcon, PBOM Code, SOW, Unit, Qty
""")

print("=" * 100)
