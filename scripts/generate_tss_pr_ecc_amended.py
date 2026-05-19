#!/usr/bin/env python3
"""
TSS PR ECC File Generation - Amendment Implementation
Amended version incorporating:
- Amendment 1: Separate contract info reference (Markdown)
- Amendment 2: Single 'details' sheet only
- Amendment 3: Sequential SN, Region→Purchasing Area, Subcon→Contract, 30-site split, fuzzy matching
"""

import argparse
import os
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from difflib import SequenceMatcher
import re


def parse_args():
    parser = argparse.ArgumentParser(description="Generate ECC files from site PR/PO view and PR model.")
    parser.add_argument('--site-data', default='Info/input/site_pr_po_view.xlsx', help='Path to daily site PR/PO view file')
    parser.add_argument('--pr-model', default='Info/input/pr_model.xlsx', help='Path to PR model Excel file')
    parser.add_argument('--template', default='Info/input/ecc_template.xls', help='Path to ECC template file')
    parser.add_argument('--mapping', default='Info/input/contract_info_reference.md', help='Path to region/subcontractor mapping markdown')
    parser.add_argument('--output', default='output', help='Output directory for generated ECC files')
    parser.add_argument('--site-code', help='Comma-separated Site Code(s) to generate PR ECC for')
    parser.add_argument('--all-sites', action='store_true', help='Generate PR ECC for all eligible sites')
    parser.add_argument('--scope', choices=['TSS', 'TI'], default='TSS', type=str.upper, help='PR scope to generate: TSS or TI')
    return parser.parse_args()


def require_file(path, description):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(
            f"{description} not found: {p}\n" \
            f"Place the correct file at this path or use the --site-data/--pr-model/--template/--mapping options."
        )
    return p


def validate_template_file(path):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"ECC Template not found: {path}")
    if p.suffix.lower() == '.xlsx':
        wb = load_workbook(path, read_only=True, data_only=True)
        if 'details' not in wb.sheetnames:
            raise ValueError(f"ECC template must contain a 'details' sheet: {path}")
    return True


args = parse_args()
output_dir = Path(args.output)
output_dir.mkdir(parents=True, exist_ok=True)

print("=" * 100)
print(f"{args.scope.upper()} PR ECC GENERATION - AMENDED IMPLEMENTATION")
print("=" * 100)
print(f"Execution Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

# ===== HELPER FUNCTIONS =====

def load_region_mapping(ref_file):
    """Load Region to Purchasing Area mapping from reference file."""
    mapping = {}
    with open(ref_file, 'r') as f:
        lines = f.readlines()
    in_section = False
    for line in lines:
        if '## Region to Purchasing Area' in line:
            in_section = True
            continue
        if in_section and line.startswith('## '):
            break
        if in_section and '|' in line:
            # Skip header row (contains "Region*") and separator row (contains "---")
            if 'Region*' in line or '---' in line:
                continue
            parts = [p.strip() for p in line.split('|')]
            # parts = ['', region, purchasing_area, '']
            if len(parts) >= 3 and parts[1] and parts[2] and parts[1] != '':
                mapping[parts[1]] = parts[2]
    return mapping

def load_subcon_mapping(ref_file):
    """Load Subcontractor to Contract mapping from reference file."""
    mapping = {}
    with open(ref_file, 'r') as f:
        lines = f.readlines()
    in_section = False
    for line in lines:
        if '## Subcontractor to Contract Number' in line:
            in_section = True
            continue
        if in_section and line.startswith('## '):
            break
        if in_section and '|' in line:
            # Skip header row (contains "Subcontractor*") and separator row (contains "---")
            if 'Subcontractor*' in line or '---' in line:
                continue
            parts = [p.strip() for p in line.split('|')]
            # parts = ['', subcon, contract_number, company_name, '']
            if len(parts) >= 4 and parts[1] and parts[2] and parts[1] != '':
                mapping[parts[1]] = {
                    'contract_number': parts[2],
                    'company_name': parts[3]
                }
    return mapping

def fuzzy_match_subcon(subcon_name, subcon_mapping, threshold=0.6):
    """
    Fuzzy match a subcontractor name against known list.
    Returns matched subcon key or None.
    """
    if subcon_name in subcon_mapping:
        return subcon_name  # Exact match
    
    best_match = None
    best_score = threshold
    
    for known_subcon in subcon_mapping.keys():
        score = SequenceMatcher(None, subcon_name.lower(), known_subcon.lower()).ratio()
        if score > best_score:
            best_score = score
            best_match = known_subcon
    
    return best_match


def parse_site_codes(site_code_str):
    if not site_code_str:
        return []
    codes = [code.strip().upper() for code in site_code_str.split(',') if code.strip()]
    return [code for code in codes if code]


def detect_site_code_column(df):
    preferred = ['Site ID', 'Site ID*', 'Site ID ', 'customer site code', 'Site Code', 'site code']
    lower_columns = {col.lower(): col for col in df.columns}
    for candidate in preferred:
        if candidate in df.columns:
            return candidate
        if candidate.lower() in lower_columns:
            return lower_columns[candidate.lower()]
    raise ValueError(
        'Site code column not found. Expected one of: Site ID, Site ID*, customer site code, Site Code.'
    )


def normalize_antenna_size(value):
    if pd.isna(value):
        return None
    raw = str(value).strip()
    if not raw:
        return None
    match = re.search(r'(\d+(?:\.\d+)?)', raw.replace(',', '.'))
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def determine_ti_antenna_category(ne_size, fe_size):
    ne_val = normalize_antenna_size(ne_size)
    fe_val = normalize_antenna_size(fe_size)
    chosen_size = None
    remark = None

    if ne_val is None and fe_val is None:
        remark = 'Missing TI antenna size - review required'
    elif ne_val is None or fe_val is None:
        chosen_size = fe_val if ne_val is None else ne_val
        remark = 'Incomplete TI antenna size - review required'
    else:
        chosen_size = max(ne_val, fe_val)
        if ne_val != fe_val:
            remark = 'TI antenna sizes differ - using larger size for matching'

    category = None
    if chosen_size is not None:
        if chosen_size <= 0.6:
            category = '0.3/0.6m'
        elif chosen_size <= 1.2:
            category = '0.9/1.2m'
        elif chosen_size <= 1.8:
            category = '1.8m'
        elif chosen_size <= 2.4:
            category = '2.4m'
        else:
            category = f'{chosen_size}m'

    return category, remark


def match_ti_models(sow, antenna_category, ti_models):
    sow_upper = sow.upper()
    candidates = []
    for item in ti_models:
        item_sow_upper = item['SOW'].upper()
        if item_sow_upper == sow_upper or item_sow_upper in sow_upper or sow_upper in item_sow_upper:
            if item['Is_Mandatory']:
                candidates.append(item)

    if not candidates:
        return [], False

    size_candidates = []
    if antenna_category:
        search_term = antenna_category.lower()
        for item in candidates:
            if search_term in item['Description'].lower() or search_term in item['SOW'].lower():
                size_candidates.append(item)
        if size_candidates:
            candidates = size_candidates

    review_required = len(candidates) > 1
    return candidates, review_required


def filter_site_rows(df_site, site_codes=None, all_sites=False):
    if site_codes and all_sites:
        raise ValueError('Use either --site-code or --all-sites, not both.')
    if not site_codes and not all_sites:
        raise ValueError('Please provide --site-code <SITE_CODE> or use --all-sites to generate all eligible sites.')
    if all_sites:
        return df_site.copy(), []

    normalized_codes = [code for code in site_codes if code]
    if not normalized_codes:
        raise ValueError('No valid site codes provided to --site-code.')

    site_code_column = detect_site_code_column(df_site)

    df_site['_normalized_site_code'] = df_site[site_code_column].astype(str).str.strip().str.upper()
    requested_set = set(normalized_codes)
    matched_df = df_site[df_site['_normalized_site_code'].isin(requested_set)].copy()
    matched_codes = set(matched_df['_normalized_site_code'].unique())
    missing_codes = [code for code in normalized_codes if code not in matched_codes]

    if len(matched_df) == 0:
        raise ValueError('No matching Site Codes found in Site PR/PO View. Please check the input site code(s).')

    if missing_codes:
        print('Warning: requested Site Code(s) not found:')
        for code in missing_codes:
            print(f'- {code}')
        print(f'✓ Continuing with {len(matched_df)} matched row(s).')

    matched_df = matched_df.drop(columns=['_normalized_site_code'])
    return matched_df, missing_codes


# ===== STEP 0: LOAD REFERENCE DATA =====
print("[STEP 0] Loading Reference Data from Markdown and validating inputs...")

site_file = require_file(args.site_data, 'Site PR/PO View')
pr_file = require_file(args.pr_model, 'PR Model')
template_file = require_file(args.template, 'ECC Template')
ref_file = require_file(args.mapping, 'Contract info mapping')
validate_template_file(template_file)

print(f"✓ Site PR/PO View: {site_file}")
print(f"✓ PR Model: {pr_file}")
print(f"✓ ECC Template: {template_file}")
print(f"✓ Mapping file: {ref_file}")

region_mapping = load_region_mapping(ref_file)
subcon_mapping = load_subcon_mapping(ref_file)

print(f"✓ Loaded {len(region_mapping)} region mappings")
print(f"  Regions: {', '.join(sorted(region_mapping.keys()))}")
print(f"✓ Loaded {len(subcon_mapping)} subcontractor mappings")
print(f"  Subcons: {', '.join(sorted(subcon_mapping.keys()))[:80]}...")

# ===== STEP 1: EXTRACT PR MODEL DATA =====
print("\n[STEP 1] Extracting PR Model Data...")

df_pr = pd.read_excel(pr_file, sheet_name="TX Line Item (After 21-Apr 26)", header=None)

# Extract TSS models (starting from row 7 = index 7)
tss_models = []

for idx in range(7, len(df_pr)):
    sow = df_pr.iloc[idx, 0]
    pbom = df_pr.iloc[idx, 1]
    desc = df_pr.iloc[idx, 2]
    unit = df_pr.iloc[idx, 3]
    qty = df_pr.iloc[idx, 4]
    rules = df_pr.iloc[idx, 5]
    
    # Stop at next section
    if pd.isna(sow) or str(sow).strip() == '':
        break
    
    # Check if mandatory
    is_mandatory = 'Mandatory' in str(rules) if pd.notna(rules) else False
    
    if pd.notna(pbom) and pd.notna(desc):
        tss_models.append({
            'SOW': str(sow).strip(),
            'PBOM_Code': str(pbom).strip(),
            'Description': str(desc).strip(),
            'Unit': str(unit).strip() if pd.notna(unit) else 'Hop',
            'Quantity': float(qty) if pd.notna(qty) else 1,
            'Is_Mandatory': is_mandatory
        })

print(f"✓ Extracted {len(tss_models)} TSS line items")

# Extract TI models (starting from TI Model header)
ti_models = []
ti_header_idx = None
for idx in range(len(df_pr)):
    cell_value = df_pr.iloc[idx, 0]
    if isinstance(cell_value, str) and 'TI Model' in cell_value:
        ti_header_idx = idx
        break

if ti_header_idx is not None:
    for idx in range(ti_header_idx + 1, len(df_pr)):
        sow = df_pr.iloc[idx, 0]
        pbom = df_pr.iloc[idx, 1]
        desc = df_pr.iloc[idx, 2]
        unit = df_pr.iloc[idx, 3]
        qty = df_pr.iloc[idx, 4]
        rules = df_pr.iloc[idx, 5]

        if pd.isna(sow) or str(sow).strip() == '':
            break

        is_mandatory = 'Mandatory' in str(rules) if pd.notna(rules) else False
        if pd.notna(pbom) and pd.notna(desc):
            ti_models.append({
                'SOW': str(sow).strip(),
                'PBOM_Code': str(pbom).strip(),
                'Description': str(desc).strip(),
                'Unit': str(unit).strip() if pd.notna(unit) else 'Hop',
                'Quantity': float(qty) if pd.notna(qty) else 1,
                'Is_Mandatory': is_mandatory
            })

    print(f"✓ Extracted {len(ti_models)} TI line items")
else:
    print('Warning: TI Model section not found in PR model sheet.')

# Group by SOW for TSS and TI models
sow_groups = {}
for item in tss_models:
    sow = item['SOW']
    if sow not in sow_groups:
        sow_groups[sow] = []
    sow_groups[sow].append(item)

print(f"✓ Found {len(sow_groups)} unique TSS SOWs:")
for sow in sorted(sow_groups.keys()):
    mandatory_count = len([x for x in sow_groups[sow] if x['Is_Mandatory']])
    print(f"  - {sow[:50]}: {len(sow_groups[sow])} items ({mandatory_count} mandatory)")

if ti_models:
    ti_sow_groups = {}
    for item in ti_models:
        sow = item['SOW']
        if sow not in ti_sow_groups:
            ti_sow_groups[sow] = []
        ti_sow_groups[sow].append(item)

    print(f"✓ Found {len(ti_sow_groups)} unique TI SOWs:")
    for sow in sorted(ti_sow_groups.keys()):
        mandatory_count = len([x for x in ti_sow_groups[sow] if x['Is_Mandatory']])
        print(f"  - {sow[:50]}: {len(ti_sow_groups[sow])} items ({mandatory_count} mandatory)")

# ===== STEP 2: LOAD SITE DATA =====
print("\n[STEP 2] Loading Site Data...")

df_site = pd.read_excel(site_file, sheet_name='data', header=3)

# Apply Site Selection filter
try:
    selected_site_codes = parse_site_codes(args.site_code)
    df_site, missing_codes = filter_site_rows(df_site, site_codes=selected_site_codes, all_sites=args.all_sites)
    if args.site_code:
        print(f"✓ Site selection: {len(selected_site_codes)} requested codes")
    if args.all_sites:
        print("✓ Site selection: all eligible sites")
    print(f"✓ Matched site rows: {len(df_site)}")
except ValueError as e:
    print(f"ERROR: {e}")
    sys.exit(1)

scope_name = args.scope.upper()
subcon_column = 'SubCon - TSS Team' if scope_name == 'TSS' else 'SubCon - TI Team'

candidates = df_site[
    (df_site[subcon_column].notna()) &
    (df_site[subcon_column].astype(str).str.strip() != '')
].copy().reset_index(drop=True)

print(f"✓ Found {len(candidates)} {scope_name} candidates")

# Display first 5 for dry run
print(f"✓ First 5 candidates:")
for i in range(min(5, len(candidates))):
    site_id = candidates.iloc[i]['customer site code']
    site_name = candidates.iloc[i]['customer site name']
    region = candidates.iloc[i]['region']
    subcon = candidates.iloc[i][subcon_column]
    sow = str(candidates.iloc[i]['Tx SOW'])
    print(f"  {i+1}. {site_id} ({site_name}) - Region: {region}, SubCon: {subcon}, SOW: {sow[:40]}")

# ===== STEP 3: BUILD ECC OUTPUT ROWS =====
print("\n[STEP 3] Building ECC Output Rows...")

ecc_rows = []
fuzzy_matches = {}

for idx in range(len(candidates)):
    row = candidates.iloc[idx]
    
    site_id = str(row['customer site code']).strip()
    site_name = str(row['customer site name']).strip()
    region = str(row['region']).strip()
    subcon = str(row[subcon_column]).strip()
    sow = str(row['Tx SOW']).strip()
    delivery_unit = str(row.get('du code', '')).strip() if pd.notna(row.get('du code')) else ''
    logical_site = str(row.get('customer site code', '')).strip() if pd.notna(row.get('customer site code')) else ''
    
    # Get Purchasing Area from Region
    purchasing_area = region_mapping.get(region, f"UNKNOWN ({region})")
    
    # Get Contract Number from Subcontractor (with fuzzy matching)
    matched_subcon = subcon
    if subcon in subcon_mapping:
        contract_info = subcon_mapping[subcon]
    else:
        fuzzy_match = fuzzy_match_subcon(subcon, subcon_mapping)
        if fuzzy_match:
            matched_subcon = fuzzy_match
            contract_info = subcon_mapping[fuzzy_match]
            fuzzy_matches[subcon] = fuzzy_match
            print(f"  ⚠ Fuzzy matched '{subcon}' → '{fuzzy_match}'")
        else:
            contract_info = {'contract_number': 'UNKNOWN', 'company_name': 'Unknown'}
            print(f"  ✗ No match for subcontractor: {subcon}")
    
    contract_number = contract_info['contract_number']
    
    matched_items = []
    remarks = ''

    if scope_name == 'TSS':
        sow_upper = sow.upper()
        for item in tss_models:
            item_sow_upper = item['SOW'].upper()
            if item['Is_Mandatory'] and (item_sow_upper == sow_upper or item_sow_upper in sow_upper or sow_upper in item_sow_upper):
                matched_items.append(item)
    else:
        antenna_category, antenna_remark = determine_ti_antenna_category(
            row.get('MW Config Antenna Size NE', ''),
            row.get('MW Config Antenna Size FE', '')
        )
        if antenna_remark:
            remarks = antenna_remark
        matched_items, review_required = match_ti_models(sow, antenna_category, ti_models)
        if review_required:
            remarks = 'REVIEW_REQUIRED' if not remarks else f"{remarks}; REVIEW_REQUIRED"

    if not matched_items:
        print(f"  ✗ No mandatory items found for SOW: {sow[:50]} (Scope: {scope_name})")
        continue
    
    for item in matched_items:
        ecc_row = {
            'Site_ID': site_id,
            'Site_Name': site_name,
            'Region': region,
            'Purchasing_Area': purchasing_area,
            'Subcontractor': matched_subcon,
            'Contract_Number': contract_number,
            'PBOM_Code': item['PBOM_Code'],
            'SOW': item['Description'],
            'Unit': item['Unit'],
            'Quantity': item['Quantity'],
            'Delivery_Unit_Code': delivery_unit,
            'Logical_Site_Name': logical_site,
            'Remarks': remarks
        }
        ecc_rows.append(ecc_row)

print(f"✓ Built {len(ecc_rows)} ECC output rows")

# ===== STEP 4: GROUP BY REGION-SUBCON AND CREATE EXCEL FILES =====
print("\n[STEP 4] Grouping and Creating Excel Files...")

# Group by (Region, Subcontractor)
grouped = {}
for ecc_row in ecc_rows:
    group_key = (ecc_row['Region'], ecc_row['Subcontractor'])
    if group_key not in grouped:
        grouped[group_key] = []
    grouped[group_key].append(ecc_row)

file_count = 0
total_rows = 0

for (region, subcon), rows in sorted(grouped.items()):
    # Check if split needed (max 30 unique sites per file)
    unique_sites = set([r['Site_ID'] for r in rows])
    num_files = (len(unique_sites) + 29) // 30  # Ceiling division
    
    print(f"\n  Group: {region}-{subcon}")
    print(f"    Unique sites: {len(unique_sites)}")
    print(f"    Total rows: {len(rows)}")
    print(f"    Files needed: {num_files}")
    
    # Distribute rows across files
    if num_files == 1:
        # Single file
        parts = [rows]
        file_parts = [1]
    else:
        # Multiple files - split by unique sites
        sites_list = sorted(list(unique_sites))
        parts = []
        file_parts = []
        
        for part_num in range(num_files):
            start_idx = part_num * 30
            end_idx = (part_num + 1) * 30
            part_sites = set(sites_list[start_idx:end_idx])
            part_rows = [r for r in rows if r['Site_ID'] in part_sites]
            parts.append(part_rows)
            file_parts.append(part_num + 1)
    
    # Create Excel file(s)
    for part_rows, part_num in zip(parts, file_parts):
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'details'
        
        # Write headers
        headers = ['SN.', 'Purchasing Area*', 'Region*', 'Site ID*', 'Site Name*', 
                   'Delivery Unit Code*', 'Logical Site Name', 'Contract Number *',
                   'Subcontractor*', 'PBOM Code*', 'SOW*', 'Unit*', 'Quantity*', 
                   'Remarks', '', 'Contract Number']
        
        # Header styling
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_font = Font(bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data rows with sequential SN
        for sn, ecc_row in enumerate(part_rows, 1):
            row_data = [
                sn,  # SN. - Sequential from 1
                ecc_row['Purchasing_Area'],
                ecc_row['Region'],
                ecc_row['Site_ID'],
                ecc_row['Site_Name'],
                ecc_row['Delivery_Unit_Code'],
                ecc_row['Logical_Site_Name'],
                ecc_row['Contract_Number'],  # Contract Number *
                ecc_row['Subcontractor'],
                ecc_row['PBOM_Code'],
                ecc_row['SOW'],
                ecc_row['Unit'],
                ecc_row['Quantity'],
                ecc_row['Remarks'],
                '',  # Column O (empty)
                ecc_row['Contract_Number']  # Column P - Contract Number (same as *)
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(2 + sn - 1, col_idx, value)
                if col_idx in [1, 13]:  # SN and Quantity - numeric
                    cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        column_widths = [5, 20, 12, 15, 20, 15, 20, 18, 15, 15, 40, 8, 10, 15, 5, 18]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d')
        if num_files == 1:
            filename = output_dir / f"{region}-{subcon} TX Mini Project {scope_name} PR {timestamp}.xls"
        else:
            filename = output_dir / f"{region}-{subcon} TX Mini Project {scope_name} PR {timestamp} Part {part_num}.xls"
        
        # Save file
        wb.save(str(filename))
        
        file_count += 1
        total_rows += len(part_rows)
        print(f"    ✓ Created: {filename} ({len(part_rows)} rows)")

# ===== SUMMARY =====
print("\n" + "=" * 100)
print("SUMMARY")
print("=" * 100)
print(f"Total files generated: {file_count}")
print(f"Total ECC rows: {total_rows}")
print(f"Fuzzy matched subcontractors: {len(fuzzy_matches)}")
if fuzzy_matches:
    for original, matched in fuzzy_matches.items():
        print(f"  - '{original}' → '{matched}'")
print(f"\nTimestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 100)
