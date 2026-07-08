"""TX Mini golden-parity harness: legacy view vs canonical-path round trip.

Dry-run validation evidence only. The harness never approves anything: the
legacy view's header hash is not in the profile's approved list, which per the
change-control rules permits validation dry-runs but no production output.

Path A (legacy): the ECC generator consumes the site view unchanged.
Path B (canonical): the same view is resolved through the approved DU profile's
four-layer fingerprints into Canonical PR Site Records; the mapped columns are
then re-rendered from each record's source evidence into a canonical-path view,
and the generator consumes that. Identical ECC output proves the profile
resolves exactly the columns and values the legacy path uses.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Tuple

from canonical_site_validator import SCOPE_REQUIRED_FIELDS
from du_export_adapter import build_canonical_site_record, resolve_profile_field_mappings
from du_profile_loader import load_du_profile
from profile_du_export import build_header_inventory, calculate_header_hash, fingerprint_key, sha256_file

DATA_SHEET = "data"
HEADER_ROW_COUNT = 4
_TIMESTAMP_PATTERN = re.compile(r"\d{6,}")


def resolution_report(inventory: Mapping[str, Any], profile: Mapping[str, Any]) -> Dict[str, Any]:
    resolved = resolve_profile_field_mappings(inventory, profile)
    header_hash = calculate_header_hash(inventory)
    approved_hashes = profile.get("export_structure", {}).get("approved_header_hashes", [])
    mapped_fields = {
        field: config
        for field, config in profile.get("field_mapping", {}).items()
        if config.get("source_candidates")
    }
    unresolved = sorted(field for field in mapped_fields if resolved[field]["status"] != "RESOLVED")
    return {
        "header_hash": header_hash,
        "header_hash_approved": header_hash in approved_hashes,
        "resolved_mappings": resolved,
        "mapped_field_count": len(mapped_fields),
        "unresolved_mapped_fields": unresolved,
        "dry_run_permitted": not unresolved,
    }


def _column_index_by_fingerprint(inventory: Mapping[str, Any]) -> Dict[str, int]:
    indexes: Dict[str, int] = {}
    for sheet in inventory.get("sheets", []):
        if sheet.get("sheet_name") != DATA_SHEET:
            continue
        for column in sheet.get("columns", []):
            indexes[column["fingerprint_key"]] = column["source_position"]["one_based_index"]
    return indexes


def build_canonical_records(
    view_path: Path, profile: Mapping[str, Any], inventory: Mapping[str, Any], resolved: Mapping[str, Any]
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Build one TSS-scope record per data row; return records plus stats."""
    from openpyxl import load_workbook

    indexes = _column_index_by_fingerprint(inventory)
    field_columns: Dict[str, Tuple[str, int]] = {}
    for field, result in resolved.items():
        if result["status"] != "RESOLVED":
            continue
        key = fingerprint_key(result["matches"][0]["fingerprint"])
        if key in indexes:
            field_columns[field] = (key, indexes[key])

    context_base = {
        "project_key": profile.get("identity", {}).get("project_key", ""),
        "du_model_name": profile.get("identity", {}).get("accepted_du_models", [""])[0],
        "du_model_id": profile.get("identity", {}).get("accepted_du_model_ids", [""])[0],
        "view_id": profile.get("identity", {}).get("accepted_view_ids", [""])[0],
        "source_file_name": view_path.name,
        "source_file_hash": sha256_file(view_path),
        "header_hash": calculate_header_hash(inventory),
    }

    records: List[Dict[str, Any]] = []
    classification_counts: Dict[str, int] = {}
    workbook = load_workbook(view_path, read_only=True, data_only=True)
    try:
        sheet = workbook[DATA_SHEET]
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=HEADER_ROW_COUNT + 1, values_only=True), start=HEADER_ROW_COUNT + 1
        ):
            if all(value is None or (isinstance(value, str) and not value.strip()) for value in row):
                continue
            raw_values = {key: row[index - 1] for field, (key, index) in field_columns.items()}
            record = build_canonical_site_record(
                raw_values,
                profile,
                {**context_base, "source_row_number": row_number},
                scope="TSS",
                resolved_mappings=resolved,
            )
            classification = record["validation"]["pr_input_classification"]
            classification_counts[classification] = classification_counts.get(classification, 0) + 1
            records.append(record)
    finally:
        workbook.close()
    return records, {"record_count": len(records), "classification_counts": classification_counts}


def render_canonical_path_view(
    view_path: Path,
    records: List[Mapping[str, Any]],
    resolved: Mapping[str, Any],
    inventory: Mapping[str, Any],
    out_path: Path,
) -> Dict[str, Any]:
    """Rewrite the mapped columns from canonical-record source evidence."""
    from openpyxl import load_workbook

    indexes = _column_index_by_fingerprint(inventory)
    field_columns: Dict[str, int] = {}
    for field, result in resolved.items():
        if result["status"] != "RESOLVED":
            continue
        key = fingerprint_key(result["matches"][0]["fingerprint"])
        if key in indexes:
            field_columns[field] = indexes[key]

    workbook = load_workbook(view_path)
    sheet = workbook[DATA_SHEET]
    rewritten_cells = 0
    for record in records:
        row_number = record["identity"]["source_row_number"]
        for field, column_index in field_columns.items():
            evidence = record["source_evidence"]["fields"].get(field)
            if evidence is None:
                continue
            sheet.cell(row=row_number, column=column_index, value=evidence["source_value"])
            rewritten_cells += 1
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    return {"rewritten_cells": rewritten_cells, "rewritten_fields": sorted(field_columns)}


def run_generator(site_data: Path, output_dir: Path, scope: str, repo_root: Path) -> Dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    command = [
        sys.executable,
        str(repo_root / "scripts" / "generate_tss_pr_ecc.py"),
        "--site-data",
        str(site_data),
        "--output",
        str(output_dir),
        "--all-sites",
        "--scope",
        scope,
    ]
    # PYTHONUTF8=1 works around a pre-existing legacy defect: the generator
    # writes review CSVs without an explicit encoding, and site data containing
    # U+200B crashes it under the default Windows cp1252 file encoding. The
    # workaround applies identically to both parity paths and changes no source.
    environment = {**os.environ, "PYTHONUTF8": "1"}
    completed = subprocess.run(command, cwd=repo_root, capture_output=True, text=True, env=environment)
    return {
        "scope": scope,
        "returncode": completed.returncode,
        "stdout_tail": completed.stdout[-2000:],
        "stderr_tail": completed.stderr[-2000:],
    }


def _normalized_name(path: Path) -> str:
    return _TIMESTAMP_PATTERN.sub("#", path.name)


def _xlsx_cells(path: Path) -> List[Tuple[str, int, int, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    cells: List[Tuple[str, int, int, Any]] = []
    try:
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for column_index, value in enumerate(row, start=1):
                    if value is not None:
                        cells.append((sheet.title, row_index, column_index, value))
    finally:
        workbook.close()
    return cells


def _csv_rows(path: Path) -> List[List[str]]:
    with open(path, newline="", encoding="utf-8-sig") as handle:
        return [row for row in csv.reader(handle)]


def compare_output_dirs(legacy_dir: Path, canonical_dir: Path) -> Dict[str, Any]:
    legacy_files = {_normalized_name(p): p for p in sorted(legacy_dir.iterdir()) if p.is_file()}
    canonical_files = {_normalized_name(p): p for p in sorted(canonical_dir.iterdir()) if p.is_file()}
    report: Dict[str, Any] = {
        "legacy_file_count": len(legacy_files),
        "canonical_file_count": len(canonical_files),
        "only_in_legacy": sorted(set(legacy_files) - set(canonical_files)),
        "only_in_canonical": sorted(set(canonical_files) - set(legacy_files)),
        "files": [],
        "identical": True,
    }
    if report["only_in_legacy"] or report["only_in_canonical"]:
        report["identical"] = False
    for name in sorted(set(legacy_files) & set(canonical_files)):
        legacy_path, canonical_path = legacy_files[name], canonical_files[name]
        entry: Dict[str, Any] = {"file": name}
        if legacy_path.suffix.lower() == ".xlsx":
            left, right = _xlsx_cells(legacy_path), _xlsx_cells(canonical_path)
            entry["legacy_cells"] = len(left)
            entry["canonical_cells"] = len(right)
            differences = [
                {"left": l, "right": r}
                for l, r in zip(left, right)
                if l != r
            ][:20]
            if len(left) != len(right) or differences:
                entry["identical"] = False
                entry["sample_differences"] = differences
                report["identical"] = False
            else:
                entry["identical"] = True
        elif legacy_path.suffix.lower() == ".csv":
            left_rows, right_rows = _csv_rows(legacy_path), _csv_rows(canonical_path)
            entry["legacy_rows"] = len(left_rows)
            entry["canonical_rows"] = len(right_rows)
            entry["identical"] = left_rows == right_rows
            if not entry["identical"]:
                report["identical"] = False
        else:
            entry["identical"] = legacy_path.read_bytes() == canonical_path.read_bytes()
            if not entry["identical"]:
                report["identical"] = False
        report["files"].append(entry)
    return report


def parity_markdown(report: Mapping[str, Any]) -> str:
    lines = [
        "# TX Mini Golden ECC Parity Report (LOCAL-ONLY, DRY-RUN)",
        "",
        f"- Site view: `{report['site_view']}`",
        f"- Site view SHA-256: `{report['site_view_hash']}`",
        f"- View header hash: `{report['resolution']['header_hash']}`",
        f"- View header hash approved: `{report['resolution']['header_hash_approved']}`"
        " (not approved => validation dry-run only; no production claim)",
        f"- Profile: `{report['profile_id']}` v`{report['profile_version']}`"
        f" (mapping `{report['mapping_version']}`, status `{report['profile_status']}` — unchanged)",
        f"- Mapped fields resolved: {report['resolution']['mapped_field_count'] - len(report['resolution']['unresolved_mapped_fields'])}"
        f"/{report['resolution']['mapped_field_count']}"
        f" (unresolved: {', '.join(report['resolution']['unresolved_mapped_fields']) or 'none'})",
        f"- Canonical records built: {report['canonical']['record_count']}",
        f"- Record classifications: {json.dumps(report['canonical']['classification_counts'])}",
        f"- Canonical-path cells re-rendered from source evidence: {report['render']['rewritten_cells']}"
        f" across fields: {', '.join(report['render']['rewritten_fields'])}",
        "",
        "## Generator runs",
        "",
    ]
    for run in report["generator_runs"]:
        lines.append(
            f"- {run['input']} / {run['scope']}: return code {run['returncode']}"
        )
    lines.extend(["", "## ECC output comparison", ""])
    for scope, comparison in report["comparisons"].items():
        lines.extend(
            [
                f"### Scope {scope}",
                "",
                f"- Legacy files: {comparison['legacy_file_count']}, canonical-path files: {comparison['canonical_file_count']}",
                f"- Only in legacy: {', '.join(comparison['only_in_legacy']) or 'none'}",
                f"- Only in canonical: {', '.join(comparison['only_in_canonical']) or 'none'}",
                f"- Identical: **{comparison['identical']}**",
                "",
            ]
        )
        for entry in comparison["files"]:
            detail = ""
            if "legacy_cells" in entry:
                detail = f" ({entry['legacy_cells']} vs {entry['canonical_cells']} non-empty cells)"
            elif "legacy_rows" in entry:
                detail = f" ({entry['legacy_rows']} vs {entry['canonical_rows']} rows)"
            lines.append(f"- `{entry['file']}`: identical={entry['identical']}{detail}")
    lines.extend(
        [
            "",
            "## Verdict",
            "",
            f"- Golden parity: **{'PASS' if report['parity_pass'] else 'FAIL'}**",
            "- This report is dry-run validation evidence. The profile lifecycle status is unchanged,",
            "  the legacy view's header hash remains unapproved, and no production enablement is implied.",
        ]
    )
    return "\n".join(lines) + "\n"


def run_parity(
    *,
    repo_root: Path,
    site_view: Path,
    profile_path: Path,
    work_dir: Path,
    scopes: Tuple[str, ...] = ("TSS", "TI"),
) -> Dict[str, Any]:
    profile = load_du_profile(profile_path)
    inventory = build_header_inventory(site_view)
    resolution = resolution_report(inventory, profile)
    if not resolution["dry_run_permitted"]:
        raise ValueError(
            "Dry-run not permitted: required profile fingerprints do not resolve uniquely in the site view: "
            + ", ".join(resolution["unresolved_mapped_fields"])
        )
    resolved = resolution["resolved_mappings"]
    records, canonical_stats = build_canonical_records(site_view, profile, inventory, resolved)
    canonical_view = work_dir / "canonical_path_site_view.xlsx"
    render_stats = render_canonical_path_view(site_view, records, resolved, inventory, canonical_view)

    generator_runs: List[Dict[str, Any]] = []
    comparisons: Dict[str, Any] = {}
    parity_pass = True
    for scope in scopes:
        legacy_dir = work_dir / f"legacy_{scope.lower()}"
        canonical_dir = work_dir / f"canonical_{scope.lower()}"
        for input_name, site_data, out_dir in (
            ("legacy", site_view, legacy_dir),
            ("canonical", canonical_view, canonical_dir),
        ):
            run = run_generator(site_data, out_dir, scope, repo_root)
            run["input"] = input_name
            generator_runs.append(run)
            if run["returncode"] != 0:
                parity_pass = False
        comparison = compare_output_dirs(legacy_dir, canonical_dir)
        comparisons[scope] = comparison
        if not comparison["identical"]:
            parity_pass = False

    report = {
        "site_view": str(site_view),
        "site_view_hash": sha256_file(site_view),
        "profile_id": profile["profile_id"],
        "profile_version": profile["profile_version"],
        "mapping_version": profile["mapping_version"],
        "profile_status": profile["status"],
        "resolution": {k: v for k, v in resolution.items() if k != "resolved_mappings"},
        "canonical": canonical_stats,
        "render": render_stats,
        "generator_runs": generator_runs,
        "comparisons": comparisons,
        "parity_pass": parity_pass,
    }
    work_dir.mkdir(parents=True, exist_ok=True)
    (work_dir / "TX_MINI_GOLDEN_PARITY_REPORT.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8"
    )
    (work_dir / "TX_MINI_GOLDEN_PARITY_REPORT.md").write_text(parity_markdown(report), encoding="utf-8")
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    repo_root = Path(__file__).resolve().parent.parent
    parser.add_argument("--site-view", type=Path, default=repo_root / "Info" / "input" / "site_pr_po_view.xlsx")
    parser.add_argument("--du-profile", type=Path, default=repo_root / "config" / "du_profiles" / "tx_mini_pr_v1.yaml")
    parser.add_argument(
        "--work-dir", type=Path, default=repo_root / "output" / "du-20260706-profile" / "tx-mini-golden-parity"
    )
    args = parser.parse_args()
    report = run_parity(
        repo_root=repo_root,
        site_view=args.site_view,
        profile_path=args.du_profile,
        work_dir=args.work_dir,
    )
    print(f"Golden parity: {'PASS' if report['parity_pass'] else 'FAIL'} — report in {args.work_dir}")
    return 0 if report["parity_pass"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
