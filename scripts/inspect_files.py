import openpyxl
import pandas as pd
from pathlib import Path

# File paths
info_folder = Path("Info")
site_data_file = info_folder / "input" / "site_pr_po_view.xlsx"
pr_model_file = info_folder / "input" / "pr_model.xlsx"

print("=" * 80)
print("FILE STRUCTURE ANALYSIS")
print("=" * 80)

# Inspect site data file structure
print("\n[1] SITE DATA FILE SHEETS:")
wb_site = openpyxl.load_workbook(site_data_file)
for sheet_name in wb_site.sheetnames:
    ws = wb_site[sheet_name]
    print(f"\n  Sheet: {sheet_name}")
    print(f"    Rows: {ws.max_row}, Cols: {ws.max_column}")
    print(f"    First 10 columns (row 1):")
    for i in range(1, min(11, ws.max_column + 1)):
        cell_value = ws.cell(1, i).value
        print(f"      Col {i}: {cell_value}")

# Inspect PR model file
print("\n\n[2] PR MODEL FILE SHEETS:")
wb_pr = openpyxl.load_workbook(pr_model_file)
for sheet_name in wb_pr.sheetnames:
    ws = wb_pr[sheet_name]
    print(f"\n  Sheet: {sheet_name}")
    print(f"    Rows: {ws.max_row}, Cols: {ws.max_column}")
    print(f"    First 15 columns (row 1):")
    for i in range(1, min(16, ws.max_column + 1)):
        cell_value = ws.cell(1, i).value
        print(f"      Col {i}: {cell_value}")

print("\n" + "=" * 80)
