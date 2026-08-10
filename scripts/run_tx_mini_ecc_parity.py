import csv
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path
import hashlib
from datetime import datetime

from typing import Any
import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path("scripts")))
from profile_du_export import build_header_inventory
from run_tx_mini_golden_parity import resolution_report, build_canonical_records, render_canonical_path_view
from canonical_generator_bridge import build_records_from_export
from du_profile_loader import load_du_profile

APPROVED_PR_MODEL_SHA256 = "6c4fda502a8998b41bd88704dd6c59d986dc6c46fe42b82947d12c0c0cd8178f"
ALLOWED_NORMALIZATION_CELLS = {("Summary", "A1"), ("Summary", "B1")}


def validate_pr_model(path: Path) -> str:
    """Validate PR model existence and immutable approved SHA-256 hash."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"PR Model not found: {p}")
    actual_sha = hashlib.sha256(p.read_bytes()).hexdigest()
    if actual_sha != APPROVED_PR_MODEL_SHA256:
        raise ValueError(f"PR_MODEL_HASH_MISMATCH: expected {APPROVED_PR_MODEL_SHA256}, got {actual_sha}")
    return actual_sha


def normalize_source_row(val: Any) -> int:
    """Normalize Source Row value to a positive integer (> 0) or raise ValueError."""
    if val is None or isinstance(val, bool):
        raise ValueError(f"Invalid Source Row value: {val!r} (must be a positive integer)")

    if isinstance(val, int):
        if val <= 0:
            raise ValueError(f"Source Row must be positive integer (> 0), got {val}")
        return val

    if isinstance(val, float):
        if val > 0 and val.is_integer():
            return int(val)
        raise ValueError(f"Invalid Source Row float value: {val!r}")

    if isinstance(val, str):
        val_str = val.strip()
        if not val_str:
            raise ValueError("Source Row cannot be blank")
        if val_str.isdigit():
            num = int(val_str)
            if num <= 0:
                raise ValueError(f"Source Row must be positive integer (> 0), got {num}")
            return num

    raise ValueError(f"Invalid non-integer Source Row value: {val!r}")


def validate_candidate_manifest(candidates: Any) -> bool:
    """Validate candidate manifest format, count, required fields, and uniqueness."""
    items = candidates.get("candidates") if isinstance(candidates, dict) and "candidates" in candidates else candidates

    if not isinstance(items, list):
        raise ValueError(f"Candidate manifest payload must be a list, got {type(items).__name__}")
    if len(items) != 12:
        raise ValueError(f"Candidate count must be exactly 12, got {len(items)}")

    cand_ids = []
    source_rows = []

    for i, c in enumerate(items):
        if not isinstance(c, dict):
            raise ValueError(f"Candidate row at index {i} must be a dictionary/object, got {type(c).__name__}")

        if "Candidate ID" not in c:
            raise ValueError(f"Candidate row at index {i} missing 'Candidate ID'")
        cand_id = c["Candidate ID"]
        if cand_id is None or (isinstance(cand_id, str) and not cand_id.strip()):
            raise ValueError(f"Candidate row at index {i} has blank 'Candidate ID'")

        if "Source Row" not in c:
            raise ValueError(f"Candidate row at index {i} missing 'Source Row'")
        norm_source_row = normalize_source_row(c["Source Row"])
        c["Source Row"] = norm_source_row  # Coerce to canonical int in place

        if "Site Code" not in c:
            raise ValueError(f"Candidate row at index {i} missing 'Site Code'")
        site_code = c["Site Code"]
        if site_code is None or (isinstance(site_code, str) and not site_code.strip()):
            raise ValueError(f"Candidate row at index {i} has blank 'Site Code'")

        cand_ids.append(str(cand_id).strip())
        source_rows.append(norm_source_row)

    if len(cand_ids) != len(set(cand_ids)):
        raise ValueError("Duplicate Candidate IDs found in manifest")

    if len(source_rows) != len(set(source_rows)):
        raise ValueError("Duplicate Source Rows found in manifest")

    return True


def _extract_color_attrs(color) -> tuple:
    """Extract comparable color attributes (type, rgb, indexed, theme, tint) from an openpyxl Color."""
    if color is None:
        return (None, None, None, None, None)
    return (
        getattr(color, 'type', None),
        getattr(color, 'rgb', None),
        getattr(color, 'indexed', None),
        getattr(color, 'theme', None),
        getattr(color, 'tint', None),
    )


def derive_expected_candidates(input_file: Path, profile_path: Path, sow_registry: Path, scope_config_path: Path, csv_in: Path) -> list:
    """Derive expected 12-candidate set from current eligibility source.

    Always rebuilds the candidate CSV from build_records_from_export to ensure
    the parity set reflects the current scope config and classifier. A stale
    existing CSV is never accepted as the current candidate set.
    """
    # Always rebuild the CSV from current classifications
    subprocess.run([
        sys.executable, "scripts/build_tx_mini_scope_uat.py",
        "--input", str(input_file),
        "--profile", str(profile_path),
        "--scope-config", str(scope_config_path),
        "--output", str(csv_in.parent)
    ], check=True)

    scope_config_data = json.loads(scope_config_path.read_text(encoding="utf-8"))
    records, _ = build_records_from_export(input_file, profile_path, "TSS", sow_registry, scope_config=scope_config_data.get("scopes", {}))
    real_sites = {r["identity"]["source_row_number"]: r["site"].get("site_code", "") for r in records}

    # Read the freshly-built CSV and cross-check against current classifications
    cands = []
    with open(csv_in, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            sr = normalize_source_row(row["Source Row"])
            cands.append({
                "Candidate ID": f"TXM-TSS-{i+1:03d}",
                "Source Row": sr,
                "Site Code": real_sites.get(sr, row["Masked Site Code"])
            })

    # Verify every CSV candidate source row exists in the classified record set
    # with a matching site code. build_records_from_export returns ALL TSS records,
    # so CSV rows (12 UAT candidates) are a subset.
    for c in cands:
        sr = c["Source Row"]
        if sr not in real_sites:
            raise ValueError(
                f"STALE_CANDIDATE_CSV: CSV candidate source row {sr} "
                f"(Site Code={c['Site Code']}) is not present in the freshly "
                f"classified record set. The candidate CSV must be rebuilt "
                f"from current classifications."
            )
        if real_sites[sr] != c["Site Code"]:
            raise ValueError(
                f"STALE_CANDIDATE_CSV: CSV candidate source row {sr} has "
                f"Site Code={c['Site Code']} but fresh classification has "
                f"Site Code={real_sites[sr]}. The candidate CSV must be "
                f"rebuilt from current classifications."
            )

    return cands


def compute_manifest_identity_hash(candidates: list) -> str:
    """Compute deterministic SHA-256 hash of sorted candidate identity set."""
    norm_identities = sorted([(str(c["Candidate ID"]).strip(), normalize_source_row(c["Source Row"]), str(c["Site Code"]).strip()) for c in candidates])
    raw = json.dumps(norm_identities, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def cross_check_candidate_manifest(cached_manifest_payload: Any, expected_candidates: list) -> bool:
    """Validate cached candidate manifest structurally and cross-check exact identity equality against expected candidate set."""
    validate_candidate_manifest(cached_manifest_payload)

    cached_items = cached_manifest_payload if isinstance(cached_manifest_payload, list) else cached_manifest_payload.get("candidates", [])

    if len(cached_items) != len(expected_candidates):
        raise ValueError(f"MANIFEST_CANDIDATE_COUNT_MISMATCH: cached candidate count ({len(cached_items)}) does not match expected ({len(expected_candidates)})")

    cached_identities = sorted([(str(c["Candidate ID"]).strip(), normalize_source_row(c["Source Row"]), str(c["Site Code"]).strip()) for c in cached_items])
    expected_identities = sorted([(str(c["Candidate ID"]).strip(), normalize_source_row(c["Source Row"]), str(c["Site Code"]).strip()) for c in expected_candidates])

    if cached_identities != expected_identities:
        raise ValueError(f"MANIFEST_CANDIDATE_IDENTITY_MISMATCH: cached manifest identities do not match current candidate set.\nCached: {cached_identities}\nExpected: {expected_identities}")

    return True


def validate_generator_result(res_legacy, leg_files, res_canon, can_files) -> bool:
    """Validate generator return code and output count."""
    legacy_ok = (res_legacy.returncode == 0) and (len(leg_files) == 1)
    canon_ok = (res_canon.returncode == 0) and (len(can_files) == 1)
    if not legacy_ok or not canon_ok:
        return False
    return True


def validate_independent_paths(legacy_input: Path, canonical_input: Path, legacy_output: Path, canonical_output: Path) -> bool:
    """Reject same input paths, same output paths, or non-existent paths."""
    leg_in_res = Path(legacy_input).resolve()
    can_in_res = Path(canonical_input).resolve()
    leg_out_res = Path(legacy_output).resolve()
    can_out_res = Path(canonical_output).resolve()

    if leg_in_res == can_in_res:
        raise ValueError(f"Legacy input path and Canonical input path are identical: {leg_in_res}")
    if leg_out_res == can_out_res:
        raise ValueError(f"Legacy output path and Canonical output path are identical: {leg_out_res}")
    if not leg_in_res.exists():
        raise FileNotFoundError(f"Legacy input missing: {leg_in_res}")
    if not can_in_res.exists():
        raise FileNotFoundError(f"Canonical input missing: {can_in_res}")
    return True


def normalize_allowed_metadata(sheet_name: str, coordinate: str, value: Any) -> str:
    """Only normalize explicitly approved sheet/coordinate metadata cells."""
    if value is None:
        return ""
    val_str = str(value)
    if (sheet_name, coordinate) in ALLOWED_NORMALIZATION_CELLS:
        return val_str.strip()
    return val_str


def compare_cell(cl, cc, sheet_name: str, coord: str) -> list:
    """Compare value, formula, data_type, number_format, font, fill, border, alignment, protection."""
    diffs = []
    if cl.value != cc.value:
        diffs.append(f"value: '{cl.value}' vs '{cc.value}'")
    if cl.data_type != cc.data_type:
        diffs.append(f"data_type: '{cl.data_type}' vs '{cc.data_type}'")
    if cl.number_format != cc.number_format:
        diffs.append(f"number_format: '{cl.number_format}' vs '{cc.number_format}'")

    # Font
    fl, fc = cl.font, cc.font
    if fl and fc:
        if (fl.name, fl.size, fl.bold, fl.italic) != (fc.name, fc.size, fc.bold, fc.italic):
            diffs.append(f"font: ({fl.name}, {fl.size}, {fl.bold}, {fl.italic}) vs ({fc.name}, {fc.size}, {fc.bold}, {fc.italic})")
    elif bool(fl) != bool(fc):
        diffs.append(f"font presence: {bool(fl)} vs {bool(fc)}")

    # Fill — compare fill_type AND foreground/background color attributes
    fill_l, fill_c = cl.fill, cc.fill
    if fill_l and fill_c:
        if fill_l.fill_type != fill_c.fill_type:
            diffs.append(f"fill_type: {fill_l.fill_type} vs {fill_c.fill_type}")
        else:
            # Same fill_type — compare foreground and background color properties
            fg_l = _extract_color_attrs(fill_l.fgColor)
            fg_c = _extract_color_attrs(fill_c.fgColor)
            if fg_l != fg_c:
                diffs.append(f"fill_fgColor: {fg_l} vs {fg_c}")
            bg_l = _extract_color_attrs(fill_l.bgColor)
            bg_c = _extract_color_attrs(fill_c.bgColor)
            if bg_l != bg_c:
                diffs.append(f"fill_bgColor: {bg_l} vs {bg_c}")
    elif bool(fill_l) != bool(fill_c):
        diffs.append(f"fill presence: {bool(fill_l)} vs {bool(fill_c)}")

    # Alignment
    al, ac = cl.alignment, cc.alignment
    if al and ac:
        if (al.horizontal, al.vertical, al.wrap_text) != (ac.horizontal, ac.vertical, ac.wrap_text):
            diffs.append(f"alignment: ({al.horizontal}, {al.vertical}, {al.wrap_text}) vs ({ac.horizontal}, {ac.vertical}, {ac.wrap_text})")
    elif bool(al) != bool(ac):
        diffs.append(f"alignment presence: {bool(al)} vs {bool(ac)}")

    # Protection
    prot_l, prot_c = cl.protection, cc.protection
    if prot_l and prot_c:
        if (prot_l.locked, prot_l.hidden) != (prot_c.locked, prot_c.hidden):
            diffs.append(f"protection: ({prot_l.locked}, {prot_l.hidden}) vs ({prot_c.locked}, {prot_c.hidden})")
    elif bool(prot_l) != bool(prot_c):
        diffs.append(f"protection presence: {bool(prot_l)} vs {bool(prot_c)}")

    # Border
    bl, bc = cl.border, cc.border
    if bl and bc:
        b_l_tuple = (getattr(bl.left, 'style', None), getattr(bl.right, 'style', None), getattr(bl.top, 'style', None), getattr(bl.bottom, 'style', None))
        b_c_tuple = (getattr(bc.left, 'style', None), getattr(bc.right, 'style', None), getattr(bc.top, 'style', None), getattr(bc.bottom, 'style', None))
        if b_l_tuple != b_c_tuple:
            diffs.append(f"border: {b_l_tuple} vs {b_c_tuple}")
    elif bool(bl) != bool(bc):
        diffs.append(f"border presence: {bool(bl)} vs {bool(bc)}")

    return diffs


def compare_workbook_structure(leg_wb, can_wb) -> list:
    """Compare sheet names, count, states, dimensions, merged ranges, freeze panes, auto filter, print area."""
    struct_diffs = []

    if leg_wb.sheetnames != can_wb.sheetnames:
        struct_diffs.append(f"Sheet names differ: legacy {leg_wb.sheetnames} vs canonical {can_wb.sheetnames}")
        return struct_diffs

    for sheet in leg_wb.sheetnames:
        ws_l = leg_wb[sheet]
        ws_c = can_wb[sheet]

        if ws_l.sheet_state != ws_c.sheet_state:
            struct_diffs.append(f"Sheet '{sheet}' state differs: {ws_l.sheet_state} vs {ws_c.sheet_state}")

        if (ws_l.max_row, ws_l.max_column) != (ws_c.max_row, ws_c.max_column):
            struct_diffs.append(f"Sheet '{sheet}' dimensions differ: ({ws_l.max_row}, {ws_l.max_column}) vs ({ws_c.max_row}, {ws_c.max_column})")

        merged_l = set(str(r) for r in ws_l.merged_cells.ranges)
        merged_c = set(str(r) for r in ws_c.merged_cells.ranges)
        if merged_l != merged_c:
            struct_diffs.append(f"Sheet '{sheet}' merged cells differ: {merged_l ^ merged_c}")

        if ws_l.freeze_panes != ws_c.freeze_panes:
            struct_diffs.append(f"Sheet '{sheet}' freeze panes differ: {ws_l.freeze_panes} vs {ws_c.freeze_panes}")

        if getattr(ws_l.auto_filter, 'ref', None) != getattr(ws_c.auto_filter, 'ref', None):
            struct_diffs.append(f"Sheet '{sheet}' auto filter ref differs: {getattr(ws_l.auto_filter, 'ref', None)} vs {getattr(ws_c.auto_filter, 'ref', None)}")

        if getattr(ws_l, 'print_title_rows', None) != getattr(ws_c, 'print_title_rows', None):
            struct_diffs.append(f"Sheet '{sheet}' print_title_rows differ: {ws_l.print_title_rows} vs {ws_c.print_title_rows}")

        # Row dimensions / heights / hidden states
        all_rows = set(ws_l.row_dimensions.keys()) | set(ws_c.row_dimensions.keys())
        for r_idx in all_rows:
            rd_l = ws_l.row_dimensions.get(r_idx)
            rd_c = ws_c.row_dimensions.get(r_idx)
            h_l = rd_l.height if rd_l else None
            h_c = rd_c.height if rd_c else None
            if h_l != h_c:
                struct_diffs.append(f"Sheet '{sheet}' row {r_idx} height differs: {h_l} vs {h_c}")
            hid_l = rd_l.hidden if rd_l else False
            hid_c = rd_c.hidden if rd_c else False
            if hid_l != hid_c:
                struct_diffs.append(f"Sheet '{sheet}' row {r_idx} hidden differs: {hid_l} vs {hid_c}")

        # Column dimensions / widths / hidden states
        all_cols = set(ws_l.column_dimensions.keys()) | set(ws_c.column_dimensions.keys())
        for col_letter in all_cols:
            cd_l = ws_l.column_dimensions.get(col_letter)
            cd_c = ws_c.column_dimensions.get(col_letter)
            w_l = cd_l.width if cd_l else None
            w_c = cd_c.width if cd_c else None
            if w_l != w_c:
                struct_diffs.append(f"Sheet '{sheet}' column {col_letter} width differs: {w_l} vs {w_c}")
            hid_l = cd_l.hidden if cd_l else False
            hid_c = cd_c.hidden if cd_c else False
            if hid_l != hid_c:
                struct_diffs.append(f"Sheet '{sheet}' column {col_letter} hidden differs: {hid_l} vs {hid_c}")

    return struct_diffs


def extract_business_fields(wb) -> dict:
    """Extract explicit ECC workbook business fields with exact names, cells, and values."""
    fields = {}
    ws = wb["details"] if "details" in wb.sheetnames else wb.active

    # 1. Header/Metadata fields
    fields["Site Code"] = {"cell": "A1", "value": str(ws["A1"].value or "").strip()}
    fields["Region"] = {"cell": "B1", "value": str(ws["B1"].value or "").strip()}
    fields["TSS subcontractor"] = {"cell": "C1", "value": str(ws["C1"].value or "").strip()}
    fields["Scope"] = {"cell": "D1", "value": str(ws["D1"].value or "").strip()}

    # 2. Line Items (Row 2 onwards)
    for r in range(2, ws.max_row + 1):
        line_item = str(ws.cell(row=r, column=1).value or "").strip()
        pbom = str(ws.cell(row=r, column=2).value or "").strip()
        desc = str(ws.cell(row=r, column=3).value or "").strip()
        qty = str(ws.cell(row=r, column=4).value or "").strip()
        unit = str(ws.cell(row=r, column=5).value or "").strip()
        price = str(ws.cell(row=r, column=6).value or "").strip()
        total = str(ws.cell(row=r, column=7).value or "").strip()
        col_o_sow = str(ws.cell(row=r, column=15).value or "").strip()

        if line_item or pbom or desc:
            fields[f"LineItem_R{r}_Code"] = {"cell": f"A{r}", "value": line_item}
            fields[f"LineItem_R{r}_PBOM"] = {"cell": f"B{r}", "value": pbom}
            fields[f"LineItem_R{r}_Desc"] = {"cell": f"C{r}", "value": desc}
            fields[f"LineItem_R{r}_Qty"] = {"cell": f"D{r}", "value": qty}
            fields[f"LineItem_R{r}_Unit"] = {"cell": f"E{r}", "value": unit}
            fields[f"LineItem_R{r}_Price"] = {"cell": f"F{r}", "value": price}
            fields[f"LineItem_R{r}_Total"] = {"cell": f"G{r}", "value": total}
            fields[f"LineItem_R{r}_ColumnO_SOW"] = {"cell": f"O{r}", "value": col_o_sow}

    return fields


def compare_business_fields(cand_id: str, leg_fields: dict, can_fields: dict) -> tuple:
    """Compare business fields explicitly and produce detailed CSV records."""
    all_keys = set(leg_fields.keys()) | set(can_fields.keys())
    csv_rows = []
    is_match = True

    for field_name in sorted(all_keys):
        leg_info = leg_fields.get(field_name, {"cell": "N/A", "value": "<MISSING>"})
        can_info = can_fields.get(field_name, {"cell": "N/A", "value": "<MISSING>"})

        match_flag = (leg_info["value"] == can_info["value"])
        if not match_flag:
            is_match = False

        csv_rows.append({
            "Candidate ID": cand_id,
            "Field Name": field_name,
            "Legacy Value": leg_info["value"],
            "Canonical Value": can_info["value"],
            "Match": str(match_flag).upper(),
            "Legacy Cell": leg_info["cell"],
            "Canonical Cell": can_info["cell"]
        })

    return is_match, csv_rows


def calculate_overall_parity(candidate_count, results_count, canonical_success, legacy_success, exact_matches, normalized_matches, business_diffs, struct_diffs, gen_failures) -> str:
    """Calculate overall parity strictly requiring all 12 conditions to pass."""
    overall_pass = (
        candidate_count == 12 and
        results_count == 12 and
        canonical_success == 12 and
        legacy_success == 12 and
        (exact_matches + normalized_matches) == 12 and
        business_diffs == 0 and
        struct_diffs == 0 and
        gen_failures == 0
    )
    return "PASS" if overall_pass else "ECC_PARITY_FAILED"


def run_parity():
    out_dir = Path("output/tx-mini-tss-ecc-parity")
    legacy_dir = out_dir / "legacy_outputs"
    canonical_dir = out_dir / "canonical_outputs"
    out_dir.mkdir(parents=True, exist_ok=True)

    input_file = Path("Info/reference/du_exports/A-P202202168750_D002-TX Mini Project-TX Mini PR_PO View-20260703160246.xlsx")
    profile_path = Path("config/du_profiles/tx_mini_pr_v1.yaml")
    pr_model = Path("Info/input/pr_model.xlsx")
    sow_registry = Path("config/registries/canonical_sow_registry.yaml")
    scope_config = Path("config/scope_eligibility/tx_mini_pr_v1.json")
    csv_in = Path("output/tx-mini-scope-eligibility-uat/TX_MINI_TSS_UAT_CANDIDATES.csv")

    expected_candidates = derive_expected_candidates(input_file, profile_path, sow_registry, scope_config, csv_in)
    expected_identity_hash = compute_manifest_identity_hash(expected_candidates)

    manifest_file = out_dir / "candidate_manifest" / "TX_MINI_TSS_CANDIDATE_MANIFEST.json"
    manifest_file.parent.mkdir(parents=True, exist_ok=True)

    if manifest_file.exists():
        try:
            cached_data = json.loads(manifest_file.read_text(encoding="utf-8"))
            cross_check_candidate_manifest(cached_data, expected_candidates)
        except Exception as err:
            raise ValueError(f"Cached manifest invalid or mismatched against current candidate set: {err}") from err

    manifest_payload = {
        "metadata": {
            "source_candidate_file": str(csv_in),
            "source_candidate_file_hash": hashlib.sha256(csv_in.read_bytes()).hexdigest() if csv_in.exists() else "",
            "identity_hash": expected_identity_hash,
            "candidate_count": len(expected_candidates)
        },
        "candidates": expected_candidates
    }
    manifest_file.write_text(json.dumps(manifest_payload, indent=2), encoding="utf-8")
    candidates = expected_candidates

    validate_candidate_manifest(candidates)
    validate_pr_model(pr_model)

    print("Building canonical view for parity check...")
    profile = load_du_profile(profile_path)
    inventory = build_header_inventory(input_file)
    resolved = resolution_report(inventory, profile)["resolved_mappings"]
    records, stats = build_canonical_records(input_file, profile, inventory, resolved)

    canonical_view = out_dir / "canonical_path_site_view.xlsx"
    render_canonical_path_view(input_file, records, resolved, inventory, canonical_view)
    print(f"Canonical view built at {canonical_view}")

    validate_independent_paths(input_file, canonical_view, legacy_dir, canonical_dir)

    results = []
    cell_differences = []
    structural_differences = []
    business_comparison_rows = []

    canonical_success = 0
    legacy_success = 0
    generator_script = Path("scripts/generate_tss_pr_ecc.py")

    for c in candidates:
        site_code = c["Site Code"]
        cand_id = c["Candidate ID"]
        print(f"Running candidate {cand_id} ({site_code})...")

        legacy_out = legacy_dir / cand_id
        canon_out = canonical_dir / cand_id

        if legacy_out.exists():
            shutil.rmtree(legacy_out)
        if canon_out.exists():
            shutil.rmtree(canon_out)

        legacy_out.mkdir(parents=True, exist_ok=True)
        canon_out.mkdir(parents=True, exist_ok=True)

        cmd_legacy = [
            sys.executable, str(generator_script),
            "--site-data", str(input_file),
            "--pr-model", str(pr_model),
            "--output", str(legacy_out),
            "--site-code", site_code,
            "--scope", "TSS"
        ]
        res_legacy = subprocess.run(cmd_legacy, capture_output=True, text=True)
        leg_files = list(legacy_out.glob("*.xlsx")) if legacy_out.exists() else []

        cmd_canon = [
            sys.executable, str(generator_script),
            "--site-data", str(canonical_view),
            "--pr-model", str(pr_model),
            "--output", str(canon_out),
            "--site-code", site_code,
            "--scope", "TSS"
        ]
        res_canon = subprocess.run(cmd_canon, capture_output=True, text=True)
        can_files = list(canon_out.glob("*.xlsx")) if canon_out.exists() else []

        gen_ok = validate_generator_result(res_legacy, leg_files, res_canon, can_files)
        if gen_ok:
            legacy_success += 1
            canonical_success += 1
        else:
            results.append({"Candidate ID": cand_id, "Site Code": site_code, "Parity Classification": "GENERATION_FAILED"})
            structural_differences.append({
                "Candidate ID": cand_id,
                "Difference": f"Generation failed. Legacy code={res_legacy.returncode}, files={len(leg_files)}. Canon code={res_canon.returncode}, files={len(can_files)}.\nLegacy stderr: {res_legacy.stderr[:200]}\nCanon stderr: {res_canon.stderr[:200]}"
            })
            continue

        leg_wb = openpyxl.load_workbook(leg_files[0], data_only=False)
        can_wb = openpyxl.load_workbook(can_files[0], data_only=False)

        struct_diffs = compare_workbook_structure(leg_wb, can_wb)
        if struct_diffs:
            results.append({"Candidate ID": cand_id, "Site Code": site_code, "Parity Classification": "STRUCTURAL_DIFFERENCE"})
            for sd in struct_diffs:
                structural_differences.append({"Candidate ID": cand_id, "Difference": sd})
            continue

        is_exact = True
        is_normalized = True
        has_cell_diff = False

        for sheet in leg_wb.sheetnames:
            ws_l = leg_wb[sheet]
            ws_c = can_wb[sheet]
            max_r = max(ws_l.max_row, ws_c.max_row)
            max_c = max(ws_l.max_column, ws_c.max_column)

            for r in range(1, max_r + 1):
                for c_idx in range(1, max_c + 1):
                    cl = ws_l.cell(row=r, column=c_idx)
                    cc = ws_c.cell(row=r, column=c_idx)
                    coord = get_column_letter(c_idx) + str(r)

                    cell_diffs = compare_cell(cl, cc, sheet, coord)
                    if cell_diffs:
                        is_exact = False
                        is_allowed = False
                        if (sheet, coord) in ALLOWED_NORMALIZATION_CELLS:
                            norm_l = normalize_allowed_metadata(sheet, coord, cl.value)
                            norm_c = normalize_allowed_metadata(sheet, coord, cc.value)
                            only_val_diffs = all(d.startswith("value:") for d in cell_diffs)
                            if only_val_diffs and norm_l == norm_c:
                                is_allowed = True

                        if not is_allowed:
                            is_normalized = False
                            has_cell_diff = True
                            cell_differences.append({
                                "Candidate ID": cand_id,
                                "Sheet": sheet,
                                "Cell": coord,
                                "Legacy": str(cl.value),
                                "Canonical": str(cc.value)
                            })

        leg_biz = extract_business_fields(leg_wb)
        can_biz = extract_business_fields(can_wb)
        biz_match, biz_csv_rows = compare_business_fields(cand_id, leg_biz, can_biz)
        business_comparison_rows.extend(biz_csv_rows)

        if not biz_match:
            results.append({"Candidate ID": cand_id, "Site Code": site_code, "Parity Classification": "BUSINESS_DIFFERENCE"})
        elif has_cell_diff:
            results.append({"Candidate ID": cand_id, "Site Code": site_code, "Parity Classification": "STRUCTURAL_DIFFERENCE"})
        elif is_exact:
            results.append({"Candidate ID": cand_id, "Site Code": site_code, "Parity Classification": "EXACT_MATCH"})
        else:
            results.append({"Candidate ID": cand_id, "Site Code": site_code, "Parity Classification": "MATCH_AFTER_ALLOWED_NORMALIZATION"})

    # Write Results CSVs
    with open(out_dir / "TX_MINI_TSS_ECC_PARITY_RESULTS.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Candidate ID", "Site Code", "Parity Classification"])
        writer.writeheader()
        writer.writerows(results)

    with open(out_dir / "TX_MINI_TSS_ECC_CELL_DIFFERENCES.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Candidate ID", "Sheet", "Cell", "Legacy", "Canonical"])
        writer.writeheader()
        writer.writerows(cell_differences)

    with open(out_dir / "TX_MINI_TSS_ECC_STRUCTURAL_DIFFERENCES.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Candidate ID", "Difference"])
        writer.writeheader()
        writer.writerows(structural_differences)

    with open(out_dir / "TX_MINI_TSS_ECC_BUSINESS_FIELD_COMPARISON.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Candidate ID", "Field Name", "Legacy Value", "Canonical Value", "Match", "Legacy Cell", "Canonical Cell"])
        writer.writeheader()
        writer.writerows(business_comparison_rows)

    model_sha = hashlib.sha256(pr_model.read_bytes()).hexdigest()
    model_id = {
        "actual_filename": pr_model.name,
        "logical_model_identity": "Celcomdigi TX PR Model & Line Item v4",
        "absolute_path": str(pr_model.resolve()),
        "sha256": model_sha,
        "sheet_names": openpyxl.load_workbook(pr_model, data_only=True).sheetnames
    }
    (out_dir / "TX_MINI_TSS_PR_MODEL_IDENTITY.json").write_text(json.dumps(model_id, indent=2))

    counts = {"EXACT_MATCH": 0, "MATCH_AFTER_ALLOWED_NORMALIZATION": 0, "BUSINESS_DIFFERENCE": 0, "STRUCTURAL_DIFFERENCE": 0, "GENERATION_FAILED": 0}
    for r in results:
        counts[r["Parity Classification"]] += 1

    overall = calculate_overall_parity(
        len(candidates),
        len(results),
        canonical_success,
        legacy_success,
        counts["EXACT_MATCH"],
        counts["MATCH_AFTER_ALLOWED_NORMALIZATION"],
        counts["BUSINESS_DIFFERENCE"],
        counts["STRUCTURAL_DIFFERENCE"],
        counts["GENERATION_FAILED"]
    )

    summary = {
        "candidate_count": len(candidates),
        "results_count": len(results),
        "canonical_generation_success_count": canonical_success,
        "legacy_generation_success_count": legacy_success,
        "exact_match_count": counts["EXACT_MATCH"],
        "normalized_match_count": counts["MATCH_AFTER_ALLOWED_NORMALIZATION"],
        "business_difference_count": counts["BUSINESS_DIFFERENCE"],
        "structural_difference_count": counts["STRUCTURAL_DIFFERENCE"],
        "generation_failed_count": counts["GENERATION_FAILED"],
        "overall_parity_result": overall
    }
    (out_dir / "TX_MINI_TSS_ECC_PARITY_SUMMARY.json").write_text(json.dumps(summary, indent=2))

    commit_sha = subprocess.check_output(['git', 'rev-parse', 'HEAD']).decode('utf-8').strip()
    manifest = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "repository_commit_sha": commit_sha,
        "pr_number": 37,
        "business_uat_status": "BUSINESS_UAT_WAIVED_BY_OWNER",
        "owner_waiver_status": "PROCEED_WITH_ACCEPTED_RISK",
        "candidate_count": len(candidates),
        "pr_model_filename": pr_model.name,
        "pr_model_sha256": model_sha,
        "generator_commit_sha": commit_sha,
        "ecc_allowed": False,
        "production_gate": "PROFILE_NOT_PRODUCTION",
        "production_output_created": False,
        "production_submission_invoked": False
    }
    (out_dir / "TX_MINI_TSS_ECC_PARITY_MANIFEST.json").write_text(json.dumps(manifest, indent=2))

    print(f"Overall parity: {overall}")
    return 0 if overall == "PASS" else 1

if __name__ == "__main__":
    sys.exit(run_parity())
