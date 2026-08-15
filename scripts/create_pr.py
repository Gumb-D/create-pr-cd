#!/usr/bin/env python3
"""Official create-pr entrypoint with audit-complete reporting."""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import date, datetime
from collections.abc import Mapping
from pathlib import Path

import create_pr_impl as _impl
from backoffice_tracker import BackofficeTrackerError, load_backoffice_tracker
from backoffice_pr_runtime import build_backoffice_entitlements, load_service_registry
from planning_pr_runtime import (
    partition_planning_records,
    planning_scope_subcontractor,
    validate_planning_candidate_contracts,
)
from pr_model_baseline import PrModelBaselineError, validate_pr_model_baseline
from renderer_reconciliation import (
    collect_renderer_reconciliation,
    snapshot_renderer_artifacts,
    touched_renderer_artifacts,
)


for _name in dir(_impl):
    if not _name.startswith("__"):
        globals().setdefault(_name, getattr(_impl, _name))

_ORIGINAL_PARTITION = _impl._partition_records
_ORIGINAL_RENDERER_ROW = _impl._renderer_row
_ORIGINAL_SCOPE_SUBCONTRACTOR = _impl._scope_subcontractor
_ORIGINAL_VALIDATE_CANDIDATE_CONTRACTS = _impl.validate_candidate_contracts
_ORIGINAL_RENDERER = Path(_impl.RENDERER)
PLANNING_RENDERER = Path(_impl.ROOT) / "scripts" / "planning_ecc_renderer.py"
BACKOFFICE_RENDERER = Path(_impl.ROOT) / "scripts" / "backoffice_ecc_renderer.py"
BACKOFFICE_SERVICE_REGISTRY = Path(_impl.ROOT) / "config" / "backoffice_service_registry.yaml"
BACKOFFICE_REQUIRED_DU_MODELS = frozenset({
    "2023 Celcomdigi BAU",
    "2024 Celcomdigi BAU",
    "CD consolidation 2023",
    "Celcomdigi USP",
    "Jendela TX Migration",
    "MW EOS Swap",
    "TX Mini Project",
    "2023 TX Rollout",
    "ZTE TX MINI",
})
_LAST_PARTITIONS = None

_ANTENNA_EVIDENCE_RENDERER_COLUMNS = (
    "Antenna Evidence Governance",
    "Antenna Size NE Mapping Status",
    "Antenna Size FE Mapping Status",
    "TX SOW Details Mapping Status",
    "NE SOW Details Mapping Status",
    "FE SOW Details Mapping Status",
)
_PLANNING_RENDERER_COLUMNS = (
    "Subcon - Planning",
    "Planning Contract Subcontractor",
    "Planning PBOM Code",
    "Planning SOW",
    "Planning Unit",
    "Planning Quantity",
)

_BACKOFFICE_RENDERER_COLUMNS = (
    "Backoffice Event Code",
    "Backoffice Trigger Date",
    "Backoffice Billing Month",
    "Backoffice PBOM Code",
    "Backoffice Unit",
    "Backoffice Quantity",
    "Backoffice Subcontractor",
    "Backoffice Contract Number",
    "Backoffice Issue Type",
    "Backoffice Warnings",
)
CANONICAL_RENDERER_COLUMNS = tuple(_impl.CANONICAL_RENDERER_COLUMNS) + tuple(
    column
    for column in (*_ANTENNA_EVIDENCE_RENDERER_COLUMNS, *_PLANNING_RENDERER_COLUMNS, *_BACKOFFICE_RENDERER_COLUMNS)
    if column not in _impl.CANONICAL_RENDERER_COLUMNS
)


def parse_args() -> argparse.Namespace:
    """Official CLI parser; Planning is enabled only through this governed wrapper."""
    parser = argparse.ArgumentParser(description="Identify DU Profile, canonicalize iEPMS data, and generate ECC.")
    parser.add_argument("--site-data", required=True, type=Path, help="Original four-header iEPMS export")
    parser.add_argument("--output", required=True, type=Path, help="ECC output directory")
    parser.add_argument("--scope", required=True, choices=["TSS", "TI", "PLANNING", "BACKOFFICE"], type=str.upper)
    parser.add_argument("--site-code", help="Comma-separated site codes")
    parser.add_argument("--all-sites", action="store_true")
    parser.add_argument("--pr-model", type=Path, default=Path(_impl.ROOT) / "Info" / "input" / "pr_model.xlsx")
    parser.add_argument("--template", type=Path, default=Path(_impl.ROOT) / "Info" / "input" / "ecc_template.xls")
    parser.add_argument("--mapping", type=Path, default=Path(_impl.ROOT) / "Info" / "input" / "contract_info_reference.md")
    parser.add_argument(
        "--subcontractor-policy",
        type=Path,
        default=_impl.PR_POLICY_PATH,
        help="Approved fail-closed subcontractor PR policy JSON",
    )
    parser.add_argument("--backoffice-tracker", type=Path, help="Authoritative TX Outsource & NOC Database.xls for BACKOFFICE duplicate/tier governance")
    parser.add_argument("--billing-month", help="BACKOFFICE billing month in YYYY-MM")
    parser.add_argument(
        "--non-production-uat",
        action="store_true",
        help=(
            "Explicitly generate visibly isolated non-production UAT ECC output "
            "for PR_INPUT_READY or PRODUCTION profiles."
        ),
    )
    return parser.parse_args()


def _renderer_for_scope(scope):
    """Route Planning only to the deterministic Planning renderer."""
    scope_name = str(scope).strip().upper()
    if scope_name == "PLANNING":
        return PLANNING_RENDERER
    if scope_name == "BACKOFFICE":
        return BACKOFFICE_RENDERER
    return _ORIGINAL_RENDERER



def _validate_backoffice_cadence(billing_month, tracker_snapshot, today=None):
    current = today or date.today()
    try:
        month = datetime.strptime(str(billing_month), "%Y-%m").date().replace(day=1)
    except ValueError as error:
        raise CreatePrError("BACKOFFICE_BILLING_MONTH_INVALID", "Backoffice billing month must use YYYY-MM.") from error
    current_month = current.replace(day=1)
    if month >= current_month:
        raise CreatePrError("BACKOFFICE_BILLING_MONTH_NOT_CLOSED", "Backoffice production issuance requires a closed billing month.")
    if str(billing_month) in tracker_snapshot.month_pbom:
        return "SUPPLEMENTARY"
    if current_month.month == 1:
        previous = date(current_month.year - 1, 12, 1)
    else:
        previous = date(current_month.year, current_month.month - 1, 1)
    if month != previous:
        raise CreatePrError("BACKOFFICE_MAIN_BILLING_MONTH_NOT_PREVIOUS", "A Backoffice Main PR can only be issued for the immediately previous calendar month.")
    return "MAIN"


def _backoffice_source_files(path, issue_type):
    source = Path(path)
    if source.is_file():
        if str(issue_type).upper() == "MAIN":
            raise CreatePrError("BACKOFFICE_MAIN_REQUIRES_SOURCE_DIRECTORY", "Backoffice Main PR requires a source directory so all supported DU exports can be aggregated before tier selection.")
        return [source]
    if not source.is_dir():
        raise CreatePrError("BACKOFFICE_SOURCE_NOT_FOUND", f"Backoffice source path not found: {source}")
    allowed = {".xlsx", ".xlsm", ".csv"}
    files = sorted(p for p in source.iterdir() if p.is_file() and p.suffix.lower() in allowed and not p.name.startswith("~$"))
    if not files:
        raise CreatePrError("BACKOFFICE_SOURCE_DIRECTORY_EMPTY", "Backoffice source directory contains no supported iEPMS exports.")
    return files

def _validate_backoffice_main_du_coverage(metadata):
    observed = {
        str(item.get("du_model_name", "") or "").strip()
        for item in metadata
        if isinstance(item, Mapping)
    }
    missing = sorted(BACKOFFICE_REQUIRED_DU_MODELS - observed)
    if missing:
        raise CreatePrError(
            "BACKOFFICE_MAIN_DU_SET_INCOMPLETE",
            "Backoffice Main issuance requires the complete supported DU-model export set before the monthly PBOM tier is calculated.",
            {"missing_du_models": missing, "observed_du_models": sorted(observed)},
        )


def _canonical_relocate_site_id(value):
    """Render the approved Decom - Relo Site ID without changing source identity."""
    text = str(value or "").strip()
    match = re.match(r"^(.*?)[_-]RELOCATE(?:_?\d+)?$", text, flags=re.IGNORECASE)
    if not match:
        return text
    base = match.group(1).rstrip("_-")
    return f"{base}_Relocate" if base else text


def _canonical_source_fields(record):
    """Return canonical source evidence only when the record carries that contract."""
    source_evidence = record.get("source_evidence")
    if not isinstance(source_evidence, Mapping):
        return None
    fields = source_evidence.get("fields")
    return fields if isinstance(fields, Mapping) else None


def _source_mapping_status(record, canonical_field):
    """Return canonical mapping approval state for renderer-side evidence gates."""
    fields = _canonical_source_fields(record)
    if fields is None:
        return ""
    evidence = fields.get(canonical_field, {})
    if not isinstance(evidence, Mapping):
        return ""
    return str(evidence.get("mapping_status", "") or "").strip().upper()


def _renderer_row(record):
    row = _ORIGINAL_RENDERER_ROW(record)
    canonical_fields = _canonical_source_fields(record)
    planning_selection = record.get("planning_selection", {})
    if not isinstance(planning_selection, Mapping):
        planning_selection = {}
    row.update(
        {
            "Antenna Evidence Governance": (
                "CANONICAL_MAPPING_STATUS" if canonical_fields is not None else ""
            ),
            "Antenna Size NE Mapping Status": _source_mapping_status(record, "antenna_size_ne"),
            "Antenna Size FE Mapping Status": _source_mapping_status(record, "antenna_size_fe"),
            "TX SOW Details Mapping Status": _source_mapping_status(record, "tx_sow_details"),
            "NE SOW Details Mapping Status": _source_mapping_status(record, "ne_sow_details"),
            "FE SOW Details Mapping Status": _source_mapping_status(record, "fe_sow_details"),
            "Subcon - Planning": record.get("pr_context", {}).get("subcontractor_planning", ""),
            "Planning Contract Subcontractor": planning_selection.get("contract_subcontractor", ""),
            "Planning PBOM Code": planning_selection.get("pbom_code", ""),
            "Planning SOW": planning_selection.get("description", ""),
            "Planning Unit": planning_selection.get("unit", ""),
            "Planning Quantity": planning_selection.get("quantity", ""),
            "Backoffice Event Code": record.get("backoffice_selection", {}).get("event_code", ""),
            "Backoffice Trigger Date": record.get("backoffice_selection", {}).get("trigger_date", ""),
            "Backoffice Billing Month": record.get("backoffice_selection", {}).get("billing_month", ""),
            "Backoffice PBOM Code": record.get("backoffice_selection", {}).get("pbom_code", ""),
            "Backoffice Unit": record.get("backoffice_selection", {}).get("unit", ""),
            "Backoffice Quantity": record.get("backoffice_selection", {}).get("quantity", ""),
            "Backoffice Subcontractor": record.get("backoffice_selection", {}).get("subcontractor", ""),
            "Backoffice Contract Number": record.get("backoffice_selection", {}).get("contract_number", ""),
            "Backoffice Issue Type": record.get("backoffice_selection", {}).get("issue_type", ""),
            "Backoffice Warnings": " | ".join(record.get("backoffice_selection", {}).get("warnings", []) or []),
        }
    )
    sow = str(record.get("pr_context", {}).get("tx_sow_normalized", "") or "").strip().upper()
    if sow == "DECOM - RELO":
        row["customer site code"] = _canonical_relocate_site_id(row.get("customer site code", ""))
    return row


def _scope_subcontractor(record, scope):
    """Return the scope-specific source subcontractor without cross-scope fallback."""
    if str(scope).strip().upper() == "PLANNING":
        return planning_scope_subcontractor(record)
    return _ORIGINAL_SCOPE_SUBCONTRACTOR(record, scope)


def validate_candidate_contracts(records, scope, contract_mappings):
    """Reuse shared contracts while normalizing Planning `_AA` only for lookup."""
    if str(scope).strip().upper() == "PLANNING":
        return validate_planning_candidate_contracts(records, contract_mappings)
    return _ORIGINAL_VALIDATE_CANDIDATE_CONTRACTS(records, scope, contract_mappings)


def _record_site_code(record):
    return str(record.get("site", {}).get("site_code", "") or "").strip()


def _record_project_key(record):
    return str(record.get("identity", {}).get("project_key", "") or "").strip()


def _assert_unique_project_site_codes(records):
    """Fail closed if the Project + Site Code business identity is duplicated."""
    seen = {}
    duplicates = {}
    for record in records:
        project_key = _record_project_key(record)
        site_code = _record_site_code(record)
        if not site_code:
            continue
        key = (project_key.casefold(), site_code.upper())
        source_row = record.get("identity", {}).get("source_row_number")
        if key not in seen:
            seen[key] = {
                "project_key": project_key,
                "site_code": site_code.upper(),
                "source_rows": [source_row] if source_row is not None else [],
            }
            continue
        duplicate = duplicates.setdefault(key, dict(seen[key]))
        if source_row is not None:
            duplicate.setdefault("source_rows", []).append(source_row)

    if duplicates:
        details = sorted(
            duplicates.values(),
            key=lambda item: (str(item.get("project_key", "")).casefold(), str(item.get("site_code", ""))),
        )
        raise CreatePrError(
            "DUPLICATE_SITE_CODE_IN_PROJECT",
            "Site Code must be unique within a Project before PR generation.",
            {
                "duplicates": details,
                "required_action": "Correct the duplicate Project + Site Code records in the source export before rerunning create-pr.",
            },
        )


def _partition_records(records, scope, policy=None):
    """Validate business identity, then apply the scope-specific partition flow."""
    global _LAST_PARTITIONS
    _assert_unique_project_site_codes(records)
    if str(scope).strip().upper() == "PLANNING":
        _LAST_PARTITIONS = partition_planning_records(records)
    else:
        _LAST_PARTITIONS = _ORIGINAL_PARTITION(records, scope, policy)
    return _LAST_PARTITIONS



def _validate_backoffice_source_identity(records):
    """Backoffice duplicate identity is governed later as Delivery Unit Code + canonical event.

    Site ID is intentionally not a uniqueness key for this scope.
    """
    return None


def _canonicalize_backoffice_sources(sources):
    all_records = []
    metadata = []
    for source in sources:
        resolution = resolve_du_profile(
            Path(source),
            profile_root=_impl.PROFILE_ROOT,
            identity_registry_path=_impl.IDENTITY_REGISTRY,
            scope="BACKOFFICE",
        )
        scope_status = str(resolution["profile"].get("scope_status", {}).get("BACKOFFICE", "")).strip().upper()
        if scope_status != "PRODUCTION":
            raise CreatePrError(
                "BACKOFFICE_PROFILE_SCOPE_NOT_PRODUCTION",
                f"DU Profile Backoffice scope is {scope_status or '(blank)'}; production Backoffice generation is blocked.",
                {"source": str(source), "scope_status": scope_status},
            )
        records, item_metadata = build_canonical_records(
            input_path=Path(source),
            profile=resolution["profile"],
            inventory=resolution["inventory"],
            header_hash=resolution["header_hash"],
            scope="BACKOFFICE",
            sow_registry_path=_impl.SOW_REGISTRY,
        )
        all_records.extend(records)
        metadata.append(item_metadata)
    return all_records, metadata


def _backoffice_identity_key(record):
    site = record.get("site", {})
    selection = record.get("backoffice_selection", {})
    du_code = str(site.get("du_key") or site.get("delivery_unit_code") or "").strip().upper()
    event_code = str(selection.get("event_code") or "").strip().upper()
    return f"{du_code}|{event_code}" if du_code and event_code else ""


def _build_backoffice_reconciliation(selected, partitions, renderer_reconciliation=None):
    """Reconcile Backoffice records by governed Delivery Unit Code + canonical event identity."""
    renderer_by_key = {
        str(item.get("identity_key", "")).strip().upper(): dict(item)
        for item in (renderer_reconciliation or {}).get("record_dispositions", [])
        if str(item.get("identity_key", "")).strip()
    }
    direct_by_object = {}
    for bucket, disposition in (
        ("review_required", "REVIEW_REQUIRED"),
        ("ignored", "IGNORED_WITH_APPROVED_REASON"),
        ("duplicates", "DUPLICATE_BLOCKED"),
    ):
        for record in partitions.get(bucket, []):
            decision = record.get("pr_generation_decision", {})
            direct_by_object[id(record)] = {
                "site_code": _record_site_code(record),
                "identity_key": _backoffice_identity_key(record),
                "disposition": disposition,
                "reason_code": decision.get("reason_code", ""),
                "reason": decision.get("reason", ""),
            }

    candidate_ids = {id(record) for record in partitions.get("candidates", [])}
    record_dispositions = []
    for record in selected:
        terminal = direct_by_object.get(id(record))
        if terminal is None and id(record) in candidate_ids:
            key = _backoffice_identity_key(record).upper()
            terminal = renderer_by_key.get(key)
            if terminal is None:
                terminal = {
                    "site_code": _record_site_code(record),
                    "identity_key": key,
                    "disposition": "FAILED",
                    "reason_code": "RENDERER_BACKOFFICE_RECORD_UNACCOUNTED",
                    "reason": "The Backoffice candidate returned no Delivery Unit + Event renderer evidence.",
                }
        if terminal is None:
            terminal = {
                "site_code": _record_site_code(record),
                "identity_key": _backoffice_identity_key(record),
                "disposition": "FAILED",
                "reason_code": "ENGINE_BACKOFFICE_RECORD_UNACCOUNTED",
                "reason": "The selected Backoffice record did not enter any terminal engine partition.",
            }
        record_dispositions.append(dict(terminal))

    counts = {name: 0 for name in (
        "GENERATED", "REVIEW_REQUIRED", "IGNORED_WITH_APPROVED_REASON", "DUPLICATE_BLOCKED", "FAILED"
    )}
    for item in record_dispositions:
        disposition = str(item.get("disposition", ""))
        if disposition in counts:
            counts[disposition] += 1
    accounted = sum(counts.values())
    return {
        "requested_count": len(selected),
        "generated_count": counts["GENERATED"],
        "review_required_count": counts["REVIEW_REQUIRED"],
        "approved_ignored_count": counts["IGNORED_WITH_APPROVED_REASON"],
        "duplicate_blocked_count": counts["DUPLICATE_BLOCKED"],
        "failed_count": counts["FAILED"],
        "unaccounted_count": max(0, len(selected) - accounted),
        "record_dispositions": record_dispositions,
    }


def _collect_backoffice_renderer_reconciliation(output, candidates, created_paths):
    """Read Backoffice ECC evidence using Delivery Unit Code + Event, never Site ID."""
    from openpyxl import load_workbook
    generated = set()
    for raw_path in created_paths or []:
        path = Path(raw_path)
        if not path.exists() or path.suffix.lower() != ".xlsx" or " PR " not in path.name.upper():
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if "details" not in workbook.sheetnames:
                continue
            worksheet = workbook["details"]
            header = [str(value or "").strip() for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            try:
                du_index = header.index("Delivery Unit Code*")
                remarks_index = header.index("Remarks")
            except ValueError:
                continue
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                du = str(row[du_index] or "").strip().upper() if du_index < len(row) else ""
                remarks = str(row[remarks_index] or "").strip() if remarks_index < len(row) else ""
                event = remarks.split(";", 1)[0].strip().upper()
                if du and event:
                    generated.add(f"{du}|{event}")
        finally:
            workbook.close()

    dispositions = []
    for record in candidates:
        key = _backoffice_identity_key(record).upper()
        if key and key in generated:
            dispositions.append({
                "site_code": _record_site_code(record),
                "identity_key": key,
                "disposition": "GENERATED",
                "reason_code": "ECC_GENERATED",
                "ecc_evidence_present": True,
            })
        else:
            dispositions.append({
                "site_code": _record_site_code(record),
                "identity_key": key,
                "disposition": "FAILED",
                "reason_code": "RENDERER_BACKOFFICE_RECORD_UNACCOUNTED",
                "ecc_evidence_present": False,
            })
    return {"record_dispositions": dispositions}


def _allocate_backoffice_summary_path(output: Path, billing_month: str, issue_type: str, issued_on: date | None = None) -> Path:
    issued_on = issued_on or date.today()
    base = f"CREATE_PR_SUMMARY_BACKOFFICE_{billing_month}_{issue_type}_{issued_on.strftime('%Y%m%d')}"
    first = output / f"{base}.json"
    if not first.exists():
        return first
    batch = 2
    while True:
        candidate = output / f"{base} Batch {batch}.json"
        if not candidate.exists():
            return candidate
        batch += 1


def _write_backoffice_summary(summary: Mapping[str, object], output: Path) -> None:
    payload = json.dumps(summary, ensure_ascii=False, indent=2)
    Path(summary["summary_path"]).write_text(payload, encoding="utf-8")
    (output / "CREATE_PR_SUMMARY_BACKOFFICE.json").write_text(payload, encoding="utf-8")


def _write_declared_summary(summary: Mapping[str, object]) -> None:
    Path(summary["summary_path"]).write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _backoffice_governance_summary(partitions):
    runtime_summary = partitions.get("summary", {}) if isinstance(partitions, Mapping) else {}
    candidates = partitions.get("candidates", []) if isinstance(partitions, Mapping) else []
    warning_distribution = {}
    warning_count = 0
    for record in candidates:
        selection = record.get("backoffice_selection", {}) if isinstance(record, Mapping) else {}
        warnings = selection.get("warnings", []) if isinstance(selection, Mapping) else []
        for warning in warnings if isinstance(warnings, (list, tuple, set)) else []:
            code = str(warning or "").strip()
            if not code:
                continue
            warning_count += 1
            warning_distribution[code] = warning_distribution.get(code, 0) + 1
    return {
        "eligible_hops": runtime_summary.get("eligible_hops", len(candidates)),
        "tier_source": runtime_summary.get("tier_source", ""),
        "warning_count": warning_count,
        "warning_distribution": dict(sorted(warning_distribution.items())),
    }


def _run_backoffice(parsed, baseline):
    global _LAST_PARTITIONS
    if bool(getattr(parsed, "non_production_uat", False)):
        raise CreatePrError("BACKOFFICE_PRODUCTION_ONLY", "Operation Backoffice PR is governed as production-only in Issue #94.")
    tracker_path = getattr(parsed, "backoffice_tracker", None)
    billing_month = str(getattr(parsed, "billing_month", "") or "").strip()
    if not tracker_path:
        raise CreatePrError("BACKOFFICE_TRACKER_REQUIRED", "--backoffice-tracker is required for Backoffice duplicate and tier governance.")
    if not billing_month:
        raise CreatePrError("BACKOFFICE_BILLING_MONTH_REQUIRED", "--billing-month YYYY-MM is required for Backoffice generation.")
    try:
        tracker_snapshot = load_backoffice_tracker(Path(tracker_path))
    except BackofficeTrackerError as error:
        raise CreatePrError(error.code, str(error)) from error
    issue_type = _validate_backoffice_cadence(billing_month, tracker_snapshot)
    if issue_type == "MAIN" and (not bool(getattr(parsed, "all_sites", False)) or str(getattr(parsed, "site_code", "") or "").strip()):
        raise CreatePrError(
            "BACKOFFICE_MAIN_REQUIRES_ALL_SITES",
            "Backoffice Main issuance must use --all-sites so the monthly cross-DU PBOM tier is complete.",
        )
    sources = _backoffice_source_files(parsed.site_data, issue_type)
    records, metadata = _canonicalize_backoffice_sources(sources)
    if issue_type == "MAIN":
        _validate_backoffice_main_du_coverage(metadata)
    selected = _impl._select_records(records, _impl._parse_site_codes(parsed.site_code), parsed.all_sites)
    _validate_backoffice_source_identity(selected)
    service_registry = load_service_registry(BACKOFFICE_SERVICE_REGISTRY)
    partitions = build_backoffice_entitlements(selected, billing_month, tracker_snapshot, service_registry)
    _LAST_PARTITIONS = partitions

    requested_output = Path(parsed.output).resolve()
    requested_output.mkdir(parents=True, exist_ok=True)
    review_path = _impl._write_review_report(requested_output, "BACKOFFICE", partitions["review_required"], RUN_MODE_PRODUCTION)
    duplicate_path = _write_ignored_report(requested_output, "BACKOFFICE_DUPLICATES", partitions["duplicates"], RUN_MODE_PRODUCTION)
    ignored_path = _write_ignored_report(requested_output, "BACKOFFICE", partitions["ignored"], RUN_MODE_PRODUCTION)
    if partitions["review_required"]:
        raise CreatePrError(
            "BACKOFFICE_REVIEW_REQUIRED",
            "Backoffice generation is blocked because one or more records require review; no partial ECC is permitted.",
            {
                "review_required_count": len(partitions["review_required"]),
                "review_required_reason_distribution": reason_distribution(partitions["review_required"]),
                "review_report": str(review_path.resolve()) if review_path else None,
            },
        )
    before_renderer = snapshot_renderer_artifacts(requested_output)

    if partitions["candidates"]:
        import subprocess
        import tempfile
        with tempfile.TemporaryDirectory(prefix="create-pr-backoffice-") as temp_dir:
            canonical_input = Path(temp_dir) / "canonical_backoffice_input.xlsx"
            _impl._write_renderer_input(canonical_input, partitions["candidates"])
            command = [
                sys.executable, str(BACKOFFICE_RENDERER),
                "--site-data", str(canonical_input),
                "--pr-model", str(parsed.pr_model),
                "--template", str(parsed.template),
                "--mapping", str(parsed.mapping),
                "--output", str(requested_output),
                "--scope", "BACKOFFICE",
                "--all-sites",
                "--du-model-name", "MULTI_DU_BACKOFFICE",
            ]
            result = subprocess.run(command, cwd=_impl.ROOT, text=True, encoding="utf-8", errors="replace", capture_output=True, check=False)
            if result.stdout:
                _safe_print(result.stdout, end="")
            if result.stderr:
                _safe_print(result.stderr, file=sys.stderr, end="")
            if result.returncode != 0:
                raise CreatePrError("ECC_RENDERER_FAILED", "Validated Backoffice candidates could not be rendered to ECC.", {"exit_code": result.returncode})

    touched = touched_renderer_artifacts(requested_output, before_renderer)
    renderer = _collect_backoffice_renderer_reconciliation(
        requested_output,
        partitions["candidates"],
        touched,
    )
    reconciliation = _build_backoffice_reconciliation(selected, partitions, renderer)
    _assert_reconciliation_success(reconciliation)
    summary = {
        "status": "SUCCESS",
        "entrypoint": "create_pr.py",
        "run_mode": RUN_MODE_PRODUCTION,
        "scope": "BACKOFFICE",
        "billing_month": billing_month,
        "issue_type": partitions["summary"]["issue_type"],
        "pbom_code": partitions["summary"]["pbom_code"],
        **_backoffice_governance_summary(partitions),
        "source_files": [str(path.resolve()) for path in sources],
        "source_file_count": len(sources),
        "source_record_count": len(records),
        "selected_record_count": len(selected),
        "candidate_count": len(partitions["candidates"]),
        "duplicate_count": len(partitions["duplicates"]),
        "ignored_count": len(partitions["ignored"]),
        "review_required_count": len(partitions["review_required"]),
        "review_required_reason_distribution": reason_distribution(partitions["review_required"]),
        "review_report": str(review_path.resolve()) if review_path else None,
        "duplicate_report": str(duplicate_path.resolve()) if duplicate_path else None,
        "ignored_report": str(ignored_path.resolve()) if ignored_path else None,
        "tracker": str(Path(tracker_path).resolve()),
        "service_registry": str(BACKOFFICE_SERVICE_REGISTRY.resolve()),
        "output_root": str(requested_output),
        "created_files": [str(path.resolve()) for path in touched if path.suffix.lower() == ".xlsx"],
        "profiles": metadata,
        "pr_model_baseline": {"baseline_id": baseline["baseline_id"], "version": baseline["version"], "sha256": baseline["actual_sha256"]},
        **reconciliation,
    }
    summary_path = _allocate_backoffice_summary_path(
        requested_output, billing_month, partitions["summary"]["issue_type"]
    )
    summary["summary_path"] = str(summary_path.resolve())
    _write_backoffice_summary(summary, requested_output)
    return summary

def _build_site_reconciliation(selected, partitions, renderer_reconciliation=None):
    """Assign exactly one terminal engine disposition to every selected site."""
    renderer_by_site = {
        str(item.get("site_code", "")).strip().upper(): dict(item)
        for item in (renderer_reconciliation or {}).get("site_dispositions", [])
        if str(item.get("site_code", "")).strip()
    }
    direct = {}
    for bucket, disposition in (
        ("review_required", "REVIEW_REQUIRED"),
        ("ignored", "IGNORED_WITH_APPROVED_REASON"),
        ("duplicates", "DUPLICATE_BLOCKED"),
    ):
        for record in partitions.get(bucket, []):
            code = _record_site_code(record)
            decision = record.get("pr_generation_decision", {})
            direct[code.upper()] = {
                "site_code": code,
                "disposition": disposition,
                "reason_code": decision.get("reason_code", ""),
                "reason": decision.get("reason", ""),
            }

    candidate_codes = {_record_site_code(record).upper() for record in partitions.get("candidates", [])}
    site_dispositions = []
    for record in selected:
        source_code = _record_site_code(record)
        key = source_code.upper()
        terminal = direct.get(key)
        if terminal is None and key in candidate_codes:
            terminal = renderer_by_site.get(key)
            if terminal is None:
                terminal = {
                    "site_code": source_code,
                    "disposition": "FAILED",
                    "reason_code": "RENDERER_SITE_UNACCOUNTED",
                    "reason": "The site entered the renderer candidate set but returned no terminal renderer disposition.",
                }
        if terminal is None:
            terminal = {
                "site_code": source_code,
                "disposition": "FAILED",
                "reason_code": "ENGINE_SITE_UNACCOUNTED",
                "reason": "The selected site did not enter any terminal engine partition.",
            }
        site_dispositions.append(terminal)

    counts = {name: 0 for name in (
        "GENERATED", "REVIEW_REQUIRED", "IGNORED_WITH_APPROVED_REASON", "DUPLICATE_BLOCKED", "FAILED"
    )}
    for item in site_dispositions:
        disposition = str(item.get("disposition", ""))
        if disposition in counts:
            counts[disposition] += 1
    accounted = sum(counts.values())
    return {
        "requested_count": len(selected),
        "generated_count": counts["GENERATED"],
        "review_required_count": counts["REVIEW_REQUIRED"],
        "approved_ignored_count": counts["IGNORED_WITH_APPROVED_REASON"],
        "duplicate_blocked_count": counts["DUPLICATE_BLOCKED"],
        "failed_count": counts["FAILED"],
        "unaccounted_count": max(0, len(selected) - accounted),
        "site_dispositions": site_dispositions,
    }


def _assert_reconciliation_success(reconciliation):
    """Fail closed only for engine-unaccounted outcomes, not valid review terminals."""
    failed_count = int(reconciliation.get("failed_count", 0) or 0)
    unaccounted_count = int(reconciliation.get("unaccounted_count", 0) or 0)
    if failed_count or unaccounted_count:
        failed_sites = [
            item
            for item in reconciliation.get("site_dispositions", [])
            if str(item.get("disposition", "")).upper() == "FAILED"
        ]
        raise CreatePrError(
            "PR_SITE_RECONCILIATION_FAILED",
            "PR generation did not produce a valid terminal engine outcome for every requested site.",
            {
                "failed_count": failed_count,
                "unaccounted_count": unaccounted_count,
                "failed_sites": failed_sites,
                "required_action": "Review renderer/engine failure evidence before rerunning create-pr.",
            },
        )


def _write_ignored_report(output, scope, records, run_mode=RUN_MODE_PRODUCTION):
    """Write one auditable row for every ignored record, including SM."""
    if not records:
        return None
    marker = f"_{UAT_MARKER}" if run_mode == RUN_MODE_NON_PRODUCTION_UAT else ""
    path = Path(output) / f"CANONICAL_IGNORED_{scope}{marker}.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_REPORT_FIELDS)
        writer.writeheader()
        for record in records:
            writer.writerow(_impl._report_row(record, scope))
    return path


def _sync_dependencies() -> None:
    """Propagate public test seams and audit/Planning wrappers."""
    for name in (
        "resolve_du_profile",
        "build_canonical_records",
        "load_subcontractor_policy",
        "load_contract_reference",
        "validate_candidate_contracts",
    ):
        public_value = globals().get(name)
        if public_value is not None:
            setattr(_impl, name, public_value)
    _impl.CANONICAL_RENDERER_COLUMNS = CANONICAL_RENDERER_COLUMNS
    _impl._partition_records = _partition_records
    _impl._renderer_row = _renderer_row
    _impl._scope_subcontractor = _scope_subcontractor


def _reconcile_summary(summary, scope, before_renderer=None):
    partitions = _LAST_PARTITIONS or {
        "candidates": [], "duplicates": [], "ignored": [], "review_required": []
    }
    selected = []
    for bucket in ("candidates", "duplicates", "ignored", "review_required"):
        selected.extend(partitions.get(bucket, []))

    touched = touched_renderer_artifacts(
        Path(summary["output_root"]),
        before_renderer or {},
    )
    renderer = collect_renderer_reconciliation(
        Path(summary["output_root"]),
        partitions.get("candidates", []),
        scope,
        lambda record: _renderer_row(record).get("customer site code", ""),
        created_paths=touched,
    )
    return _build_site_reconciliation(selected, partitions, renderer)


def _persist_reconciliation_artifact_error(summary, error):
    """Replace a previously successful summary when reconciliation evidence cannot be read."""
    diagnostic = {
        "exception_type": type(error).__name__,
        "message": str(error),
    }
    summary["status"] = "ERROR"
    summary["code"] = "PR_RECONCILIATION_ARTIFACT_READ_FAILED"
    summary["reconciliation_error"] = diagnostic
    _write_declared_summary(summary)
    raise CreatePrError(
        "PR_RECONCILIATION_ARTIFACT_READ_FAILED",
        "Renderer output could not be read for site reconciliation.",
        {
            "summary_path": summary.get("summary_path"),
            "reconciliation_error": diagnostic,
            "required_action": "Inspect the renderer output artifact for corruption or incomplete write, then rerun create-pr.",
        },
    ) from error


def run(parsed):
    """Run the implementation only after validating the single approved PR Model baseline."""
    global _LAST_PARTITIONS
    _LAST_PARTITIONS = None
    baseline = validate_pr_model_baseline(getattr(parsed, "pr_model", None))
    if not hasattr(parsed, "pr_model"):
        parsed.pr_model = baseline["path"]
    _sync_dependencies()
    if str(parsed.scope).strip().upper() == "BACKOFFICE":
        return _run_backoffice(parsed, baseline)
    _impl.RENDERER = _renderer_for_scope(parsed.scope)
    before_renderer = snapshot_renderer_artifacts(Path(parsed.output))
    summary = _impl.run(parsed)
    summary["pr_model_baseline"] = {
        "baseline_id": baseline["baseline_id"],
        "version": baseline["version"],
        "sha256": baseline["actual_sha256"],
    }
    ignored_records = (_LAST_PARTITIONS or {}).get("ignored", [])
    ignored_path = _write_ignored_report(
        Path(summary["output_root"]), parsed.scope, ignored_records, summary["run_mode"]
    )
    summary["ignored_report"] = str(ignored_path.resolve()) if ignored_path else None
    try:
        reconciliation = _reconcile_summary(summary, parsed.scope, before_renderer)
    except Exception as error:
        _persist_reconciliation_artifact_error(summary, error)

    summary.update(reconciliation)

    if reconciliation.get("failed_count", 0) or reconciliation.get("unaccounted_count", 0):
        summary["status"] = "ERROR"
        summary["code"] = "PR_SITE_RECONCILIATION_FAILED"

    _write_declared_summary(summary)
    _assert_reconciliation_success(reconciliation)
    return summary


def main() -> int:
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure:
            reconfigure(errors="backslashreplace")
    try:
        result = run(parse_args())
        _safe_print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    except PrModelBaselineError as error:
        payload = {"status": "ERROR", "code": error.code, "message": str(error), "details": error.details}
    except DuProfileResolutionError as error:
        payload = {"status": "ERROR", "code": error.code, "message": str(error), "details": error.details}
    except (CreatePrError, SafetyControlError) as error:
        payload = {"status": "ERROR", "code": error.code, "message": str(error), "details": error.details}
    except Exception as error:
        payload = {"status": "ERROR", "code": "CREATE_PR_FAILED", "message": str(error)}
    _safe_print(json.dumps(payload, ensure_ascii=False, indent=2), file=sys.stderr)
    return 1


if __name__ == "__main__":
    sys.exit(main())