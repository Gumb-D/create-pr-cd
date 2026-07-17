#!/usr/bin/env python3
"""Build local-only generator-compatible UAT packets from approved DU exports.

This module never imports or invokes the ECC generator. Every output row is
permanently marked ECC Allowed = False.
"""
from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook

from du_export_adapter import (
    PR_STATUS_EXISTS,
    PR_STATUS_NONE,
    PR_STATUS_NOT_REQUIRED,
    build_canonical_site_record,
    resolve_profile_field_mappings,
)
from du_profile_loader import load_du_profile
from profile_du_export import build_header_inventory, calculate_header_hash, fingerprint_key

ALLOWED_UAT_PROFILE_STATUSES = {"PR_INPUT_READY", "PRODUCTION"}
UAT_CLASSIFICATIONS = ("UAT_CANDIDATE", "DUPLICATE_BLOCKED", "NO_PR_REQUIRED", "REVIEW_REQUIRED")

GENERATOR_COLUMNS = (
    "customer site code",
    "customer site name",
    "du code",
    "Tx SOW",
    "region",
    "Province/State",
    "SubCon - TSS Team",
    "SubCon - TI Team",
    "Subcon PR - TSS",
    "Subcon PR - TI",
    "TX Upgrade Scope",
    "Latitude (North Plus South Minus)",
    "Longitude (East Plus West Minus)",
    "MW Config Antenna Size NE",
    "MW Config Antenna Size FE",
    "BOQ Configuration",
    "TX SOW Details",
    "NE SOW Details",
    "FE SOW Details",
    "Scope Actual End Date",
    "Source Row Number",
    "DU Profile ID",
    "DU Profile Version",
    "Mapping Version",
    "Header Hash",
    "UAT Classification",
    "UAT Blocking Reasons",
    "ECC Allowed",
)


def _load_json_or_yaml(path: Path) -> Any:
    text = Path(path).read_text(encoding="utf-8")
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            import yaml  # type: ignore
        except ModuleNotFoundError as error:
            raise ValueError(f"{path} requires PyYAML or JSON-compatible YAML.") from error
        return yaml.safe_load(text)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _scope_date_field(scope: str) -> str:
    return f"{scope.lower()}_actual_end_date"


def classify_uat_record(record: Mapping[str, Any], scope: str) -> tuple[str, list[str]]:
    """Classify one canonical record for non-production UAT.

    Existing PR and explicit no-PR-required status take precedence. When a
    profile has an approved scope actual-end gate, a blank actual end means the
    scope is not yet eligible and must be ignored rather than generated early.
    """
    scope = str(scope).upper()
    if scope not in {"TSS", "TI"}:
        raise ValueError("scope must be TSS or TI")

    validation = record.get("validation", {})
    reasons = [str(value) for value in validation.get("blocking_reasons", []) if str(value).strip()]
    status_field = "existing_tss_pr_status" if scope == "TSS" else "existing_ti_pr_status"
    status = record.get("pr_context", {}).get(status_field, PR_STATUS_NONE)

    if status == PR_STATUS_EXISTS:
        return "DUPLICATE_BLOCKED", reasons + [f"{status_field}:PR_EXISTS"]
    if status == PR_STATUS_NOT_REQUIRED:
        return "NO_PR_REQUIRED", reasons + [f"{status_field}:NO_PR_REQUIRED"]

    date_field = _scope_date_field(scope)
    date_evidence = record.get("source_evidence", {}).get("fields", {}).get(date_field)
    if isinstance(date_evidence, Mapping) and _is_blank(date_evidence.get("source_value")):
        return "NO_PR_REQUIRED", reasons + [f"{date_field}:ACTUAL_END_MISSING"]

    if validation.get("pr_input_classification") not in {"PR_INPUT_READY", "PR_INPUT_READY_WITH_REVIEW"}:
        return "REVIEW_REQUIRED", reasons or ["CANONICAL_RECORD_NOT_READY"]

    sow_evidence = record.get("source_evidence", {}).get("fields", {}).get("tx_sow_normalized", {})
    normalization_status = sow_evidence.get("normalization_status")
    if normalization_status not in {"APPROVED", "APPROVED_NO_OUTPUT"}:
        return "REVIEW_REQUIRED", reasons + [f"SOW_NORMALIZATION:{normalization_status or 'MISSING'}"]
    if normalization_status == "APPROVED_NO_OUTPUT":
        return "NO_PR_REQUIRED", reasons + ["SOW_CLASSIFICATION:NO_PR_TRIGGER"]

    return "UAT_CANDIDATE", reasons


def _legacy_pr_reference_value(value: Any) -> str:
    if value in (None, "", PR_STATUS_NONE):
        return ""
    return str(value)


def canonical_record_to_generator_row(record: Mapping[str, Any], scope: str) -> dict[str, Any]:
    classification, reasons = classify_uat_record(record, scope)
    identity = record.get("identity", {})
    site = record.get("site", {})
    context = record.get("pr_context", {})
    technical = record.get("technical_context", {})
    validation = record.get("validation", {})
    date_evidence = record.get("source_evidence", {}).get("fields", {}).get(_scope_date_field(str(scope).upper()), {})
    return {
        "customer site code": site.get("site_code", ""),
        "customer site name": site.get("site_name", ""),
        "du code": site.get("du_key", ""),
        "Tx SOW": context.get("tx_sow_normalized") or context.get("tx_sow_raw", ""),
        "region": context.get("region", ""),
        "Province/State": context.get("state", ""),
        "SubCon - TSS Team": context.get("subcontractor_tss", ""),
        "SubCon - TI Team": context.get("subcontractor_ti", ""),
        "Subcon PR - TSS": _legacy_pr_reference_value(context.get("existing_tss_pr_status")),
        "Subcon PR - TI": _legacy_pr_reference_value(context.get("existing_ti_pr_status")),
        "TX Upgrade Scope": context.get("tx_upgrade_scope_raw", ""),
        "Latitude (North Plus South Minus)": technical.get("latitude"),
        "Longitude (East Plus West Minus)": technical.get("longitude"),
        "MW Config Antenna Size NE": technical.get("antenna_size_ne", ""),
        "MW Config Antenna Size FE": technical.get("antenna_size_fe", ""),
        "BOQ Configuration": technical.get("boq_configuration", ""),
        "TX SOW Details": technical.get("tx_sow_details", ""),
        "NE SOW Details": technical.get("ne_sow_details", ""),
        "FE SOW Details": technical.get("fe_sow_details", ""),
        "Scope Actual End Date": date_evidence.get("source_value") if isinstance(date_evidence, Mapping) else "",
        "Source Row Number": identity.get("source_row_number"),
        "DU Profile ID": validation.get("profile_id", ""),
        "DU Profile Version": validation.get("profile_version", ""),
        "Mapping Version": validation.get("mapping_version", ""),
        "Header Hash": identity.get("header_hash", ""),
        "UAT Classification": classification,
        "UAT Blocking Reasons": " | ".join(reasons),
        "ECC Allowed": False,
    }


def _append_rows(worksheet, rows: Iterable[Mapping[str, Any]], columns: Iterable[str]) -> None:
    columns = list(columns)
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column) for column in columns])


def _append_generator_data_sheet(worksheet, rows: list[Mapping[str, Any]]) -> None:
    worksheet.append(["NON-PRODUCTION UAT BRIDGE OUTPUT"])
    worksheet.append(["This sheet contains UAT_CANDIDATE rows only."])
    worksheet.append(["ECC Allowed is permanently false; use explicit generator CLI arguments for any later local UAT."])
    _append_rows(worksheet, rows, GENERATOR_COLUMNS)


def write_uat_packet(
    records: list[Mapping[str, Any]],
    metadata: Mapping[str, Any],
    output_dir: Path,
    scope: str,
) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    scope = str(scope).upper()
    rows = [canonical_record_to_generator_row(record, scope) for record in records]
    counts = Counter(row["UAT Classification"] for row in rows)

    summary = {
        "scope": scope,
        "profile_id": metadata.get("profile_id", ""),
        "profile_version": metadata.get("profile_version", ""),
        "mapping_version": metadata.get("mapping_version", ""),
        "source_file_name": metadata.get("source_file_name", ""),
        "source_file_hash": metadata.get("source_file_hash", ""),
        "header_hash": metadata.get("header_hash", ""),
        "record_count": len(rows),
        "counts": {name: counts.get(name, 0) for name in UAT_CLASSIFICATIONS},
        "generator_data_row_count": counts.get("UAT_CANDIDATE", 0),
        "ecc_allowed": False,
    }

    workbook_path = output_dir / f"canonical_generator_uat_{scope.lower()}.xlsx"
    summary_path = output_dir / f"canonical_generator_uat_{scope.lower()}_summary.json"
    workbook = Workbook()
    workbook.remove(workbook.active)

    candidate_rows = [row for row in rows if row["UAT Classification"] == "UAT_CANDIDATE"]
    _append_generator_data_sheet(workbook.create_sheet("data"), candidate_rows)

    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.append(["Key", "Value"])
    for key, value in summary.items():
        summary_sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value])

    partitions = {
        "uat_candidates": candidate_rows,
        "duplicate_blocked": [row for row in rows if row["UAT Classification"] == "DUPLICATE_BLOCKED"],
        "no_pr_required": [row for row in rows if row["UAT Classification"] == "NO_PR_REQUIRED"],
        "review_required": [row for row in rows if row["UAT Classification"] == "REVIEW_REQUIRED"],
    }
    for sheet_name, partition_rows in partitions.items():
        _append_rows(workbook.create_sheet(sheet_name), partition_rows, GENERATOR_COLUMNS)

    traceability_columns = (
        "customer site code",
        "Scope Actual End Date",
        "Source Row Number",
        "DU Profile ID",
        "DU Profile Version",
        "Mapping Version",
        "Header Hash",
        "UAT Classification",
        "UAT Blocking Reasons",
        "ECC Allowed",
    )
    _append_rows(workbook.create_sheet("traceability"), rows, traceability_columns)
    workbook.save(workbook_path)
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"workbook": workbook_path, "summary_json": summary_path}


def _iter_source_rows(input_path: Path, sheet_name: str, positions: Mapping[str, int]):
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(input_path, read_only=True, data_only=False)
        worksheet = workbook[sheet_name]
        try:
            for row_number, values in enumerate(worksheet.iter_rows(min_row=5, values_only=True), start=5):
                yield row_number, {
                    key: values[index - 1] if index <= len(values) else None
                    for key, index in positions.items()
                }
        finally:
            workbook.close()
        return
    if suffix == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for _ in range(4):
                next(reader, None)
            for row_number, values in enumerate(reader, start=5):
                yield row_number, {
                    key: values[index - 1] if index <= len(values) else None
                    for key, index in positions.items()
                }
        return
    raise ValueError("Only .xlsx, .xlsm, and .csv exports are supported.")


def _select_source_sheet(
    inventory: Mapping[str, Any],
    resolved: Mapping[str, Any],
    profile: Mapping[str, Any],
) -> Mapping[str, Any]:
    required_fields = [name for name, config in profile.get("field_mapping", {}).items() if config.get("required")]
    required_sheets = {
        match["sheet_name"]
        for name in required_fields
        for match in resolved.get(name, {}).get("matches", [])
    }
    if len(required_sheets) != 1:
        raise ValueError("REQUIRED_MAPPINGS_SPAN_MULTIPLE_OR_NO_SHEETS")
    selected_name = next(iter(required_sheets))
    for sheet in inventory.get("sheets", []):
        if sheet.get("sheet_name") == selected_name:
            return sheet
    raise ValueError("RESOLVED_SOURCE_SHEET_NOT_FOUND")


def _load_scope_eligibility(profile: Mapping[str, Any], profile_path: Path) -> Mapping[str, Any]:
    config_path = profile_path.parent.parent / "scope_eligibility" / f"{profile.get('profile_id', '')}.json"
    if not config_path.exists():
        return {}
    config = _load_json_or_yaml(config_path)
    if config.get("profile_id") != profile.get("profile_id"):
        raise ValueError("SCOPE_ELIGIBILITY_PROFILE_MISMATCH")
    return config.get("scopes", {})


def _attach_scope_eligibility(
    record: dict[str, Any],
    raw_values: Mapping[str, Any],
    scope: str,
    scope_config: Mapping[str, Any],
) -> dict[str, Any]:
    config = scope_config.get(scope, {}) if isinstance(scope_config, Mapping) else {}
    if not config:
        return record
    if config.get("rule") != "actual_end_required":
        raise ValueError("UNSUPPORTED_SCOPE_ELIGIBILITY_RULE")
    fingerprint = config.get("actual_end_fingerprint")
    if not isinstance(fingerprint, Mapping):
        raise ValueError("MISSING_SCOPE_ACTUAL_END_FINGERPRINT")
    key = fingerprint_key(fingerprint)
    if key not in raw_values:
        raise ValueError("SCOPE_ACTUAL_END_HEADER_NOT_FOUND")
    record["source_evidence"]["fields"][_scope_date_field(scope)] = {
        "source_header_fingerprint": dict(fingerprint),
        "source_value": raw_values.get(key),
        "transformation": "none",
        "mapping_status": "APPROVED",
    }
    return record


def build_records_from_export(
    input_path: Path,
    profile_path: Path,
    scope: str,
    sow_registry_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    input_path = Path(input_path)
    profile_path = Path(profile_path)
    scope = str(scope).upper()
    profile = load_du_profile(profile_path)
    if profile.get("status") not in ALLOWED_UAT_PROFILE_STATUSES:
        raise ValueError("DU profile must be PR_INPUT_READY or PRODUCTION for non-production UAT bridging.")

    inventory = build_header_inventory(input_path)
    header_hash = calculate_header_hash(inventory)
    approved_hashes = profile.get("export_structure", {}).get("approved_header_hashes", [])
    if header_hash not in approved_hashes:
        raise ValueError("HEADER_HASH_REVALIDATION_REQUIRED")

    resolved = resolve_profile_field_mappings(inventory, profile)
    required_missing = [
        name
        for name, config in profile.get("field_mapping", {}).items()
        if config.get("required") and resolved.get(name, {}).get("status") != "RESOLVED"
    ]
    if required_missing:
        raise ValueError("MISSING_OR_AMBIGUOUS_REQUIRED_MAPPING:" + ",".join(sorted(required_missing)))

    source_sheet = _select_source_sheet(inventory, resolved, profile)
    positions = {
        column["fingerprint_key"]: column["source_position"]["one_based_index"]
        for column in source_sheet.get("columns", [])
    }
    scope_config = _load_scope_eligibility(profile, profile_path)
    configured = scope_config.get(scope, {}) if isinstance(scope_config, Mapping) else {}
    if configured:
        expected_key = fingerprint_key(configured.get("actual_end_fingerprint", {}))
        if expected_key not in positions:
            raise ValueError("SCOPE_ACTUAL_END_HEADER_NOT_FOUND")

    sow_registry = _load_json_or_yaml(Path(sow_registry_path))
    identity = profile["identity"]
    metadata = {
        "profile_id": profile["profile_id"],
        "profile_version": profile["profile_version"],
        "mapping_version": profile["mapping_version"],
        "source_file_name": input_path.name,
        "source_file_hash": inventory["source"]["source_file_hash"],
        "header_hash": header_hash,
    }
    context_base = {
        "project_key": identity["project_key"],
        "du_model_name": identity["accepted_du_models"][0],
        "du_model_id": identity["accepted_du_model_ids"][0],
        "view_id": identity["accepted_view_ids"][0],
        **metadata,
    }

    records = []
    for row_number, raw_values in _iter_source_rows(input_path, source_sheet["sheet_name"], positions):
        if not any(value not in (None, "") for value in raw_values.values()):
            continue
        record = build_canonical_site_record(
            raw_values,
            profile,
            {**context_base, "source_row_number": row_number},
            scope=scope,
            resolved_mappings=resolved,
            sow_registry=sow_registry,
        )
        records.append(_attach_scope_eligibility(record, raw_values, scope, scope_config))
    return records, metadata


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a non-production canonical-to-generator UAT packet.")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--profile", required=True, type=Path)
    parser.add_argument("--scope", required=True, choices=["TSS", "TI"], type=str.upper)
    parser.add_argument("--sow-registry", default=Path("config/registries/canonical_sow_registry.yaml"), type=Path)
    parser.add_argument("--output", default=Path("output/canonical_generator_uat"), type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    records, metadata = build_records_from_export(args.input, args.profile, args.scope, args.sow_registry)
    outputs = write_uat_packet(records, metadata, args.output, args.scope)
    print(
        json.dumps(
            {
                "records": len(records),
                "ecc_allowed": False,
                **{key: str(value) for key, value in outputs.items()},
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())