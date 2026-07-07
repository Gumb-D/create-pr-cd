"""Build a local-only mapping-decision review workbook for one profiled DU export.

This is a manual-review packaging tool only. It never edits DU profiles, never
changes lifecycle status, and never approves a mapping. The generated workbook
and summary are written under ``output/`` (gitignored) because they contain
masked sample values taken from a local, uncommitted source export.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence

from canonical_site_validator import SCOPE_REQUIRED_FIELDS

DECISION_OPTIONS = ("APPROVE", "REJECT", "NEEDS_BUSINESS_DECISION", "NO_SOURCE_FIELD")

FLAG_NO_SOURCE_CANDIDATE = "NO_SOURCE_CANDIDATE"
FLAG_MULTIPLE_CANDIDATES = "MULTIPLE_CANDIDATES"
FLAG_SINGLE_CANDIDATE_VERIFY = "SINGLE_CANDIDATE_VERIFY"
FLAG_DERIVED_FIELD = "DERIVED_FIELD"
FLAG_TRANSFORMATION_UNCERTAIN = "TRANSFORMATION_UNCERTAIN"
NOISE_NOTE = "LIKELY_NOISE: keyword hit outside the data sheet or without a display header."

DEFAULT_REVIEW_FIELDS = (
    "site_code",
    "site_name",
    "du_key",
    "tx_sow_raw",
    "tx_sow_normalized",
    "tx_upgrade_scope_raw",
    "region",
    "state",
    "subcontractor_ti",
    "existing_tss_pr_status",
    "existing_ti_pr_status",
    "latitude",
    "longitude",
    "antenna_size_ne",
    "antenna_size_fe",
)

DERIVED_FIELDS = {
    "tx_sow_normalized": (
        "Derived field: controlled normalization of the approved tx_sow_raw value. "
        "It has no direct source column; the normalization value map itself requires "
        "business approval."
    ),
}

# Fields whose sample values identify a site, a person, or a customer partner.
MASK_RULES = {
    "site_code": "code",
    "site_name": "name",
    "du_key": "code",
    "latitude": "coordinate",
    "longitude": "coordinate",
    "subcontractor_ti": "partner",
    "subcontractor_planning": "partner",
}

# Column-level masking by display header, so a sensitive source column is masked
# consistently no matter which canonical field surfaced it as a candidate.
HEADER_MASK_HINTS = (
    ("person", "partner"),
    ("company", "partner"),
    ("subcon", "partner"),
    ("site name", "name"),
    ("site code", "code"),
    ("site id", "code"),
    ("du code", "code"),
    ("latitude", "coordinate"),
    ("longitude", "coordinate"),
)

# Map canonical fields to their discovery shortlist so candidate ranking uses
# scores for the same field, never a lookalike field's scores.
CANONICAL_TO_SHORTLIST_FIELD = {
    "site_code": "site_id",
    "site_name": "site_name",
    "du_key": "du_code",
    "region": "region",
    "tx_sow_raw": "tx_sow",
    "antenna_size_ne": "antenna_size_ne",
    "antenna_size_fe": "antenna_size_fe",
    "subcontractor_ti": "subcon_ti_team",
    "subcontractor_planning": "subcon_planning",
    "existing_tss_pr_status": "existing_tss_pr",
    "existing_ti_pr_status": "existing_ti_pr",
}

UNRESOLVED_EXTRA_NOTES = {
    "existing_tss_pr_status": (
        "Keyword discovery treats every current TX Mini candidate as a weak false positive. "
        "`2023 TX Rollout` is the profiled donor reference for this duplicate-prevention status "
        "(see config/registries/mw_du_missing_field_bridge_review.yaml). If no TX Mini source "
        "column is confirmed, record NO_SOURCE_FIELD and choose the missing-field treatment."
    ),
    "existing_ti_pr_status": (
        "Keyword discovery treats every current TX Mini candidate as a weak false positive. "
        "`2023 TX Rollout` is the profiled donor reference for this duplicate-prevention status "
        "(see config/registries/mw_du_missing_field_bridge_review.yaml). If no TX Mini source "
        "column is confirmed, record NO_SOURCE_FIELD and choose the missing-field treatment."
    ),
}

# Site-code style tokens (letters followed by digits) and long digit runs such as
# employee or PR reference numbers, scrubbed from otherwise unmasked free text.
_SITE_TOKEN_PATTERN = re.compile(r"\b([A-Za-z]{1,4})(\d{4,})\b")
_LONG_DIGIT_PATTERN = re.compile(r"\b\d{7,}\b")

SAMPLE_COUNT = 5
SAMPLE_SCAN_LIMIT = 4000


def requirement_class(field: str) -> str:
    in_tss = field in SCOPE_REQUIRED_FIELDS["TSS"]
    in_ti = field in SCOPE_REQUIRED_FIELDS["TI"]
    if in_tss and in_ti:
        return "REQUIRED (TSS + TI)"
    if in_tss:
        return "REQUIRED (TSS)"
    if in_ti:
        return "REQUIRED (TI)"
    return "CONDITIONAL"


def mask_value(value: Any, rule: Optional[str]) -> str:
    text = str(value).strip()
    if not text:
        return ""
    if rule is None:
        text = _SITE_TOKEN_PATTERN.sub(lambda m: m.group(1)[:2] + "***", text)
        text = _LONG_DIGIT_PATTERN.sub("#####", text)
        return text if len(text) <= 80 else text[:77] + "..."
    if rule == "coordinate":
        sign = "-" if text.startswith("-") else ""
        body = text.lstrip("+-")
        integer_part = body.split(".", 1)[0] if body else ""
        if integer_part.isdigit():
            return f"{sign}{integer_part}.***"
        return "***"
    if rule == "name":
        return text[:3] + "***" if len(text) > 3 else "***"
    if rule == "partner":
        return text[:4] + "***" if len(text) > 4 else text[:2] + "***"
    # "code": keep the first two and last one characters, mask the middle.
    if len(text) <= 3:
        return "*" * len(text)
    return text[:2] + "*" * (len(text) - 3) + text[-1]


def mask_rule_for(field: str, display_header: str) -> Optional[str]:
    """Field-level rule first, then column-level hints from the display header."""
    rule = MASK_RULES.get(field)
    if rule:
        return rule
    display = str(display_header).strip().lower()
    for hint, hint_rule in HEADER_MASK_HINTS:
        if hint in display:
            return hint_rule
    return None


def _fingerprint_tuple(fingerprint: Mapping[str, Any]) -> tuple:
    return tuple(str(fingerprint.get(k, "")) for k in ("field_code", "wbs_stage", "task_name", "display_header"))


def load_column_positions(header_inventory: Mapping[str, Any]) -> Dict[tuple, Dict[str, Any]]:
    """Map (sheet_name, four-layer fingerprint) to the profiler-observed position."""
    positions: Dict[tuple, Dict[str, Any]] = {}
    for sheet in header_inventory.get("sheets", []):
        sheet_name = sheet.get("sheet_name", "")
        for column in sheet.get("columns", []):
            key = (sheet_name,) + _fingerprint_tuple(column.get("fingerprint", {}))
            positions[key] = dict(column.get("source_position", {}))
    return positions


def load_shortlist_scores(
    shortlist_registry: Optional[Mapping[str, Any]], observed_header_hash: str
) -> Dict[str, Dict[tuple, Dict[str, Any]]]:
    """Map shortlist field -> four-layer fingerprint -> score/reason for this export."""
    scores: Dict[str, Dict[tuple, Dict[str, Any]]] = {}
    if not shortlist_registry:
        return scores
    for entry in shortlist_registry.get("entries", []):
        if entry.get("observed_header_hash") != observed_header_hash:
            continue
        for shortlist_field, shortlist in entry.get("skill_field_shortlists", {}).items():
            field_scores = scores.setdefault(shortlist_field, {})
            for candidate in shortlist:
                key = _fingerprint_tuple(candidate.get("fingerprint", {}))
                existing = field_scores.get(key)
                if existing is None or int(candidate.get("score", 0)) > int(existing.get("score", 0)):
                    field_scores[key] = {"score": candidate.get("score"), "reason": candidate.get("reason", "")}
    return scores


def collect_sample_values(
    source_export_path: Optional[Path],
    sheet_name: str,
    one_based_index: Optional[int],
    header_row_count: int,
    mask_rule: Optional[str],
) -> Dict[str, Any]:
    """Read-only sampling of one source column; values are masked before return."""
    result: Dict[str, Any] = {"samples": [], "non_blank_count": 0, "distinct_count": 0, "note": ""}
    if source_export_path is None:
        result["note"] = "Source export not supplied; samples unavailable."
        return result
    if not one_based_index:
        result["note"] = "Column position not found in the profiler header inventory."
        return result
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(source_export_path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - environment/IO failure path
        result["note"] = f"Could not open source export read-only: {exc}"
        return result
    try:
        sheet = None
        for name in workbook.sheetnames:
            if name.strip().lower() == sheet_name.strip().lower():
                sheet = workbook[name]
                break
        if sheet is None:
            result["note"] = f"Sheet '{sheet_name}' not present in source export."
            return result
        raw_samples: List[str] = []
        distinct: set = set()
        non_blank = 0
        for row_number, row in enumerate(
            sheet.iter_rows(
                min_row=header_row_count + 1,
                min_col=one_based_index,
                max_col=one_based_index,
                values_only=True,
            )
        ):
            if row_number >= SAMPLE_SCAN_LIMIT:
                break
            value = row[0]
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            non_blank += 1
            text = str(value).strip()
            distinct.add(text)
            if len(raw_samples) < SAMPLE_COUNT and text not in raw_samples:
                raw_samples.append(text)
        # Pad with repeats when the column has fewer distinct values than requested.
        if raw_samples and len(raw_samples) < SAMPLE_COUNT and non_blank >= SAMPLE_COUNT:
            while len(raw_samples) < SAMPLE_COUNT:
                raw_samples.append(raw_samples[len(raw_samples) % len(distinct or {1})])
        result["samples"] = [mask_value(v, mask_rule) for v in raw_samples]
        result["non_blank_count"] = non_blank
        result["distinct_count"] = len(distinct)
        if not raw_samples:
            result["note"] = "Column is present but has no non-blank values in the scanned rows."
    finally:
        workbook.close()
    return result


def _rank_candidates(
    candidates: Sequence[Mapping[str, Any]], shortlist_scores: Mapping[tuple, Mapping[str, Any]]
) -> List[Dict[str, Any]]:
    enriched: List[Dict[str, Any]] = []
    for order, candidate in enumerate(candidates):
        fingerprint = candidate.get("fingerprint", {})
        key = _fingerprint_tuple(fingerprint)
        shortlist = shortlist_scores.get(key)
        sheet_name = candidate.get("sheet_name", "")
        is_noise = sheet_name.strip().lower() != "data" or not str(fingerprint.get("display_header", "")).strip()
        enriched.append(
            {
                "sheet_name": sheet_name,
                "fingerprint": dict(fingerprint),
                "mapping_status": candidate.get("mapping_status", "UNVERIFIED"),
                "shortlist_score": shortlist.get("score") if shortlist else None,
                "shortlist_reason": shortlist.get("reason", "") if shortlist else "",
                "is_noise": is_noise,
                "_order": order,
            }
        )
    enriched.sort(
        key=lambda item: (
            item["is_noise"],
            -(item["shortlist_score"] if item["shortlist_score"] is not None else -1),
            item["_order"],
        )
    )
    for rank, item in enumerate(enriched, start=1):
        item["rank"] = rank
        del item["_order"]
    return enriched


def build_review_model(
    *,
    candidates_report: Mapping[str, Any],
    header_inventory: Mapping[str, Any],
    observed_header_hash: str,
    du_profile: Mapping[str, Any],
    shortlist_registry: Optional[Mapping[str, Any]] = None,
    source_export_path: Optional[Path] = None,
    review_fields: Sequence[str] = DEFAULT_REVIEW_FIELDS,
) -> Dict[str, Any]:
    """Assemble the full review model. Read-only over every input."""
    positions = load_column_positions(header_inventory)
    shortlist_scores = load_shortlist_scores(shortlist_registry, observed_header_hash)
    header_row_count = int(header_inventory.get("source", {}).get("header_row_count", 4))
    profile_mapping = du_profile.get("field_mapping", {})
    candidate_fields = candidates_report.get("fields", {})

    fields: List[Dict[str, Any]] = []
    for field in review_fields:
        profile_entry = profile_mapping.get(field, {})
        report_entry = candidate_fields.get(field, {})
        derived_note = DERIVED_FIELDS.get(field, "")
        raw_candidates = [] if derived_note else list(report_entry.get("candidates", []))
        field_shortlist_scores = shortlist_scores.get(CANONICAL_TO_SHORTLIST_FIELD.get(field, field), {})
        ranked = _rank_candidates(raw_candidates, field_shortlist_scores)
        plausible = [c for c in ranked if not c["is_noise"]]

        for candidate in ranked:
            key = (candidate["sheet_name"],) + _fingerprint_tuple(candidate["fingerprint"])
            position = positions.get(key, {})
            candidate["excel_column"] = position.get("excel_column", "")
            candidate["one_based_index"] = position.get("one_based_index")
            candidate["mask_rule"] = mask_rule_for(field, candidate["fingerprint"].get("display_header", ""))
            if candidate["is_noise"]:
                candidate["sampling"] = {
                    "samples": [],
                    "non_blank_count": 0,
                    "distinct_count": 0,
                    "note": "Not sampled: flagged as likely keyword noise.",
                }
            else:
                candidate["sampling"] = collect_sample_values(
                    source_export_path,
                    candidate["sheet_name"],
                    candidate["one_based_index"],
                    header_row_count,
                    candidate["mask_rule"],
                )

        transforms = list(profile_entry.get("transforms", []))
        if derived_note:
            proposed_transformation = derived_note
            flag = FLAG_DERIVED_FIELD
            discovery_status = "DERIVED"
        else:
            proposed_transformation = " -> ".join(transforms) if transforms else "(none proposed)"
            discovery_status = report_entry.get("status", "NOT_DISCOVERED" if not ranked else "UNVERIFIED")
            if not plausible:
                flag = FLAG_NO_SOURCE_CANDIDATE
            elif len(plausible) > 1:
                flag = FLAG_MULTIPLE_CANDIDATES
            else:
                flag = FLAG_SINGLE_CANDIDATE_VERIFY

        seeded = profile_entry.get("source_candidates", [])
        seeded_fp = seeded[0].get("fingerprint", {}) if seeded else {}
        fields.append(
            {
                "canonical_field": field,
                "requirement": requirement_class(field),
                "profile_required_flag": bool(profile_entry.get("required", False)),
                "discovery_status": discovery_status,
                "flag": flag,
                "candidates": ranked,
                "plausible_candidate_count": len(plausible),
                "proposed_transformation": proposed_transformation,
                "profile_seeded_fingerprint": " | ".join(_fingerprint_tuple(seeded_fp)) if seeded_fp else "(none)",
                "derived_note": derived_note,
                "mask_rule": MASK_RULES.get(field, ""),
            }
        )

    source_info = header_inventory.get("source", {})
    return {
        "profile_id": du_profile.get("profile_id", ""),
        "profile_version": du_profile.get("profile_version", ""),
        "mapping_version": du_profile.get("mapping_version", ""),
        "profile_status": du_profile.get("status", ""),
        "source_file_name": source_info.get("file_name", ""),
        "source_file_hash": source_info.get("source_file_hash", ""),
        "observed_header_hash": observed_header_hash,
        "fields": fields,
    }


def _fingerprint_columns(candidate: Mapping[str, Any]) -> List[str]:
    fp = candidate.get("fingerprint", {})
    return [str(fp.get(k, "")) for k in ("field_code", "wbs_stage", "task_name", "display_header")]


def write_decision_workbook(model: Mapping[str, Any], workbook_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E79")
    body_font = Font(name="Arial")
    flag_fill = PatternFill("solid", start_color="FFF2CC")

    workbook = Workbook()

    def add_sheet(title: str, headers: Sequence[str], widths: Sequence[int]) -> Any:
        sheet = workbook.create_sheet(title)
        sheet.append(list(headers))
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
            cell = sheet.cell(row=1, column=index)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        return sheet

    workbook.remove(workbook.active)

    instructions = workbook.create_sheet("Review_Instructions")
    instructions.column_dimensions["A"].width = 130
    instruction_lines = [
        "TX-style DU mapping decision workbook (LOCAL-ONLY REVIEW ARTIFACT)",
        "",
        f"DU profile: {model['profile_id']} v{model['profile_version']} "
        f"(mapping {model['mapping_version']}, lifecycle status {model['profile_status']})",
        f"Source export: {model['source_file_name']}",
        f"Source file SHA-256: {model['source_file_hash']}",
        f"Observed header hash: {model['observed_header_hash']}",
        "",
        "PURPOSE",
        "Convert discovery-only source-field candidates into explicit human mapping decisions.",
        "Nothing in this workbook is an approval by itself, and no profile file has been changed by it.",
        "",
        "RULES",
        "1. Every mapping decision must reference the exact four-layer fingerprint",
        "   (field code / WBS stage / task name / display header), never a column position.",
        "2. Source-column positions shown in this workbook are NON-AUTHORITATIVE review aids only.",
        "3. Sample values are masked where they identify sites, coordinates, people, or partners.",
        "4. The DU profile stays DRAFT and ECC output stays blocked until decisions here are",
        "   approved and applied through the controlled profile-change process.",
        "5. Do not commit this workbook or any source-export values to Git.",
        "",
        "HOW TO REVIEW",
        "1. Read Required_Fields for the per-field status and flags.",
        "2. Check Candidate_Fingerprints for every ranked candidate and its exact fingerprint.",
        "3. Use Sample_Source_Values to confirm the value domain of each candidate column.",
        "4. Record one decision per canonical field in Decision_Log:",
        f"   decision must be one of {', '.join(DECISION_OPTIONS)};",
        "   when the decision is APPROVE, also fill the selected candidate rank.",
        "5. Unresolved_Fields lists every field that cannot be approved without more evidence",
        "   or a business decision; resolve or explicitly defer each one.",
    ]
    for line in instruction_lines:
        instructions.append([line])
    instructions["A1"].font = Font(name="Arial", bold=True, size=14)
    for row in instructions.iter_rows(min_row=2):
        row[0].font = body_font
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    required_sheet = add_sheet(
        "Required_Fields",
        [
            "Canonical field",
            "Required for",
            "Profile required flag",
            "Current discovery status",
            "Plausible candidates",
            "Review flag",
            "Proposed transformation",
            "Profile-seeded fingerprint (DRAFT, UNVERIFIED)",
        ],
        [24, 20, 16, 20, 14, 26, 44, 60],
    )
    for field in model["fields"]:
        required_sheet.append(
            [
                field["canonical_field"],
                field["requirement"],
                "yes" if field["profile_required_flag"] else "no",
                field["discovery_status"],
                field["plausible_candidate_count"],
                field["flag"],
                field["proposed_transformation"],
                field["profile_seeded_fingerprint"],
            ]
        )

    candidate_sheet = add_sheet(
        "Candidate_Fingerprints",
        [
            "Canonical field",
            "Candidate rank",
            "Sheet",
            "Field code / ID",
            "WBS stage",
            "Task name",
            "Display header",
            "Source column (NON-AUTHORITATIVE, review only)",
            "Shortlist score",
            "Shortlist reason",
            "Mapping status",
            "Noise flag",
        ],
        [24, 10, 12, 44, 30, 26, 32, 24, 12, 44, 14, 40],
    )
    for field in model["fields"]:
        if field["derived_note"]:
            candidate_sheet.append(
                [field["canonical_field"], "-", "-", "(derived field)", "-", "-", "-", "-", "", field["derived_note"], "DERIVED", ""]
            )
            continue
        if not field["candidates"]:
            candidate_sheet.append(
                [field["canonical_field"], "-", "-", "(no source candidate discovered)", "-", "-", "-", "-", "", "", "NOT_DISCOVERED", ""]
            )
            continue
        for candidate in field["candidates"]:
            position = candidate["excel_column"] or "?"
            index = candidate["one_based_index"]
            fingerprint_columns = _fingerprint_columns(candidate)
            if candidate["is_noise"]:
                # Noise fingerprints can embed person names / employee IDs from
                # approval routing lists; they are not proposed mappings, so
                # their field code is masked instead of shown verbatim.
                fingerprint_columns[0] = mask_value(fingerprint_columns[0], "partner")
            candidate_sheet.append(
                [
                    field["canonical_field"],
                    candidate["rank"],
                    candidate["sheet_name"],
                    *fingerprint_columns,
                    f"{position} (col #{index})" if index else position,
                    candidate["shortlist_score"] if candidate["shortlist_score"] is not None else "",
                    candidate["shortlist_reason"],
                    candidate["mapping_status"],
                    NOISE_NOTE if candidate["is_noise"] else "",
                ]
            )

    sample_headers = (
        ["Canonical field", "Candidate rank", "Display header", "Masking rule"]
        + [f"Sample {i}" for i in range(1, SAMPLE_COUNT + 1)]
        + ["Non-blank rows", "Distinct values", "Sampling note"]
    )
    sample_sheet = add_sheet(
        "Sample_Source_Values",
        sample_headers,
        [24, 10, 32, 14] + [22] * SAMPLE_COUNT + [14, 14, 46],
    )
    for field in model["fields"]:
        for candidate in field["candidates"]:
            if candidate["is_noise"]:
                continue
            sampling = candidate["sampling"]
            samples = list(sampling["samples"]) + [""] * (SAMPLE_COUNT - len(sampling["samples"]))
            sample_sheet.append(
                [
                    field["canonical_field"],
                    candidate["rank"],
                    candidate["fingerprint"].get("display_header", ""),
                    candidate["mask_rule"] or "(none)",
                    *samples[:SAMPLE_COUNT],
                    sampling["non_blank_count"],
                    sampling["distinct_count"],
                    sampling["note"],
                ]
            )
        if field["derived_note"] or not any(not c["is_noise"] for c in field["candidates"]):
            sample_sheet.append(
                [
                    field["canonical_field"],
                    "-",
                    "(no sampled source column)",
                    field["mask_rule"] or "(none)",
                    *[""] * SAMPLE_COUNT,
                    "",
                    "",
                    field["derived_note"] or "No plausible source candidate to sample.",
                ]
            )

    decision_sheet = add_sheet(
        "Decision_Log",
        [
            "Canonical field",
            "Required for",
            "Current discovery status",
            "Review flag",
            "Proposed transformation",
            "Reviewer decision",
            "Reviewer-selected candidate rank",
            "Reviewer rationale",
            "Reviewer name",
            "Review date (YYYY-MM-DD)",
        ],
        [24, 20, 20, 26, 44, 28, 18, 60, 22, 20],
    )
    for field in model["fields"]:
        decision_sheet.append(
            [
                field["canonical_field"],
                field["requirement"],
                field["discovery_status"],
                field["flag"],
                field["proposed_transformation"],
                "",
                "",
                "",
                "",
                "",
            ]
        )
    validation = DataValidation(
        type="list",
        formula1=f"\"{','.join(DECISION_OPTIONS)}\"",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid decision",
        error=f"Decision must be one of: {', '.join(DECISION_OPTIONS)}",
    )
    decision_sheet.add_data_validation(validation)
    validation.add(f"F2:F{decision_sheet.max_row}")
    for row in decision_sheet.iter_rows(min_row=2, min_col=6, max_col=10):
        for cell in row:
            cell.fill = flag_fill

    unresolved_sheet = add_sheet(
        "Unresolved_Fields",
        ["Canonical field", "Required for", "Review flag", "Why it is unresolved", "Evidence or decision still required"],
        [24, 20, 26, 60, 60],
    )
    for field in model["fields"]:
        if field["flag"] == FLAG_SINGLE_CANDIDATE_VERIFY:
            continue
        if field["flag"] == FLAG_NO_SOURCE_CANDIDATE:
            why = "No plausible source column was discovered in this export."
            needs = (
                "Business decision: confirm the field is absent from this DU model and choose the "
                "missing-field treatment (donor reference data, manual resolution, or blocking)."
            )
        elif field["flag"] == FLAG_MULTIPLE_CANDIDATES:
            why = f"{field['plausible_candidate_count']} plausible source columns match this field."
            needs = "Reviewer must select exactly one four-layer fingerprint; the others must be rejected."
        elif field["flag"] == FLAG_DERIVED_FIELD:
            why = "Derived field with no direct source column."
            needs = field["derived_note"]
        else:  # pragma: no cover - defensive default
            why = "Transformation or mapping evidence is uncertain."
            needs = "Confirm the controlled transformation with business evidence."
        extra_note = UNRESOLVED_EXTRA_NOTES.get(field["canonical_field"])
        if extra_note:
            needs = f"{needs} {extra_note}"
        unresolved_sheet.append([field["canonical_field"], field["requirement"], field["flag"], why, needs])

    for sheet in (required_sheet, candidate_sheet, sample_sheet, decision_sheet, unresolved_sheet):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.font is None or not cell.font.name:
                    cell.font = body_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)


def decision_summary_markdown(model: Mapping[str, Any]) -> str:
    lines = [
        "# TX Mini Mapping Decision Summary (LOCAL-ONLY)"
        if model["profile_id"] == "tx_mini_pr_v1"
        else f"# {model['profile_id']} Mapping Decision Summary (LOCAL-ONLY)",
        "",
        "This package converts discovery-only profiler evidence into a human decision worksheet.",
        "It grants no approval, changes no profile file, and must not be committed to Git.",
        "",
        f"- DU profile: `{model['profile_id']}` v`{model['profile_version']}` "
        f"(mapping `{model['mapping_version']}`, status `{model['profile_status']}` — unchanged)",
        f"- Source export: `{model['source_file_name']}` (read-only, external, uncommitted)",
        f"- Source file SHA-256: `{model['source_file_hash']}`",
        f"- Observed header hash: `{model['observed_header_hash']}`",
        "- Source-column positions in the workbook are non-authoritative review aids;",
        "  the four-layer fingerprint is the only valid mapping reference.",
        "",
        "## Field review status",
        "",
        "| Canonical field | Required for | Discovery status | Plausible candidates | Review flag |",
        "|---|---|---|---|---|",
    ]
    for field in model["fields"]:
        lines.append(
            "| `{name}` | {req} | {status} | {count} | {flag} |".format(
                name=field["canonical_field"],
                req=field["requirement"],
                status=field["discovery_status"],
                count=field["plausible_candidate_count"],
                flag=field["flag"],
            )
        )

    single = [f for f in model["fields"] if f["flag"] == FLAG_SINGLE_CANDIDATE_VERIFY]
    multiple = [f for f in model["fields"] if f["flag"] == FLAG_MULTIPLE_CANDIDATES]
    missing = [f for f in model["fields"] if f["flag"] == FLAG_NO_SOURCE_CANDIDATE]
    derived = [f for f in model["fields"] if f["flag"] == FLAG_DERIVED_FIELD]

    def field_list(items: List[Dict[str, Any]]) -> str:
        return ", ".join(f"`{item['canonical_field']}`" for item in items) if items else "(none)"

    lines.extend(
        [
            "",
            "## Decisions required from the reviewer",
            "",
            f"- Verify-and-decide single candidates: {field_list(single)}",
            f"- Select one of multiple plausible candidates: {field_list(multiple)}",
            f"- No source candidate — business treatment decision needed: {field_list(missing)}",
            f"- Derived fields needing an approved value map: {field_list(derived)}",
            "",
            "## Standing constraints",
            "",
            "- Profile lifecycle status stays `DRAFT`; no approval is inferred from this package.",
            "- No mapping decision is written back to the profile until human approval is recorded.",
            "- ECC output remains blocked by the existing fail-closed guards.",
        ]
    )
    return "\n".join(lines) + "\n"


def build_decision_package(
    *,
    profiler_dir: Path,
    du_profile_path: Path,
    output_dir: Path,
    workbook_name: str,
    summary_name: str,
    shortlist_registry_path: Optional[Path] = None,
    source_export_path: Optional[Path] = None,
    review_fields: Sequence[str] = DEFAULT_REVIEW_FIELDS,
) -> Dict[str, Any]:
    candidates_report = json.loads((profiler_dir / "canonical_field_candidates.json").read_text(encoding="utf-8"))
    header_inventory = json.loads((profiler_dir / "header_inventory.json").read_text(encoding="utf-8"))
    observed_header_hash = (profiler_dir / "header_hash.txt").read_text(encoding="utf-8").strip()
    du_profile = json.loads(du_profile_path.read_text(encoding="utf-8"))
    shortlist_registry = None
    if shortlist_registry_path is not None and shortlist_registry_path.exists():
        shortlist_registry = json.loads(shortlist_registry_path.read_text(encoding="utf-8"))

    if source_export_path is not None:
        digest = hashlib.sha256(source_export_path.read_bytes()).hexdigest()
        expected = header_inventory.get("source", {}).get("source_file_hash", "")
        if expected and digest != expected:
            raise ValueError(
                "Source export hash does not match the profiled export; refusing to build a review "
                f"package from mismatched evidence (expected {expected}, got {digest})."
            )

    model = build_review_model(
        candidates_report=candidates_report,
        header_inventory=header_inventory,
        observed_header_hash=observed_header_hash,
        du_profile=du_profile,
        shortlist_registry=shortlist_registry,
        source_export_path=source_export_path,
        review_fields=review_fields,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    write_decision_workbook(model, output_dir / workbook_name)
    (output_dir / summary_name).write_text(decision_summary_markdown(model), encoding="utf-8")
    return model


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--profiler-dir", required=True, type=Path)
    parser.add_argument("--du-profile", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--workbook-name", default="MAPPING_DECISION_WORKBOOK.xlsx")
    parser.add_argument("--summary-name", default="MAPPING_DECISION_SUMMARY.md")
    parser.add_argument(
        "--shortlist-registry",
        type=Path,
        default=Path("config/registries/mw_du_priority_skill_field_shortlists.yaml"),
    )
    parser.add_argument(
        "--source-export",
        type=Path,
        default=None,
        help="Local, uncommitted source export used read-only for masked sample values.",
    )
    args = parser.parse_args()
    model = build_decision_package(
        profiler_dir=args.profiler_dir,
        du_profile_path=args.du_profile,
        output_dir=args.output_dir,
        workbook_name=args.workbook_name,
        summary_name=args.summary_name,
        shortlist_registry_path=args.shortlist_registry,
        source_export_path=args.source_export,
    )
    print(
        f"Wrote local-only mapping-decision package for {model['profile_id']} "
        f"({len(model['fields'])} canonical fields) to {args.output_dir}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
