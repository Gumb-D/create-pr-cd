#!/usr/bin/env python3
"""Compare current and candidate PR Model workbooks for production compatibility."""
from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SUPPORTED_SHEET_PREFIX = "TX Line Item"
SECTION_HEADERS = {
    "TSS Model": "TSS",
    "TI Model": "TI",
    "Add PR Line Item Model": "ADDITIONAL",
}


def _norm(value: Any) -> Any:
    if isinstance(value, str):
        return " ".join(value.split())
    return value


def _sheet(workbook):
    matches = [name for name in workbook.sheetnames if str(name).startswith(SUPPORTED_SHEET_PREFIX)]
    if len(matches) != 1:
        raise ValueError(f"PR_MODEL_STRUCTURE_INVALID: expected exactly one '{SUPPORTED_SHEET_PREFIX}' sheet, found {matches}")
    return workbook[matches[0]]


def extract_business_rows(path: Path | str) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _sheet(workbook)
        section = None
        rows: list[tuple[Any, ...]] = []
        for values in ws.iter_rows(values_only=True):
            padded = list(values[:8]) + [None] * max(0, 8 - len(values))
            first = str(padded[0]).strip() if padded[0] is not None else ""
            if first in SECTION_HEADERS:
                section = SECTION_HEADERS[first]
                continue
            if section is None:
                continue
            code = padded[1]
            model = padded[0]
            if model in (None, "") or code in (None, ""):
                continue
            rows.append((section, *tuple(_norm(value) for value in padded[:8])))
        return rows
    finally:
        workbook.close()


def _row_payload(row: tuple[Any, ...]) -> dict[str, Any]:
    section, sow, code, description, unit, quantity, rules, remarks, remarks2 = row
    return {
        "section": section,
        "sow": sow,
        "code": str(code),
        "description": description,
        "unit": unit,
        "quantity": quantity,
        "rules": rules,
        "remarks": remarks,
        "remarks2": remarks2,
    }


def analyze_pr_model_change(current_path: Path | str, candidate_path: Path | str) -> dict[str, Any]:
    current_rows = extract_business_rows(current_path)
    candidate_rows = extract_business_rows(candidate_path)
    current_counter = Counter(current_rows)
    candidate_counter = Counter(candidate_rows)

    removed: list[tuple[Any, ...]] = []
    added: list[tuple[Any, ...]] = []
    for row, count in (current_counter - candidate_counter).items():
        removed.extend([row] * count)
    for row, count in (candidate_counter - current_counter).items():
        added.extend([row] * count)

    current_sows = {str(row[1]) for row in current_rows if row[0] in {"TSS", "TI"}}
    candidate_sows = {str(row[1]) for row in candidate_rows if row[0] in {"TSS", "TI"}}
    new_sows = sorted(candidate_sows - current_sows)

    reason_codes: list[str] = []
    if removed:
        reason_codes.append("REMOVED_BUSINESS_ROWS")
    if new_sows:
        reason_codes.append("NEW_SOW")

    status = "REVIEW_REQUIRED" if reason_codes else "COMPATIBLE"
    return {
        "status": status,
        "reason_codes": reason_codes,
        "current_row_count": len(current_rows),
        "candidate_row_count": len(candidate_rows),
        "removed_count": len(removed),
        "added_count": len(added),
        "new_sows": new_sows,
        "removed_rows": [_row_payload(row) for row in removed],
        "added_rows": [_row_payload(row) for row in added],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare current and candidate PR Model workbooks.")
    parser.add_argument("current")
    parser.add_argument("candidate")
    parser.add_argument("--output")
    args = parser.parse_args()
    try:
        report = analyze_pr_model_change(Path(args.current), Path(args.candidate))
    except Exception as exc:
        report = {"status": "INVALID", "reason_codes": ["PR_MODEL_STRUCTURE_INVALID"], "message": str(exc)}
    rendered = json.dumps(report, ensure_ascii=False, indent=2)
    if args.output:
        Path(args.output).write_text(rendered + "\n", encoding="utf-8")
    print(rendered)
    return 0 if report["status"] == "COMPATIBLE" else 2


if __name__ == "__main__":
    raise SystemExit(main())
