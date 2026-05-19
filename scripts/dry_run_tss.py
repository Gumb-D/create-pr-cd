import openpyxl
import pandas as pd
from pathlib import Path
import json

# File paths
info_folder = Path("Info")
site_data_file = info_folder / "input" / "site_pr_po_view.xlsx"
pr_model_file = info_folder / "input" / "pr_model.xlsx"
ecc_template_file = info_folder / "input" / "ecc_template.xls"

print("=" * 80)
print("DRY RUN: TSS PR Generation")
print("=" * 80)

# Load site data
print("\n[1] Loading Site Data Reference...")
try:
    site_df = pd.read_excel(site_data_file, sheet_name=0)
    print(f"✓ Loaded {len(site_df)} site records")
    print(f"\nColumn names:")
    for i, col in enumerate(site_df.columns, 1):
        print(f"  {i}. {col}")
except Exception as e:
    print(f"✗ Error loading site data: {e}")
    exit(1)

# Load PR Model Reference
print("\n[2] Loading PR Model Reference...")
try:
    wb = openpyxl.load_workbook(pr_model_file)
    print(f"✓ Available sheets: {wb.sheetnames}")
    
    # Try to load TSS sheet
    if "TSS" in wb.sheetnames:
        tss_sheet = wb["TSS"]
        print(f"\n✓ TSS Sheet found with {tss_sheet.max_row - 1} rows")
        print(f"  First row (headers):")
        for i, cell in enumerate(tss_sheet[1], 1):
            if cell.value:
                print(f"    {i}. {cell.value}")
except Exception as e:
    print(f"✗ Error loading PR model: {e}")

# Load ECC Template
print("\n[3] Loading ECC Template...")
try:
    ecc_wb = openpyxl.load_workbook(ecc_template_file)
    ecc_sheet = ecc_wb.active
    print(f"✓ ECC Template sheet: {ecc_sheet.title}")
    print(f"  ECC columns:")
    for i, cell in enumerate(ecc_sheet[1], 1):
        if cell.value:
            print(f"    {i}. {cell.value}")
except Exception as e:
    print(f"✗ Error loading ECC template: {e}")

# Identify TSS trigger columns
print("\n[4] Identifying TSS Trigger Columns...")
target_cols = ['SubCon - TSS Team', 'SubCon - TI Team', 'Subcon - Planning', 'TX Integrated actual end date']
found_cols = [col for col in target_cols if col in site_df.columns]
print(f"✓ Found columns: {found_cols}")

# Find TSS PR status/number column
tss_pr_cols = [col for col in site_df.columns if 'TSS' in col and ('PR' in col or 'Status' in col or 'Number' in col)]
print(f"✓ TSS PR status/number columns: {tss_pr_cols}")

# Display first 5 sites with TSS trigger
print("\n[5] Identifying TSS PR Triggers (First 5 sites with TSS trigger)...")
print("\nKey columns to check:")
print("  - SubCon - TSS Team (must be non-blank)")
print("  - Tx SOW (for SOW matching)")
print("  - Region")
print("  - Site ID, Site Name")

# Display sample data
display_cols = ['Site ID', 'Site Name', 'Region', 'Tx SOW', 'SubCon - TSS Team'] + tss_pr_cols
available_cols = [col for col in display_cols if col in site_df.columns]

print("\n" + "=" * 80)
print("SAMPLE DATA - First 5 records:")
print("=" * 80)
print(site_df[available_cols].head(10).to_string())

# Filter TSS candidates (non-blank SubCon - TSS Team)
print("\n" + "=" * 80)
print("TSS PR CANDIDATES (SubCon - TSS Team is not blank):")
print("=" * 80)

if 'SubCon - TSS Team' in site_df.columns:
    tss_candidates = site_df[site_df['SubCon - TSS Team'].notna() & (site_df['SubCon - TSS Team'] != '')]
    print(f"\nTotal TSS candidates: {len(tss_candidates)}")
    
    if len(tss_candidates) > 0:
        print("\nFirst 5 TSS candidates:")
        for idx, (i, row) in enumerate(tss_candidates.head(5).iterrows(), 1):
            print(f"\n  [{idx}] Site ID: {row.get('Site ID', 'N/A')}")
            print(f"      Site Name: {row.get('Site Name', 'N/A')}")
            print(f"      Region: {row.get('Region', 'N/A')}")
            print(f"      Tx SOW: {row.get('Tx SOW', 'N/A')}")
            print(f"      SubCon - TSS Team: {row.get('SubCon - TSS Team', 'N/A')}")
            if tss_pr_cols:
                for col in tss_pr_cols:
                    status = row.get(col, 'N/A')
                    print(f"      {col}: {status if status else '[EMPTY - Can generate]'}")

print("\n" + "=" * 80)
print("DRY RUN SUMMARY")
print("=" * 80)
print("\nKey findings:")
print(f"  - Total site records: {len(site_df)}")
print(f"  - Total TSS candidates: {len(tss_candidates) if 'SubCon - TSS Team' in site_df.columns else 'N/A'}")
print(f"  - TSS PR status column: {tss_pr_cols[0] if tss_pr_cols else 'Not found'}")
print("\nNext steps:")
print("  1. Match Tx SOW to TSS PR models")
print("  2. Extract mandatory line items for each matched model")
print("  3. Retrieve contract number and purchasing area")
print("  4. Generate ECC output rows grouped by Region + SubCon")
print("  5. Create output files: <Region>-<Subcon> TX Mini Project TSS PR <YYYYMMDD>.xls")
