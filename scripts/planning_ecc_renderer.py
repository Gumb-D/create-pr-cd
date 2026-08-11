#!/usr/bin/env python3
"""Deterministic ECC renderer for the Issue #34 Planning scope.

The official create-pr entrypoint supplies already-gated canonical candidates.
This renderer reuses the shared Planning selector at the subprocess boundary so
no PBOM decision is duplicated or trusted from a flat workbook without
verification.
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from planning_pr_selector import select_planning_item


ECC_HEADERS = [
    "SN.",
    "Purchasing Area*",
    "Region*",
    "Site ID*",
    "Site Name*",
    "Delivery Unit Code*",
    "Logical Site Name",
    "Contract Number *",
    "Subcontractor*",
    "PBOM Code*",
    "SOW*",
    "Unit*",
    "Quantity*",
    "Remarks",
    "",
    "Contract Number",
]

REQUIRED_INPUT_COLUMNS = {
    "customer site code",
    "customer site name",
    "du code",
    "region",
    "Subcon - Planning",
}

DISALLOWED_CONTRACT_VALUES = {"", "N/A", "NA", "NONE", "NULL", "TBD", "TO BE CONFIRMED"}
MAX_SITES_PER_FILE = 30


class PlanningRendererError(RuntimeError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render deterministic Planning PR ECC output.")
    parser.add_argument("--site-data", required=True, type=Path)
    parser.add_argument("--mapping", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--scope", required=True, type=str)
    parser.add_argument("--du-model-name", required=True, type=str)
    parser.add_argument("--all-sites", action="store_true")
    parser.add_argument("--site-code")
    # Compatibility with create_pr_impl renderer invocation. Planning does not
    # read the PR Model or ECC template to select its deterministic line item.
    parser.add_argument("--pr-model", type=Path)
    parser.add_argument("--template", type=Path)
    return parser.parse_args()


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split())


def _parse_markdown_tables(path: Path) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    if not path.exists():
        raise PlanningRendererError("MAPPING_FILE_NOT_FOUND", f"Mapping reference not found: {path}")
    region_mapping: dict[str, str] = {}
    contract_mapping: dict[str, dict[str, str]] = {}
    section = ""
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if line.startswith("## Region to Purchasing Area"):
            section = "region"
            continue
        if line.startswith("## Subcontractor to Contract Number"):
            section = "contract"
            continue
        if line.startswith("## "):
            section = ""
            continue
        if not line.startswith("|") or "---" in line:
            continue
        parts = [part.strip() for part in line.strip("|").split("|")]
        if section == "region" and len(parts) >= 2 and parts[0] != "Region*":
            if parts[0] and parts[1]:
                region_mapping[_normalize(parts[0]).casefold()] = parts[1]
        elif section == "contract" and len(parts) >= 2 and parts[0] != "Subcontractor*":
            if parts[0] and parts[1]:
                contract_mapping[_normalize(parts[0]).casefold()] = {
                    "subcontractor": _normalize(parts[0]),
                    "contract_number": _normalize(parts[1]),
                }
    return region_mapping, contract_mapping


def _find_header_row(worksheet) -> tuple[int, list[str]]:
    for row_number in range(1, min(10, worksheet.max_row) + 1):
        values = ["" if value is None else str(value).strip() for value in next(
            worksheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True)
        )]
        if REQUIRED_INPUT_COLUMNS.issubset(set(values)):
            return row_number, values
    raise PlanningRendererError(
        "CANONICAL_INPUT_SCHEMA_ERROR",
        "Planning renderer input does not contain the required canonical Planning columns.",
    )


def _load_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise PlanningRendererError("SITE_DATA_NOT_FOUND", f"Site data not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "data" not in workbook.sheetnames:
            raise PlanningRendererError("CANONICAL_INPUT_SCHEMA_ERROR", "Planning renderer requires a data sheet.")
        worksheet = workbook["data"]
        header_row, headers = _find_header_row(worksheet)
        index = {header: position for position, header in enumerate(headers) if header}
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            record = {
                header: values[position] if position < len(values) else None
                for header, position in index.items()
            }
            if not any(_normalize(value) for value in record.values()):
                continue
            rows.append(record)
        return rows
    finally:
        workbook.close()


def _filter_sites(rows: list[dict[str, Any]], all_sites: bool, raw_site_codes: str | None) -> list[dict[str, Any]]:
    requested = list(dict.fromkeys(
        token.strip().upper()
        for token in str(raw_site_codes or "").split(",")
        if token.strip()
    ))
    if all_sites == bool(requested):
        raise PlanningRendererError("INVALID_SITE_SELECTION", "Use exactly one of --all-sites or --site-code.")
    if all_sites:
        return rows
    available = {_normalize(row.get("customer site code")).upper() for row in rows}
    missing = [code for code in requested if code not in available]
    if missing:
        raise PlanningRendererError("SITE_CODES_NOT_FOUND", "Requested site codes were not found: " + ", ".join(missing))
    wanted = set(requested)
    return [row for row in rows if _normalize(row.get("customer site code")).upper() in wanted]


def _verify_optional_pipeline_value(row: Mapping[str, Any], column: str, expected: Any) -> None:
    if column not in row:
        return
    actual = _normalize(row.get(column))
    if not actual:
        return
    if _normalize(expected) != actual:
        raise PlanningRendererError(
            "PLANNING_SELECTOR_PIPELINE_MISMATCH",
            f"{column} does not match the governed Planning selector result.",
        )


def _build_ecc_rows(
    rows: list[dict[str, Any]],
    du_model_name: str,
    region_mapping: Mapping[str, str],
    contract_mapping: Mapping[str, Mapping[str, str]],
) -> list[dict[str, Any]]:
    rendered: list[dict[str, Any]] = []
    for row in rows:
        site_id = _normalize(row.get("customer site code"))
        region = _normalize(row.get("region"))
        source_subcontractor = _normalize(row.get("Subcon - Planning"))
        if not site_id:
            raise PlanningRendererError("SITE_ID_MISSING", "Planning candidate has a blank Site ID.")
        if not region:
            raise PlanningRendererError("REGION_MISSING", f"Planning candidate {site_id} has a blank Region.")

        selection = select_planning_item(du_model_name, source_subcontractor)
        if selection.status != "RESOLVED":
            raise PlanningRendererError(
                str(selection.reason_code or "PLANNING_SELECTION_UNRESOLVED"),
                f"Planning selection is not resolved for site {site_id}.",
            )

        purchasing_area = region_mapping.get(region.casefold())
        if not purchasing_area:
            raise PlanningRendererError(
                "PURCHASING_AREA_NOT_FOUND",
                f"No Purchasing Area mapping exists for Region {region!r} (site {site_id}).",
            )

        contract_subcontractor = str(selection.contract_subcontractor or "")
        contract = contract_mapping.get(contract_subcontractor.casefold())
        contract_number = _normalize((contract or {}).get("contract_number"))
        if not contract or contract_number.upper() in DISALLOWED_CONTRACT_VALUES:
            raise PlanningRendererError(
                "CONTRACT_MAPPING_MISSING",
                f"No approved contract mapping exists for {contract_subcontractor!r} (site {site_id}).",
            )

        _verify_optional_pipeline_value(row, "Planning Contract Subcontractor", contract_subcontractor)
        _verify_optional_pipeline_value(row, "Planning PBOM Code", selection.pbom_code)
        _verify_optional_pipeline_value(row, "Planning SOW", selection.description)
        _verify_optional_pipeline_value(row, "Planning Unit", selection.unit)
        _verify_optional_pipeline_value(row, "Planning Quantity", selection.quantity)

        rendered.append(
            {
                "Purchasing_Area": purchasing_area,
                "Region": region,
                "Site_ID": site_id,
                "Site_Name": _normalize(row.get("customer site name")),
                "Delivery_Unit_Code": _normalize(row.get("du code")),
                "Logical_Site_Name": _normalize(row.get("customer site name")),
                "Contract_Number": contract_number,
                "Subcontractor": contract_subcontractor,
                "PBOM_Code": str(selection.pbom_code),
                "SOW": str(selection.description),
                "Unit": str(selection.unit),
                "Quantity": int(selection.quantity or 0),
                "Remarks": "",
                "Source_Tx_SOW": "",
            }
        )
    return rendered


def _safe_filename_component(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*]+', " ", str(value)).strip().strip(".")


def _split_group(rows: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    site_ids = sorted({str(row["Site_ID"]) for row in rows})
    parts: list[list[dict[str, Any]]] = []
    for offset in range(0, len(site_ids), MAX_SITES_PER_FILE):
        sites = set(site_ids[offset : offset + MAX_SITES_PER_FILE])
        parts.append([row for row in rows if str(row["Site_ID"]) in sites])
    return parts


def _write_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "details"
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    for column, header in enumerate(ECC_HEADERS, 1):
        cell = worksheet.cell(1, column, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for serial, row in enumerate(rows, 1):
        values = [
            serial,
            row["Purchasing_Area"],
            row["Region"],
            row["Site_ID"],
            row["Site_Name"],
            row["Delivery_Unit_Code"],
            row["Logical_Site_Name"],
            row["Contract_Number"],
            row["Subcontractor"],
            row["PBOM_Code"],
            row["SOW"],
            row["Unit"],
            row["Quantity"],
            row["Remarks"],
            row["Source_Tx_SOW"],
            row["Contract_Number"],
        ]
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(serial + 1, column, value)
            if column in {1, 13}:
                cell.alignment = Alignment(horizontal="center")

    widths = [5, 20, 12, 15, 20, 15, 20, 18, 15, 15, 40, 8, 10, 15, 5, 18]
    for column, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    workbook.save(path)
    workbook.close()


def render(args: argparse.Namespace) -> list[Path]:
    if _normalize(args.scope).upper() != "PLANNING":
        raise PlanningRendererError("INVALID_SCOPE", "Planning renderer accepts only --scope Planning.")
    du_model_name = _normalize(args.du_model_name)
    if not du_model_name:
        raise PlanningRendererError("DU_MODEL_NAME_MISSING", "Resolved DU Model name is blank.")

    rows = _filter_sites(_load_rows(args.site_data), args.all_sites, args.site_code)
    region_mapping, contract_mapping = _parse_markdown_tables(args.mapping)
    ecc_rows = _build_ecc_rows(rows, du_model_name, region_mapping, contract_mapping)

    # Validate the complete candidate set before creating any file. A bad row
    # therefore cannot leave partial Planning ECC output behind.
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in ecc_rows:
        grouped[(row["Region"], row["Subcontractor"])].append(row)

    output = Path(args.output)
    output.mkdir(parents=True, exist_ok=True)
    created: list[Path] = []
    date_stamp = datetime.now().strftime("%Y%m%d")
    safe_du = _safe_filename_component(du_model_name)
    for (region, subcontractor), group_rows in sorted(grouped.items()):
        parts = _split_group(group_rows)
        for part_number, part_rows in enumerate(parts, 1):
            suffix = f" Part {part_number}" if len(parts) > 1 else ""
            filename = (
                f"{_safe_filename_component(region)}-{_safe_filename_component(subcontractor)} "
                f"{safe_du} Planning PR {date_stamp}{suffix}.xlsx"
            )
            path = output / filename
            _write_workbook(path, part_rows)
            created.append(path)
    return created


def main() -> int:
    try:
        created = render(parse_args())
        for path in created:
            print(f"[OK] Created: {path}")
        return 0
    except PlanningRendererError as error:
        print(f"ERROR: {error.code}: {error}", file=sys.stderr)
        return 1
    except Exception as error:
        print(f"ERROR: PLANNING_RENDERER_FAILED: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
