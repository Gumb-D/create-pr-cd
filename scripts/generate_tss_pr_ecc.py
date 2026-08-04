#!/usr/bin/env python3
"""
TSS/TI PR ECC File Generation - Amendment Implementation
Amended version incorporating:
- Amendment 1: Separate contract info reference (Markdown)
- Amendment 2: Single 'details' sheet only
- Amendment 3: Sequential SN, Region->Purchasing Area, Subcon->Contract, 30-site split, fuzzy matching
- TI Phase 1: Trigger hardening, antenna parser, REVIEW_REQUIRED framework, duplicate prevention
"""

import argparse
import os
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from difflib import SequenceMatcher
import re
import csv
from geography_resolver import GeographyResolver


def _console_safe_text(value: object, encoding: str | None = None) -> str:
    text = str(value)
    encoding = encoding or getattr(sys.stdout, "encoding", None)
    if not encoding:
        return text
    try:
        text.encode(encoding)
        return text
    except UnicodeEncodeError:
        return text.encode(encoding, errors="backslashreplace").decode(encoding)


def _safe_print(*args, file=None, **kwargs):
    if file is None:
        file = sys.stdout
    encoding = getattr(file, "encoding", None) or getattr(sys.stdout, "encoding", None)
    safe_args = tuple(_console_safe_text(arg, encoding) for arg in args)
    print(*safe_args, file=file, **kwargs)
from pr_helpers import (
    normalize_pbom_code,
    normalize_ti_sow,
    ti_sow_matches_model,
    is_mw_reroute_row,
    parse_mw_new_link_reroute,
    filter_tss_mw_new_link_reroute_items,
    select_tss_items_for_site,
    has_duplicate_pbom,
    validate_required_pbom_selection,
    filter_failed_migration_decisions,
    optional_cell_text,
    load_pr_model_items,
)


def parse_args():
    parser = argparse.ArgumentParser(description="Generate ECC files from site PR/PO view and PR model.")
    parser.add_argument('--site-data', default='Info/input/site_pr_po_view.xlsx', help='Path to daily site PR/PO view file')
    parser.add_argument('--pr-model', default='Info/input/pr_model.xlsx', help='Path to PR model Excel file')
    parser.add_argument('--template', default='Info/input/ecc_template.xls', help='Path to ECC template file')
    parser.add_argument('--mapping', default='Info/input/contract_info_reference.md', help='Path to region/subcontractor mapping markdown')
    default_output = str(Path(__file__).resolve().parent.parent / 'output')
    parser.add_argument('--output', default=default_output, help='Output directory for generated ECC files (default: <skill-root>/output)')
    parser.add_argument('--site-code', help='Comma-separated Site Code(s) to generate PR ECC for')
    parser.add_argument('--all-sites', action='store_true', help='Generate PR ECC for all eligible sites')
    parser.add_argument('--scope', choices=['TSS', 'TI'], default='TSS', type=str.upper, help='PR scope to generate: TSS or TI')
    parser.add_argument('--du-model-name', default='TX Mini Project', help='Resolved DU model name for ECC output filenames')
    return parser.parse_args()


import hashlib

APPROVED_PR_MODEL_SHA256 = "d3cc64664fc147f8c560688e41264753592eb0b8cdc513d7ebe2d9b989e8aefd"


def validate_pr_model_file(path, description="PR Model"):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"{description} not found: {p}")
    actual_sha = hashlib.sha256(p.read_bytes()).hexdigest()
    if actual_sha != APPROVED_PR_MODEL_SHA256:
        raise ValueError(
            f"PR_MODEL_HASH_MISMATCH: {description} content hash mismatch!\n"
            f"Expected: {APPROVED_PR_MODEL_SHA256}\n"
            f"Actual:   {actual_sha}\n"
            f"Use the officially approved PR Model v4 workbook."
        )
    return p


def require_file(path, description):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(
            f"{description} not found: {p}\n" \
            f"Place the correct file at this path or use the --site-data/--pr-model/--template/--mapping options."
        )
    return p


def validate_template_file(path):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"ECC Template not found: {path}")
    if p.suffix.lower() == '.xlsx':
        wb = load_workbook(path, read_only=True, data_only=True)
        if 'details' not in wb.sheetnames:
            raise ValueError(f"ECC template must contain a 'details' sheet: {path}")
    return True


args = parse_args()
output_dir = Path(args.output)
output_dir.mkdir(parents=True, exist_ok=True)

print("=" * 100)
print(f"{args.scope.upper()} PR ECC GENERATION - AMENDED IMPLEMENTATION")
print("=" * 100)
print(f"Execution Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

# ===== HELPER FUNCTIONS =====

def load_region_mapping(ref_file):
    """Load Region to Purchasing Area mapping from reference file."""
    mapping = {}
    with open(ref_file, 'r') as f:
        lines = f.readlines()
    in_section = False
    for line in lines:
        if '## Region to Purchasing Area' in line:
            in_section = True
            continue
        if in_section and line.startswith('## '):
            break
        if in_section and '|' in line:
            # Skip header row (contains "Region*") and separator row (contains "---")
            if 'Region*' in line or '---' in line:
                continue
            parts = [p.strip() for p in line.split('|')]
            # parts = ['', region, purchasing_area, '']
            if len(parts) >= 3 and parts[1] and parts[2] and parts[1] != '':
                mapping[parts[1]] = parts[2]
    return mapping

def load_subcon_mapping(ref_file):
    """Load Subcontractor to Contract mapping from reference file."""
    mapping = {}
    with open(ref_file, 'r') as f:
        lines = f.readlines()
    in_section = False
    for line in lines:
        if '## Subcontractor to Contract Number' in line:
            in_section = True
            continue
        if in_section and line.startswith('## '):
            break
        if in_section and '|' in line:
            # Skip header row (contains "Subcontractor*") and separator row (contains "---")
            if 'Subcontractor*' in line or '---' in line:
                continue
            parts = [p.strip() for p in line.split('|')]
            # parts = ['', subcon, contract_number, company_name, '']
            if len(parts) >= 4 and parts[1] and parts[2] and parts[1] != '':
                mapping[parts[1]] = {
                    'contract_number': parts[2],
                    'company_name': parts[3]
                }
    return mapping

def fuzzy_match_subcon(subcon_name, subcon_mapping, threshold=0.6):
    """
    Fuzzy match a subcontractor name against known list.
    Returns matched subcon key or None.
    """
    if subcon_name in subcon_mapping:
        return subcon_name  # Exact match
    
    best_match = None
    best_score = threshold
    
    for known_subcon in subcon_mapping.keys():
        score = SequenceMatcher(None, subcon_name.lower(), known_subcon.lower()).ratio()
        if score > best_score:
            best_score = score
            best_match = known_subcon
    
    return best_match


def is_mw_reengineering_sow(sow):
    normalized = str(sow or '').strip().lower()
    return 'mw re-engineering' in normalized or normalized.startswith('mw re-eng')


def make_review_reason(reason_code, category='PR Model Matching', description=None, action=None, technical_detail=None):
    descriptions = {
        'MISSING_TX_SOW': 'The site does not have a Tx SOW value, so the TI PR model cannot be selected.',
        'MISSING_TI_ANTENNA_SIZE': 'The site does not have enough TI antenna size information to select the mandatory antenna PR item.',
        'NO_MATCHING_TI_PR_MODEL_ITEM': 'No TI PR model item matches the site Tx SOW.',
        'NO_MANDATORY_TI_ITEM_FOUND': 'The matching TI PR model rows do not contain mandatory items that can be generated automatically.',
        'NO_MATCHING_ANTENNA_GROUP_ITEM': 'The mandatory antenna choose-one group could not be matched to the site antenna size.',
        'MW_REROUTE_INSTALL_ANTENNA_SIZE_MISSING': 'The MW Reroute install antenna size could not be determined from the site data.',
        'MW_REROUTE_DECOM_ANTENNA_SIZE_AMBIGUOUS': 'The MW Reroute dismantle antenna size is ambiguous in the site data.',
        'MW_REROUTE_DECOM_ANTENNA_SIZE_MISSING': 'The MW Reroute dismantle antenna size could not be determined from the site data.',
        'MW_REROUTE_INSTALL_ITEM_NOT_MATCHED': 'The MW Reroute install mandatory PR item could not be matched.',
        'MW_REROUTE_DISMANTLE_ITEM_NOT_MATCHED': 'The MW Reroute dismantle mandatory PR item could not be matched.',
        'MATERIAL_CODE_NOT_FOUND': 'The geography route was resolved, but the resolved material code is not present in the matching PR model group.',
        'DUPLICATE_MATERIAL_CODE': 'The geography route was resolved, but more than one PR model item has the resolved material code.',
        'COORDINATE_RESOLUTION_UNSUPPORTED': 'The site is in Sabah or Sarawak, but latitude/longitude cannot yet be converted into a confirmed city or warehouse route.',
        'MISSING_COORDINATES': 'Latitude or longitude is missing from the iEPMS site record.',
        'INVALID_COORDINATES': 'Latitude or longitude is non-numeric or outside the valid range.',
        'COORDINATE_OUTSIDE_SUPPORTED_BOUNDARY': 'The coordinate does not fall within a supported Sabah or Sarawak administrative boundary.',
        'AMBIGUOUS_DISTRICT_BOUNDARY': 'The coordinate falls on or close to multiple administrative boundaries.',
        'RESOLVED_STATE_MISMATCH': 'The coordinate resolves to a different state from the declared site region or state.',
        'ROUTE_MAPPING_MISSING': 'The district was resolved geographically, but no confirmed business route exists.',
        'WAREHOUSE_MAPPING_MISSING': 'The route was resolved, but the warehouse mapping is missing or unconfirmed.',
        'MATERIAL_CODE_MAPPING_MISSING': 'The route was resolved, but the material code mapping is missing or unconfirmed.',
        'LAWAS_SIMPLE_PACKING_UNCONFIRMED': 'Lawas was resolved geographically, but its Simple Packing business mapping is not confirmed.'
    }
    actions = {
        'MISSING_TX_SOW': 'Confirm the Tx SOW for this site.',
        'MISSING_TI_ANTENNA_SIZE': 'Confirm the TI antenna size and rerun generation.',
        'NO_MATCHING_TI_PR_MODEL_ITEM': 'Confirm the Tx SOW and update the PR model if this SOW should be supported.',
        'NO_MANDATORY_TI_ITEM_FOUND': 'Confirm the mandatory TI PR model rows for this SOW.',
        'NO_MATCHING_ANTENNA_GROUP_ITEM': 'Confirm the antenna size and mandatory antenna group selection.',
        'MW_REROUTE_INSTALL_ANTENNA_SIZE_MISSING': 'Confirm the MW Reroute install antenna size.',
        'MW_REROUTE_DECOM_ANTENNA_SIZE_AMBIGUOUS': 'Confirm the MW Reroute dismantle antenna size.',
        'MW_REROUTE_DECOM_ANTENNA_SIZE_MISSING': 'Confirm the MW Reroute dismantle antenna size.',
        'MW_REROUTE_INSTALL_ITEM_NOT_MATCHED': 'Confirm the install antenna size and matching PR model row.',
        'MW_REROUTE_DISMANTLE_ITEM_NOT_MATCHED': 'Confirm the dismantle antenna size and matching PR model row.',
        'MATERIAL_CODE_NOT_FOUND': 'Confirm the route mapping and PR model material code.',
        'DUPLICATE_MATERIAL_CODE': 'Confirm which duplicate PR model material row should be used.',
        'COORDINATE_RESOLUTION_UNSUPPORTED': 'Confirm the site city/district and applicable warehouse/material code.',
        'MISSING_COORDINATES': 'Update site coordinates and rerun PR generation.',
        'INVALID_COORDINATES': 'Correct the site latitude and longitude values and rerun PR generation.',
        'COORDINATE_OUTSIDE_SUPPORTED_BOUNDARY': 'Confirm the site coordinates and supported administrative boundary coverage.',
        'AMBIGUOUS_DISTRICT_BOUNDARY': 'Confirm the correct district or city before generating PR.',
        'RESOLVED_STATE_MISMATCH': 'Confirm the declared region/state and the site coordinates.',
        'ROUTE_MAPPING_MISSING': 'Obtain a confirmed city/district-to-route bucket mapping from the business owner.',
        'WAREHOUSE_MAPPING_MISSING': 'Obtain a confirmed warehouse mapping from the business owner.',
        'MATERIAL_CODE_MAPPING_MISSING': 'Obtain a confirmed material code from the business owner.',
        'LAWAS_SIMPLE_PACKING_UNCONFIRMED': 'Obtain confirmed Lawas Simple Packing warehouse and material code from the business owner.'
    }
    return {
        'Reason_Category': category,
        'Reason_Code': reason_code,
        'Reason_Description': description or descriptions.get(reason_code, reason_code.replace('_', ' ').title()),
        'Required_Action': action or actions.get(reason_code, 'Review the site input and PR model mapping.'),
        'Technical_Detail': technical_detail or reason_code,
        'Review_Reason': technical_detail or reason_code
    }


def normalize_reason_code(text):
    normalized = re.sub(r'[^A-Z0-9]+', '_', str(text or '').upper()).strip('_')
    replacements = {
        'MISSING_TI_ANTENNA_SIZE_REVIEW_REQUIRED': 'MISSING_TI_ANTENNA_SIZE',
        'INCOMPLETE_TI_ANTENNA_SIZE_REVIEW_REQUIRED': 'MISSING_TI_ANTENNA_SIZE',
        'NO_MATCHING_TI_PR_MODEL_ITEM': 'NO_MATCHING_TI_PR_MODEL_ITEM',
        'NO_MANDATORY_TI_ITEM_FOUND': 'NO_MANDATORY_TI_ITEM_FOUND',
        'NO_MATCHING_ANTENNA_GROUP_ITEM': 'NO_MATCHING_ANTENNA_GROUP_ITEM',
        'MW_REROUTE_DECOM_ANTENNA_SIZE_AMBIGUOUS': 'MW_REROUTE_DECOM_ANTENNA_SIZE_AMBIGUOUS',
        'MW_REROUTE_DECOM_ANTENNA_SIZE_MISSING': 'MW_REROUTE_DECOM_ANTENNA_SIZE_MISSING',
        'MW_REROUTE_INSTALL_ANTENNA_SIZE_MISSING': 'MW_REROUTE_INSTALL_ANTENNA_SIZE_MISSING',
        'MW_REROUTE_INSTALL_ITEM_NOT_MATCHED': 'MW_REROUTE_INSTALL_ITEM_NOT_MATCHED',
        'MW_REROUTE_DISMANTLE_ITEM_NOT_MATCHED': 'MW_REROUTE_DISMANTLE_ITEM_NOT_MATCHED'
    }
    return replacements.get(normalized, normalized or 'REVIEW_REQUIRED')


def route_error_technical_detail(err, row):
    parts = [
        f"route_type={err.get('route_type')}",
        f"reason={err.get('reason_code')}",
        f"bucket={err.get('bucket') or ''}",
        f"region={row.get('region', '') if row is not None else ''}"
    ]
    if err.get('state'):
        parts.append(f"resolved_state={err.get('state')}")
    if err.get('city_or_district'):
        parts.append(f"resolved_city_or_district={err.get('city_or_district')}")
    if err.get('warehouse'):
        parts.append(f"warehouse={err.get('warehouse')}")
    if err.get('material_code'):
        parts.append(f"material_code={err.get('material_code')}")
    if row is not None:
        lat = row.get('Latitude (North Plus South Minus)', '')
        lon = row.get('Longitude (East Plus West Minus)', '')
        if pd.notna(lat) or pd.notna(lon):
            parts.append(f"coordinates=({lat}, {lon})")
    return '; '.join(parts)


def make_route_review_reason(err, row):
    reason_code = err.get('reason_code') or 'ROUTE_RESOLVER_REVIEW_REQUIRED'
    category = 'Geography Mapping'
    technical_detail = route_error_technical_detail(err, row)
    if reason_code == 'MATERIAL_CODE_NOT_FOUND':
        technical_detail = f"ROUTE_RESOLVER_PR_MODEL_MISMATCH: {technical_detail}"
        category = 'PR Model Matching'
    elif reason_code == 'DUPLICATE_MATERIAL_CODE':
        technical_detail = f"ROUTE_RESOLVER_PR_MODEL_DUPLICATE: {technical_detail}"
        category = 'PR Model Matching'
    else:
        technical_detail = f"ROUTE_RESOLVER_REVIEW_REQUIRED: {technical_detail}"
    return make_review_reason(
        reason_code,
        category=category,
        description=err.get('reason_description'),
        action=err.get('required_action'),
        technical_detail=technical_detail
    )


def make_generic_review_reason(text, category='PR Model Matching'):
    code = normalize_reason_code(text)
    return make_review_reason(code, category=category, technical_detail=str(text or code))


def add_review_fields(base, reason):
    enriched = dict(base)
    enriched.update(reason)
    return enriched


def add_route_context_fields(review_item, row, resolver=None):
    enriched = dict(review_item)
    if row is not None:
        enriched['Latitude'] = row.get('Latitude (North Plus South Minus)', '')
        enriched['Longitude'] = row.get('Longitude (East Plus West Minus)', '')
        enriched['Declared_Region'] = row.get('region', '')
        enriched['Declared_State'] = row.get('Province/State', '')
    err = getattr(resolver, 'last_error', None) if resolver is not None else None
    if err:
        enriched['Resolved_State'] = err.get('state', '')
        enriched['Resolved_City_District'] = err.get('city_or_district', '')
        enriched['Route_Type'] = err.get('route_type', '')
        enriched['Route_Bucket'] = err.get('bucket', '')
        enriched['Warehouse'] = err.get('warehouse', '')
        enriched['Material_Code'] = err.get('material_code', '')
    return enriched


def parse_site_codes(site_code_str):
    if not site_code_str:
        return []
    codes = [code.strip().upper() for code in site_code_str.split(',') if code.strip()]
    return [code for code in codes if code]


def detect_site_code_column(df):
    preferred = ['Site ID', 'Site ID*', 'Site ID ', 'customer site code', 'Site Code', 'site code']
    lower_columns = {col.lower(): col for col in df.columns}
    for candidate in preferred:
        if candidate in df.columns:
            return candidate
        if candidate.lower() in lower_columns:
            return lower_columns[candidate.lower()]
    raise ValueError(
        'Site code column not found. Expected one of: Site ID, Site ID*, customer site code, Site Code.'
    )


def parse_antenna_sizes(antenna_string):
    """
    Extract all antenna sizes from a string and return them as a list of floats.
    Handles formats like:
    - "18G_1.2M(MAC)+OMT x1" => [1.2]
    - "0.6m, 1.2m" => [0.6, 1.2]
    - "1.2m" => [1.2]
    
    Returns: list of floats (sorted)
    """
    if pd.isna(antenna_string):
        return []
    
    raw = str(antenna_string).strip()
    if not raw:
        return []
    
    # Find all numeric patterns with optional decimals
    matches = re.findall(r'(\d+(?:\.\d+)?)', raw.replace(',', '.'))
    sizes = []
    for match in matches:
        try:
            sizes.append(float(match))
        except ValueError:
            pass
    
    return sorted(list(set(sizes)))  # unique, sorted


def get_max_antenna_size(antenna_string):
    """
    Extract antenna sizes and return the maximum.
    Returns: float or None
    """
    sizes = parse_antenna_sizes(antenna_string)
    return max(sizes) if sizes else None


def normalize_antenna_size(value):
    if pd.isna(value):
        return None
    raw = str(value).strip()
    if not raw:
        return None
    normalized = raw.replace(',', '.')
    sizes = parse_antenna_sizes(normalized)
    sizes = [size for size in sizes if size <= 5.0]
    return max(sizes) if sizes else None


def determine_ti_antenna_category(ne_size, fe_size):
    ne_val = normalize_antenna_size(ne_size)
    fe_val = normalize_antenna_size(fe_size)
    chosen_size = None
    remark = None

    if ne_val is None and fe_val is None:
        remark = 'Missing TI antenna size - review required'
    elif ne_val is None or fe_val is None:
        chosen_size = fe_val if ne_val is None else ne_val
        remark = 'Incomplete TI antenna size - review required'
    else:
        chosen_size = max(ne_val, fe_val)
        if ne_val != fe_val:
            remark = 'TI antenna sizes differ - using larger size for matching'

    category = None
    if chosen_size is not None:
        if chosen_size <= 0.6:
            category = '0.3/0.6m'
        elif chosen_size <= 1.2:
            category = '0.9/1.2m'
        elif chosen_size <= 1.8:
            category = '1.8m'
        elif chosen_size <= 2.4:
            category = '2.4m'
        else:
            category = f'{chosen_size}m'

    return category, remark, chosen_size



def get_text_value(row, column):
    value = row.get(column, '')
    if pd.isna(value):
        return ''
    return str(value).strip()


def extract_antenna_size_tokens(text):
    if not text or pd.isna(text):
        return []

    normalized = str(text).replace(',', '.')
    tokens = []
    pattern = re.compile(r'(?<![A-Za-z0-9])(\d+(?:\.\d+)?)(?=\s*(?:m(?=\b|[^A-Za-z0-9])|/|-))', re.IGNORECASE)
    for match in pattern.finditer(normalized):
        try:
            size = float(match.group(1))
        except ValueError:
            continue
        if 0 < size <= 5.0:
            tokens.append({'size': size, 'start': match.start(), 'end': match.end()})
    return tokens


def size_context_window(text, token, window=70):
    segment_start = max(text.rfind('\n', 0, token['start']), text.rfind(';', 0, token['start']), text.rfind('.', 0, token['start']))
    segment_end_candidates = [
        position for position in [
            text.find('\n', token['end']),
            text.find(';', token['end']),
            text.find('.', token['end'])
        ] if position != -1
    ]
    segment_start = 0 if segment_start == -1 else segment_start + 1
    segment_end = min(segment_end_candidates) if segment_end_candidates else len(text)
    start = max(segment_start, token['start'] - window)
    end = min(segment_end, token['end'] + window)
    return text[start:end].lower()


def size_context_before(text, token, window=70):
    segment_start = max(text.rfind('\n', 0, token['start']), text.rfind(';', 0, token['start']), text.rfind('.', 0, token['start']))
    segment_start = 0 if segment_start == -1 else segment_start + 1
    start = max(segment_start, token['start'] - window)
    return text[start:token['start']].lower()


def extract_contextual_sizes(text, context_keywords, ambiguous_keywords=None, require_keyword_before=False):
    if not text or pd.isna(text):
        return [], False

    raw = str(text)
    tokens = extract_antenna_size_tokens(raw)
    if not tokens:
        return [], False

    sizes = []
    ambiguous_found = False
    ambiguous_keywords = ambiguous_keywords or []
    for token in tokens:
        window = size_context_window(raw, token)
        before_window = size_context_before(raw, token)
        context_text = before_window if require_keyword_before else window
        has_context = any(keyword in context_text for keyword in context_keywords)
        has_ambiguous_context = any(keyword in context_text for keyword in ambiguous_keywords)
        if has_context:
            sizes.append(token['size'])
        elif has_ambiguous_context:
            ambiguous_found = True

    return sorted(set(sizes)), ambiguous_found


def extract_mw_reroute_install_size(row):
    ne_size = normalize_antenna_size(row.get('MW Config Antenna Size NE', ''))
    fe_size = normalize_antenna_size(row.get('MW Config Antenna Size FE', ''))
    config_sizes = [size for size in [ne_size, fe_size] if size is not None]
    if config_sizes:
        return {
            'size': max(config_sizes),
            'source': 'MW Config Antenna Size NE/FE',
            'confidence': 'HIGH',
            'reason': None
        }

    install_keywords = ['target', 'new', 'install', 'upgrade', 'build']
    source_groups = [
        ('BOQ Configuration', ['BOQ Configuration']),
        ('TX SOW Details', ['TX SOW Details']),
        ('NE/FE SOW Details', ['NE SOW Details', 'FE SOW Details'])
    ]
    for source_name, columns in source_groups:
        sizes = []
        for column in columns:
            column_sizes, _ = extract_contextual_sizes(get_text_value(row, column), install_keywords)
            sizes.extend(column_sizes)
        if sizes:
            return {
                'size': max(sizes),
                'source': source_name,
                'confidence': 'MEDIUM',
                'reason': None
            }

    return {
        'size': None,
        'source': None,
        'confidence': 'LOW',
        'reason': 'MW Reroute install antenna size missing'
    }


def extract_mw_reroute_decom_size(row):
    decom_keywords = ['decom', 'dismant', 'remove', 'removal']
    existing_keywords = ['existing']
    source_columns = [
        ('BOQ Configuration', 'BOQ Configuration'),
        ('TX SOW Details', 'TX SOW Details'),
        ('NE SOW Details', 'NE SOW Details'),
        ('FE SOW Details', 'FE SOW Details')
    ]

    for source_name, column in source_columns:
        text = get_text_value(row, column)
        sizes, ambiguous_reuse = extract_contextual_sizes(
            text,
            decom_keywords,
            ambiguous_keywords=['reuse existing'],
            require_keyword_before=True
        )
        existing_sizes, _ = extract_contextual_sizes(text, existing_keywords, require_keyword_before=True)

        # "Reuse existing" often describes retained equipment, not a decom target.
        lower_text = text.lower()
        if 'reuse existing' in lower_text:
            existing_sizes = []
            ambiguous_reuse = True

        combined = sorted(set(sizes + existing_sizes))
        if len(combined) == 1:
            return {
                'size': combined[0],
                'source': source_name,
                'confidence': 'MEDIUM',
                'reason': None
            }
        if len(combined) > 1 or ambiguous_reuse:
            return {
                'size': None,
                'source': source_name,
                'confidence': 'LOW',
                'reason': 'MW Reroute decom antenna size ambiguous'
            }

    return {
        'size': None,
        'source': None,
        'confidence': 'LOW',
        'reason': 'MW Reroute decom antenna size missing'
    }


def classify_mw_reroute_model_item(item):
    text = ' '.join([str(item.get('SOW', '')), str(item.get('Description', '')), str(item.get('Rules', ''))]).lower()
    if 'new - mw link' in text:
        return 'install'
    if 'mw dismantling' in text and 'antenna' in text:
        return 'dismantle'
    return 'other'


def text_matches_size_bucket(text, chosen_size):
    if chosen_size is None:
        return False
    normalized = str(text).replace(',', '.').lower()

    if '>3.2' in normalized:
        return chosen_size > 3.2

    for first, second in re.findall(r'(\d+(?:\.\d+)?)\s*[-/]\s*(\d+(?:\.\d+)?)\s*m?', normalized):
        low = float(first)
        high = float(second)
        if low <= chosen_size <= high:
            return True

    sizes = extract_antenna_size_tokens(normalized)
    return any(abs(token['size'] - chosen_size) < 1e-6 for token in sizes)


def select_mw_reroute_item(items, chosen_size):
    matched = []
    for item in items:
        text = ' '.join([str(item.get('SOW', '')), str(item.get('Description', '')), str(item.get('Rules', ''))])
        if text_matches_size_bucket(text, chosen_size):
            matched.append(item)
    if len(matched) == 1:
        return matched[0], None
    if len(matched) > 1:
        return None, 'ambiguous'
    return None, 'missing'


def match_mw_reroute_models(row, region, ti_models):
    sow = str(row.get('Tx SOW', '')).strip()
    sow_upper = sow.upper()
    candidates = []
    for item in ti_models:
        item_sow_upper = item['SOW'].upper()
        if item['Is_Mandatory'] and ('REROUTE' in item_sow_upper or item_sow_upper == 'MW REROUTE'):
            candidates.append(item)

    install_items = [item for item in candidates if classify_mw_reroute_model_item(item) == 'install']
    dismantle_items = [item for item in candidates if classify_mw_reroute_model_item(item) == 'dismantle']

    selected_items = []
    review_reasons = []
    install_result = extract_mw_reroute_install_size(row)
    decom_result = extract_mw_reroute_decom_size(row)

    if install_result['size'] is None:
        return [], ['MW Reroute install antenna size missing'], install_result, decom_result

    install_item, install_status = select_mw_reroute_item(install_items, install_result['size'])
    if install_item:
        selected_items.append(install_item)
    else:
        review_reasons.append('MW Reroute install item not matched')

    if decom_result['size'] is None:
        review_reasons.append(decom_result['reason'])
    else:
        dismantle_item, dismantle_status = select_mw_reroute_item(dismantle_items, decom_result['size'])
        if dismantle_item:
            selected_items.append(dismantle_item)
        elif dismantle_status == 'ambiguous':
            review_reasons.append('MW Reroute dismantle item not matched')
        else:
            review_reasons.append('MW Reroute dismantle item not matched')

    return selected_items, review_reasons, install_result, decom_result


def parse_choice_rule(rule_text):
    if not rule_text or pd.isna(rule_text):
        return None
    rule_text = str(rule_text).lower()
    match = re.search(r'(\d+)\s*choose\s*1', rule_text)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def normalize_choice_category(text):
    if not text:
        return None
    lower = str(text).lower()
    if 'simple packing' in lower:
        return 'outbound_route'
    if 'partial material transportation' in lower:
        return 'material_route'
    if 'antenna' in lower and parse_antenna_sizes(lower):
        return 'antenna'
    if 'inland transportation' in lower:
        return 'inbound_route'
    if 'dismantling' in lower and 'antenna' in lower:
        return 'antenna'
    return 'choose'

def ti_model_requires_antenna(sow, ti_models):
    """
    Returns True if the matching TI PR model contains
    a mandatory antenna-dependent choose-one group.

    This makes antenna validation conditional instead
    of universally mandatory.
    """

    sow_upper = normalize_ti_sow(sow)

    if not sow_upper:
        return False

    candidates = []

    for item in ti_models:
        if item["Is_Mandatory"] and ti_sow_matches_model(sow_upper, item.get("SOW", "")):
            candidates.append(item)

    for item in candidates:

        rules = str(item.get("Rules", "")).lower()

        if "choose" not in rules:
            continue

        category = normalize_choice_category(
            " ".join([
                str(item.get("SOW", "")),
                str(item.get("Description", "")),
                str(item.get("Rules", ""))
            ])
        )

        if category == "antenna":
            return True

    return False

def get_region_search_terms(region):
    if not region:
        return []
    region_key = str(region).strip().lower()
    mapping = {
        'northern': ['north region', 'perlis', 'kedah', 'penang', 'perak'],
        'southern': ['south region', 'negeri sembilan', 'malacca', 'johor'],
        'eastern': ['east region', 'pahang', 'terengganu', 'kelantan'],
        'sabah': ['sabah'],
        'sarawak': ['sarawak', 'salawak', 'kuching', 'sibu', 'bintulu', 'miri', 'limbang', 'lawas', 'sri aman'],
        'central': ['kv region', 'kv warehouse', 'kuantan', 'kk']
    }
    terms = mapping.get(region_key, [region_key])
    return [term for term in terms if term]


def item_matches_region(item, region):
    if not region:
        return False
    text = ' '.join([str(item.get('SOW', '')), str(item.get('Description', '')), str(item.get('Rules', ''))]).lower()
    for term in get_region_search_terms(region):
        if term in text:
            return True
    return False


def item_matches_chosen_size(item, chosen_size):
    if chosen_size is None:
        return False
    text = ' '.join([str(item.get('SOW', '')), str(item.get('Description', '')), str(item.get('Rules', ''))])
    sizes = parse_antenna_sizes(text)
    return any(abs(size - chosen_size) < 1e-6 for size in sizes)


def filter_choose_group_items(group_items, chosen_size, region, row=None, resolver=None):
    """
    Filter choose-1 group items.
    
    Returns:
        (list, bool): (filtered_items, ambiguous)
        - If exactly one item matches: return that single item, ambiguous=False
        - If multiple items match: return empty list, ambiguous=True (do not write to ECC)
        - If zero items match: return empty list, ambiguous=True (do not write to ECC)
    """
    if len(group_items) <= 1:
        return group_items, False

    category = normalize_choice_category(' '.join([group_items[0].get('SOW', ''), group_items[0].get('Description', ''), group_items[0].get('Rules', '')]))
    if category == "antenna":

        if chosen_size is None:
            return [], True

        matched = [
            item
            for item in group_items
            if item_matches_chosen_size(item, chosen_size)
        ]

        if len(matched) == 1:
            return matched, False

        return [], True
    
    if category == 'outbound_route':
        if row is not None and resolver is not None:
            res = resolver.resolve_material_code(row, "outbound_route")
            if res["status"] == "RESOLVED":
                material_code = res["material_code"]
                matched = [item for item in group_items if item.get("PBOM_Code") == material_code]
                if len(matched) == 1:
                    return matched, False
                resolver.last_error = {
                    "route_type": "outbound_route",
                    "reason_code": "MATERIAL_CODE_NOT_FOUND" if len(matched) == 0 else "DUPLICATE_MATERIAL_CODE",
                    "bucket": res.get("bucket"),
                    "state": res.get("state"),
                    "city_or_district": res.get("city_or_district"),
                    "warehouse": res.get("warehouse"),
                    "material_code": material_code
                }
                return [], True
            return [], True
        return [], True

    if category == 'inbound_route':
        if row is not None and resolver is not None:
            res = resolver.resolve_material_code(row, "inbound_route")
            if res["status"] == "RESOLVED":
                material_code = res["material_code"]
                matched = [item for item in group_items if item.get("PBOM_Code") == material_code]
                if len(matched) == 1:
                    return matched, False
                resolver.last_error = {
                    "route_type": "inbound_route",
                    "reason_code": "MATERIAL_CODE_NOT_FOUND" if len(matched) == 0 else "DUPLICATE_MATERIAL_CODE",
                    "bucket": res.get("bucket"),
                    "state": res.get("state"),
                    "city_or_district": res.get("city_or_district"),
                    "warehouse": res.get("warehouse"),
                    "material_code": material_code
                }
                return [], True
            return [], True
        return [], True

    if category in {'material_route', 'choose'}:
        for term in get_region_search_terms(region):
            matched = [
                item for item in group_items
                if term in ' '.join([str(item.get('SOW', '')), str(item.get('Description', '')), str(item.get('Rules', ''))]).lower()
            ]
            if len(matched) == 1:
                return matched, False
        # Zero or multiple matched -> return empty, flag as ambiguous
        return [], True

    # Unknown category -> return empty, flag as ambiguous
    return [], True


def is_mw_hardware_cutover_item(item):
    """
    Check if a PR model item represents MW Hardware Cutover.
    """
    sow = str(item.get('SOW', '')).strip().lower()
    desc = str(item.get('Description', '')).strip().lower()
    rules = str(item.get('Rules', '')).strip().lower()
    return "mw hardware cutover" in sow or "mw hardware cutover" in desc or "mw hardware cutover" in rules


def match_ti_models(sow, antenna_category, chosen_size, region, ti_models, row=None, resolver=None):
    sow_upper = normalize_ti_sow(sow)
    if not sow_upper:
        return [], False
    candidates = []
    for item in ti_models:
        if ti_sow_matches_model(sow_upper, item.get('SOW', '')):
            if item['Is_Mandatory']:
                if not is_mw_hardware_cutover_item(item):
                     candidates.append(item)

    if not candidates:
        return [], False

    grouped_candidates = {}
    for item in candidates:
        rules = str(item.get('Rules', '')).strip().lower()
        if 'choose' in rules:
            group_key = (
                item['SOW'],
                item['Rules'],
                normalize_choice_category(' '.join([item.get('SOW', ''), item.get('Description', ''), item.get('Rules', '')]))
            )
        else:
            group_key = (item['PBOM_Code'],)
        grouped_candidates.setdefault(group_key, []).append(item)

    requires_antenna = ti_model_requires_antenna(sow, ti_models)

    selected_items = []
    review_required = False
    choose_group_ambiguous = False
    for group_items in grouped_candidates.values():
        if len(group_items) == 1:
            selected_items.extend(group_items)
            continue

        rules = str(group_items[0].get('Rules', '')).lower()
        if 'choose' in rules:
            category = normalize_choice_category(
                " ".join([
                    str(group_items[0].get("SOW", "")),
                    str(group_items[0].get("Description", "")),
                    str(group_items[0].get("Rules", ""))
                ])
            )

            if category == "antenna" and chosen_size is None:
                review_required = True
                choose_group_ambiguous = True
                continue

            chosen_items, ambiguous = filter_choose_group_items(
                group_items,
                chosen_size,
                region,
                row=row,
                resolver=resolver
            )            
            selected_items.extend(chosen_items)
            if ambiguous:
                review_required = True
                choose_group_ambiguous = True
        else:
            selected_items.extend(group_items)
            if len(group_items) > 1:
                review_required = True

    if choose_group_ambiguous:
        return [], True

    return selected_items, review_required


def filter_site_rows(df_site, site_codes=None, all_sites=False):
    if site_codes and all_sites:
        raise ValueError('Use either --site-code or --all-sites, not both.')
    if not site_codes and not all_sites:
        raise ValueError('Please provide --site-code <SITE_CODE> or use --all-sites to generate all eligible sites.')
    if all_sites:
        return df_site.copy(), []

    normalized_codes = [code for code in site_codes if code]
    if not normalized_codes:
        raise ValueError('No valid site codes provided to --site-code.')

    site_code_column = detect_site_code_column(df_site)

    df_site['_normalized_site_code'] = df_site[site_code_column].astype(str).str.strip().str.upper()
    requested_set = set(normalized_codes)
    matched_df = df_site[df_site['_normalized_site_code'].isin(requested_set)].copy()
    matched_codes = set(matched_df['_normalized_site_code'].unique())
    missing_codes = [code for code in normalized_codes if code not in matched_codes]

    if len(matched_df) == 0:
        raise ValueError('No matching Site Codes found in Site PR/PO View. Please check the input site code(s).')

    if missing_codes:
        print('Warning: requested Site Code(s) not found:')
        for code in missing_codes:
            print(f'- {code}')
        print(f'[OK] Continuing with {len(matched_df)} matched row(s).')

    matched_df = matched_df.drop(columns=['_normalized_site_code'])
    return matched_df, missing_codes


# ===== STEP 0: LOAD REFERENCE DATA =====
print("[STEP 0] Loading Reference Data from Markdown and validating inputs...")

site_file = require_file(args.site_data, 'Site PR/PO View')
pr_file = validate_pr_model_file(args.pr_model, 'PR Model')
template_file = require_file(args.template, 'ECC Template')
ref_file = require_file(args.mapping, 'Contract info mapping')
validate_template_file(template_file)

print(f"[OK] Site PR/PO View: {site_file}")
print(f"[OK] PR Model: {pr_file}")
print(f"[OK] ECC Template: {template_file}")
print(f"[OK] Mapping file: {ref_file}")

region_mapping = load_region_mapping(ref_file)
subcon_mapping = load_subcon_mapping(ref_file)
resolver = GeographyResolver()

print(f"[OK] Loaded {len(region_mapping)} region mappings")
print(f"  Regions: {', '.join(sorted(region_mapping.keys()))}")
print(f"[OK] Loaded {len(subcon_mapping)} subcontractor mappings")
print(f"  Subcons: {', '.join(sorted(subcon_mapping.keys()))[:80]}...")

# ===== STEP 1: EXTRACT PR MODEL DATA =====
print("\n[STEP 1] Extracting PR Model Data...")

tss_models, ti_models = load_pr_model_items(pr_file)

if ti_models:
    print(f"[OK] Extracted {len(ti_models)} TI line items")
else:
    print('Warning: TI Model section not found in PR model sheet.')

# Group by SOW for TSS and TI models
sow_groups = {}
for item in tss_models:
    sow = item['SOW']
    if sow not in sow_groups:
        sow_groups[sow] = []
    sow_groups[sow].append(item)

print(f"[OK] Found {len(sow_groups)} unique TSS SOWs:")
for sow in sorted(sow_groups.keys()):
    mandatory_count = len([x for x in sow_groups[sow] if x['Is_Mandatory']])
    print(f"  - {sow[:50]}: {len(sow_groups[sow])} items ({mandatory_count} mandatory)")

if ti_models:
    ti_sow_groups = {}
    for item in ti_models:
        sow = item['SOW']
        if sow not in ti_sow_groups:
            ti_sow_groups[sow] = []
        ti_sow_groups[sow].append(item)


def get_ti_sow_matches(sow, ti_models):
    normalized_sow = normalize_ti_sow(sow)
    if not normalized_sow:
        return []
    return [
        item for item in ti_models
        if ti_sow_matches_model(normalized_sow, item.get('SOW', ''))
    ]

    print(f"[OK] Found {len(ti_sow_groups)} unique TI SOWs:")
    for sow in sorted(ti_sow_groups.keys()):
        mandatory_count = len([x for x in ti_sow_groups[sow] if x['Is_Mandatory']])
        print(f"  - {sow[:50]}: {len(ti_sow_groups[sow])} items ({mandatory_count} mandatory)")

# ===== STEP 2: LOAD SITE DATA =====
print("\n[STEP 2] Loading Site Data...")

df_site = pd.read_excel(site_file, sheet_name='data', header=3)

# Apply Site Selection filter
try:
    selected_site_codes = parse_site_codes(args.site_code)
    df_site, missing_codes = filter_site_rows(df_site, site_codes=selected_site_codes, all_sites=args.all_sites)
    if args.site_code:
        print(f"[OK] Site selection: {len(selected_site_codes)} requested codes")
    if args.all_sites:
        print("[OK] Site selection: all eligible sites")
    print(f"[OK] Matched site rows: {len(df_site)}")
except ValueError as e:
    print(f"ERROR: {e}")
    sys.exit(1)

scope_name = args.scope.upper()
subcon_column = 'SubCon - TSS Team' if scope_name == 'TSS' else 'SubCon - TI Team'
if subcon_column not in df_site.columns:
    print(
        f"ERROR: CANONICAL_INPUT_SCHEMA_ERROR: internal renderer input is missing "
        f"'{subcon_column}'. Use scripts/create_pr.py for original iEPMS exports."
    )
    sys.exit(1)
pr_status_column = 'Subcon PR - TSS' if scope_name == 'TSS' else 'Subcon PR - TI'
du_model_name = re.sub(r'[<>:"/\\|?*]+', ' ', str(args.du_model_name)).strip().strip('.')
if not du_model_name:
    print("ERROR: CANONICAL_INPUT_SCHEMA_ERROR: resolved DU model name is blank.")
    sys.exit(1)

# Track review-required and skipped items for TI Phase 1
review_required_items = []
duplicates_skipped = []
warnings = []

# TI Phase 1: Hardened candidate filtering with duplicate prevention
if scope_name == 'TI':
    # For TI: must have SubCon - TI Team AND blank Subcon PR - TI
    candidates_all = df_site[
        (df_site[subcon_column].notna()) &
        (df_site[subcon_column].astype(str).str.strip() != '')
    ].copy()
    
    print(f"[OK] Found {len(candidates_all)} rows with {subcon_column}")
    
    # Split into candidates and duplicates
    candidates_list = []
    for idx, row in candidates_all.iterrows():
        site_id = str(row['customer site code']).strip()
        region = str(row['region']).strip()
        subcon = str(row[subcon_column]).strip()
        sow = str(row.get('Tx SOW', '')).strip() if pd.notna(row.get('Tx SOW')) else ''
        pr_status = str(row.get(pr_status_column, '')).strip() if pd.notna(row.get(pr_status_column)) else ''
        
        # Duplicate prevention: check if PR already exists
        if pr_status and pr_status != '':
            duplicates_skipped.append({
                'Site_ID': site_id,
                'Region': region,
                'SubCon_TI': subcon,
                'Tx_SOW': sow,
                'Existing_PR': pr_status,
                'Reason': 'Duplicate - PR already exists'
            })
            continue
        
        # Check for missing Tx SOW
        if not sow or sow == '':
            review_required_items.append(add_review_fields({
                'Site_ID': site_id,
                'Region': region,
                'SubCon_TI': subcon,
                'Tx_SOW': '(missing)'
            }, make_review_reason('MISSING_TX_SOW', technical_detail='Missing Tx SOW')))
            continue
        
        # All checks passed - add as candidate
        candidates_list.append(idx)
    
    candidates = df_site.loc[candidates_list].copy().reset_index(drop=True)
    print(f"[OK] Candidates after TI Phase 1 filtering: {len(candidates)}")
    print(f"  - Duplicates skipped: {len(duplicates_skipped)}")
    print(f"  - Review-required flagged: {len(review_required_items)}")

else:
    # TSS: simple filtering (no TI Phase 1 changes)
    candidates = df_site[
        (df_site[subcon_column].notna()) &
        (df_site[subcon_column].astype(str).str.strip() != '')
    ].copy().reset_index(drop=True)
    print(f"[OK] Found {len(candidates)} {scope_name} candidates")


# Display first 5 for dry run
_safe_print(f"[OK] First 5 candidates:")
for i in range(min(5, len(candidates))):
    site_id = candidates.iloc[i]['customer site code']
    site_name = candidates.iloc[i]['customer site name']
    region = candidates.iloc[i]['region']
    subcon = candidates.iloc[i][subcon_column]
    sow = str(candidates.iloc[i]['Tx SOW'])
    _safe_print(
        f"  {i+1}. {site_id} ({site_name}) - Region: {region}, SubCon: {subcon}, SOW: {sow[:40]}"
    )

# ===== STEP 3: BUILD ECC OUTPUT ROWS =====
print("\n[STEP 3] Building ECC Output Rows...")

ecc_rows = []
fuzzy_matches = {}
unmatched_ti_items = []  # Phase 2B-1: capture unmatched TI candidates
failed_migration_decisions = set()

for idx in range(len(candidates)):
    row = candidates.iloc[idx]
    if resolver is not None:
        resolver.last_error = None
    
    site_id = str(row['customer site code']).strip()
    site_name = str(row['customer site name']).strip()
    region = str(row['region']).strip()
    subcon = str(row[subcon_column]).strip()
    sow = str(row['Tx SOW']).strip()
    migration_decision_id = optional_cell_text(row.get('Migration Decision ID', ''))
    migration_work_item = optional_cell_text(row.get('Migration Work Item', ''))
    required_pbom_codes = [
        code.strip()
        for code in optional_cell_text(row.get('Required PBOM Codes', '')).split('|')
        if code.strip()
    ]
    delivery_unit = str(row.get('du code', '')).strip() if pd.notna(row.get('du code')) else ''
    logical_site = str(row.get('customer site code', '')).strip() if pd.notna(row.get('customer site code')) else ''
    
    # Get Purchasing Area from Region
    purchasing_area = region_mapping.get(region, f"UNKNOWN ({region})")
    
    # Get Contract Number from Subcontractor (with fuzzy matching)
    matched_subcon = subcon
    if subcon in subcon_mapping:
        contract_info = subcon_mapping[subcon]
    else:
        fuzzy_match = fuzzy_match_subcon(subcon, subcon_mapping)
        if fuzzy_match:
            matched_subcon = fuzzy_match
            contract_info = subcon_mapping[fuzzy_match]
            fuzzy_matches[subcon] = fuzzy_match
            print(f"  [WARN] Fuzzy matched '{subcon}' -> '{fuzzy_match}'")
        else:
            contract_info = {'contract_number': 'UNKNOWN', 'company_name': 'Unknown'}
            print(f"  [FAIL] No match for subcontractor: {subcon}")
    
    contract_number = contract_info['contract_number']
    
    matched_items = []
    review_required = False
    remarks = ''
    unmatched_reason = None
    review_base = {
        'Site_ID': site_id,
        'Region': region,
        'SubCon_TI': subcon,
        'Tx_SOW': sow
    }

    if scope_name == 'TSS':
        upgrade_scope = str(row.get('TX Upgrade Scope', '')).strip()
        matched_items = select_tss_items_for_site(site_id, sow, upgrade_scope, tss_models)
    else:
        if is_mw_reroute_row(row):
            matched_items, mw_review_reasons, install_result, decom_result = match_mw_reroute_models(
                row,
                region,
                ti_models
            )
            if mw_review_reasons:
                remarks = 'REVIEW_REQUIRED'
                for review_reason in mw_review_reasons:
                    unmatched_ti_items.append(add_review_fields(
                        review_base,
                        make_generic_review_reason(review_reason)
                    ))
        elif is_mw_reengineering_sow(sow):
            matched_items, review_required = match_ti_models(
                sow,
                None,
                None,
                region,
                ti_models,
                row=row,
                resolver=resolver
            )
            if review_required:
                remarks = 'REVIEW_REQUIRED' if not remarks else f"{remarks}; REVIEW_REQUIRED"

            if not matched_items:
                sow_matches = get_ti_sow_matches(sow, ti_models)
                if not sow_matches:
                    unmatched_reason = make_review_reason(
                        'NO_MATCHING_TI_PR_MODEL_ITEM',
                        technical_detail='No matching TI PR model item'
                    )
                elif not any(m['Is_Mandatory'] for m in sow_matches):
                    unmatched_reason = make_review_reason(
                        'NO_MANDATORY_TI_ITEM_FOUND',
                        technical_detail='No mandatory TI item found'
                    )
                else:
                    unmatched_reason = make_review_reason(
                        'NO_MANDATORY_TI_ITEM_FOUND',
                        technical_detail='No valid mandatory TI PR model items matched'
                    )
        else:
            requires_antenna = ti_model_requires_antenna(
                sow,
                ti_models
            )

            antenna_category = None
            antenna_size = None
            antenna_remark = None

            if requires_antenna:

                antenna_category, antenna_remark, antenna_size = determine_ti_antenna_category(
                    row.get("MW Config Antenna Size NE", ""),
                    row.get("MW Config Antenna Size FE", "")
                )

                if antenna_remark:
                    remarks = antenna_remark

                if antenna_remark in [
                    "Missing TI antenna size - review required",
                    "Incomplete TI antenna size - review required"
                ]:

                    unmatched_ti_items.append(
                        add_review_fields(
                            review_base,
                            make_generic_review_reason(antenna_remark)
                        )
                    )

                    if migration_decision_id:
                        failed_migration_decisions.add(migration_decision_id)
                    continue

            matched_items, review_required = match_ti_models(
                sow,
                antenna_category,
                antenna_size,
                region,
                ti_models,
                row=row,
                resolver=resolver
            )

            if review_required:
                remarks = "REVIEW_REQUIRED" if not remarks else f"{remarks}; REVIEW_REQUIRED"
            # Phase 2B-1: capture reason for unmatched TI candidates
            if not matched_items:
                if resolver is not None and getattr(resolver, 'last_error', None) is not None:
                    err = resolver.last_error
                    unmatched_reason = make_route_review_reason(err, row)
                else:
                    # Determine the reason for no matching items
                    sow_matches = get_ti_sow_matches(sow, ti_models)
                    if not sow_matches:
                        unmatched_reason = make_review_reason(
                            'NO_MATCHING_TI_PR_MODEL_ITEM',
                            technical_detail='No matching TI PR model item'
                        )
                    elif not any(m['Is_Mandatory'] for m in sow_matches):
                        unmatched_reason = make_review_reason(
                            'NO_MANDATORY_TI_ITEM_FOUND',
                            technical_detail='No mandatory TI item found'
                        )
                    elif requires_antenna:
                        unmatched_reason = make_review_reason(
                            'NO_MATCHING_ANTENNA_GROUP_ITEM',
                            technical_detail='No matching antenna group item'
                        )
                    else:
                        unmatched_reason = make_review_reason(
                            'NO_MATCHING_TI_PR_MODEL_ITEM',
                            technical_detail='No matching TI PR model item'
                        )

    if scope_name == 'TI' and migration_decision_id and review_required:
        failed_migration_decisions.add(migration_decision_id)
        matched_items = []
        if unmatched_reason is None:
            unmatched_reason = make_review_reason(
                'JENDELA_PR_MODEL_ITEM_NOT_FOUND',
                technical_detail=f'Non-unique PR model selection for migration work item: {migration_work_item}'
            )

    if matched_items and required_pbom_codes:
        valid_pboms, pbom_reason = validate_required_pbom_selection(matched_items, required_pbom_codes)
        if not valid_pboms:
            failed_migration_decisions.add(migration_decision_id)
            matched_items = []
            unmatched_reason = make_review_reason(
                pbom_reason or 'JENDELA_PR_MODEL_ITEM_NOT_FOUND',
                technical_detail=(
                    f'work_item={migration_work_item}; required_pboms={required_pbom_codes}'
                )
            )

    if not matched_items:
        if migration_decision_id:
            failed_migration_decisions.add(migration_decision_id)
        if scope_name == 'TI' and unmatched_reason:
            # Phase 2B-1: Add to review-required instead of silently dropping
            unmatched_ti_items.append(add_route_context_fields(
                add_review_fields(review_base, unmatched_reason),
                row,
                resolver=resolver
            ))
        print(f"  [FAIL] No mandatory items found for SOW: {sow[:50]} (Scope: {scope_name})")
        continue
    
    for item in matched_items:
        ecc_row = {
            'Site_ID': site_id,
            'Site_Name': site_name,
            'Region': region,
            'Purchasing_Area': purchasing_area,
            'Subcontractor': matched_subcon,
            'Contract_Number': contract_number,
            'PBOM_Code': item['PBOM_Code'],

            # PR Model description (Column K)
            'SOW': item['Description'],

            # NEW - Original Tx SOW from site data (Column O)
            'Source_Tx_SOW': sow,

            'Unit': item['Unit'],
            'Quantity': item['Quantity'],
            'Delivery_Unit_Code': delivery_unit,
            'Logical_Site_Name': logical_site,
            'Remarks': remarks
        }
        if migration_decision_id:
            ecc_row['Migration_Decision_ID'] = migration_decision_id
        ecc_rows.append(ecc_row)

ecc_rows = filter_failed_migration_decisions(ecc_rows, failed_migration_decisions)
print(f"[OK] Built {len(ecc_rows)} ECC output rows")

# ===== STEP 4: GROUP BY REGION-SUBCON AND CREATE EXCEL FILES =====
print("\n[STEP 4] Grouping and Creating Excel Files...")

# Group by (Region, Subcontractor)
grouped = {}
for ecc_row in ecc_rows:
    group_key = (ecc_row['Region'], ecc_row['Subcontractor'])
    if group_key not in grouped:
        grouped[group_key] = []
    grouped[group_key].append(ecc_row)

file_count = 0
total_rows = 0

for (region, subcon), rows in sorted(grouped.items()):
    # Check if split needed (max 30 unique sites per file)
    unique_sites = set([r['Site_ID'] for r in rows])
    num_files = (len(unique_sites) + 29) // 30  # Ceiling division
    
    print(f"\n  Group: {region}-{subcon}")
    print(f"    Unique sites: {len(unique_sites)}")
    print(f"    Total rows: {len(rows)}")
    print(f"    Files needed: {num_files}")
    
    # Distribute rows across files
    if num_files == 1:
        # Single file
        parts = [rows]
        file_parts = [1]
    else:
        # Multiple files - split by unique sites
        sites_list = sorted(list(unique_sites))
        parts = []
        file_parts = []
        
        for part_num in range(num_files):
            start_idx = part_num * 30
            end_idx = (part_num + 1) * 30
            part_sites = set(sites_list[start_idx:end_idx])
            part_rows = [r for r in rows if r['Site_ID'] in part_sites]
            parts.append(part_rows)
            file_parts.append(part_num + 1)
    
    # Create Excel file(s)
    for part_rows, part_num in zip(parts, file_parts):
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'details'
        
        # Write headers
        headers = ['SN.', 'Purchasing Area*', 'Region*', 'Site ID*', 'Site Name*', 
                   'Delivery Unit Code*', 'Logical Site Name', 'Contract Number *',
                   'Subcontractor*', 'PBOM Code*', 'SOW*', 'Unit*', 'Quantity*', 
                   'Remarks', '', 'Contract Number']
        
        # Header styling
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_font = Font(bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data rows with sequential SN
        for sn, ecc_row in enumerate(part_rows, 1):
            row_data = [
                sn,  # SN. - Sequential from 1
                ecc_row['Purchasing_Area'],
                ecc_row['Region'],
                ecc_row['Site_ID'],
                ecc_row['Site_Name'],
                ecc_row['Delivery_Unit_Code'],
                ecc_row['Logical_Site_Name'],
                ecc_row['Contract_Number'],  # Contract Number *
                ecc_row['Subcontractor'],
                ecc_row['PBOM_Code'],
                ecc_row['SOW'],
                ecc_row['Unit'],
                ecc_row['Quantity'],
                ecc_row['Remarks'],
                ecc_row['Source_Tx_SOW'],         # Column O
                ecc_row['Contract_Number']  # Column P - Contract Number (same as *)
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(2 + sn - 1, col_idx, value)
                if col_idx in [1, 13]:  # SN and Quantity - numeric
                    cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        column_widths = [5, 20, 12, 15, 20, 15, 20, 18, 15, 15, 40, 8, 10, 15, 5, 18]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d')
        if num_files == 1:
            filename = output_dir / f"{region}-{subcon} {du_model_name} {scope_name} PR {timestamp}.xlsx"
        else:
            filename = output_dir / f"{region}-{subcon} {du_model_name} {scope_name} PR {timestamp} Part {part_num}.xlsx"
        
        # Save file
        wb.save(str(filename))
        
        file_count += 1
        total_rows += len(part_rows)
        print(f"    [OK] Created: {filename} ({len(part_rows)} rows)")

# ===== STEP 5: CREATE REVIEW-REQUIRED OUTPUT (TI PHASE 1 + PHASE 2B-1) =====
if scope_name == 'TI' and (review_required_items or duplicates_skipped or unmatched_ti_items):
    print("\n[STEP 5] Creating Review Output Files (TI Phase 1 + Phase 2B-1)...")
    
    # Create review-required CSV (includes Phase 1 + Phase 2B-1 items)
    combined_review = review_required_items + unmatched_ti_items
    if combined_review:
        timestamp = datetime.now().strftime('%Y%m%d')
        review_file = output_dir / f"REVIEW_REQUIRED_TI_{timestamp}.csv"
        
        review_fieldnames = [
            'Site_ID',
            'Region',
            'SubCon_TI',
            'Tx_SOW',
            'Reason_Category',
            'Reason_Code',
            'Reason_Description',
            'Required_Action',
            'Technical_Detail',
            'Review_Reason',
            'Latitude',
            'Longitude',
            'Declared_Region',
            'Declared_State',
            'Resolved_State',
            'Resolved_City_District',
            'Route_Type',
            'Route_Bucket',
            'Warehouse',
            'Material_Code',
            'Source_Scope'
        ]
        # utf-8-sig: site data can contain characters (e.g. U+200B) that the
        # Windows default cp1252 encoding cannot write, and the BOM keeps the
        # file readable in Excel.
        with open(review_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=review_fieldnames)
            writer.writeheader()
            for item in combined_review:
                row = {
                    'Site_ID': item['Site_ID'],
                    'Region': item['Region'],
                    'SubCon_TI': item['SubCon_TI'],
                    'Tx_SOW': item['Tx_SOW'],
                    'Reason_Category': item.get('Reason_Category', 'PR Model Matching'),
                    'Reason_Code': item.get('Reason_Code', normalize_reason_code(item.get('Review_Reason'))),
                    'Reason_Description': item.get('Reason_Description', item.get('Review_Reason', 'Review required.')),
                    'Required_Action': item.get('Required_Action', 'Review the site input and PR model mapping.'),
                    'Technical_Detail': item.get('Technical_Detail', item.get('Review_Reason', '')),
                    'Review_Reason': item.get('Review_Reason', item.get('Technical_Detail', '')),
                    'Latitude': item.get('Latitude', ''),
                    'Longitude': item.get('Longitude', ''),
                    'Declared_Region': item.get('Declared_Region', item.get('Region', '')),
                    'Declared_State': item.get('Declared_State', ''),
                    'Resolved_State': item.get('Resolved_State', ''),
                    'Resolved_City_District': item.get('Resolved_City_District', ''),
                    'Route_Type': item.get('Route_Type', ''),
                    'Route_Bucket': item.get('Route_Bucket', ''),
                    'Warehouse': item.get('Warehouse', ''),
                    'Material_Code': item.get('Material_Code', ''),
                    'Source_Scope': 'TI'
                }
                writer.writerow(row)
        
        print(f"  [OK] Created review-required file: {review_file} ({len(combined_review)} items)")
        print(f"    - Phase 1 review-required: {len(review_required_items)}")
        print(f"    - Phase 2B-1 unmatched TI items: {len(unmatched_ti_items)}")
    
    # Create duplicates-skipped CSV
    if duplicates_skipped:
        timestamp = datetime.now().strftime('%Y%m%d')
        dups_file = output_dir / f"DUPLICATES_SKIPPED_TI_{timestamp}.csv"
        
        with open(dups_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=['Site_ID', 'Region', 'SubCon_TI', 'Tx_SOW', 'Existing_PR', 'Reason'])
            writer.writeheader()
            for item in duplicates_skipped:
                writer.writerow(item)
        
        print(f"  [OK] Created duplicates-skipped file: {dups_file} ({len(duplicates_skipped)} items)")

# ===== SUMMARY =====
print("\n" + "=" * 100)
print("SUMMARY")
print("=" * 100)
print(f"Scope: {scope_name}")
print(f"Total files generated: {file_count}")
print(f"Total ECC rows: {total_rows}")
_safe_print(f"Fuzzy matched subcontractors: {len(fuzzy_matches)}")
if fuzzy_matches:
    for original, matched in fuzzy_matches.items():
        _safe_print(f"  - '{original}' -> '{matched}'")

# TI Phase 1 summary
if scope_name == 'TI':
    print(f"\nTI Phase 1 Summary:")
    print(f"  Review-required items: {len(review_required_items)}")
    print(f"  Duplicates skipped: {len(duplicates_skipped)}")
    if review_required_items:
        review_reasons = {}
        for item in review_required_items:
            reason = item.get('Reason_Code') or item.get('Review_Reason')
            review_reasons[reason] = review_reasons.get(reason, 0) + 1
        for reason, count in sorted(review_reasons.items()):
            print(f"    - {reason}: {count}")

print(f"\nTimestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 100)
