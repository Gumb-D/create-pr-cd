#!/usr/bin/env python3
"""Convert an original DU export to validated canonical PR site records."""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from canonical_site_validator import SCOPE_REQUIRED_FIELDS
from du_export_adapter import build_canonical_site_record, resolve_profile_field_mappings
from profile_du_export import (
    extract_du_identities,
    fingerprint_key,
    resolve_approved_header_structure,
)


def _load_document(path: Path) -> Any:
    text = Path(path).read_text(encoding="utf-8")
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            import yaml  # type: ignore
        except ModuleNotFoundError as error:
            raise ValueError(f"{path} requires PyYAML or JSON-compatible YAML.") from error
        return yaml.safe_load(text)


def _select_source_sheet(
    inventory: Mapping[str, Any],
    resolved: Mapping[str, Any],
    required_fields: Iterable[str],
) -> Mapping[str, Any]:
    required_sheets = {
        match["sheet_name"]
        for name in required_fields
        for match in resolved.get(name, {}).get("matches", [])
    }
    if len(required_sheets) != 1:
        raise ValueError("REQUIRED_MAPPINGS_SPAN_MULTIPLE_OR_NO_SHEETS")
    selected_name = next(iter(required_sheets))
    return next(
        sheet for sheet in inventory.get("sheets", []) if sheet.get("sheet_name") == selected_name
    )


def _iter_source_rows(input_path: Path, sheet_name: str, positions: Mapping[str, int]):
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(input_path, read_only=True, data_only=False)
        worksheet = workbook[sheet_name]
        try:
            for row_number, values in enumerate(worksheet.iter_rows(min_row=5, values_only=True), start=5):
                yield row_number, {
                    key: values[index - 1] if index <= len(values) else None
                    for key, index in positions.items()
                }
        finally:
            workbook.close()
        return
    if suffix == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for _ in range(4):
                next(reader, None)
            for row_number, values in enumerate(reader, start=5):
                yield row_number, {
                    key: values[index - 1] if index <= len(values) else None
                    for key, index in positions.items()
                }
        return
    raise ValueError("Only .xlsx, .xlsm, and .csv exports are supported.")


def build_canonical_records(
    *,
    input_path: Path,
    profile: Mapping[str, Any],
    inventory: Mapping[str, Any],
    header_hash: str,
    scope: str,
    sow_registry_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    scope = str(scope).upper()
    resolved = resolve_profile_field_mappings(inventory, profile)
    mapped_scope_fields = {
        field for field in SCOPE_REQUIRED_FIELDS[scope] if field != "tx_sow_normalized"
    }
    profile_required = {
        field
        for field, config in profile.get("field_mapping", {}).items()
        if config.get("required")
    }
    required_fields = mapped_scope_fields | profile_required
    missing = [
        field
        for field in sorted(required_fields)
        if resolved.get(field, {}).get("status") != "RESOLVED"
    ]
    if missing:
        raise ValueError("MISSING_OR_AMBIGUOUS_REQUIRED_MAPPING:" + ",".join(missing))

    identities = extract_du_identities(inventory)
    if len(identities) != 1:
        raise ValueError("DU_IDENTITY_NOT_UNIQUE")
    runtime_identity = identities[0]

    header_validation = resolve_approved_header_structure(inventory, profile)
    if not header_validation["approved"]:
        raise ValueError("HEADER_HASH_REVALIDATION_REQUIRED")
    if header_hash and str(header_hash) != str(header_validation["raw_header_hash"]):
        raise ValueError("HEADER_HASH_CONTEXT_MISMATCH")

    source_sheet = _select_source_sheet(inventory, resolved, required_fields)
    positions = {
        column["fingerprint_key"]: column["source_position"]["one_based_index"]
        for column in source_sheet.get("columns", [])
    }
    sow_registry = _load_document(sow_registry_path)
    identity = profile["identity"]
    metadata = {
        "profile_id": profile["profile_id"],
        "profile_version": profile["profile_version"],
        "mapping_version": profile["mapping_version"],
        "project_key": identity["project_key"],
        "du_model_name": identity["accepted_du_models"][0],
        "du_model_id": str(runtime_identity["du_model_id"]),
        "view_id": str(runtime_identity["view_id"]),
        "source_file_name": input_path.name,
        "source_file_hash": inventory["source"]["source_file_hash"],
        "header_hash": header_validation["raw_header_hash"],
        "raw_header_hash": header_validation["raw_header_hash"],
        "structural_header_hash": header_validation["structural_header_hash"],
        "approved_header_hash": header_validation["approved_header_hash"],
        "header_hash_approval_basis": header_validation["approval_basis"],
    }
    records = []
    for row_number, raw_values in _iter_source_rows(input_path, source_sheet["sheet_name"], positions):
        if not any(value not in (None, "") for value in raw_values.values()):
            continue
        records.append(
            build_canonical_site_record(
                raw_values,
                profile,
                {**metadata, "source_row_number": row_number},
                scope=scope,
                resolved_mappings=resolved,
                sow_registry=sow_registry,
            )
        )
    return records, metadata