#!/usr/bin/env python3
"""
Regression tests for strict TI SOW matching.

These tests exercise both the shared production helper contract and the
production TI generator CLI path with controlled temporary Excel fixtures.
"""

import csv
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
from pr_helpers import normalize_ti_sow, ti_sow_matches_model


REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPT_PATH = REPO_ROOT / "scripts" / "generate_tss_pr_ecc.py"
PR_MODEL_PATH = REPO_ROOT / "Info" / "input" / "pr_model.xlsx"
TEMPLATE_PATH = REPO_ROOT / "Info" / "input" / "ecc_template.xls"
MAPPING_PATH = REPO_ROOT / "Info" / "input" / "contract_info_reference.md"


SITE_COLUMNS = [
    "customer site code",
    "customer site name",
    "du code",
    "region",
    "Province/State",
    "Latitude (North Plus South Minus)",
    "Longitude (East Plus West Minus)",
    "TX Upgrade Scope",
    "Tx SOW",
    "MW Config Antenna Size NE",
    "MW Config Antenna Size FE",
    "SubCon - TI Team",
    "Subcon PR - TI",
]


class TestStrictTiSowMatcher(unittest.TestCase):
    def test_shared_helper_matches_exact_sow_only(self):
        self.assertEqual(normalize_ti_sow("  mw   idu  patching "), "MW IDU PATCHING")
        self.assertEqual(normalize_ti_sow(""), "")
        self.assertFalse(ti_sow_matches_model("BBU Patching Extended", "BBU Patching"))
        self.assertFalse(ti_sow_matches_model("MW Parallel Link Extended", "MW Parallel Link"))
        self.assertTrue(ti_sow_matches_model("  bbu   patching  ", "BBU Patching"))


class TestProductionTiSowMatching(unittest.TestCase):
    def _write_site_workbook(self, path, rows):
        df = pd.DataFrame(rows, columns=SITE_COLUMNS)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="data", index=False, startrow=3)

    def _run_ti_generator(self, site_rows):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            site_path = tmpdir_path / "site_data.xlsx"
            output_dir = tmpdir_path / "output"
            output_dir.mkdir()
            self._write_site_workbook(site_path, site_rows)

            command = [
                sys.executable,
                str(SCRIPT_PATH),
                "--site-data",
                str(site_path),
                "--pr-model",
                str(PR_MODEL_PATH),
                "--template",
                str(TEMPLATE_PATH),
                "--mapping",
                str(MAPPING_PATH),
                "--output",
                str(output_dir),
                "--scope",
                "TI",
                "--all-sites",
            ]
            result = subprocess.run(
                command,
                cwd=REPO_ROOT,
                text=True,
                capture_output=True,
                check=False,
            )

            workbook_rows = []
            for workbook_path in sorted(output_dir.glob("*.xlsx")):
                wb = load_workbook(workbook_path, data_only=True)
                ws = wb["details"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    workbook_rows.append(
                        {
                            "Site_ID": row[3],
                            "PBOM_Code": str(row[9]),
                            "SOW": row[10],
                            "Remarks": row[13] or "",
                            "Source_Tx_SOW": row[14],
                        }
                    )

            review_rows = []
            review_files = sorted(output_dir.glob("REVIEW_REQUIRED_TI_*.csv"))
            if review_files:
                # utf-8-sig: review CSVs carry a BOM since the approved
                # 2026-07-08 encoding fix.
                with review_files[0].open(newline="", encoding="utf-8-sig") as handle:
                    review_rows = list(csv.DictReader(handle))

            return result, workbook_rows, review_rows

    def test_ti_generator_requires_exact_sow_matches(self):
        site_rows = [
            {
                "customer site code": "TI_BBU_OK",
                "customer site name": "BBU Positive",
                "du code": "DU_BBU_OK",
                "region": "Sarawak",
                "Province/State": "Sarawak",
                "Latitude (North Plus South Minus)": 1.171995,
                "Longitude (East Plus West Minus)": 110.566047,
                "TX Upgrade Scope": "TSS+TI",
                "Tx SOW": "  bbu   patching  ",
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "SubCon - TI Team": "Trintel",
                "Subcon PR - TI": "",
            },
            {
                "customer site code": "TI_IDU_OK",
                "customer site name": "IDU Positive",
                "du code": "DU_IDU_OK",
                "region": "Northern",
                "Province/State": "Kedah",
                "Latitude (North Plus South Minus)": 6.4171,
                "Longitude (East Plus West Minus)": 100.49,
                "TX Upgrade Scope": "TSS+TI",
                "Tx SOW": "mw idu patching",
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "SubCon - TI Team": "GTSB",
                "Subcon PR - TI": "",
            },
            {
                "customer site code": "TI_BBU_EXT",
                "customer site name": "BBU Extended Negative",
                "du code": "DU_BBU_EXT",
                "region": "Sarawak",
                "Province/State": "Sarawak",
                "Latitude (North Plus South Minus)": 2.131453,
                "Longitude (East Plus West Minus)": 111.485124,
                "TX Upgrade Scope": "TSS+TI",
                "Tx SOW": "BBU Patching Extended",
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "SubCon - TI Team": "Trintel",
                "Subcon PR - TI": "",
            },
            {
                "customer site code": "TI_MW_ANT_REVIEW",
                "customer site name": "MW Antenna Review",
                "du code": "DU_MW_ANT_REVIEW",
                "region": "Sabah",
                "Province/State": "Sabah",
                "Latitude (North Plus South Minus)": 4.414137,
                "Longitude (East Plus West Minus)": 118.522111,
                "TX Upgrade Scope": "TSS+AA+TI",
                "Tx SOW": "MW Parallel Link",
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "SubCon - TI Team": "NR services",
                "Subcon PR - TI": "",
            },
            {
                "customer site code": "TI_MW_ANT_OK",
                "customer site name": "MW Antenna Positive",
                "du code": "DU_MW_ANT_OK",
                "region": "Sabah",
                "Province/State": "Sabah",
                "Latitude (North Plus South Minus)": 4.414137,
                "Longitude (East Plus West Minus)": 118.522111,
                "TX Upgrade Scope": "TSS+AA+TI",
                "Tx SOW": "MW Parallel Link",
                "MW Config Antenna Size NE": "1.2m",
                "MW Config Antenna Size FE": "1.2m",
                "SubCon - TI Team": "NR services",
                "Subcon PR - TI": "",
            },
            {
                "customer site code": "TI_MW_EXT",
                "customer site name": "MW Extended Negative",
                "du code": "DU_MW_EXT",
                "region": "Sabah",
                "Province/State": "Sabah",
                "Latitude (North Plus South Minus)": 4.414137,
                "Longitude (East Plus West Minus)": 118.522111,
                "TX Upgrade Scope": "TSS+AA+TI",
                "Tx SOW": "MW Parallel Link Extended",
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "SubCon - TI Team": "NR services",
                "Subcon PR - TI": "",
            },
        ]

        result, workbook_rows, review_rows = self._run_ti_generator(site_rows)

        self.assertEqual(
            result.returncode,
            0,
            msg=f"stdout:\n{result.stdout}\n\nstderr:\n{result.stderr}",
        )

        rows_by_site = {}
        for row in workbook_rows:
            rows_by_site.setdefault(row["Site_ID"], []).append(row)

        self.assertIn("TI_BBU_OK", rows_by_site)
        self.assertEqual(
            {row["PBOM_Code"] for row in rows_by_site["TI_BBU_OK"]},
            {"350001095420"},
        )
        self.assertEqual(
            {row["Source_Tx_SOW"] for row in rows_by_site["TI_BBU_OK"]},
            {"bbu   patching"},
        )

        self.assertIn("TI_IDU_OK", rows_by_site)
        self.assertEqual(
            {row["PBOM_Code"] for row in rows_by_site["TI_IDU_OK"]},
            {"350001095420"},
        )

        self.assertIn("TI_MW_ANT_OK", rows_by_site)
        self.assertIn(
            "350001095410",
            {row["PBOM_Code"] for row in rows_by_site["TI_MW_ANT_OK"]},
        )

        self.assertNotIn("TI_BBU_EXT", rows_by_site)
        self.assertNotIn("TI_MW_ANT_REVIEW", rows_by_site)
        self.assertNotIn("TI_MW_EXT", rows_by_site)

        review_by_site = {row["Site_ID"]: row for row in review_rows}
        self.assertEqual(review_by_site["TI_BBU_EXT"]["Reason_Code"], "NO_MATCHING_TI_PR_MODEL_ITEM")
        self.assertEqual(review_by_site["TI_MW_ANT_REVIEW"]["Reason_Code"], "MISSING_TI_ANTENNA_SIZE")
        self.assertEqual(review_by_site["TI_MW_EXT"]["Reason_Code"], "NO_MATCHING_TI_PR_MODEL_ITEM")


if __name__ == "__main__":
    unittest.main()
