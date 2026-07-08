import json
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from build_mapping_decision_workbook import (
    DECISION_OPTIONS,
    FLAG_DERIVED_FIELD,
    FLAG_MULTIPLE_CANDIDATES,
    FLAG_NO_SOURCE_CANDIDATE,
    FLAG_SINGLE_CANDIDATE_VERIFY,
    build_decision_package,
    build_review_model,
    decision_summary_markdown,
    mask_rule_for,
    mask_value,
    requirement_class,
)

REVIEW_FIELDS = (
    "site_code",
    "tx_sow_raw",
    "tx_sow_normalized",
    "existing_tss_pr_status",
    "latitude",
)


def _column(field_code, wbs, task, display, excel_column, index):
    return {
        "source_position": {"excel_column": excel_column, "one_based_index": index},
        "fingerprint": {
            "field_code": field_code,
            "wbs_stage": wbs,
            "task_name": task,
            "display_header": display,
        },
    }


def _candidate(field_code, wbs, task, display, sheet_name="data"):
    return {
        "sheet_name": sheet_name,
        "fingerprint": {
            "field_code": field_code,
            "wbs_stage": wbs,
            "task_name": task,
            "display_header": display,
        },
        "mapping_status": "UNVERIFIED",
    }


def _synthetic_inputs():
    header_inventory = {
        "schema_version": "1.0",
        "source": {
            "file_name": "synthetic-tx-mini.xlsx",
            "source_file_hash": "",
            "format": "xlsx",
            "header_row_count": 4,
        },
        "sheets": [
            {
                "sheet_name": "data",
                "header_row_count": 4,
                "columns": [
                    _column("site|fix00012|1|2", "Site Basic Info", "Site Basic Info", "customer site code", "A", 1),
                    _column("docata|SOW1", "Installation", "Microwave", "Tx SOW", "B", 2),
                    _column("docata|SOW2", "TX Solution", "TX SOW Details", "TX SOW Details", "C", 3),
                    _column("site|fix00013", "Site Basic Info", "Site Basic Info", "Latitude", "D", 4),
                ],
            }
        ],
    }
    candidates_report = {
        "schema_version": "1.0",
        "discovery_only": True,
        "fields": {
            "site_code": {
                "status": "UNVERIFIED",
                "candidates": [
                    _candidate("site|fix00012|1|2", "Site Basic Info", "Site Basic Info", "customer site code")
                ],
            },
            "tx_sow_raw": {
                "status": "AMBIGUOUS",
                "candidates": [
                    _candidate("docata|SOW1", "Installation", "Microwave", "Tx SOW"),
                    _candidate("docata|SOW2", "TX Solution", "TX SOW Details", "TX SOW Details"),
                    _candidate("PERSON NAME123", "Approval Stage", "", "", sheet_name="drop_down"),
                ],
            },
            "existing_tss_pr_status": {"status": "AMBIGUOUS", "candidates": []},
            "latitude": {
                "status": "UNVERIFIED",
                "candidates": [
                    _candidate("site|fix00013", "Site Basic Info", "Site Basic Info", "Latitude")
                ],
            },
        },
    }
    du_profile = {
        "profile_id": "synthetic_pr_v1",
        "profile_version": "0.1.0",
        "mapping_version": "discovery-synthetic-v1",
        "status": "DRAFT",
        "field_mapping": {
            "site_code": {"required": True, "source_candidates": [], "transforms": ["trim", "uppercase"]},
            "tx_sow_raw": {"required": True, "source_candidates": [], "transforms": ["trim"]},
            "existing_tss_pr_status": {"required": True, "source_candidates": [], "transforms": ["trim"]},
            "latitude": {"required": False, "source_candidates": [], "transforms": ["parse_decimal"]},
        },
    }
    shortlist_registry = {
        "entries": [
            {
                "observed_header_hash": "synthetic-hash",
                "skill_field_shortlists": {
                    "tx_sow": [
                        {
                            "score": 100,
                            "reason": "Direct Tx SOW field.",
                            "fingerprint": {
                                "field_code": "docata|SOW1",
                                "wbs_stage": "Installation",
                                "task_name": "Microwave",
                                "display_header": "Tx SOW",
                            },
                        },
                        {
                            "score": 45,
                            "reason": "SOW details field.",
                            "fingerprint": {
                                "field_code": "docata|SOW2",
                                "wbs_stage": "TX Solution",
                                "task_name": "TX SOW Details",
                                "display_header": "TX SOW Details",
                            },
                        },
                    ]
                },
            }
        ]
    }
    return header_inventory, candidates_report, du_profile, shortlist_registry


def _write_source_export(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(["site|fix00012|1|2", "docata|SOW1", "docata|SOW2", "site|fix00013"])
    sheet.append(["Site Basic Info", "Installation", "TX Solution", "Site Basic Info"])
    sheet.append(["Site Basic Info", "Microwave", "TX SOW Details", "Site Basic Info"])
    sheet.append(["customer site code", "Tx SOW", "TX SOW Details", "Latitude"])
    rows = [
        ["KUL0001", "MW Upgrade", "Upgrade 1+0 to 2+0", 3.152893],
        ["KUL0002", "MW New Link", "New 1+0 link", 3.98211],
        ["SBH0003", "MW Upgrade", "Upgrade 2+0 to 4+0", 5.42371],
        ["SWK0004", "MW Dismantle", "Dismantle link", 1.55512],
        ["PNG0005", "MW New Link", "New 2+0 link", -2.31441],
        ["JHR0006", "MW Upgrade", "Upgrade 1+1", 4.10009],
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class TestRequirementClass(unittest.TestCase):
    def test_scope_classification_follows_validator_contract(self):
        self.assertEqual(requirement_class("site_code"), "REQUIRED (TSS + TI)")
        self.assertEqual(requirement_class("existing_tss_pr_status"), "REQUIRED (TSS)")
        self.assertEqual(requirement_class("existing_ti_pr_status"), "REQUIRED (TI)")
        self.assertEqual(requirement_class("latitude"), "CONDITIONAL")


class TestMasking(unittest.TestCase):
    def test_code_masking_hides_middle_characters(self):
        masked = mask_value("KUL0001", "code")
        self.assertEqual(masked, "KU****1")
        self.assertNotIn("KUL0001", masked)

    def test_coordinate_masking_drops_decimals(self):
        self.assertEqual(mask_value("3.152893", "coordinate"), "3.***")
        self.assertEqual(mask_value("-2.31441", "coordinate"), "-2.***")

    def test_partner_and_name_masking(self):
        self.assertEqual(mask_value("EXAMPLE SUBCON SDN BHD", "partner"), "EXAM***")
        self.assertEqual(mask_value("Kampung Tower Alpha", "name"), "Kam***")

    def test_unmasked_rule_passes_value_through(self):
        self.assertEqual(mask_value("MW Upgrade", None), "MW Upgrade")

    def test_short_partner_values_keep_two_characters(self):
        self.assertEqual(mask_value("GTSB", "partner"), "GT***")

    def test_unmasked_text_scrubs_site_tokens_and_long_digit_runs(self):
        scrubbed = mask_value("MW SWAP A01073-A00607 WITH 18G_2+0", None)
        self.assertNotIn("A01073", scrubbed)
        self.assertNotIn("A00607", scrubbed)
        self.assertIn("18G_2+0", scrubbed)
        self.assertNotIn("6798000131", mask_value("APPROVED BY 6798000131", None))

    def test_column_level_mask_hints_override_unmasked_fields(self):
        self.assertEqual(mask_rule_for("existing_tss_pr_status", "responsible person"), "partner")
        self.assertEqual(mask_rule_for("existing_tss_pr_status", "Subcon PR - TSS"), "partner")
        self.assertEqual(mask_rule_for("existing_tss_pr_status", "responsible company"), "partner")
        self.assertIsNone(mask_rule_for("existing_tss_pr_status", "activity status"))
        # Field-level rules still win over header hints.
        self.assertEqual(mask_rule_for("site_code", "customer site code"), "code")


class TestReviewModel(unittest.TestCase):
    def setUp(self):
        self.header_inventory, self.candidates_report, self.du_profile, self.shortlist_registry = _synthetic_inputs()
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp.name)
        self.source_export = self.tmp_path / "synthetic-tx-mini.xlsx"
        _write_source_export(self.source_export)

    def tearDown(self):
        self._tmp.cleanup()

    def _model(self):
        return build_review_model(
            candidates_report=self.candidates_report,
            header_inventory=self.header_inventory,
            observed_header_hash="synthetic-hash",
            du_profile=self.du_profile,
            shortlist_registry=self.shortlist_registry,
            source_export_path=self.source_export,
            review_fields=REVIEW_FIELDS,
        )

    def _field(self, model, name):
        return next(f for f in model["fields"] if f["canonical_field"] == name)

    def test_flags_cover_single_multiple_missing_and_derived(self):
        model = self._model()
        self.assertEqual(self._field(model, "site_code")["flag"], FLAG_SINGLE_CANDIDATE_VERIFY)
        self.assertEqual(self._field(model, "tx_sow_raw")["flag"], FLAG_MULTIPLE_CANDIDATES)
        self.assertEqual(self._field(model, "existing_tss_pr_status")["flag"], FLAG_NO_SOURCE_CANDIDATE)
        self.assertEqual(self._field(model, "tx_sow_normalized")["flag"], FLAG_DERIVED_FIELD)

    def test_candidates_are_ranked_by_shortlist_score_with_noise_last(self):
        model = self._model()
        candidates = self._field(model, "tx_sow_raw")["candidates"]
        self.assertEqual(len(candidates), 3)
        self.assertEqual(candidates[0]["fingerprint"]["display_header"], "Tx SOW")
        self.assertEqual(candidates[0]["rank"], 1)
        self.assertEqual(candidates[0]["shortlist_score"], 100)
        self.assertEqual(candidates[1]["fingerprint"]["display_header"], "TX SOW Details")
        self.assertTrue(candidates[2]["is_noise"])
        self.assertEqual(candidates[2]["sampling"]["samples"], [])

    def test_sample_values_resolved_by_fingerprint_and_masked(self):
        model = self._model()
        site_candidate = self._field(model, "site_code")["candidates"][0]
        self.assertEqual(site_candidate["excel_column"], "A")
        samples = site_candidate["sampling"]["samples"]
        self.assertEqual(len(samples), 5)
        self.assertIn("KU****1", samples)
        for raw in ("KUL0001", "KUL0002", "SBH0003"):
            self.assertNotIn(raw, samples)
        latitude_samples = self._field(model, "latitude")["candidates"][0]["sampling"]["samples"]
        self.assertIn("3.***", latitude_samples)
        self.assertNotIn("3.152893", latitude_samples)

    def test_unmasked_categorical_values_and_counts(self):
        model = self._model()
        sow_sampling = self._field(model, "tx_sow_raw")["candidates"][0]["sampling"]
        self.assertIn("MW Upgrade", sow_sampling["samples"])
        self.assertEqual(sow_sampling["non_blank_count"], 6)
        self.assertEqual(sow_sampling["distinct_count"], 3)

    def test_no_decision_is_prefilled_and_summary_lists_decision_fields(self):
        model = self._model()
        summary = decision_summary_markdown(model)
        self.assertIn("`existing_tss_pr_status`", summary)
        self.assertIn("`tx_sow_raw`", summary)
        self.assertIn("status `DRAFT` — unchanged", summary)
        self.assertNotIn("APPROVED", summary)


class TestDecisionPackage(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp.name)
        header_inventory, candidates_report, du_profile, shortlist_registry = _synthetic_inputs()
        self.profiler_dir = self.tmp_path / "profiler"
        self.profiler_dir.mkdir()
        (self.profiler_dir / "header_inventory.json").write_text(json.dumps(header_inventory), encoding="utf-8")
        (self.profiler_dir / "canonical_field_candidates.json").write_text(
            json.dumps(candidates_report), encoding="utf-8"
        )
        (self.profiler_dir / "header_hash.txt").write_text("synthetic-hash\n", encoding="utf-8")
        self.du_profile_path = self.tmp_path / "profile.yaml"
        self.du_profile_path.write_text(json.dumps(du_profile), encoding="utf-8")
        self.shortlist_path = self.tmp_path / "shortlists.yaml"
        self.shortlist_path.write_text(json.dumps(shortlist_registry), encoding="utf-8")
        self.source_export = self.tmp_path / "synthetic-tx-mini.xlsx"
        _write_source_export(self.source_export)
        self.output_dir = self.tmp_path / "review"

    def tearDown(self):
        self._tmp.cleanup()

    def test_package_writes_workbook_with_required_sheets_and_summary(self):
        from openpyxl import load_workbook

        build_decision_package(
            profiler_dir=self.profiler_dir,
            du_profile_path=self.du_profile_path,
            output_dir=self.output_dir,
            workbook_name="WORKBOOK.xlsx",
            summary_name="SUMMARY.md",
            shortlist_registry_path=self.shortlist_path,
            source_export_path=self.source_export,
            review_fields=REVIEW_FIELDS,
        )
        workbook_path = self.output_dir / "WORKBOOK.xlsx"
        self.assertTrue(workbook_path.exists())
        self.assertTrue((self.output_dir / "SUMMARY.md").exists())

        workbook = load_workbook(workbook_path)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Review_Instructions",
                "Required_Fields",
                "Candidate_Fingerprints",
                "Sample_Source_Values",
                "Decision_Log",
                "Unresolved_Fields",
            ],
        )
        decision_sheet = workbook["Decision_Log"]
        # One row per review field, no prefilled reviewer decisions.
        self.assertEqual(decision_sheet.max_row, 1 + len(REVIEW_FIELDS))
        for row in decision_sheet.iter_rows(min_row=2, min_col=6, max_col=10, values_only=True):
            self.assertTrue(all(cell in (None, "") for cell in row))
        validations = decision_sheet.data_validations.dataValidation
        self.assertEqual(len(validations), 1)
        for option in DECISION_OPTIONS:
            self.assertIn(option, validations[0].formula1)

        candidate_sheet = workbook["Candidate_Fingerprints"]
        header = [cell.value for cell in candidate_sheet[1]]
        self.assertIn("Field code / ID", header)
        self.assertIn("WBS stage", header)
        self.assertIn("Task name", header)
        self.assertIn("Display header", header)
        position_header = header[7]
        self.assertIn("NON-AUTHORITATIVE", position_header)

        unresolved_sheet = workbook["Unresolved_Fields"]
        unresolved_fields = [row[0].value for row in unresolved_sheet.iter_rows(min_row=2)]
        self.assertIn("existing_tss_pr_status", unresolved_fields)
        self.assertIn("tx_sow_raw", unresolved_fields)
        self.assertIn("tx_sow_normalized", unresolved_fields)
        self.assertNotIn("site_code", unresolved_fields)

        # No raw sensitive site codes or noise-candidate person codes may leak into any sheet.
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        self.assertNotIn("KUL0001", cell)
                        self.assertNotIn("PERSON NAME123", cell)

    def test_package_refuses_mismatched_source_export(self):
        header_inventory, _, _, _ = _synthetic_inputs()
        header_inventory["source"]["source_file_hash"] = "0" * 64
        (self.profiler_dir / "header_inventory.json").write_text(json.dumps(header_inventory), encoding="utf-8")
        with self.assertRaises(ValueError):
            build_decision_package(
                profiler_dir=self.profiler_dir,
                du_profile_path=self.du_profile_path,
                output_dir=self.output_dir,
                workbook_name="WORKBOOK.xlsx",
                summary_name="SUMMARY.md",
                shortlist_registry_path=self.shortlist_path,
                source_export_path=self.source_export,
                review_fields=REVIEW_FIELDS,
            )


if __name__ == "__main__":
    unittest.main()
