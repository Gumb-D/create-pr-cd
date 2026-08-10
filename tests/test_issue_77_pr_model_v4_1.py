import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
CANONICAL = ROOT / "Info" / "input" / "pr_model.xlsx"
STARLINK_SOW = "Starlink Dismanle"
MW_NEW_LINK_SOW = "MW New Link / Reroute"


class TestIssue77PrModelV41(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        workbook = load_workbook(CANONICAL, read_only=True, data_only=True)
        worksheet = workbook["TX Line Item (After 21-Apr 26)"]
        ti_header_row = next(
            row[0].row
            for row in worksheet.iter_rows()
            if isinstance(row[0].value, str) and "TI Model" in row[0].value
        )
        cls.rows = [
            tuple(cell.value for cell in row[:7])
            for row in worksheet.iter_rows(min_row=ti_header_row + 1)
            if row[1].value is not None
        ]
        workbook.close()
        relevant_names = {
            STARLINK_SOW,
            "MW Dismantle",
            MW_NEW_LINK_SOW,
            "BBU Patching",
            "MW IDU Patching",
        }
        cls.relevant_rows = [row for row in cls.rows if row[0] in relevant_names]
        cls.relevant_sows = sorted({str(row[0]) for row in cls.relevant_rows})
        print("ISSUE77_V41_RELEVANT_ROWS=" + repr(cls.relevant_rows))

    def _rows_for_sow(self, sow):
        return [row for row in self.rows if row[0] == sow]

    def test_starlink_dismantle_sow_exists(self):
        rows = self._rows_for_sow(STARLINK_SOW)
        self.assertTrue(rows, f"Missing v4.1 SOW: {STARLINK_SOW}; relevant={self.relevant_sows}")
        self.assertTrue(all(row[1] is not None for row in rows), rows)

    def test_mw_dismantle_sow_exists(self):
        self.assertTrue(self._rows_for_sow("MW Dismantle"), f"Missing v4.1 SOW: MW Dismantle; relevant={self.relevant_sows}")

    def test_mw_new_link_reroute_sow_exists(self):
        self.assertTrue(
            self._rows_for_sow(MW_NEW_LINK_SOW),
            f"Missing v4.1 SOW: {MW_NEW_LINK_SOW}; relevant={self.relevant_sows}",
        )

    def test_patching_models_exist_as_exact_split_models(self):
        self.assertTrue(self._rows_for_sow("BBU Patching"), "Missing v4.1 SOW: BBU Patching")
        self.assertTrue(self._rows_for_sow("MW IDU Patching"), "Missing v4.1 SOW: MW IDU Patching")

    def test_removed_v4_0_jendela_sow_names_are_not_reintroduced(self):
        self.assertFalse(self._rows_for_sow("Starlink Dismantle (Return/MRCF included) & Migration"))
        self.assertFalse(self._rows_for_sow("MW Installation"))


if __name__ == "__main__":
    unittest.main()
