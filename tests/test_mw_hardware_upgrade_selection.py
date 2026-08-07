#!/usr/bin/env python3
"""Regression coverage for Issue #69 MW Hardware Upgrade TI selection."""

from __future__ import annotations

import csv
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from canonical_generator_bridge_impl import canonical_record_to_generator_row  # noqa: E402
from mw_hardware_upgrade_selector import resolve_mw_hardware_upgrade_subtype  # noqa: E402


SCRIPT_PATH = SCRIPTS / "generate_tss_pr_ecc.py"
PR_MODEL_PATH = ROOT / "Info" / "input" / "pr_model.xlsx"
SITE_DATA_PATH = ROOT / "Info" / "input" / "site_pr_po_view.xlsx"
TEMPLATE_PATH = ROOT / "Info" / "input" / "ecc_template.xls"
MAPPING_PATH = ROOT / "Info" / "input" / "contract_info_reference.md"
PROFILE_DIR = ROOT / "config" / "du_profiles"

SITE_COLUMNS = [
    "customer site code",
    "customer site name",
    "du code",
    "region",
    "Province/State",
    "Latitude (North Plus South Minus)",
    "Longitude (East Plus West Minus)",
    "TX Upgrade Scope",
    "Tx SOW",
    "MW Config Antenna Size NE",
    "MW Config Antenna Size FE",
    "BOQ Configuration",
    "TX SOW Details",
    "NE SOW Details",
    "FE SOW Details",
    "SubCon - TI Team",
    "Subcon PR - TI",
    "DU Profile ID",
]

SUBTYPE_PBOMS = {
    "IDU_WITHOUT_SITE_SURVEY": "350001095419",
    "ODU_WITHOUT_SITE_SURVEY": "350001095418",
    "ODU_WITH_SITE_SURVEY": "350001095417",
}


def _site(
    site_code: str,
    upgrade_scope: str,
    details: str,
    profile_id: str = "tx_mini_pr_v1",
) -> dict[str, object]:
    return {
        "customer site code": site_code,
        "customer site name": f"Issue 69 {site_code}",
        "du code": f"DU-{site_code}",
        "region": "Central",
        "Province/State": "Selangor",
        "Latitude (North Plus South Minus)": 3.0738,
        "Longitude (East Plus West Minus)": 101.5183,
        "TX Upgrade Scope": upgrade_scope,
        "Tx SOW": "  mw   hardware   upgrade  ",
        "MW Config Antenna Size NE": "",
        "MW Config Antenna Size FE": "",
        "BOQ Configuration": details,
        "TX SOW Details": details,
        "NE SOW Details": details,
        "FE SOW Details": details,
        "SubCon - TI Team": "GTSB",
        "Subcon PR - TI": "",
        "DU Profile ID": profile_id,
    }


class MwHardwareUpgradeGeneratorHarness(unittest.TestCase):
    def _write_site_workbook(self, path: Path, rows: list[dict[str, object]]) -> None:
        dataframe = pd.DataFrame(rows, columns=SITE_COLUMNS)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name="data", index=False, startrow=3)

    def _run_generator(self, rows: list[dict[str, object]]):
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_root = Path(temporary_directory)
            site_path = temporary_root / "issue-69-site-data.xlsx"
            output_dir = temporary_root / "output"
            output_dir.mkdir()
            self._write_site_workbook(site_path, rows)

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    "--site-data",
                    str(site_path),
                    "--pr-model",
                    str(PR_MODEL_PATH),
                    "--template",
                    str(TEMPLATE_PATH),
                    "--mapping",
                    str(MAPPING_PATH),
                    "--output",
                    str(output_dir),
                    "--scope",
                    "TI",
                    "--all-sites",
                ],
                cwd=ROOT,
                text=True,
                capture_output=True,
                check=False,
            )

            output_rows: list[dict[str, object]] = []
            for workbook_path in sorted(output_dir.glob("*.xlsx")):
                workbook = load_workbook(workbook_path, read_only=True, data_only=True)
                worksheet = workbook["details"]
                for values in worksheet.iter_rows(min_row=2, values_only=True):
                    if not values[0]:
                        continue
                    output_rows.append(
                        {
                            "Site_ID": values[3],
                            "PBOM_Code": str(values[9]),
                            "Description": values[10],
                            "Unit": values[11],
                            "Quantity": values[12],
                        }
                    )
                workbook.close()

            review_rows: list[dict[str, str]] = []
            review_files = sorted(output_dir.glob("REVIEW_REQUIRED_TI_*.csv"))
            if review_files:
                with review_files[0].open(newline="", encoding="utf-8-sig") as handle:
                    review_rows = list(csv.DictReader(handle))

            return result, output_rows, review_rows

    def _assert_success(self, result: subprocess.CompletedProcess[str]) -> None:
        self.assertEqual(
            result.returncode,
            0,
            msg=f"stdout:\n{result.stdout}\n\nstderr:\n{result.stderr}",
        )


class TestMwHardwareUpgradePositiveSelection(MwHardwareUpgradeGeneratorHarness):
    def test_each_approved_mandatory_subtype_selects_exact_pr_model_row(self):
        cases = [
            (
                _site(
                    "HW_IDU",
                    "TSS+AA+TI",
                    "New ISM8 x1. Re-use existing RTN910AF IDU. Re-use existing XMC-5D ODU.",
                ),
                "IDU_WITHOUT_SITE_SURVEY",
            ),
            (
                _site(
                    "HW_ODU_NO_TSS",
                    "AA+TI",
                    "Re-use existing RTN910AF IDU. New XMC-3E ODU x1.",
                ),
                "ODU_WITHOUT_SITE_SURVEY",
            ),
            (
                _site(
                    "HW_ODU_TSS",
                    "TSS+AA+TI",
                    "Re-use existing RTN910AF IDU. New XMC-3E ODU x1.",
                ),
                "ODU_WITH_SITE_SURVEY",
            ),
        ]

        result, output_rows, review_rows = self._run_generator([row for row, _ in cases])
        self._assert_success(result)

        rows_by_site: dict[str, list[dict[str, object]]] = {}
        for row in output_rows:
            rows_by_site.setdefault(str(row["Site_ID"]), []).append(row)

        for source_row, subtype in cases:
            site_code = str(source_row["customer site code"])
            expected_pbom = SUBTYPE_PBOMS[subtype]
            with self.subTest(site=site_code):
                self.assertIn(site_code, rows_by_site)
                subtype_rows = [
                    row
                    for row in rows_by_site[site_code]
                    if row["PBOM_Code"] in set(SUBTYPE_PBOMS.values())
                ]
                self.assertEqual(len(subtype_rows), 1)
                self.assertEqual(subtype_rows[0]["PBOM_Code"], expected_pbom)
                self.assertEqual(subtype_rows[0]["Unit"], "Hop")
                self.assertEqual(float(subtype_rows[0]["Quantity"]), 1.0)

        self.assertEqual(review_rows, [])
        self.assertIn("[MW_HARDWARE_UPGRADE_SELECTOR]", result.stdout)
        self.assertIn("matched_model_rows=3", result.stdout)

    def test_repository_q02210_fixture_resolves_to_idu_pbom(self):
        dataframe = pd.read_excel(SITE_DATA_PATH, sheet_name="data", header=3)
        source = dataframe[
            dataframe["customer site code"].astype(str).str.upper().eq("Q02210_AD")
        ]
        self.assertEqual(len(source), 1, "Repository fixture must contain Q02210_AD exactly once")
        row = source.iloc[0].to_dict()
        row["SubCon - TI Team"] = "Allstar"
        row["Subcon PR - TI"] = ""
        row["DU Profile ID"] = "tx_mini_pr_v1"

        result, output_rows, review_rows = self._run_generator([row])
        self._assert_success(result)

        subtype_rows = [
            output_row
            for output_row in output_rows
            if output_row["Site_ID"] == "Q02210_AD"
            and output_row["PBOM_Code"] in set(SUBTYPE_PBOMS.values())
        ]
        self.assertEqual(len(subtype_rows), 1)
        self.assertEqual(
            subtype_rows[0]["PBOM_Code"],
            SUBTYPE_PBOMS["IDU_WITHOUT_SITE_SURVEY"],
        )
        self.assertFalse(
            any(
                review["Site_ID"] == "Q02210_AD"
                and review["Reason_Code"] == "NO_MATCHING_TI_PR_MODEL_ITEM"
                for review in review_rows
            )
        )


class TestMwHardwareUpgradeFailClosed(MwHardwareUpgradeGeneratorHarness):
    def _assert_unresolved(self, row: dict[str, object]) -> None:
        result, output_rows, review_rows = self._run_generator([row])
        self._assert_success(result)
        site_code = str(row["customer site code"])
        self.assertFalse(any(output["Site_ID"] == site_code for output in output_rows))
        review_by_site = {review["Site_ID"]: review for review in review_rows}
        self.assertEqual(
            review_by_site[site_code]["Reason_Code"],
            "MW_HARDWARE_UPGRADE_TYPE_UNRESOLVED",
        )
        self.assertNotEqual(
            review_by_site[site_code]["Reason_Code"],
            "NO_MATCHING_TI_PR_MODEL_ITEM",
        )

    def test_missing_component_evidence_blocks_whole_site(self):
        self._assert_unresolved(_site("HW_MISSING", "TSS+AA+TI", ""))

    def test_ambiguous_idu_and_odu_evidence_blocks_whole_site(self):
        self._assert_unresolved(
            _site(
                "HW_AMBIGUOUS",
                "TSS+AA+TI",
                "New ISM8 IDU x1 and new XMC-3E ODU x1.",
            )
        )

    def test_odu_without_tss_or_ti_scope_is_not_guessed(self):
        self._assert_unresolved(
            _site(
                "HW_ODU_UNQUALIFIED",
                "AA",
                "Re-use existing RTN910AF IDU. New XMC-3E ODU x1.",
            )
        )


class TestMwHardwareUpgradeCanonicalEvidence(unittest.TestCase):
    def test_bridge_preserves_all_selector_evidence_fields(self):
        record = {
            "identity": {"source_row_number": 2, "header_hash": "hash"},
            "site": {"site_code": "HW_BRIDGE", "site_name": "Bridge", "du_key": "DU-1"},
            "pr_context": {
                "tx_sow_raw": "MW Hardware Upgrade",
                "tx_sow_normalized": "MW Hardware Upgrade",
                "tx_upgrade_scope_raw": "TSS+AA+TI",
                "region": "Central",
                "state": "Selangor",
                "subcontractor_ti": "GTSB",
                "existing_ti_pr_status": "",
            },
            "technical_context": {
                "boq_configuration": "New ISM8",
                "tx_sow_details": "Re-use existing ODU",
                "ne_sow_details": "NE evidence",
                "fe_sow_details": "FE evidence",
            },
            "source_evidence": {
                "fields": {
                    "tx_sow_normalized": {"normalization_status": "APPROVED"},
                    "ti_actual_end_date": {"source_value": "2026-08-07"},
                }
            },
            "validation": {
                "profile_id": "tx_mini_pr_v1",
                "profile_version": "1.0",
                "mapping_version": "test",
                "pr_input_classification": "PR_INPUT_READY",
            },
        }

        row = canonical_record_to_generator_row(record, "TI")

        self.assertEqual(row["TX Upgrade Scope"], "TSS+AA+TI")
        self.assertEqual(row["BOQ Configuration"], "New ISM8")
        self.assertEqual(row["TX SOW Details"], "Re-use existing ODU")
        self.assertEqual(row["NE SOW Details"], "NE evidence")
        self.assertEqual(row["FE SOW Details"], "FE evidence")

    def test_shared_selector_is_not_bound_to_one_du_profile(self):
        evidence = _site(
            "HW_SHARED",
            "TSS+AA+TI",
            "New ISM8 x1. Re-use existing RTN910AF IDU and XMC-5D ODU.",
        )
        for profile_id in (
            "tx_mini_pr_v1",
            "mw_eos_swap_pr_v1",
            "zte_tx_mini_pr_v1",
            "tx_rollout_2023_pr_v1",
        ):
            with self.subTest(profile_id=profile_id):
                evidence["DU Profile ID"] = profile_id
                result = resolve_mw_hardware_upgrade_subtype(evidence)
                self.assertEqual(result["status"], "RESOLVED")
                self.assertEqual(result["pbom_code"], "350001095419")

    def test_profile_evidence_is_never_promoted_from_unverified_candidates(self):
        approved_profiles = []
        for profile_path in sorted(PROFILE_DIR.glob("*.yaml")):
            profile = json.loads(profile_path.read_text(encoding="utf-8"))
            if profile.get("status") != "PRODUCTION":
                continue
            field_mapping = profile.get("field_mapping", {})
            approved_fields = set()
            for field_name in ("tx_upgrade_scope_raw", "tx_sow_details"):
                candidates = field_mapping.get(field_name, {}).get("source_candidates", [])
                if any(candidate.get("mapping_status") == "APPROVED" for candidate in candidates):
                    approved_fields.add(field_name)
            if approved_fields == {"tx_upgrade_scope_raw", "tx_sow_details"}:
                approved_profiles.append(profile["profile_id"])

        self.assertIn("tx_mini_pr_v1", approved_profiles)


if __name__ == "__main__":
    unittest.main()
