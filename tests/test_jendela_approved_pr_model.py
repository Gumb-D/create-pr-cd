import csv
import shutil
import subprocess
import sys
import tempfile
import unittest
from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import create_pr
import pr_helpers


CANONICAL_PR_MODEL = ROOT / "Info" / "input" / "pr_model.xlsx"
APPROVED_V4_SHA256 = "6c4fda502a8998b41bd88704dd6c59d986dc6c46fe42b82947d12c0c0cd8178f"
STARLINK_SOW = "Starlink Dismanle"
MW_NEW_LINK_SOW = "MW New Link / Reroute"


class TestApprovedJendelaPrModel(unittest.TestCase):
    def _candidate_record(self, work_items):
        return {
            "identity": {"source_row_number": 7},
            "site": {"site_code": "JENDELA-TEST", "site_name": "Jendela Test", "du_key": "DU-JENDELA"},
            "pr_context": {
                "region": "Southern",
                "state": "Johor",
                "subcontractor_ti": "GTSB",
                "migration_decision": {
                    "classification": "APPROVED",
                    "decision_code": "JENDELA_TI_WORK_PLAN",
                    "work_items": work_items,
                },
            },
            "technical_context": {"antenna_size_ne": "0.6m", "antenna_size_fe": "0.6m"},
            "approved_contract": {"scope": "TI", "subcontractor": "GTSB"},
            "validation": {"profile_id": "jendela_tx_migration_pr_v1"},
        }

    def _run_generator(self, candidate, pr_model=CANONICAL_PR_MODEL):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        temp_path = Path(temp_dir.name)
        renderer_input = temp_path / "renderer.xlsx"
        output = temp_path / "output"
        output.mkdir()
        create_pr._write_renderer_input(renderer_input, [candidate])
        result = subprocess.run(
            [
                sys.executable,
                str(SCRIPTS / "generate_tss_pr_ecc.py"),
                "--site-data", str(renderer_input),
                "--pr-model", str(pr_model),
                "--template", str(ROOT / "Info" / "input" / "ecc_template.xls"),
                "--mapping", str(ROOT / "Info" / "input" / "contract_info_reference.md"),
                "--output", str(output),
                "--scope", "TI",
                "--all-sites",
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=False,
        )
        return result, output

    def test_canonical_workbook_is_the_exact_approved_v4_bytes(self):
        self.assertEqual(sha256(CANONICAL_PR_MODEL.read_bytes()).hexdigest(), APPROVED_V4_SHA256)

    def test_canonical_workbook_contains_exact_v4_1_jendela_rows(self):
        workbook = load_workbook(CANONICAL_PR_MODEL, read_only=True, data_only=True)
        worksheet = workbook["TX Line Item (After 21-Apr 26)"]
        ti_header_row = next(
            row[0].row
            for row in worksheet.iter_rows()
            if isinstance(row[0].value, str) and "TI Model" in row[0].value
        )
        rows = [
            tuple(cell.value for cell in row[:7])
            for row in worksheet.iter_rows(min_row=ti_header_row + 1)
            if row[1].value is not None
        ]
        workbook.close()

        starlink_rows = [row for row in rows if row[0] == STARLINK_SOW]
        self.assertEqual([str(row[1]) for row in starlink_rows], ["350000597850", "350000597852"])
        self.assertTrue(all(row[5] == "Mandatory" for row in starlink_rows))

        self.assertTrue(any(row[0] == "MW Dismantle" for row in rows))
        self.assertTrue(any(row[0] == MW_NEW_LINK_SOW for row in rows))
        self.assertTrue(any(row[0] == "BBU Patching" and str(row[1]) == "350001095420" for row in rows))
        self.assertTrue(any(row[0] == "MW IDU Patching" and str(row[1]) == "350001095420" for row in rows))

        self.assertFalse(any(row[0] == "Starlink Dismantle (Return/MRCF included) & Migration" for row in rows))
        self.assertFalse(any(row[0] == "MW Installation" for row in rows))

    def test_loader_separates_v4_1_new_link_and_reroute_rows(self):
        _, ti_models = pr_helpers.load_pr_model_items(CANONICAL_PR_MODEL)
        starlink = [item for item in ti_models if item["SOW"] == STARLINK_SOW]
        self.assertEqual([item["PBOM_Code"] for item in starlink], ["350000597850", "350000597852"])
        self.assertTrue(all(item["Is_Mandatory"] for item in starlink))

        new_link = [item for item in ti_models if item["SOW"] == MW_NEW_LINK_SOW]
        reroute = [item for item in ti_models if item["SOW"] == "MW Reroute"]
        self.assertTrue(new_link)
        self.assertTrue(reroute)
        self.assertTrue(all(item.get("Remarks", "").casefold() != "reroute" for item in new_link))
        self.assertTrue(all(item.get("Remarks", "").casefold() == "reroute" for item in reroute))

    def test_modified_canonical_candidate_fails_hash_validation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            modified = Path(temp_dir) / "modified.xlsx"
            shutil.copy2(CANONICAL_PR_MODEL, modified)
            workbook = load_workbook(modified)
            workbook["TX Line Item (After 21-Apr 26)"]["A1"] = "unapproved mutation"
            workbook.save(modified)
            workbook.close()
            candidate = self._candidate_record(
                [
                    {
                        "work_item": "Dismantle Starlink",
                        "model_sow": STARLINK_SOW,
                        "required_pbom_codes": ["350000597850", "350000597852"],
                    }
                ]
            )
            result, output = self._run_generator(candidate, modified)
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("PR_MODEL_HASH_MISMATCH", result.stderr + result.stdout)
        self.assertEqual(list(output.glob("*.xlsx")), [])

    def test_fixed_v4_1_starlink_and_patching_pboms_are_unique(self):
        _, ti_models = pr_helpers.load_pr_model_items(CANONICAL_PR_MODEL)

        starlink = [item for item in ti_models if item["SOW"] == STARLINK_SOW]
        self.assertEqual(
            pr_helpers.validate_required_pbom_selection(
                starlink,
                ["350000597850", "350000597852"],
            ),
            (True, None),
        )

        for sow in ("BBU Patching", "MW IDU Patching"):
            with self.subTest(sow=sow):
                mandatory = [item for item in ti_models if item["SOW"] == sow and item["Is_Mandatory"]]
                self.assertEqual([item["PBOM_Code"] for item in mandatory], ["350001095420"])
                self.assertEqual(
                    pr_helpers.validate_required_pbom_selection(mandatory, ["350001095420"]),
                    (True, None),
                )

    def test_new_business_work_plans_generate_complete_atomic_ecc(self):
        cases = [
            (
                [
                    {
                        "work_item": "Dismantle Starlink",
                        "model_sow": STARLINK_SOW,
                        "required_pbom_codes": ["350000597850", "350000597852"],
                    },
                    {
                        "work_item": "BBU Patching / MW IDU Patching",
                        "model_sow": "BBU Patching",
                        "required_pbom_codes": ["350001095420"],
                    },
                ],
                {"350000597850", "350000597852", "350001095420"},
            ),
            (
                [
                    {"work_item": "Dismantle MW", "model_sow": "MW Dismantle", "required_pbom_codes": []},
                    {
                        "work_item": "BBU Patching / MW IDU Patching",
                        "model_sow": "MW IDU Patching",
                        "required_pbom_codes": ["350001095420"],
                    },
                ],
                {"350000589265", "350001095413", "350001095420"},
            ),
            (
                [
                    {
                        "work_item": "Dismantle Starlink",
                        "model_sow": STARLINK_SOW,
                        "required_pbom_codes": ["350000597850", "350000597852"],
                    },
                    {"work_item": "MW New Link", "model_sow": MW_NEW_LINK_SOW, "required_pbom_codes": []},
                ],
                {"350000597850", "350000597852", "350000214932", "350001095409"},
            ),
            (
                [
                    {"work_item": "Dismantle MW", "model_sow": "MW Dismantle", "required_pbom_codes": []},
                    {"work_item": "MW New Link", "model_sow": MW_NEW_LINK_SOW, "required_pbom_codes": []},
                ],
                {"350000589265", "350001095413", "350000214932", "350001095409"},
            ),
        ]
        for work_items, expected_pboms in cases:
            with self.subTest(work_items=[item["work_item"] for item in work_items]):
                result, output = self._run_generator(self._candidate_record(work_items))
                self.assertEqual(result.returncode, 0, result.stderr + result.stdout)
                output_files = list(output.glob("*.xlsx"))
                self.assertEqual(len(output_files), 1, result.stdout + result.stderr)
                self.assertEqual(list(output.glob("REVIEW_REQUIRED_TI_*.csv")), [])
                workbook = load_workbook(output_files[0], read_only=True, data_only=True)
                worksheet = workbook["details"]
                actual_pboms = {str(row[9]) for row in worksheet.iter_rows(min_row=2, values_only=True)}
                workbook.close()
                self.assertEqual(actual_pboms, expected_pboms)

    def test_missing_fixed_pbom_rejects_entire_atomic_decision(self):
        candidate = self._candidate_record(
            [
                {
                    "work_item": "Dismantle Starlink",
                    "model_sow": STARLINK_SOW,
                    "required_pbom_codes": ["350000597850", "350000597852", "MISSING-PBOM"],
                },
                {
                    "work_item": "BBU Patching / MW IDU Patching",
                    "model_sow": "BBU Patching",
                    "required_pbom_codes": ["350001095420"],
                },
            ]
        )
        result, output = self._run_generator(candidate)
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(list(output.glob("*.xlsx")), [])
        review_files = list(output.glob("REVIEW_REQUIRED_TI_*.csv"))
        self.assertEqual(len(review_files), 1)
        with review_files[0].open(encoding="utf-8-sig", newline="") as handle:
            review_rows = list(csv.DictReader(handle))
        self.assertEqual({row["Site_ID"] for row in review_rows}, {"JENDELA-TEST"})


if __name__ == "__main__":
    unittest.main()
