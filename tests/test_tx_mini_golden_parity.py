import json
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from run_tx_mini_golden_parity import (
    _normalized_name,
    build_canonical_records,
    compare_output_dirs,
    render_canonical_path_view,
    resolution_report,
)
from profile_du_export import build_header_inventory


def _write_view(path: Path, site_values, sow_values) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(["site|fix00012|1|2", "docata|SOW1"])
    sheet.append(["Site Basic Info", "Installation"])
    sheet.append(["Site Basic Info", "Microwave"])
    sheet.append(["customer site code", "Tx SOW"])
    for site, sow in zip(site_values, sow_values):
        sheet.append([site, sow])
    workbook.save(path)


def _profile():
    return {
        "profile_id": "synthetic_pr_v1",
        "profile_version": "0.1.0",
        "mapping_version": "synthetic-v1",
        "status": "DRAFT",
        "identity": {
            "project_key": "Synthetic",
            "accepted_du_models": ["Synthetic DU"],
            "accepted_du_model_ids": ["1"],
            "accepted_view_ids": ["2"],
        },
        "export_structure": {
            "header_rows": [0, 1, 2, 3],
            "header_hash_policy": "strict",
            "approved_header_hashes": [],
        },
        "field_mapping": {
            "site_code": {
                "required": True,
                "source_candidates": [
                    {
                        "fingerprint": {
                            "field_code": "site|fix00012|1|2",
                            "wbs_stage": "Site Basic Info",
                            "task_name": "Site Basic Info",
                            "display_header": "customer site code",
                        },
                        "mapping_status": "APPROVED",
                    }
                ],
                "transforms": ["trim", "uppercase"],
            },
            "tx_sow_raw": {
                "required": True,
                "source_candidates": [
                    {
                        "fingerprint": {
                            "field_code": "docata|SOW1",
                            "wbs_stage": "Installation",
                            "task_name": "Microwave",
                            "display_header": "Tx SOW",
                        },
                        "mapping_status": "APPROVED",
                    }
                ],
                "transforms": ["trim"],
            },
        },
    }


class TestNormalizedName(unittest.TestCase):
    def test_strips_timestamps_for_pairing(self):
        self.assertEqual(
            _normalized_name(Path("Sabah-GTSB TX Mini Project TSS PR 20260708093015.xlsx")),
            _normalized_name(Path("Sabah-GTSB TX Mini Project TSS PR 20260708101122.xlsx")),
        )
        self.assertNotEqual(
            _normalized_name(Path("Sabah-GTSB TSS PR 1.xlsx")),
            _normalized_name(Path("Sabah-OTHER TSS PR 1.xlsx")),
        )


class TestResolutionAndRoundTrip(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp.name)
        self.view_path = self.tmp_path / "view.xlsx"
        _write_view(self.view_path, [" a0001 ", "A0002", "A0003"], ["MW Swap", "MW Dismantle", "MW Swap"])
        self.profile = _profile()
        self.inventory = build_header_inventory(self.view_path)

    def tearDown(self):
        self._tmp.cleanup()

    def test_resolution_report_flags_unresolved_mapped_fields(self):
        report = resolution_report(self.inventory, self.profile)
        self.assertTrue(report["dry_run_permitted"])
        self.assertEqual(report["unresolved_mapped_fields"], [])
        self.assertFalse(report["header_hash_approved"])

        broken = json.loads(json.dumps(self.profile))
        broken["field_mapping"]["tx_sow_raw"]["source_candidates"][0]["fingerprint"]["display_header"] = "Nope"
        broken_report = resolution_report(self.inventory, broken)
        self.assertFalse(broken_report["dry_run_permitted"])
        self.assertEqual(broken_report["unresolved_mapped_fields"], ["tx_sow_raw"])

    def test_canonical_records_carry_raw_evidence_per_row(self):
        report = resolution_report(self.inventory, self.profile)
        records, stats = build_canonical_records(
            self.view_path, self.profile, self.inventory, report["resolved_mappings"]
        )
        self.assertEqual(stats["record_count"], 3)
        first = records[0]
        self.assertEqual(first["identity"]["source_row_number"], 5)
        self.assertEqual(first["site"]["site_code"], "A0001")
        self.assertEqual(first["source_evidence"]["fields"]["site_code"]["source_value"], " a0001 ")
        self.assertEqual(first["pr_context"]["tx_sow_raw"], "MW Swap")

    def test_render_round_trip_reproduces_mapped_column_values(self):
        from openpyxl import load_workbook

        report = resolution_report(self.inventory, self.profile)
        records, _ = build_canonical_records(
            self.view_path, self.profile, self.inventory, report["resolved_mappings"]
        )
        out_path = self.tmp_path / "canonical.xlsx"
        render_stats = render_canonical_path_view(
            self.view_path, records, report["resolved_mappings"], self.inventory, out_path
        )
        self.assertEqual(render_stats["rewritten_cells"], 6)
        original = load_workbook(self.view_path)["data"]
        rendered = load_workbook(out_path)["data"]
        for row in range(1, 8):
            for column in (1, 2):
                self.assertEqual(
                    rendered.cell(row=row, column=column).value,
                    original.cell(row=row, column=column).value,
                )


class TestCompareOutputDirs(unittest.TestCase):
    def test_detects_identical_and_differing_outputs(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            legacy = tmp_path / "legacy"
            canonical = tmp_path / "canonical"
            legacy.mkdir()
            canonical.mkdir()

            for directory, quantity in ((legacy, 5), (canonical, 5)):
                workbook = Workbook()
                workbook.active.append(["Site", "Quantity"])
                workbook.active.append(["A0001", quantity])
                workbook.save(directory / "Sabah-GTSB TSS PR 20260708.xlsx")
            (legacy / "REVIEW_20260708.csv").write_text("a,b\n1,2\n", encoding="utf-8")
            (canonical / "REVIEW_20260708.csv").write_text("a,b\n1,2\n", encoding="utf-8")

            report = compare_output_dirs(legacy, canonical)
            self.assertTrue(report["identical"])

            workbook = Workbook()
            workbook.active.append(["Site", "Quantity"])
            workbook.active.append(["A0001", 999])
            workbook.save(canonical / "Sabah-GTSB TSS PR 20260708.xlsx")
            report = compare_output_dirs(legacy, canonical)
            self.assertFalse(report["identical"])
            xlsx_entry = next(e for e in report["files"] if e["file"].endswith(".xlsx"))
            self.assertFalse(xlsx_entry["identical"])
            self.assertTrue(xlsx_entry["sample_differences"])

    def test_flags_missing_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            legacy = tmp_path / "legacy"
            canonical = tmp_path / "canonical"
            legacy.mkdir()
            canonical.mkdir()
            (legacy / "ONLY_LEGACY_20260708.csv").write_text("a\n", encoding="utf-8")
            report = compare_output_dirs(legacy, canonical)
            self.assertFalse(report["identical"])
            self.assertEqual(report["only_in_legacy"], ["ONLY_LEGACY_#.csv"])


if __name__ == "__main__":
    unittest.main()
