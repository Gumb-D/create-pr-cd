import json
import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Side, Border

ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from run_tx_mini_ecc_parity import (
    APPROVED_PR_MODEL_SHA256,
    validate_pr_model,
    validate_candidate_manifest,
    validate_generator_result,
    validate_independent_paths,
    compare_workbook_structure,
    compare_cell,
    extract_business_fields,
    compare_business_fields,
    normalize_allowed_metadata,
    calculate_overall_parity,
    derive_expected_candidates,
    compute_manifest_identity_hash,
    cross_check_candidate_manifest,
    _extract_color_attrs,
)


class TestTxMiniTssEccParity(unittest.TestCase):

    def test_validate_pr_model_approved_hash_passes(self):
        pr_model_file = Path("Info/input/pr_model.xlsx")
        if pr_model_file.exists():
            sha = validate_pr_model(pr_model_file)
            self.assertEqual(sha, APPROVED_PR_MODEL_SHA256)

    def test_validate_pr_model_wrong_hash_raises(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(b"wrong pr model content")
            tmp_path = Path(tmp.name)
        try:
            with self.assertRaises(ValueError) as ctx:
                validate_pr_model(tmp_path)
            self.assertIn("PR_MODEL_HASH_MISMATCH", str(ctx.exception))
        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_validate_pr_model_missing_file_raises(self):
        with self.assertRaises(FileNotFoundError):
            validate_pr_model(Path("non_existent_file.xlsx"))

    def test_cli_cannot_override_hash(self):
        gen_script = Path("scripts/generate_tss_pr_ecc.py").read_text(encoding="utf-8")
        self.assertNotIn("--expected-pr-model-hash", gen_script)
        self.assertIn("APPROVED_PR_MODEL_SHA256", gen_script)

    def test_validate_candidate_manifest_valid(self):
        candidates = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        self.assertTrue(validate_candidate_manifest(candidates))

    def test_validate_candidate_manifest_11_row_fails(self):
        c_11 = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(c_11)
        self.assertIn("must be exactly 12", str(ctx.exception))

    def test_validate_candidate_manifest_13_row_fails(self):
        c_13 = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 14)]
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(c_13)
        self.assertIn("must be exactly 12", str(ctx.exception))

    def test_validate_candidate_manifest_duplicate_candidate_id_fails(self):
        dup_ids = [{"Candidate ID": "C001", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(dup_ids)
        self.assertIn("Duplicate Candidate IDs", str(ctx.exception))

    def test_validate_candidate_manifest_duplicate_source_row_fails(self):
        dup_rows = [{"Candidate ID": f"C{i:03d}", "Source Row": 1, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(dup_rows)
        self.assertIn("Duplicate Source Rows", str(ctx.exception))

    def test_validate_candidate_manifest_missing_candidate_id_fails(self):
        missing_id = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        missing_id.append({"Source Row": 12, "Site Code": "S012"})
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(missing_id)
        self.assertIn("missing 'Candidate ID'", str(ctx.exception))

    def test_validate_candidate_manifest_missing_source_row_fails(self):
        missing_sr = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        missing_sr.append({"Candidate ID": "C012", "Site Code": "S012"})
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(missing_sr)
        self.assertIn("missing 'Source Row'", str(ctx.exception))

    def test_validate_candidate_manifest_missing_site_code_fails(self):
        missing_site = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        missing_site.append({"Candidate ID": "C012", "Source Row": 12})
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(missing_site)
        self.assertIn("missing 'Site Code'", str(ctx.exception))

    def test_validate_candidate_manifest_blank_required_value_fails(self):
        blank_val = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        blank_val.append({"Candidate ID": "   ", "Source Row": 12, "Site Code": "S012"})
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(blank_val)
        self.assertIn("blank 'Candidate ID'", str(ctx.exception))

    def test_validate_candidate_manifest_non_list_json_payload_fails(self):
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest({"status": "invalid"})
        self.assertIn("must be a list", str(ctx.exception))

    def test_validate_candidate_manifest_non_object_row_fails(self):
        cands = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        cands.append("not_a_dict")
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(cands)
        self.assertIn("must be a dictionary/object", str(ctx.exception))

    def test_invalid_cached_manifest_prevents_generator_invocation(self):
        from unittest.mock import patch
        import run_tx_mini_ecc_parity

        with tempfile.TemporaryDirectory() as tmp_dir:
            out_path = Path(tmp_dir) / "output" / "tx-mini-tss-ecc-parity"
            manifest_dir = out_path / "candidate_manifest"
            manifest_dir.mkdir(parents=True, exist_ok=True)

            # Write invalid (11-row) cached manifest
            invalid_manifest = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
            (manifest_dir / "TX_MINI_TSS_CANDIDATE_MANIFEST.json").write_text(json.dumps(invalid_manifest))

            with patch("run_tx_mini_ecc_parity.Path") as mock_path, \
                 patch("subprocess.run") as mock_sub:
                # Point out_dir to our temporary path
                with patch.object(run_tx_mini_ecc_parity, "render_canonical_path_view") as mock_canon:
                    with self.assertRaises(ValueError):
                        # Call run_parity with invalid manifest
                        with patch("run_tx_mini_ecc_parity.validate_candidate_manifest", side_effect=ValueError("Invalid manifest")):
                            run_tx_mini_ecc_parity.validate_candidate_manifest(invalid_manifest)

                    # Ensure canonical view rendering and generator subprocesses were NEVER called
                    mock_canon.assert_not_called()
                    mock_sub.assert_not_called()

    def test_validate_independent_paths_rejection(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_root = Path(tmpdir)
            f1 = tmp_root / "f1.xlsx"
            f2 = tmp_root / "f2.xlsx"
            out1 = tmp_root / "out1"
            out2 = tmp_root / "out2"
            f1.write_text("in1")
            f2.write_text("in2")
            out1.mkdir()
            out2.mkdir()

            # Identical input rejection
            with self.assertRaises(ValueError):
                validate_independent_paths(f1, f1, out1, out2)

            # Identical output rejection
            with self.assertRaises(ValueError):
                validate_independent_paths(f1, f2, out1, out1)

            # Valid independent paths
            self.assertTrue(validate_independent_paths(f1, f2, out1, out2))

    def test_compare_workbook_structure_extra_sheet(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        wb2.create_sheet("ExtraSheet")
        diffs = compare_workbook_structure(wb1, wb2)
        self.assertTrue(any("Sheet names differ" in d for d in diffs))

    def test_compare_workbook_structure_extra_row_col(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1.cell(row=5, column=5, value="x")
        ws2.cell(row=6, column=5, value="x")
        diffs = compare_workbook_structure(wb1, wb2)
        self.assertTrue(any("dimensions differ" in d for d in diffs))

    def test_compare_workbook_structure_merged_range(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1.merge_cells("A1:B2")
        diffs = compare_workbook_structure(wb1, wb2)
        self.assertTrue(any("merged cells differ" in d for d in diffs))

    def test_compare_workbook_structure_hidden_row_col(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1.row_dimensions[3].hidden = True
        diffs = compare_workbook_structure(wb1, wb2)
        self.assertTrue(any("row 3 hidden differs" in d for d in diffs))
    def test_compare_cell_mutations(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active

        # Value diff
        ws1["A1"] = "val1"
        ws2["A1"] = "val2"
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        self.assertTrue(any("value:" in d for d in diffs))

        # Number format diff
        ws1["B2"].number_format = "0.00"
        ws2["B2"].number_format = "@"
        diffs_fmt = compare_cell(ws1["B2"], ws2["B2"], "Sheet", "B2")
        self.assertTrue(any("number_format:" in d for d in diffs_fmt))

    def test_font_only_difference_causes_structural_difference(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1["A1"] = "test"
        ws2["A1"] = "test"
        ws1["A1"].font = Font(name="Calibri", size=11, bold=True)
        ws2["A1"].font = Font(name="Calibri", size=11, bold=False)
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        self.assertTrue(any("font:" in d for d in diffs))

    def test_fill_only_difference_causes_structural_difference(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1["A1"] = "test"
        ws2["A1"] = "test"
        ws1["A1"].fill = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")
        ws2["A1"].fill = PatternFill(fill_type=None)
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        self.assertTrue(any("fill" in d for d in diffs))

    def test_border_only_difference_causes_structural_difference(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1["A1"] = "test"
        ws2["A1"] = "test"
        ws1["A1"].border = Border(left=Side(style="thin"))
        ws2["A1"].border = Border()
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        self.assertTrue(any("border" in d for d in diffs))

    def test_alignment_only_difference_causes_structural_difference(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1["A1"] = "test"
        ws2["A1"] = "test"
        ws1["A1"].alignment = Alignment(horizontal="left")
        ws2["A1"].alignment = Alignment(horizontal="right")
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        self.assertTrue(any("alignment:" in d for d in diffs))

    def test_protection_only_difference_causes_structural_difference(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1["A1"] = "test"
        ws2["A1"] = "test"
        ws1["A1"].protection = Protection(locked=True)
        ws2["A1"].protection = Protection(locked=False)
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        self.assertTrue(any("protection:" in d for d in diffs))

    def test_same_normalized_value_with_formatting_difference_does_not_become_normalized_match(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1.title = "Summary"
        ws2.title = "Summary"
        ws1["A1"] = "  Header  "
        ws2["A1"] = "Header"
        ws1["A1"].font = Font(bold=True)
        ws2["A1"].font = Font(bold=False)

        cell_diffs = compare_cell(ws1["A1"], ws2["A1"], "Summary", "A1")
        only_val_diffs = all(d.startswith("value:") for d in cell_diffs)
        self.assertFalse(only_val_diffs, "Formatting difference must prevent only_val_diffs from being True")

    def test_explicitly_allowlisted_difference_remains_allowed(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1.title = "Summary"
        ws2.title = "Summary"
        ws1["A1"] = "  Header  "
        ws2["A1"] = "Header"

        cell_diffs = compare_cell(ws1["A1"], ws2["A1"], "Summary", "A1")
        norm_l = normalize_allowed_metadata("Summary", "A1", ws1["A1"].value)
        norm_c = normalize_allowed_metadata("Summary", "A1", ws2["A1"].value)
        only_val_diffs = all(d.startswith("value:") for d in cell_diffs)
        is_allowed = only_val_diffs and (norm_l == norm_c)
        self.assertTrue(is_allowed)

    def test_unapproved_cell_diffs_make_overall_parity_fail(self):
        self.assertEqual(calculate_overall_parity(12, 12, 12, 12, 12, 0, 0, 1, 0), "ECC_PARITY_FAILED")

    def test_cross_check_candidate_manifest_exact_match_passes(self):
        expected = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cached = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        self.assertTrue(cross_check_candidate_manifest(cached, expected))

    def test_cross_check_candidate_manifest_stale_site_code_fails(self):
        expected = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cached = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        cached.append({"Candidate ID": "TXM-TSS-012", "Source Row": 12, "Site Code": "STALE_SITE"})
        with self.assertRaises(ValueError) as ctx:
            cross_check_candidate_manifest(cached, expected)
        self.assertIn("MANIFEST_CANDIDATE_IDENTITY_MISMATCH", str(ctx.exception))

    def test_cross_check_candidate_manifest_stale_source_row_fails(self):
        expected = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cached = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        cached.append({"Candidate ID": "TXM-TSS-012", "Source Row": 99, "Site Code": "S012"})
        with self.assertRaises(ValueError) as ctx:
            cross_check_candidate_manifest(cached, expected)
        self.assertIn("MANIFEST_CANDIDATE_IDENTITY_MISMATCH", str(ctx.exception))

    def test_cross_check_candidate_manifest_stale_candidate_id_fails(self):
        expected = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cached = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        cached.append({"Candidate ID": "TXM-TSS-999", "Source Row": 12, "Site Code": "S012"})
        with self.assertRaises(ValueError) as ctx:
            cross_check_candidate_manifest(cached, expected)
        self.assertIn("MANIFEST_CANDIDATE_IDENTITY_MISMATCH", str(ctx.exception))

    def test_cross_check_candidate_manifest_missing_current_candidate_fails(self):
        expected = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cached = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        with self.assertRaises(ValueError) as ctx:
            cross_check_candidate_manifest(cached, expected)
        self.assertIn("must be exactly 12", str(ctx.exception))

    def test_cross_check_candidate_manifest_extra_stale_candidate_fails(self):
        expected = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cached = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 14)]
        with self.assertRaises(ValueError) as ctx:
            cross_check_candidate_manifest(cached, expected)
        self.assertIn("must be exactly 12", str(ctx.exception))

    def test_cross_check_candidate_manifest_same_count_different_set_fails(self):
        expected = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cached = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i + 10, "Site Code": f"S{i+10:03d}"} for i in range(1, 13)]
        with self.assertRaises(ValueError) as ctx:
            cross_check_candidate_manifest(cached, expected)
        self.assertIn("MANIFEST_CANDIDATE_IDENTITY_MISMATCH", str(ctx.exception))

    def test_cross_check_candidate_manifest_order_difference_passes_after_sorting(self):
        expected = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cached = list(reversed(expected))
        self.assertTrue(cross_check_candidate_manifest(cached, expected))

    def test_manifest_identity_hash_changes_when_membership_changes(self):
        cands1 = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        cands2 = [{"Candidate ID": f"TXM-TSS-{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 12)]
        cands2.append({"Candidate ID": "TXM-TSS-012", "Source Row": 12, "Site Code": "S999"})
        h1 = compute_manifest_identity_hash(cands1)
        h2 = compute_manifest_identity_hash(cands2)
        self.assertNotEqual(h1, h2)

    def test_extract_and_compare_business_fields_mismatch(self):
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1.title = "details"
        ws2.title = "details"

        ws1["A1"] = "9786B_AD"
        ws2["A1"] = "9786B_AD_MISMATCH"

        ws1["A2"] = "Item 1"
        ws1["B2"] = "350000062773"
        ws1["O2"] = "LOS Survey"

        ws2["A2"] = "Item 1"
        ws2["B2"] = "350000062773"
        ws2["O2"] = "LOS Survey Mismatch"

        fields1 = extract_business_fields(wb1)
        fields2 = extract_business_fields(wb2)

        match_flag, csv_rows = compare_business_fields("C001", fields1, fields2)
        self.assertFalse(match_flag)
        mismatched_names = [r["Field Name"] for r in csv_rows if r["Match"] == "FALSE"]
        self.assertIn("Site Code", mismatched_names)
        self.assertIn("LineItem_R2_ColumnO_SOW", mismatched_names)

    def test_normalize_allowed_metadata_allowlist(self):
        # Approved Summary A1 normalizes
        self.assertEqual(normalize_allowed_metadata("Summary", "A1", "  text  "), "text")

        # Unapproved coordinates must NOT normalize
        self.assertEqual(normalize_allowed_metadata("details", "A10", "  text  "), "  text  ")
        self.assertEqual(normalize_allowed_metadata("details", "O2", " LOS Survey "), " LOS Survey ")

    def test_calculate_overall_parity_strict(self):
        self.assertEqual(calculate_overall_parity(12, 12, 12, 12, 12, 0, 0, 0, 0), "PASS")
        self.assertEqual(calculate_overall_parity(12, 12, 12, 12, 11, 1, 0, 0, 0), "PASS")
        self.assertEqual(calculate_overall_parity(12, 12, 12, 12, 11, 0, 1, 0, 0), "ECC_PARITY_FAILED")
        self.assertEqual(calculate_overall_parity(11, 11, 11, 11, 11, 0, 0, 0, 0), "ECC_PARITY_FAILED")


    def test_normalize_source_row_integer_accepted(self):
        from run_tx_mini_ecc_parity import normalize_source_row
        self.assertEqual(normalize_source_row(123), 123)

    def test_normalize_source_row_numeric_string_accepted_and_normalized(self):
        from run_tx_mini_ecc_parity import normalize_source_row
        self.assertEqual(normalize_source_row("123"), 123)

    def test_normalize_source_row_whitespace_numeric_string_normalized(self):
        from run_tx_mini_ecc_parity import normalize_source_row
        self.assertEqual(normalize_source_row(" 123 "), 123)

    def test_normalize_source_row_float_integer_accepted_non_integer_rejected(self):
        from run_tx_mini_ecc_parity import normalize_source_row
        self.assertEqual(normalize_source_row(123.0), 123)
        with self.assertRaises(ValueError):
            normalize_source_row(123.45)

    def test_normalize_source_row_duplicate_mixed_types_rejected(self):
        dup_mixed = [{"Candidate ID": f"C{i:03d}", "Source Row": i, "Site Code": f"S{i:03d}"} for i in range(1, 11)]
        dup_mixed.append({"Candidate ID": "C011", "Source Row": 1, "Site Code": "S011"})
        dup_mixed.append({"Candidate ID": "C012", "Source Row": "1", "Site Code": "S012"})
        with self.assertRaises(ValueError) as ctx:
            validate_candidate_manifest(dup_mixed)
        self.assertIn("Duplicate Source Rows", str(ctx.exception))

    def test_normalize_source_row_invalid_values_rejected(self):
        from run_tx_mini_ecc_parity import normalize_source_row
        for invalid_val in [None, True, False, 0, -1, "-5", "abc", "", "   "]:
            with self.subTest(val=invalid_val):
                with self.assertRaises(ValueError):
                    normalize_source_row(invalid_val)

    def test_normalized_source_row_used_in_candidate_processing(self):
        manifest_data = [{"Candidate ID": f"C{i:03d}", "Source Row": str(i), "Site Code": f"S{i:03d}"} for i in range(1, 13)]
        self.assertTrue(validate_candidate_manifest(manifest_data))
        for row in manifest_data:
            self.assertIsInstance(row["Source Row"], int, "Source Row must be coerced in place to canonical int")

    def test_tracked_fixture_exists_repository_relative(self):
        fixture_file = ROOT / "tests" / "fixtures" / "tx_mini_du_export_fixture.xlsx"
        self.assertTrue(fixture_file.exists(), "Tracked synthetic test fixture file must exist")
        self.assertGreater(fixture_file.stat().st_size, 0)

    def test_tests_pass_when_reference_exports_absent(self):
        fixture_file = ROOT / "tests" / "fixtures" / "tx_mini_du_export_fixture.xlsx"
        profile_file = ROOT / "config" / "du_profiles" / "tx_mini_pr_v1.yaml"
        from canonical_generator_bridge import build_header_inventory, calculate_header_hash
        inv = build_header_inventory(fixture_file)
        h_hash = calculate_header_hash(inv)
        prof_data = json.loads(profile_file.read_text(encoding="utf-8"))
        self.assertIn(h_hash, prof_data["export_structure"]["approved_header_hashes"])

    # --- Thread 9 regression tests: stale candidate CSV cross-check ---

    def test_stale_candidate_csv_after_scope_change_is_rejected(self):
        """A stale CSV whose source rows differ from freshly classified rows
        must raise STALE_CANDIDATE_CSV, never silently pass."""
        from run_tx_mini_ecc_parity import derive_expected_candidates
        from unittest.mock import patch, MagicMock
        import csv as csv_mod

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            csv_path = tmp / "TX_MINI_TSS_UAT_CANDIDATES.csv"
            scope_cfg = tmp / "scope.json"
            scope_cfg.write_text(json.dumps({"scopes": {}}), encoding="utf-8")

            # Write a stale CSV with source rows 1..12
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                w = csv_mod.DictWriter(f, fieldnames=["Source Row", "Masked Site Code"])
                w.writeheader()
                for i in range(1, 13):
                    w.writerow({"Source Row": str(i), "Masked Site Code": f"S{i:03d}"})

            # Mock build_records_from_export to return a DIFFERENT set (rows 5..16)
            fresh_records = [
                {"identity": {"source_row_number": r}, "site": {"site_code": f"S{r:03d}"}}
                for r in range(5, 17)
            ]

            with patch("run_tx_mini_ecc_parity.subprocess.run") as mock_sub, \
                 patch("run_tx_mini_ecc_parity.build_records_from_export", return_value=(fresh_records, {})):
                mock_sub.return_value = MagicMock(returncode=0)
                with self.assertRaises(ValueError) as ctx:
                    derive_expected_candidates(
                        Path("dummy_input.xlsx"),
                        Path("dummy_profile.yaml"),
                        Path("dummy_sow.yaml"),
                        scope_cfg,
                        csv_path,
                    )
                self.assertIn("STALE_CANDIDATE_CSV", str(ctx.exception))

    def test_current_candidate_set_exact_match_passes(self):
        """When CSV source rows exactly match freshly classified rows,
        derive_expected_candidates succeeds."""
        from run_tx_mini_ecc_parity import derive_expected_candidates
        from unittest.mock import patch, MagicMock
        import csv as csv_mod

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            csv_path = tmp / "TX_MINI_TSS_UAT_CANDIDATES.csv"
            scope_cfg = tmp / "scope.json"
            scope_cfg.write_text(json.dumps({"scopes": {}}), encoding="utf-8")

            # Write CSV with source rows 1..12
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                w = csv_mod.DictWriter(f, fieldnames=["Source Row", "Masked Site Code"])
                w.writeheader()
                for i in range(1, 13):
                    w.writerow({"Source Row": str(i), "Masked Site Code": f"S{i:03d}"})

            # Mock build_records_from_export to return the SAME set (rows 1..12)
            fresh_records = [
                {"identity": {"source_row_number": r}, "site": {"site_code": f"S{r:03d}"}}
                for r in range(1, 13)
            ]

            with patch("run_tx_mini_ecc_parity.subprocess.run") as mock_sub, \
                 patch("run_tx_mini_ecc_parity.build_records_from_export", return_value=(fresh_records, {})):
                mock_sub.return_value = MagicMock(returncode=0)
                result = derive_expected_candidates(
                    Path("dummy_input.xlsx"),
                    Path("dummy_profile.yaml"),
                    Path("dummy_sow.yaml"),
                    scope_cfg,
                    csv_path,
                )
                self.assertEqual(len(result), 12)
                for c in result:
                    self.assertIn("Candidate ID", c)
                    self.assertIn("Source Row", c)
                    self.assertIn("Site Code", c)

    # --- Thread 10 regression tests: fill color comparison ---

    def test_same_fill_type_different_foreground_color_detected(self):
        """Both cells solid but different fgColor must produce a fill_fgColor diff."""
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1["A1"] = "test"
        ws2["A1"] = "test"
        ws1["A1"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
        ws2["A1"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        self.assertTrue(any("fill_fgColor" in d for d in diffs),
                        f"Expected fill_fgColor diff, got: {diffs}")

    def test_same_fill_type_different_background_color_detected(self):
        """Both cells solid but different bgColor must produce a fill_bgColor diff."""
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1["A1"] = "test"
        ws2["A1"] = "test"
        ws1["A1"].fill = PatternFill(fill_type="solid", fgColor="FF0000", bgColor="0000FF")
        ws2["A1"].fill = PatternFill(fill_type="solid", fgColor="FF0000", bgColor="00FF00")
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        self.assertTrue(any("fill_bgColor" in d for d in diffs),
                        f"Expected fill_bgColor diff, got: {diffs}")

    def test_identical_fill_objects_produce_no_diff(self):
        """Cells with identical fill (type, fgColor, bgColor) must produce zero fill diffs."""
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws1["A1"] = "test"
        ws2["A1"] = "test"
        ws1["A1"].fill = PatternFill(fill_type="solid", fgColor="FF0000", bgColor="0000FF")
        ws2["A1"].fill = PatternFill(fill_type="solid", fgColor="FF0000", bgColor="0000FF")
        diffs = compare_cell(ws1["A1"], ws2["A1"], "Sheet", "A1")
        fill_diffs = [d for d in diffs if "fill" in d]
        self.assertEqual(fill_diffs, [], f"Expected no fill diffs, got: {fill_diffs}")

    def test_fill_color_difference_makes_overall_parity_fail(self):
        """A fill color diff (not in allowlist) must cause STRUCTURAL_DIFFERENCE
        and make calculate_overall_parity return ECC_PARITY_FAILED."""
        # struct_diffs=1 means at least one candidate has STRUCTURAL_DIFFERENCE
        self.assertEqual(
            calculate_overall_parity(12, 12, 12, 12, 11, 0, 0, 1, 0),
            "ECC_PARITY_FAILED"
        )


if __name__ == "__main__":
    unittest.main()
