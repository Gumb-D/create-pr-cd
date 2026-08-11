#!/usr/bin/env python3
"""Raw four-header all-DU Planning PR end-to-end regression for Issue #34."""
from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
CREATE_PR = ROOT / "scripts" / "create_pr.py"
PROFILE_ROOT = ROOT / "config" / "du_profiles"

PROFILE_EXPECTED_PBOM = {
    "tx_rollout_2023_pr_v1": "350001143904",
    "tx_mini_pr_v1": "350001143905",
    "celcomdigi_bau_2023_pr_v1": "350001143904",
    "celcomdigi_bau_2024_pr_v1": "350001143904",
    "celcomdigi_usp_pr_v1": "350001143904",
    "jendela_tx_migration_pr_v1": "350001143904",
    "mw_eos_swap_pr_v1": "350001143905",
    "zte_tx_mini_pr_v1": "350001143905",
}


def _load_profile(profile_id: str) -> dict[str, Any]:
    return json.loads((PROFILE_ROOT / f"{profile_id}.yaml").read_text(encoding="utf-8"))


def _selected_fingerprints(profile: dict[str, Any]) -> dict[str, dict[str, str]]:
    required_for_resolver = {
        name
        for name, config in profile["field_mapping"].items()
        if config.get("required")
    }
    fields = required_for_resolver | {
        "subcontractor_planning",
        "existing_planning_pr_status",
    }
    selected: dict[str, dict[str, str]] = {}
    for field in fields:
        candidates = profile["field_mapping"].get(field, {}).get("source_candidates", [])
        approved = [candidate for candidate in candidates if candidate.get("mapping_status") == "APPROVED"]
        if not approved:
            raise AssertionError(f"{profile['profile_id']} missing approved source for {field}")
        selected[field] = approved[0]["fingerprint"]
    return selected


def _field_value(field: str, site_code: str, planning_subcon: str) -> Any:
    values: dict[str, Any] = {
        "site_code": site_code,
        "site_name": f"Planning {site_code}",
        "du_key": f"DU-{site_code}",
        "region": "Central",
        "state": "Selangor",
        "tx_sow_raw": "",
        "tx_before_migration": "",
        "subcontractor_tss": "GCI",
        "subcontractor_ti": "GCI",
        "subcontractor_planning": planning_subcon,
        "existing_tss_pr_status": "",
        "existing_ti_pr_status": "",
        "existing_planning_pr_status": "",
    }
    return values.get(field, "")


def _write_raw_export(path: Path, profile: dict[str, Any], rows: list[tuple[str, str]]) -> None:
    fingerprints = _selected_fingerprints(profile)
    ordered_fields = sorted(fingerprints)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "data"
    for column, field in enumerate(ordered_fields, start=1):
        fingerprint = fingerprints[field]
        worksheet.cell(1, column, fingerprint["field_code"])
        worksheet.cell(2, column, fingerprint["wbs_stage"])
        worksheet.cell(3, column, fingerprint["task_name"])
        worksheet.cell(4, column, fingerprint["display_header"])
    for row_number, (site_code, planning_subcon) in enumerate(rows, start=5):
        for column, field in enumerate(ordered_fields, start=1):
            worksheet.cell(row_number, column, _field_value(field, site_code, planning_subcon))
    workbook.save(path)
    workbook.close()


def _ecc_rows(output: Path) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    for path in sorted(output.glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook["details"]
            rows.extend(worksheet.iter_rows(min_row=2, values_only=True))
        finally:
            workbook.close()
    return rows


class PlanningAllDuEndToEndTest(unittest.TestCase):
    def test_all_eight_profiles_render_standard_and_aa_planning(self) -> None:
        for index, (profile_id, expected_standard_pbom) in enumerate(PROFILE_EXPECTED_PBOM.items(), start=1):
            with self.subTest(profile_id=profile_id):
                profile = _load_profile(profile_id)
                with tempfile.TemporaryDirectory() as temp_dir:
                    root = Path(temp_dir)
                    source = root / f"planning-{profile_id}.xlsx"
                    output = root / "output"
                    standard_site = f"P{index:03d}A"
                    aa_site = f"P{index:03d}B"
                    _write_raw_export(
                        source,
                        profile,
                        [(standard_site, "GCI"), (aa_site, "GCI_AA")],
                    )
                    result = subprocess.run(
                        [
                            sys.executable,
                            str(CREATE_PR),
                            "--site-data",
                            str(source),
                            "--output",
                            str(output),
                            "--scope",
                            "Planning",
                            "--all-sites",
                        ],
                        cwd=ROOT,
                        text=True,
                        capture_output=True,
                        check=False,
                    )
                    self.assertEqual(
                        result.returncode,
                        0,
                        msg=f"{profile_id}\nstdout:\n{result.stdout}\nstderr:\n{result.stderr}",
                    )
                    summary_path = output / "CREATE_PR_SUMMARY_PLANNING.json"
                    self.assertTrue(summary_path.exists(), msg=f"missing summary for {profile_id}")
                    summary = json.loads(summary_path.read_text(encoding="utf-8"))
                    self.assertEqual(summary["status"], "SUCCESS")
                    self.assertEqual(summary["profile_id"], profile_id)
                    self.assertEqual(summary["requested_count"], 2)
                    self.assertEqual(summary["generated_count"], 2)
                    self.assertEqual(summary["failed_count"], 0)
                    self.assertEqual(summary["unaccounted_count"], 0)

                    rendered = _ecc_rows(output)
                    self.assertEqual(len(rendered), 2)
                    by_site = {str(row[3]): row for row in rendered}
                    self.assertEqual(str(by_site[standard_site][9]), expected_standard_pbom)
                    self.assertEqual(str(by_site[aa_site][9]), "350001042321")
                    self.assertEqual(by_site[standard_site][8], "GCI")
                    self.assertEqual(by_site[aa_site][8], "GCI")
                    self.assertEqual(by_site[standard_site][7], "S1MY2024071002WBF1")
                    self.assertEqual(by_site[aa_site][7], "S1MY2024071002WBF1")


if __name__ == "__main__":
    unittest.main()
