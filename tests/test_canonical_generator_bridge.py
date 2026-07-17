import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from canonical_generator_bridge import (
    canonical_record_to_generator_row,
    classify_uat_record,
    write_uat_packet,
)
from canonical_site_validator import empty_canonical_site_record


class TestCanonicalGeneratorBridge(unittest.TestCase):
    def make_record(self, *, scope="TSS", status="NO_PR", classification="PR_INPUT_READY", date_val="2026-07-17"):
        record = empty_canonical_site_record()
        record["identity"].update(
            {
                "source_file_name": "source.xlsx",
                "source_file_hash": "source-hash",
                "header_hash": "header-hash",
                "source_row_number": 5,
            }
        )
        record["site"].update({"site_code": "A0001", "site_name": "Synthetic", "du_key": "DU0001"})
        record["pr_context"].update(
            {
                "tx_sow_raw": "MW Swap",
                "tx_sow_normalized": "MW SWAP",
                "region": "Northern",
                "state": "Penang",
                "subcontractor_tss": "GTSB",
                "subcontractor_ti": "GTSB",
                "existing_tss_pr_status": status if scope == "TSS" else "NO_PR",
                "existing_ti_pr_status": status if scope == "TI" else "NO_PR",
            }
        )
        record["technical_context"].update(
            {
                "latitude": 5.1,
                "longitude": 100.1,
                "antenna_size_ne": "0.6m",
                "antenna_size_fe": "0.6m",
            }
        )
        record["source_evidence"]["fields"]["tx_sow_normalized"] = {
            "source_header_fingerprint": {
                "field_code": "TX_SOW",
                "wbs_stage": "Network Planning",
                "task_name": "Microwave",
                "display_header": "Microwave Tx SOW",
            },
            "normalization_status": "APPROVED",
            "sow_classification": "PR_TRIGGER",
        }
        date_field = "tss_actual_end_date" if scope == "TSS" else "ti_actual_end_date"
        record["source_evidence"]["fields"][date_field] = {
            "source_header_fingerprint": {
                "field_code": "ACTUAL_END",
                "wbs_stage": "Stage",
                "task_name": "Task",
                "display_header": "actual end time",
            },
            "source_value": date_val,
            "mapping_status": "APPROVED",
        }
        record["validation"].update(
            {
                "profile_id": "zte_tx_mini_pr_v1",
                "profile_version": "0.2.0",
                "mapping_version": "approved-v1",
                "pr_input_classification": classification,
                "blocking_reasons": [] if classification == "PR_INPUT_READY" else ["MISSING_PR_CRITICAL_FIELD:tx_sow_raw"],
            }
        )
        return record

    def test_candidate_maps_to_exact_legacy_generator_columns_and_locks_ecc(self):
        row = canonical_record_to_generator_row(self.make_record(), "TSS")
        self.assertEqual(row["customer site code"], "A0001")
        self.assertEqual(row["customer site name"], "Synthetic")
        self.assertEqual(row["du code"], "DU0001")
        self.assertEqual(row["Tx SOW"], "MW SWAP")
        self.assertEqual(row["SubCon - TSS Team"], "GTSB")
        self.assertEqual(row["Subcon PR - TSS"], "")
        self.assertEqual(row["MW Config Antenna Size NE"], "0.6m")
        self.assertEqual(row["UAT Classification"], "UAT_CANDIDATE")
        self.assertFalse(row["ECC Allowed"])

    def test_precedence_2_unresolved_evidence_over_3_duplicate(self):
        duplicate_unresolved = self.make_record(status="PR_EXISTS", classification="PR_INPUT_INCOMPLETE")
        self.assertEqual(classify_uat_record(duplicate_unresolved, "TSS")[0], "REVIEW_REQUIRED")

    def test_precedence_3_duplicate_over_4_blank_actual_end(self):
        duplicate_blank = self.make_record(status="PR_EXISTS", date_val="")
        self.assertEqual(classify_uat_record(duplicate_blank, "TSS")[0], "DUPLICATE_BLOCKED")

    def test_precedence_4_blank_actual_end_over_5_malformed_actual_end(self):
        # Blank is NO_PR_OR_IGNORED, malformed is REVIEW_REQUIRED. 
        # A blank date is caught before malformed checks.
        blank_date = self.make_record(date_val=None)
        self.assertEqual(classify_uat_record(blank_date, "TSS")[0], "NO_PR_OR_IGNORED")

    def test_precedence_5_malformed_actual_end_over_6_sow_no_pr_trigger(self):
        malformed_and_no_pr_trigger = self.make_record(date_val="invalid-date")
        malformed_and_no_pr_trigger["source_evidence"]["fields"]["tx_sow_normalized"]["normalization_status"] = "APPROVED_NO_OUTPUT"
        self.assertEqual(classify_uat_record(malformed_and_no_pr_trigger, "TSS")[0], "REVIEW_REQUIRED")

    def test_precedence_6_sow_no_pr_trigger_over_7_sow_review_state(self):
        # If it's APPROVED_NO_OUTPUT, it's NO_PR_OR_IGNORED, instead of REVIEW_REQUIRED.
        no_pr_trigger = self.make_record()
        no_pr_trigger["source_evidence"]["fields"]["tx_sow_normalized"]["normalization_status"] = "APPROVED_NO_OUTPUT"
        self.assertEqual(classify_uat_record(no_pr_trigger, "TSS")[0], "NO_PR_OR_IGNORED")

    def test_precedence_7_sow_review_state_over_8_uat_candidate(self):
        unapproved_sow = self.make_record()
        unapproved_sow["source_evidence"]["fields"]["tx_sow_normalized"]["normalization_status"] = "UNVERIFIED"
        self.assertEqual(classify_uat_record(unapproved_sow, "TSS")[0], "REVIEW_REQUIRED")

    def test_tss_duplicate_cannot_block_ti(self):
        record = self.make_record(scope="TSS", status="PR_EXISTS")
        self.assertEqual(classify_uat_record(record, "TSS")[0], "DUPLICATE_BLOCKED")
        self.assertEqual(classify_uat_record(record, "TI")[0], "UAT_CANDIDATE")

    def test_ti_duplicate_cannot_block_tss(self):
        record = self.make_record(scope="TI", status="PR_EXISTS")
        self.assertEqual(classify_uat_record(record, "TI")[0], "DUPLICATE_BLOCKED")
        self.assertEqual(classify_uat_record(record, "TSS")[0], "UAT_CANDIDATE")

    def test_tss_date_cannot_activate_ti(self):
        record = self.make_record(scope="TSS", date_val="2026-07-17")
        # Ensure TI date is blank
        record["source_evidence"]["fields"]["ti_actual_end_date"] = {"source_value": ""}
        self.assertEqual(classify_uat_record(record, "TI")[0], "NO_PR_OR_IGNORED")

    def test_ti_date_cannot_activate_tss(self):
        record = self.make_record(scope="TI", date_val="2026-07-17")
        # Ensure TSS date is blank
        record["source_evidence"]["fields"]["tss_actual_end_date"] = {"source_value": ""}
        self.assertEqual(classify_uat_record(record, "TSS")[0], "NO_PR_OR_IGNORED")

    def test_writer_creates_generator_compatible_data_sheet_and_partitions(self):
        records = [
            self.make_record(),
            self.make_record(status="PR_EXISTS"),
            self.make_record(date_val=None),
            self.make_record(classification="PR_INPUT_INCOMPLETE"),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            outputs = write_uat_packet(
                records,
                {"profile_id": "zte_tx_mini_pr_v1", "header_hash": "header-hash"},
                Path(temp_dir),
                "TSS",
            )
            workbook = load_workbook(outputs["workbook"], read_only=True, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "data",
                    "summary",
                    "uat_candidates",
                    "duplicate_blocked",
                    "no_pr_or_ignored",
                    "review_required",
                    "traceability",
                ],
            )
            data = workbook["data"]
            headers = [cell.value for cell in next(data.iter_rows(min_row=4, max_row=4))]
            self.assertEqual(headers[0], "customer site code")
            self.assertIn("SubCon - TSS Team", headers)
            self.assertIn("Subcon PR - TI", headers)
            self.assertEqual(data.max_row, 5)
            workbook.close()

            summary = json.loads(outputs["summary_json"].read_text(encoding="utf-8"))
            self.assertEqual(summary["counts"]["UAT_CANDIDATE"], 1)
            self.assertEqual(summary["counts"]["DUPLICATE_BLOCKED"], 1)
            self.assertEqual(summary["counts"]["NO_PR_OR_IGNORED"], 1)
            self.assertEqual(summary["counts"]["REVIEW_REQUIRED"], 1)
            self.assertEqual(summary["generator_data_row_count"], 1)
            self.assertFalse(summary["ecc_allowed"])


if __name__ == "__main__":
    unittest.main()
