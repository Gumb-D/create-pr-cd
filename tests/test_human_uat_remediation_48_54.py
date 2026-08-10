import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import create_pr
import pr_helpers
from du_export_adapter import build_canonical_site_record, resolve_profile_field_mappings
from du_profile_loader import load_du_profile
from geography_resolver import GeographyResolver
from profile_du_export import fingerprint_key


JENDELA_PROFILE_PATH = ROOT / "config" / "du_profiles" / "jendela_tx_migration_pr_v1.yaml"
TX_MINI_PROFILE_PATH = ROOT / "config" / "du_profiles" / "tx_mini_pr_v1.yaml"
SOW_REGISTRY_PATH = ROOT / "config" / "registries" / "canonical_sow_registry.yaml"


class CanonicalRecordFactory:
    def __init__(self, profile_path):
        self.profile = load_du_profile(profile_path)
        self.inventory = {
            "sheets": [
                {
                    "sheet_name": "DU Export",
                    "columns": [
                        {
                            "fingerprint": candidate["fingerprint"],
                            "fingerprint_key": fingerprint_key(candidate["fingerprint"]),
                        }
                        for config in self.profile["field_mapping"].values()
                        for candidate in config.get("source_candidates", [])
                    ],
                }
            ]
        }
        self.resolved = resolve_profile_field_mappings(self.inventory, self.profile)

    def build(self, values, *, scope="TI", site_code="GENERIC-1", sow_registry=None):
        defaults = {
            "site_code": site_code,
            "site_name": "Generic Site",
            "du_key": "DU-1",
            "tx_sow_raw": "-",
            "tx_upgrade_scope_raw": "TSS+TI",
            "region": "Northern",
            "state": "Kedah",
            "subcontractor_tss": "GTSB",
            "subcontractor_ti": "GTSB",
            "existing_tss_pr_status": "",
            "existing_ti_pr_status": "",
            "antenna_size_ne": "0.6m",
            "antenna_size_fe": "0.6m",
            "tx_before_migration": "Starlink",
            "final_backhaul": "ignored",
        }
        defaults.update(values)
        raw = {}
        for field_name, config in self.profile["field_mapping"].items():
            if field_name not in defaults:
                continue
            for candidate in config.get("source_candidates", []):
                raw[fingerprint_key(candidate["fingerprint"])] = defaults[field_name]
        identity = self.profile["identity"]
        return build_canonical_site_record(
            raw,
            self.profile,
            {
                "project_key": identity["project_key"],
                "du_model_name": identity["accepted_du_models"][0],
                "du_model_id": identity["accepted_du_model_ids"][0],
                "view_id": identity["accepted_view_ids"][0],
                "source_file_name": "synthetic-export.xlsx",
                "source_file_hash": "source-hash",
                "header_hash": self.profile["export_structure"]["approved_header_hashes"][0],
                "source_row_number": 7,
            },
            scope=scope,
            resolved_mappings=self.resolved,
            sow_registry=sow_registry,
        )


class TestJendelaMigrationDecision(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.factory = CanonicalRecordFactory(JENDELA_PROFILE_PATH)
        cls.sow_registry = json.loads(SOW_REGISTRY_PATH.read_text(encoding="utf-8"))

    def _decision(self, *, before="Starlink", tx_sow="-", final_backhaul="ignored"):
        record = self.factory.build(
            {
                "tx_before_migration": before,
                "tx_sow_raw": tx_sow,
                "final_backhaul": final_backhaul,
            },
            sow_registry=self.sow_registry,
        )
        return record, record["pr_context"]["migration_decision"]

    def test_profile_keeps_final_backhaul_as_optional_audit_evidence(self):
        before = self.factory.profile["field_mapping"]["tx_before_migration"]
        final = self.factory.profile["field_mapping"]["final_backhaul"]
        self.assertTrue(before["required"])
        self.assertFalse(final["required"])
        self.assertEqual(
            before["source_candidates"][0]["fingerprint"]["display_header"],
            "TX Before Migration",
        )
        self.assertEqual(
            final["source_candidates"][0]["fingerprint"]["display_header"],
            "Final Backhaul",
        )

    def test_tx_before_migration_controls_dismantle_only(self):
        cases = {
            "Starlink": ["Dismantle Starlink"],
            "MW": ["Dismantle MW"],
            "Microwave": ["Dismantle MW"],
        }
        for before, expected in cases.items():
            with self.subTest(before=before):
                _, decision = self._decision(before=before, tx_sow="-")
                self.assertEqual(decision["classification"], "APPROVED")
                self.assertEqual([item["work_item"] for item in decision["work_items"]], expected)

        _, no_work = self._decision(before="Fiber Own Build", tx_sow="-")
        self.assertEqual(no_work["classification"], "APPROVED_NO_OUTPUT")
        self.assertEqual(no_work["reason_code"], "JENDELA_TI_NO_WORK_REQUIRED")
        self.assertEqual(no_work["work_items"], [])

    def test_tx_sow_controls_additional_work_independently(self):
        cases = [
            ("BBU Patching", ["BBU Patching / MW IDU Patching"], "BBU Patching"),
            ("MW IDU Patching", ["BBU Patching / MW IDU Patching"], "MW IDU Patching"),
            ("BBU Patching / MW IDU Patching", ["BBU Patching / MW IDU Patching"], "BBU Patching"),
            ("MW New Link / Reroute", ["MW New Link"], "MW New Link / Reroute"),
            ("MW by others", [], None),
            ("-", [], None),
            ("", [], None),
        ]
        for tx_sow, expected_items, expected_model_sow in cases:
            with self.subTest(tx_sow=tx_sow):
                _, decision = self._decision(before="Fiber Own Build", tx_sow=tx_sow)
                expected_classification = "APPROVED" if expected_items else "APPROVED_NO_OUTPUT"
                self.assertEqual(decision["classification"], expected_classification)
                self.assertEqual([item["work_item"] for item in decision["work_items"]], expected_items)
                if expected_model_sow:
                    self.assertEqual(decision["work_items"][0]["model_sow"], expected_model_sow)

    def test_dismantle_and_additional_work_combine_atomically(self):
        cases = [
            ("Starlink", "BBU Patching", ["Dismantle Starlink", "BBU Patching / MW IDU Patching"]),
            ("MW", "MW IDU Patching", ["Dismantle MW", "BBU Patching / MW IDU Patching"]),
            ("Starlink", "MW New Link / Reroute", ["Dismantle Starlink", "MW New Link"]),
            ("MW", "MW New Link / Reroute", ["Dismantle MW", "MW New Link"]),
        ]
        for before, tx_sow, expected in cases:
            with self.subTest(before=before, tx_sow=tx_sow):
                record, decision = self._decision(before=before, tx_sow=tx_sow)
                self.assertEqual(decision["classification"], "APPROVED")
                self.assertEqual([item["work_item"] for item in decision["work_items"]], expected)
                self.assertEqual(record["validation"]["pr_input_classification"], "PR_INPUT_READY")

    def test_final_backhaul_never_changes_jendela_ti_work_plan(self):
        observed = []
        for final_backhaul in ("", "Fiber Own Build", "Fiber LL", "Fiber OB/LL", "MW", "unexpected"):
            _, decision = self._decision(
                before="Starlink",
                tx_sow="BBU Patching",
                final_backhaul=final_backhaul,
            )
            observed.append(
                (
                    decision["classification"],
                    [item["work_item"] for item in decision["work_items"]],
                )
            )
        self.assertTrue(all(value == observed[0] for value in observed))
        self.assertEqual(observed[0], ("APPROVED", ["Dismantle Starlink", "BBU Patching / MW IDU Patching"]))

    def test_missing_or_unknown_business_inputs_fail_closed_without_partial_work(self):
        cases = [
            ("", "-", "JENDELA_TX_BEFORE_MIGRATION_MISSING"),
            ("Satellite", "-", "JENDELA_TX_BEFORE_MIGRATION_NOT_APPROVED"),
            ("Starlink", "Some New Work", "JENDELA_TX_SOW_NOT_APPROVED"),
        ]
        for before, tx_sow, reason in cases:
            with self.subTest(before=before, tx_sow=tx_sow):
                record, decision = self._decision(before=before, tx_sow=tx_sow)
                self.assertEqual(decision["classification"], "REVIEW_REQUIRED")
                self.assertEqual(decision["reason_code"], reason)
                self.assertEqual(decision["work_items"], [])
                self.assertIn(reason, record["validation"]["blocking_reasons"])

    def test_mw_normalization_is_exact(self):
        for before in ("MW link", "BMW", "MW-1"):
            with self.subTest(before=before):
                _, decision = self._decision(before=before, tx_sow="-")
                self.assertEqual(decision["classification"], "REVIEW_REQUIRED")
                self.assertEqual(decision["reason_code"], "JENDELA_TX_BEFORE_MIGRATION_NOT_APPROVED")

    def test_cancel_drop_hard_stop_precedes_migration_output(self):
        cases = [
            ("Starlink", "APPROVED"),
            ("Fiber Own Build", "APPROVED_NO_OUTPUT"),
        ]
        for before, expected_decision_class in cases:
            with self.subTest(before=before):
                record, decision = self._decision(before=before, tx_sow="Cancel / Drop")
                self.assertEqual(decision["classification"], expected_decision_class)
                partitions = create_pr._partition_records([record], "TI")
                self.assertEqual(partitions["ignored"], [record])
                self.assertEqual(partitions["candidates"], [])
                self.assertEqual(record["pr_generation_decision"]["reason_code"], "APPROVED_NO_OUTPUT")

    def test_empty_jendela_work_plan_is_intentional_no_output(self):
        record, decision = self._decision(before="Fiber Own Build", tx_sow="-")
        self.assertEqual(decision["classification"], "APPROVED_NO_OUTPUT")
        self.assertEqual(record["validation"]["pr_input_classification"], "PR_INPUT_READY")
        partitions = create_pr._partition_records([record], "TI")
        self.assertEqual(partitions["ignored"], [record])
        self.assertEqual(partitions["candidates"], [])
        self.assertEqual(record["pr_generation_decision"]["reason_code"], "JENDELA_TI_NO_WORK_REQUIRED")

    def test_tss_and_non_jendela_do_not_use_issue_77_exception(self):
        tss_record = self.factory.build(
            {"tx_sow_raw": "New Starlink"},
            scope="TSS",
            sow_registry=self.sow_registry,
        )
        self.assertNotEqual(
            tss_record["pr_context"].get("migration_decision", {}).get("classification"),
            "APPROVED",
        )

        tx_mini = CanonicalRecordFactory(TX_MINI_PROFILE_PATH).build(
            {"tx_sow_raw": "New Starlink"},
            scope="TI",
            sow_registry=self.sow_registry,
        )
        self.assertNotIn("migration_decision", tx_mini["pr_context"])

    def test_site_specific_identity_does_not_change_business_plan(self):
        values = {
            "tx_before_migration": "Starlink",
            "tx_sow_raw": "BBU Patching",
            "final_backhaul": "ignored",
        }
        first = self.factory.build(values, site_code="8048R", sow_registry=self.sow_registry)
        second = self.factory.build(values, site_code="ANOTHER-SITE", sow_registry=self.sow_registry)
        self.assertEqual(
            first["pr_context"]["migration_decision"],
            second["pr_context"]["migration_decision"],
        )


class TestJendelaRendererAndPbomAtomicity(unittest.TestCase):
    def _candidate_record(self):
        return {
            "identity": {"source_row_number": 7},
            "site": {"site_code": "8048R", "site_name": "Jendela Site", "du_key": "DU-8048R"},
            "pr_context": {
                "region": "Southern",
                "state": "Johor",
                "subcontractor_ti": "GTSB",
                "migration_decision": {
                    "classification": "APPROVED",
                    "decision_code": "JENDELA_TI_WORK_PLAN",
                    "work_items": [
                        {
                            "work_item": "Dismantle Starlink",
                            "model_sow": "Starlink Dismanle",
                            "required_pbom_codes": ["350000597850", "350000597852"],
                        },
                        {
                            "work_item": "BBU Patching / MW IDU Patching",
                            "model_sow": "BBU Patching",
                            "required_pbom_codes": ["350001095420"],
                        },
                    ],
                },
            },
            "technical_context": {"antenna_size_ne": "0.6m", "antenna_size_fe": "0.6m"},
            "approved_contract": {"scope": "TI", "subcontractor": "GTSB"},
            "validation": {"profile_id": "jendela_tx_migration_pr_v1"},
        }

    def test_renderer_consumes_structured_v4_1_decision_without_reclassifying(self):
        rows = create_pr._renderer_rows(self._candidate_record())
        self.assertEqual(
            [row["Migration Work Item"] for row in rows],
            ["Dismantle Starlink", "BBU Patching / MW IDU Patching"],
        )
        self.assertEqual([row["Tx SOW"] for row in rows], ["Starlink Dismanle", "BBU Patching"])
        self.assertEqual(rows[0]["Required PBOM Codes"], "350000597850|350000597852")
        self.assertEqual(rows[1]["Required PBOM Codes"], "350001095420")
        self.assertEqual(rows[0]["Migration Decision ID"], rows[1]["Migration Decision ID"])

    def test_renderer_workbook_keeps_atomic_decision_metadata(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "renderer.xlsx"
            create_pr._write_renderer_input(path, [self._candidate_record()])
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook["data"]
            rows = list(worksheet.iter_rows(min_row=4, values_only=True))
            workbook.close()
        headers = list(rows[0])
        data = [dict(zip(headers, row)) for row in rows[1:]]
        self.assertEqual(len(data), 2)
        self.assertEqual(
            {row["Migration Work Item"] for row in data},
            {"Dismantle Starlink", "BBU Patching / MW IDU Patching"},
        )

    def test_starlink_fixed_pboms_are_required_once_each(self):
        required = ["350000597850", "350000597852"]
        valid = [{"PBOM_Code": "350000597850"}, {"PBOM_Code": "350000597852"}]
        self.assertEqual(pr_helpers.validate_required_pbom_selection(valid, required), (True, None))
        for invalid in (
            [{"PBOM_Code": "350000597850"}],
            [{"PBOM_Code": "350000597852"}],
            valid + [{"PBOM_Code": "350000597852"}],
        ):
            with self.subTest(invalid=invalid):
                ok, reason = pr_helpers.validate_required_pbom_selection(invalid, required)
                self.assertFalse(ok)
                self.assertEqual(reason, "JENDELA_REQUIRED_PBOM_NOT_UNIQUE")

    def test_patching_fixed_pbom_rejects_unexpected_extra_selection(self):
        required = ["350001095420"]
        self.assertEqual(
            pr_helpers.validate_required_pbom_selection(
                [{"PBOM_Code": "350001095420"}],
                required,
            ),
            (True, None),
        )
        ok, reason = pr_helpers.validate_required_pbom_selection(
            [
                {"PBOM_Code": "350001095420"},
                {"PBOM_Code": "350000062748"},
            ],
            required,
        )
        self.assertFalse(ok)
        self.assertEqual(reason, "JENDELA_REQUIRED_PBOM_NOT_UNIQUE")

    def test_any_failed_component_removes_entire_atomic_decision(self):
        rows = [
            {"Site_ID": "8048R", "PBOM_Code": "350000597850", "Migration_Decision_ID": "decision-7"},
            {"Site_ID": "8048R", "PBOM_Code": "350000597852", "Migration_Decision_ID": "decision-7"},
            {"Site_ID": "8048R", "PBOM_Code": "350001095420", "Migration_Decision_ID": "decision-7"},
            {"Site_ID": "OTHER", "PBOM_Code": "KEEP", "Migration_Decision_ID": ""},
        ]
        self.assertEqual(pr_helpers.filter_failed_migration_decisions(rows, {"decision-7"}), [rows[-1]])

    def test_jendela_new_link_row_is_not_legacy_reroute(self):
        renderer_row = create_pr._renderer_rows(
            {
                **self._candidate_record(),
                "pr_context": {
                    **self._candidate_record()["pr_context"],
                    "migration_decision": {
                        "classification": "APPROVED",
                        "decision_code": "JENDELA_TI_WORK_PLAN",
                        "work_items": [
                            {
                                "work_item": "MW New Link",
                                "model_sow": "MW New Link / Reroute",
                                "required_pbom_codes": [],
                            }
                        ],
                    },
                },
            }
        )[0]
        self.assertFalse(pr_helpers.is_mw_reroute_row(renderer_row))


class TestGlobalGeographyCorrection(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.factory = CanonicalRecordFactory(TX_MINI_PROFILE_PATH)

    def test_northern_pahang_becomes_perak_with_audit_evidence(self):
        record = self.factory.build({"region": " Northern ", "state": " pahang "}, scope="TI")
        self.assertEqual(record["pr_context"]["state"], "Perak")
        self.assertEqual(
            record["source_evidence"]["geography_corrections"],
            [
                {
                    "reason_code": "NORTHERN_PAHANG_CORRECTED_TO_PERAK",
                    "field": "state",
                    "original_value": "pahang",
                    "corrected_value": "Perak",
                }
            ],
        )

    def test_other_state_or_other_region_is_not_corrected(self):
        cases = [("Northern", "Kedah"), ("Eastern", "Pahang")]
        for region, state in cases:
            with self.subTest(region=region, state=state):
                record = self.factory.build({"region": region, "state": state}, scope="TI")
                self.assertEqual(record["pr_context"]["state"], state)
                self.assertEqual(record["source_evidence"].get("geography_corrections", []), [])

    def test_incomplete_or_malformed_geography_is_not_silently_corrected(self):
        for region, state in (("Northern", ""), ("Northern", "Atlantis"), ("", "Pahang")):
            with self.subTest(region=region, state=state):
                record = self.factory.build({"region": region, "state": state}, scope="TI")
                self.assertNotEqual(record["pr_context"].get("state"), "Perak")
                self.assertEqual(record["source_evidence"].get("geography_corrections", []), [])


if __name__ == "__main__":
    unittest.main()
