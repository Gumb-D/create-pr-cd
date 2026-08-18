import sys
import tempfile
import unittest
from argparse import Namespace
from pathlib import Path
from unittest.mock import patch
from openpyxl import Workbook, load_workbook
sys.path.insert(0,str(Path(__file__).resolve().parents[1]/'scripts'))
from backoffice_ecc_renderer import BackofficeRendererError, load_backoffice_model_item, render

LOW='350000592793'
HIGH='350000592794'

def make_model(path, duplicate=False):
    wb=Workbook(); ws=wb.active; ws.title='TX Line Item (After 21-Apr 26)'
    ws.append(['SOW','PBOM Code','Description','Unit','Quantity','Rules'])
    ws.append(['Operation Back Office',LOW,'Microwave Back office commisioning and troubleshooting service(Less Than 800 Hops for each month)','Hop',1,'2 choose 1 (Mandatory)'])
    ws.append(['Operation Back Office',HIGH,'Microwave Back office commisioning and troubleshooting service(More Than 800 Hops for each month)','Hop',1,'2 choose 1 (Mandatory)'])
    if duplicate: ws.append(['Operation Back Office',LOW,'duplicate','Hop',1,'2 choose 1 (Mandatory)'])
    wb.save(path); wb.close()

def make_input(path, pbom=LOW, issue='MAIN', contract='S1MY2024042501WBF1'):
    headers=['customer site code','customer site name','du code','region','Backoffice Event Code','Backoffice Trigger Date','Backoffice Billing Month','Backoffice PBOM Code','Backoffice Unit','Backoffice Quantity','Backoffice Subcontractor','Backoffice Contract Number','Backoffice Issue Type','Backoffice Warnings']
    wb=Workbook(); ws=wb.active; ws.title='data'; ws.append(['CANONICAL CREATE-PR-CD INPUT']); ws.append(['x']); ws.append(['x']); ws.append(headers)
    ws.append(['S1','Site 1','DU1','Central','TX_MINI_INTEGRATION','2026-07-15','2026-07',pbom,'Hop',1,'Allstar',contract,issue,''])
    wb.save(path); wb.close()

def make_mapping(path):
    path.write_text('''## Region to Purchasing Area\n| Region* | Purchasing Area* |\n|---|---|\n| Central | Central |\n''',encoding='utf-8')

class TestBackofficeRenderer(unittest.TestCase):
    def test_model_item_is_read_from_operation_back_office_rows(self):
        with tempfile.TemporaryDirectory() as d:
            p=Path(d)/'m.xlsx'; make_model(p)
            item=load_backoffice_model_item(p,LOW)
            self.assertEqual(LOW,item['pbom_code'])
            self.assertEqual('Hop',item['unit'])
            self.assertIn('Less Than 800 Hops',item['description'])

    def test_duplicate_pbom_in_model_fails_closed(self):
        with tempfile.TemporaryDirectory() as d:
            p=Path(d)/'m.xlsx'; make_model(p,True)
            with self.assertRaises(BackofficeRendererError) as cm: load_backoffice_model_item(p,LOW)
            self.assertEqual('BACKOFFICE_PR_MODEL_PBOM_AMBIGUOUS',cm.exception.code)

    def test_main_ecc_has_governed_values(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; inp=root/'i.xlsx'; mapping=root/'map.md'; out=root/'out'
            make_model(model); make_input(inp); make_mapping(mapping)
            paths=render(Namespace(site_data=inp,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None))
            self.assertEqual(1,len(paths)); self.assertIn('Backoffice MAIN 2026-07',paths[0].name)
            wb=load_workbook(paths[0],read_only=True,data_only=True); ws=wb['details']; row=[ws.cell(2,c).value for c in range(1,17)]; wb.close()
            self.assertEqual('DU1',row[5]); self.assertEqual('S1MY2024042501WBF1',row[7]); self.assertEqual('Allstar',row[8]); self.assertEqual(LOW,str(row[9])); self.assertIn('Less Than 800 Hops',row[10]); self.assertEqual('Hop',row[11]); self.assertEqual(1,row[12])

    def test_filename_contains_pr_marker_for_terminal_reconciliation(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; inp=root/'i.xlsx'; mapping=root/'map.md'; out=root/'out'
            make_model(model); make_input(inp); make_mapping(mapping)
            paths=render(Namespace(site_data=inp,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None))
            self.assertIn(' PR ',paths[0].name)

    def test_supplementary_name_is_auditable(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; inp=root/'i.xlsx'; mapping=root/'map.md'; out=root/'out'
            make_model(model); make_input(inp,issue='SUPPLEMENTARY'); make_mapping(mapping)
            paths=render(Namespace(site_data=inp,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None))
            self.assertIn('Backoffice SUPPLEMENTARY 2026-07',paths[0].name)

    def test_renderer_splits_more_than_30_unique_sites_into_parts(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; inp=root/'i.xlsx'; mapping=root/'map.md'; out=root/'out'
            make_model(model); make_input(inp); make_mapping(mapping)
            wb=load_workbook(inp); ws=wb['data']
            base=[ws.cell(5,c).value for c in range(1,ws.max_column+1)]
            for i in range(2,32):
                row=list(base); row[0]=f'S{i}'; row[1]=f'Site {i}'; row[2]=f'DU{i}'
                ws.append(row)
            wb.save(inp); wb.close()
            paths=render(Namespace(site_data=inp,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None))
            self.assertEqual(2,len(paths))
            self.assertTrue(any(' Part 1.xlsx' in p.name for p in paths))
            self.assertTrue(any(' Part 2.xlsx' in p.name for p in paths))
            seen=set(); total_rows=0
            for path in paths:
                wb=load_workbook(path,read_only=True,data_only=True); ws=wb['details']
                ids={str(ws.cell(r,4).value) for r in range(2,ws.max_row+1)}
                total_rows += ws.max_row-1; seen.update(ids); wb.close()
                self.assertLessEqual(len(ids),30)
            self.assertEqual(31,total_rows)
            self.assertEqual(31,len(seen))

    def test_renderer_rejects_blank_required_site_name(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; inp=root/'i.xlsx'; mapping=root/'map.md'; out=root/'out'
            make_model(model); make_input(inp); make_mapping(mapping)
            wb=load_workbook(inp); ws=wb['data']
            headers=[ws.cell(4,c).value for c in range(1,ws.max_column+1)]
            ws.cell(5,headers.index('customer site name')+1,'')
            wb.save(inp); wb.close()
            with self.assertRaises(BackofficeRendererError) as cm:
                render(Namespace(site_data=inp,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None))
            self.assertEqual('BACKOFFICE_CANDIDATE_IDENTITY_MISSING',cm.exception.code)
            self.assertFalse(out.exists() and any(out.iterdir()))
    def test_same_day_supplementary_runs_do_not_overwrite_prior_workbook(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; mapping=root/'map.md'; out=root/'out'; first=root/'first.xlsx'; second=root/'second.xlsx'
            make_model(model); make_mapping(mapping); make_input(first,issue='SUPPLEMENTARY'); make_input(second,issue='SUPPLEMENTARY')
            wb=load_workbook(second); ws=wb['data']; headers=[ws.cell(4,c).value for c in range(1,ws.max_column+1)]
            ws.cell(5,headers.index('customer site code')+1,'S2'); ws.cell(5,headers.index('customer site name')+1,'Site 2'); ws.cell(5,headers.index('du code')+1,'DU2'); wb.save(second); wb.close()
            args=lambda p: Namespace(site_data=p,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None)
            first_paths=render(args(first)); second_paths=render(args(second))
            self.assertEqual(1,len(first_paths)); self.assertEqual(1,len(second_paths))
            self.assertNotEqual(first_paths[0],second_paths[0])
            self.assertTrue(first_paths[0].exists()); self.assertTrue(second_paths[0].exists())
            wb=load_workbook(first_paths[0],read_only=True,data_only=True); first_du=wb['details'].cell(2,6).value; wb.close()
            wb=load_workbook(second_paths[0],read_only=True,data_only=True); second_du=wb['details'].cell(2,6).value; wb.close()
            self.assertEqual('DU1',first_du); self.assertEqual('DU2',second_du)

    def test_effective_provider_transition_partitions_into_separate_workbooks(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; inp=root/'i.xlsx'; mapping=root/'map.md'; out=root/'out'
            make_model(model); make_input(inp,contract='OLD-CONTRACT'); make_mapping(mapping)
            wb=load_workbook(inp); ws=wb['data']; headers=[ws.cell(4,c).value for c in range(1,ws.max_column+1)]
            provider_col=headers.index('Backoffice Subcontractor')+1; contract_col=headers.index('Backoffice Contract Number')+1; trigger_col=headers.index('Backoffice Trigger Date')+1
            ws.cell(5,provider_col,'OldVendor'); ws.cell(5,trigger_col,'2026-07-10')
            base=[ws.cell(5,c).value for c in range(1,ws.max_column+1)]
            row=list(base); row[0]='S2'; row[1]='Site 2'; row[2]='DU2'; row[trigger_col-1]='2026-07-20'; row[provider_col-1]='NewVendor'; row[contract_col-1]='NEW-CONTRACT'; ws.append(row)
            wb.save(inp); wb.close()
            registry={'services':[
                {'effective_from':'2024-01-01','effective_to':'2026-07-15','subcontractor':'OldVendor','contract_number':'OLD-CONTRACT'},
                {'effective_from':'2026-07-16','effective_to':None,'subcontractor':'NewVendor','contract_number':'NEW-CONTRACT'},
            ]}
            with patch('backoffice_ecc_renderer.load_service_registry',return_value=registry):
                paths=render(Namespace(site_data=inp,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None))
            self.assertEqual(2,len(paths))
            self.assertEqual({'OldVendor','NewVendor'},{'OldVendor' if 'OldVendor' in p.name else 'NewVendor' for p in paths})
            workbook_contracts=[]
            for path in paths:
                wb=load_workbook(path,read_only=True,data_only=True); ws=wb['details']
                providers={ws.cell(r,9).value for r in range(2,ws.max_row+1)}; contracts={ws.cell(r,8).value for r in range(2,ws.max_row+1)}; wb.close()
                self.assertEqual(1,len(providers)); self.assertEqual(1,len(contracts)); workbook_contracts.append(next(iter(contracts)))
            self.assertEqual({'OLD-CONTRACT','NEW-CONTRACT'},set(workbook_contracts))
    def test_pipeline_pbom_or_contract_mismatch_leaves_no_output(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; inp=root/'i.xlsx'; mapping=root/'map.md'; out=root/'out'
            make_model(model); make_input(inp,contract='WRONG'); make_mapping(mapping)
            with self.assertRaises(BackofficeRendererError): render(Namespace(site_data=inp,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None))
            self.assertFalse(out.exists() and any(out.iterdir()))

    def test_filename_uses_validated_provider_not_hardcoded_allstar(self):
        with tempfile.TemporaryDirectory() as d:
            root=Path(d); model=root/'m.xlsx'; inp=root/'i.xlsx'; mapping=root/'map.md'; out=root/'out'
            make_model(model); make_input(inp,contract='FUTURE-CONTRACT'); make_mapping(mapping)
            wb=load_workbook(inp); ws=wb['data']; headers=[ws.cell(4,c).value for c in range(1,ws.max_column+1)]; ws.cell(5,headers.index('Backoffice Subcontractor')+1,'FutureVendor'); wb.save(inp); wb.close()
            registry={'services':[{'effective_from':'2020-01-01','effective_to':None,'subcontractor':'FutureVendor','contract_number':'FUTURE-CONTRACT'}]}
            with patch('backoffice_ecc_renderer.load_service_registry',return_value=registry):
                paths=render(Namespace(site_data=inp,pr_model=model,mapping=mapping,output=out,scope='BACKOFFICE',du_model_name='TX Mini Project',all_sites=True,site_code=None))
            self.assertIn('TX Outsource-FutureVendor Backoffice MAIN 2026-07 PR ',paths[0].name)
            self.assertNotIn('Allstar',paths[0].name)

if __name__=='__main__': unittest.main()
