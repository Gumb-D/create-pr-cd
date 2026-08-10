#!/usr/bin/env python3
"""Regression coverage for Issue #82 MW EOS MW Swap antenna evidence."""

from __future__ import annotations

import csv
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from antenna_evidence_resolver import resolve_installation_antenna_evidence


SCRIPT_PATH = REPO_ROOT / "scripts" / "generate_tss_pr_ecc.py"
PR_MODEL_PATH = REPO_ROOT / "Info" / "input" / "pr_model.xlsx"
TEMPLATE_PATH = REPO_ROOT / "Info" / "input" / "ecc_template.xls"
MAPPING_PATH = REPO_ROOT / "Info" / "input" / "contract_info_reference.md"

SITE_COLUMNS = [
    "customer site code",
    "customer site name",
    "du code",
    "region",
    "Province/State",
    "Latitude (North Plus South Minus)",
    "Longitude (East Plus West Minus)",
    "TX Upgrade Scope",
    "Tx SOW",
    "MW Config Antenna Size NE",
    "MW Config Antenna Size FE",
    "SubCon - TI Team",
    "Subcon PR - TI",
    "BOQ Configuration",
    "TX SOW Details",
    "NE SOW Details",
    "FE SOW Details",
]


class TestInstallationAntennaEvidenceResolver(unittest.TestCase):
    def test_direct_fields_take_priority_over_sow_details(self):
        result = resolve_installation_antenna_evidence(
            {
                "MW Config Antenna Size NE": "0.6m",
                "MW Config Antenna Size FE": "1.2",
                "TX SOW Details": "Install new antenna 3.0m",
                "NE SOW Details": "Install new antenna 1.8m",
                "FE SOW Details": "Target antenna 2.4m",
            }
        )
        self.assertEqual(result["status"], "RESOLVED")
        self.assertEqual(result["ne_size"], 0.6)
        self.assertEqual(result["fe_size"], 1.2)
        self.assertEqual(result["selected_size"], 1.2)
        self.assertEqual(result["ne_source"], "MW Config Antenna Size NE")
        self.assertEqual(result["fe_source"], "MW Config Antenna Size FE")
        self.assertEqual(result["group_source"], "ENDPOINT_EVIDENCE")

    def test_falls_back_to_endpoint_sow_details(self):
        result = resolve_installation_antenna_evidence(
            {
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "TX SOW Details": "",
                "NE SOW Details": "Install new antenna 0.6m",
                "FE SOW Details": "Target antenna 1.2m",
            }
        )
        self.assertEqual(result["status"], "RESOLVED")
        self.assertEqual(result["ne_size"], 0.6)
        self.assertEqual(result["fe_size"], 1.2)
        self.assertEqual(result["selected_size"], 1.2)
        self.assertEqual(result["ne_source"], "NE SOW Details")
        self.assertEqual(result["fe_source"], "FE SOW Details")

    def test_falls_back_to_common_tx_sow_details_for_group_decision(self):
        result = resolve_installation_antenna_evidence(
            {
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "TX SOW Details": "MW swap install antenna 0.6m and target antenna 1.2m",
                "NE SOW Details": "",
                "FE SOW Details": "",
            }
        )
        self.assertEqual(result["status"], "RESOLVED_COMMON")
        self.assertIsNone(result["ne_size"])
        self.assertIsNone(result["fe_size"])
        self.assertEqual(result["selected_size"], 1.2)
        self.assertEqual(result["common_size"], 1.2)
        self.assertEqual(result["group_source"], "TX SOW Details")
        self.assertEqual(result["evidence"][0]["source"], "TX SOW Details")

    def test_normalizes_bare_and_m_suffix_formats(self):
        result = resolve_installation_antenna_evidence(
            {
                "MW Config Antenna Size NE": "0.6",
                "MW Config Antenna Size FE": "0.6m",
                "TX SOW Details": "",
                "NE SOW Details": "",
                "FE SOW Details": "",
            }
        )
        self.assertEqual(result["status"], "RESOLVED")
        self.assertEqual(result["ne_size"], 0.6)
        self.assertEqual(result["fe_size"], 0.6)
        self.assertEqual(result["selected_size"], 0.6)

    def test_detail_parser_ignores_unrelated_large_numbers(self):
        result = resolve_installation_antenna_evidence(
            {
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "TX SOW Details": "",
                "NE SOW Details": "Install new antenna 0.6m; VLAN 1234; 18G radio",
                "FE SOW Details": "Build target antenna 1.2m; capacity 833Mbps",
            }
        )
        self.assertEqual(result["status"], "RESOLVED")
        self.assertEqual(result["ne_size"], 0.6)
        self.assertEqual(result["fe_size"], 1.2)
        self.assertEqual(result["selected_size"], 1.2)

    def test_one_endpoint_only_remains_incomplete(self):
        result = resolve_installation_antenna_evidence(
            {
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "TX SOW Details": "",
                "NE SOW Details": "Install new antenna 0.6m",
                "FE SOW Details": "",
            }
        )
        self.assertEqual(result["status"], "INCOMPLETE")
        self.assertEqual(result["ne_size"], 0.6)
        self.assertIsNone(result["fe_size"])
        self.assertEqual(result["selected_size"], 0.6)

    def test_missing_evidence_remains_fail_closed(self):
        result = resolve_installation_antenna_evidence(
            {
                "MW Config Antenna Size NE": "",
                "MW Config Antenna Size FE": "",
                "TX SOW Details": "",
                "NE SOW Details": "",
                "FE SOW Details": "",
            }
        )
        self.assertEqual(result["status"], "MISSING")
        self.assertIsNone(result["ne_size"])
        self.assertIsNone(result["fe_size"])
        self.assertIsNone(result["selected_size"])


class TestMwSwapGeneratorFallback(unittest.TestCase):
    def _write_site_workbook(self, path: Path, rows: list[dict]):
        frame = pd.DataFrame(rows, columns=SITE_COLUMNS)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            frame.to_excel(writer, sheet_name="data", index=False, startrow=3)

    def _run_generator(self, rows: list[dict]):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            site_path = root / "site_data.xlsx"
            output = root / "output"
            output.mkdir()
            self._write_site_workbook(site_path, rows)
            command = [
                sys.executable,
                str(SCRIPT_PATH),
                "--site-data",
                str(site_path),
                "--pr-model",
                str(PR_MODEL_PATH),
                "--template",
                str(TEMPLATE_PATH),
                "--mapping",
                str(MAPPING_PATH),
                "--output",
                str(output),
                "--scope",
                "TI",
                "--all-sites",
            ]
            result = subprocess.run(
                command,
                cwd=REPO_ROOT,
                text=True,
                capture_output=True,
                check=False,
            )

            generated_sites = set()
            for workbook_path in output.glob("*.xlsx"):
                workbook = load_workbook(workbook_path, data_only=True)
                worksheet = workbook["details"]
                for values in worksheet.iter_rows(min_row=2, values_only=True):
                    if values[0]:
                        generated_sites.add(str(values[3]))

            review_rows = []
            review_files = sorted(output.glob("REVIEW_REQUIRED_TI_*.csv"))
            if review_files:
                with review_files[0].open(newline="", encoding="utf-8-sig") as handle:
                    review_rows = list(csv.DictReader(handle))
            return result, generated_sites, review_rows

    def test_mw_swap_uses_governed_tx_sow_detail_when_direct_sizes_blank(self):
        site_code = "ISSUE82_MW_EOS_SWAP"
        row = {
            "customer site code": site_code,
            "customer site name": "Issue 82 MW EOS regression",
            "du code": "DU_ISSUE82",
            "region": "Sabah",
            "Province/State": "Sabah",
            "Latitude (North Plus South Minus)": 5.98,
            "Longitude (East Plus West Minus)": 116.07,
            "TX Upgrade Scope": "TSS+AA+TI",
            "Tx SOW": "MW SWAP",
            "MW Config Antenna Size NE": "",
            "MW Config Antenna Size FE": "",
            "SubCon - TI Team": "Seri Pancar",
            "Subcon PR - TI": "",
            "BOQ Configuration": "",
            "TX SOW Details": "MW swap install antenna 0.6m and target antenna 1.2m",
            "NE SOW Details": "",
            "FE SOW Details": "",
        }

        result, generated_sites, review_rows = self._run_generator([row])
        self.assertEqual(
            result.returncode,
            0,
            msg=f"stdout:\n{result.stdout}\n\nstderr:\n{result.stderr}",
        )
        self.assertIn(site_code, generated_sites)
        review_by_site = {item.get("Site_ID"): item for item in review_rows}
        self.assertNotIn(site_code, review_by_site)


if __name__ == "__main__":
    unittest.main()
