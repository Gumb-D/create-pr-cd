import argparse
import csv
import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import run_all_du_ecc_uat as batch


CONTRACT_PATH = ROOT / "Info" / "input" / "contract_info_reference.md"
POLICY_PATH = ROOT / "config" / "subcontractor_pr_policy.json"


def write_ecc(path: Path, rows: list[dict]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "details"
    worksheet.append(
        [
            "SN.",
            "Purchasing Area*",
            "Region*",
            "Site ID*",
            "Site Name*",
            "Delivery Unit Code*",
            "Logical Site Name",
            "Contract Number *",
            "Subcontractor*",
            "PBOM Code*",
            "SOW*",
            "Unit*",
            "Quantity*",
            "Remarks",
            "",
            "Contract Number",
        ]
    )
    for index, row in enumerate(rows, 1):
        worksheet.append(
            [
                index,
                "Malaysia_Central Region",
                row.get("region", "Central"),
                row["site_code"],
                row.get("site_name", row["site_code"]),
                row.get("du", "DU-1"),
                row["site_code"],
                row.get("contract", "S1MY2024071004WBF1"),
                row.get("subcontractor", "CCSMY"),
                "PBOM-1",
                "Model Description",
                "SITE",
                1,
                "",
                row.get("tx_sow", "MW NEW LINK"),
                row.get("contract", "S1MY2024071004WBF1"),
            ]
        )
    workbook.save(path)
    return path


def scope_summary(profile_id: str, scope: str, output_root: Path, candidate_count: int = 1) -> dict:
    return {
        "status": "SUCCESS",
        "entrypoint": "create_pr.py",
        "run_mode": "NON_PRODUCTION_UAT",
        "profile_status": "PR_INPUT_READY",
        "non_production_uat": True,
        "production_ecc_allowed": False,
        "requested_output": str(output_root),
        "output_root": str(output_root),
        "run_id": "child",
        "scope": scope,
        "profile_id": profile_id,
        "profile_version": "1.0.0",
        "mapping_version": "approved-v1",
        "project_key": "project",
        "du_model_name": "Test DU",
        "du_model_id": "1",
        "view_id": "2",
        "header_hash": "abc",
        "source_record_count": candidate_count,
        "selected_record_count": candidate_count,
        "pre_contract_candidate_count": candidate_count,
        "candidate_count": candidate_count,
        "duplicate_count": 0,
        "ignored_count": 0,
        "review_required_count": 0,
        "contract_mapping_missing_count": 0,
        "sm_excluded_count": 0,
        "created_files": [],
        "review_report": None,
        "contract_mapping_review_report": None,
        "summary_path": str(output_root / "summary.json"),
    }


class TestAllDuEccUat(unittest.TestCase):
    def test_manifest_loader_requires_unique_profile_ids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "manifest.json"
            path.write_text(
                json.dumps(
                    {
                        "schema_version": "1.0",
                        "profiles": [
                            {"profile_id": "a", "source_export": "a.xlsx"},
                            {"profile_id": "a", "source_export": "b.xlsx"},
                        ],
                    }
                ),
                encoding="utf-8",
            )
            with self.assertRaises(batch.BatchUatError) as context:
                batch.load_input_manifest(path)
        self.assertEqual(context.exception.code, "UAT_INPUT_MANIFEST_INVALID")

    def test_deterministic_review_sampling_preserves_sow_and_subcontractor_categories(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ecc = write_ecc(
                Path(temp_dir) / "sample.xlsx",
                [
                    {"site_code": "A", "region": "Central", "subcontractor": "GTSB", "tx_sow": "SOW-1"},
                    {"site_code": "B", "region": "Northern", "subcontractor": "GTSB", "tx_sow": "SOW-1"},
                    {
                        "site_code": "C",
                        "region": "Central",
                        "subcontractor": "Nera",
                        "contract": "S1MY2023083002WBF1",
                        "tx_sow": "SOW-2",
                    },
                ],
            )
            self.assertEqual(batch.deterministic_review_site_codes([ecc], 10), ["A", "B", "C"])
            self.assertEqual(batch.deterministic_review_site_codes([ecc], 2), ["A", "C"])
            with self.assertRaises(batch.BatchUatError) as context:
                batch.deterministic_review_site_codes([ecc], 1)
        self.assertEqual(context.exception.code, "REVIEW_SAMPLE_CAP_TOO_LOW")

    def test_materialize_scope_artifacts_uses_extended_windows_paths(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            scope_dir = root / "FULL_PACK" / "TSS"
            engine_root = scope_dir / "_ENGINE"
            source = engine_root / "NON_PRODUCTION_UAT" / "child" / "Central-Nera Test TSS PR.xlsx"
            source.parent.mkdir(parents=True, exist_ok=True)
            source.write_bytes(b"uat")
            summary = scope_summary("profile_a", "TSS", engine_root)
            rename_calls: list[tuple[str, str]] = []
            real_os_rename = batch.os.rename

            def strip_extended_path(value: str) -> str:
                prefix = "\\\\?\\"
                unc_prefix = "\\\\?\\UNC\\"
                if value.startswith(unc_prefix):
                    return "\\\\" + value[len(unc_prefix) :]
                if value.startswith(prefix):
                    return value[len(prefix) :]
                return value

            def fake_rename(raw_source: str, raw_target: str) -> None:
                rename_calls.append((raw_source, raw_target))
                self.assertTrue(raw_source.startswith("\\\\?\\"))
                self.assertTrue(raw_target.startswith("\\\\?\\"))
                real_os_rename(strip_extended_path(raw_source), strip_extended_path(raw_target))

            with mock.patch.object(batch, "_is_windows", return_value=True, create=True), mock.patch.object(
                batch.os, "rename", side_effect=fake_rename
            ), mock.patch.object(
                batch.shutil, "move", side_effect=AssertionError("materialisation must not use shutil.move")
            ):
                adjusted, generated = batch.materialize_scope_artifacts(
                    engine_root,
                    scope_dir,
                    summary,
                    "FULL_PACK",
                    "RUN-1",
                )

            expected = scope_dir / "Central-Nera Test TSS PR_FULL_PACK_NON_PRODUCTION_UAT_RUN-1.xlsx"
            self.assertEqual(len(rename_calls), 1)
            self.assertFalse(source.exists())
            self.assertTrue(expected.is_file())
            self.assertEqual(generated, [expected.resolve()])
            self.assertEqual(adjusted["created_files"], [])
            self.assertFalse(engine_root.exists())

    def test_master_manifest_initializes_pending_business_status(self):
        row = {column: "" for column in batch.MANIFEST_COLUMNS}
        row.update(
            {
                "Run ID": "RUN-1",
                "Pack Type": "FULL_PACK",
                "Profile ID": "profile_a",
                "Scope": "TSS",
                "Verification Status": "PENDING",
            }
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "manifest.xlsx"
            batch.write_master_manifest(path, [row])
            workbook = load_workbook(path, data_only=True)
            worksheet = workbook["ECC Files"]
            headers = [cell.value for cell in worksheet[1]]
            status_column = headers.index("Verification Status") + 1
            self.assertEqual(worksheet.cell(2, status_column).value, "PENDING")
            self.assertIn("Instructions", workbook.sheetnames)
            workbook.close()

    def test_run_batch_isolates_scope_failure_and_reports_blocked_profiles(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "profile_a.xlsx"
            source.write_bytes(b"source")
            manifest = root / "manifest.json"
            manifest.write_text(
                json.dumps(
                    {
                        "schema_version": "1.0",
                        "profiles": [
                            {"profile_id": "profile_a", "source_export": str(source)},
                            {"profile_id": "profile_b", "source_export": ""},
                            {"profile_id": "profile_c", "source_export": str(root / "missing.xlsx")},
                        ],
                    }
                ),
                encoding="utf-8",
            )
            profiles = {
                "profile_a": {"profile_id": "profile_a", "status": "PR_INPUT_READY"},
                "profile_b": {"profile_id": "profile_b", "status": "DRAFT"},
                "profile_c": {"profile_id": "profile_c", "status": "PR_INPUT_READY"},
            }

            def fake_resolve(path, **_kwargs):
                self.assertEqual(Path(path), source.resolve())
                return {"profile": profiles["profile_a"], "header_hash": "abc", "inventory": []}

            def fake_scope_pack(
                source_export,
                scope_dir,
                scope,
                pack_type,
                site_codes,
                profile_id,
                batch_run_id,
                args,
            ):
                self.assertEqual(profile_id, "profile_a")
                if scope == "TI" and pack_type == "FULL_PACK":
                    raise batch.BatchUatError("SIMULATED_TI_FAILURE", "simulated isolated TI failure")
                path = write_ecc(
                    scope_dir
                    / f"Central-CCSMY Test DU {scope} PR_{pack_type}_NON_PRODUCTION_UAT_{batch_run_id}.xlsx",
                    [{"site_code": "A", "subcontractor": "CCSMY", "tx_sow": "MW NEW LINK"}],
                )
                summary = scope_summary(profile_id, scope, scope_dir)
                summary["pack_type"] = pack_type
                summary["created_files"] = [str(path.resolve())]
                summary_path = scope_dir / f"summary-{pack_type}-{scope}.json"
                summary["summary_path"] = str(summary_path.resolve())
                summary_path.write_text(json.dumps(summary), encoding="utf-8")
                return summary, [path.resolve()]

            args = argparse.Namespace(
                manifest=manifest,
                output=root / "output",
                review_max_combinations=500,
                run_id="RUN-1",
                pr_model=Path("pr_model.xlsx"),
                template=Path("template.xls"),
                mapping=CONTRACT_PATH,
                subcontractor_policy=POLICY_PATH,
            )
            with mock.patch.object(batch, "load_structured_profiles", return_value=profiles), mock.patch.object(
                batch, "resolve_du_profile", side_effect=fake_resolve
            ), mock.patch.object(batch, "run_scope_pack", side_effect=fake_scope_pack):
                summary = batch.run_batch(args)

            self.assertEqual(summary["status"], "COMPLETED_WITH_BLOCKS")
            self.assertEqual(summary["eligible_profile_count"], 1)
            self.assertEqual(summary["blocked_profile_count"], 2)
            self.assertEqual(summary["successful_scope_runs"], 1)
            self.assertEqual(summary["failed_scope_runs"], 1)
            self.assertEqual(summary["generated_ecc_file_count"], 2)
            self.assertEqual(summary["generated_ecc_row_count"], 2)
            self.assertEqual(summary["candidate_count"], 1)
            self.assertTrue(summary["manifest_reconciliation_ok"])
            self.assertEqual(summary["unsafe_manifest_row_count"], 0)
            reasons = {row["Reason Code"] for row in summary["blocked_profiles"]}
            self.assertEqual(reasons, {"PROFILE_STATUS_BLOCKED", "MISSING_SOURCE_EXPORT"})

            output_root = Path(summary["output_root"])
            self.assertTrue((output_root / "UAT_MASTER_MANIFEST.xlsx").is_file())
            self.assertTrue((output_root / "UAT_MASTER_SUMMARY.json").is_file())
            self.assertTrue((output_root / "UAT_BLOCKED_PROFILES.csv").is_file())
            with (output_root / "UAT_BLOCKED_PROFILES.csv").open(encoding="utf-8-sig", newline="") as handle:
                blocked_rows = list(csv.DictReader(handle))
            self.assertEqual(len(blocked_rows), 2)

            with mock.patch.object(batch, "load_structured_profiles", return_value=profiles):
                with self.assertRaises(batch.BatchUatError) as context:
                    batch.run_batch(args)
            self.assertEqual(context.exception.code, "UAT_RUN_ALREADY_EXISTS")

    def test_tx_rollout_profile_preserves_mapping_and_approves_both_hashes(self):
        path = ROOT / "config" / "du_profiles" / "tx_rollout_2023_pr_v1.yaml"
        profile = json.loads(path.read_text(encoding="utf-8"))
        old_hash = "8aab4c2da2dc133e0a65b9203c62e6db1ebeb30430f9f63f5c5de1673703c320"
        new_hash = "e61b834994eeef30e7d8249f87616cb04d60598eea323feea50178fc4292c162"
        self.assertEqual(profile["profile_version"], "0.1.1")
        self.assertEqual(profile["mapping_version"], "approved-2026-07-10-2023-tx-rollout-v2")
        self.assertEqual(profile["status"], "PR_INPUT_READY")
        self.assertEqual(profile["export_structure"]["observed_header_hash"], new_hash)
        self.assertEqual(profile["export_structure"]["approved_header_hashes"], [old_hash, new_hash])
        for field in (
            "existing_ti_pr_status",
            "existing_tss_pr_status",
            "region",
            "site_code",
            "subcontractor_ti",
            "subcontractor_tss",
            "tx_sow_raw",
        ):
            candidates = profile["field_mapping"][field]["source_candidates"]
            self.assertTrue(candidates)
            self.assertTrue(all(candidate["mapping_status"] == "APPROVED" for candidate in candidates))


if __name__ == "__main__":
    unittest.main()
