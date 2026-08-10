import argparse
import json
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import create_pr


CONTRACT_PATH = ROOT / "Info" / "input" / "contract_info_reference.md"
POLICY_PATH = ROOT / "config" / "subcontractor_pr_policy.json"


def record(site: str, subcontractor: str) -> dict:
    return {
        "identity": {"source_row_number": 10},
        "site": {"site_code": site, "site_name": f"Site {site}", "du_key": f"DU-{site}"},
        "pr_context": {
            "subcontractor_tss": subcontractor,
            "subcontractor_ti": subcontractor,
            "existing_tss_pr_status": "NO_PR",
            "existing_ti_pr_status": "NO_PR",
            "tx_sow_normalized": "MW NEW LINK",
            "region": "Central",
        },
        "technical_context": {},
        "source_evidence": {
            "fields": {"tx_sow_normalized": {"normalization_status": "APPROVED"}}
        },
        "validation": {
            "profile_id": "test_profile",
            "pr_input_classification": "PR_INPUT_READY",
            "blocking_reasons": [],
        },
    }


def write_fake_ecc(path: Path, site_code: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "details"
    worksheet.append(["SN.", "Site ID*"])
    worksheet.append([1, site_code])
    workbook.save(path)
    workbook.close()


class TestCreatePrSafetyIntegration(unittest.TestCase):
    def test_mixed_input_sends_only_policy_and_contract_safe_record_to_renderer(self):
        records = [
            record("SM-SITE", " sm "),
            record("UNKNOWN-SITE", "Unknown Vendor"),
            record("CCSMY-SITE", "ccsmy"),
        ]
        metadata = {
            "profile_id": "test_profile",
            "profile_version": "1.0.0",
            "mapping_version": "approved-v1",
            "project_key": "project",
            "du_model_name": "Test DU",
            "du_model_id": "1",
            "view_id": "2",
            "header_hash": "abc",
        }
        resolution = {
            "profile": {"profile_id": "test_profile", "status": "PR_INPUT_READY"},
            "inventory": [],
            "header_hash": "abc",
        }

        seen_renderer_sites = []
        seen_renderer_subcontractors = []

        def fake_renderer(command, **_kwargs):
            canonical_input = Path(command[command.index("--site-data") + 1])
            output = Path(command[command.index("--output") + 1])
            workbook = load_workbook(canonical_input, data_only=True)
            worksheet = workbook["data"]
            headers = [cell.value for cell in worksheet[4]]
            site_col = headers.index("customer site code") + 1
            subcon_col = headers.index("SubCon - TSS Team") + 1
            for row_number in range(5, worksheet.max_row + 1):
                seen_renderer_sites.append(worksheet.cell(row_number, site_col).value)
                seen_renderer_subcontractors.append(worksheet.cell(row_number, subcon_col).value)
            workbook.close()
            write_fake_ecc(output / "Central-CCSMY Test DU TSS PR 20260729.xlsx", "CCSMY-SITE")
            return SimpleNamespace(returncode=0, stdout="", stderr="")

        with tempfile.TemporaryDirectory() as temp_dir:
            parsed = argparse.Namespace(
                site_data=Path("input.xlsx"),
                output=Path(temp_dir),
                scope="TSS",
                site_code=None,
                all_sites=True,
                pr_model=Path("pr_model.xlsx"),
                template=Path("template.xls"),
                mapping=CONTRACT_PATH,
                subcontractor_policy=POLICY_PATH,
                non_production_uat=True,
                uat_run_id="20260729T120000000000Z",
            )
            with mock.patch.object(create_pr, "resolve_du_profile", return_value=resolution), mock.patch.object(
                create_pr,
                "build_canonical_records",
                return_value=(records, metadata),
            ), mock.patch.object(create_pr.subprocess, "run", side_effect=fake_renderer):
                summary = create_pr.run(parsed)

            self.assertEqual(seen_renderer_sites, ["CCSMY-SITE"])
            self.assertEqual(seen_renderer_subcontractors, ["CCSMY"])
            self.assertEqual(summary["pre_contract_candidate_count"], 2)
            self.assertEqual(summary["candidate_count"], 1)
            self.assertEqual(summary["ignored_count"], 1)
            self.assertEqual(summary["sm_excluded_count"], 1)
            self.assertEqual(summary["contract_mapping_missing_count"], 1)
            self.assertEqual(summary["contract_mapping_missing_subcontractors"], ["Unknown Vendor"])
            self.assertEqual(summary["review_required_count"], 1)
            self.assertEqual(
                summary["ignored_reason_distribution"],
                {"PR_NOT_REQUIRED_OUTSOURCED_TO_OTHER_VENDOR": 1},
            )
            self.assertEqual(
                summary["review_required_reason_distribution"],
                {"CONTRACT_MAPPING_NOT_FOUND": 1},
            )
            self.assertTrue(Path(summary["review_report"]).is_file())
            self.assertTrue(Path(summary["contract_mapping_review_report"]).is_file())
            self.assertEqual(len(summary["created_files"]), 1)
            self.assertIn("NON_PRODUCTION_UAT", Path(summary["created_files"][0]).name)
            self.assertFalse(any("SM-SITE" in path for path in summary["created_files"]))

            persisted = json.loads(Path(summary["summary_path"]).read_text(encoding="utf-8"))
            self.assertEqual(persisted["contract_mapping_missing_count"], 1)
            self.assertEqual(persisted["sm_excluded_count"], 1)

    def test_all_unknown_contract_candidates_generate_reports_but_no_ecc(self):
        records = [record("UNKNOWN-ONLY", "Unknown Vendor")]
        metadata = {
            "profile_id": "test_profile",
            "profile_version": "1.0.0",
            "mapping_version": "approved-v1",
            "project_key": "project",
            "du_model_name": "Test DU",
            "du_model_id": "1",
            "view_id": "2",
            "header_hash": "abc",
        }
        resolution = {
            "profile": {"profile_id": "test_profile", "status": "PR_INPUT_READY"},
            "inventory": [],
            "header_hash": "abc",
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            parsed = argparse.Namespace(
                site_data=Path("input.xlsx"),
                output=Path(temp_dir),
                scope="TI",
                site_code=None,
                all_sites=True,
                pr_model=Path("pr_model.xlsx"),
                template=Path("template.xls"),
                mapping=CONTRACT_PATH,
                subcontractor_policy=POLICY_PATH,
                non_production_uat=True,
                uat_run_id="20260729T120000000001Z",
            )
            with mock.patch.object(create_pr, "resolve_du_profile", return_value=resolution), mock.patch.object(
                create_pr,
                "build_canonical_records",
                return_value=(records, metadata),
            ), mock.patch.object(create_pr.subprocess, "run") as renderer:
                summary = create_pr.run(parsed)

            renderer.assert_not_called()
            self.assertEqual(summary["candidate_count"], 0)
            self.assertEqual(summary["contract_mapping_missing_count"], 1)
            self.assertEqual(summary["created_files"], [])
            self.assertTrue(Path(summary["contract_mapping_review_report"]).is_file())


if __name__ == "__main__":
    unittest.main()
