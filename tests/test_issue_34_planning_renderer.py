#!/usr/bin/env python3
"""Planning ECC renderer regression coverage for Issue #34."""
from __future__ import annotations

import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
RENDERER = SCRIPTS / "planning_ecc_renderer.py"
MAPPING = ROOT / "Info" / "input" / "contract_info_reference.md"

INPUT_HEADERS = [
    "customer site code",
    "customer site name",
    "du code",
    "region",
    "Subcon - Planning",
    "Planning Contract Subcontractor",
    "Planning PBOM Code",
    "Planning SOW",
    "Planning Unit",
    "Planning Quantity",
]

ECC_HEADERS = [
    "SN.",
    "Purchasing Area*",
    "Region*",
    "Site ID*",
    "Site Name*",
    "Delivery Unit Code*",
    "Logical Site Name",
    "Contract Number *",
    "Subcontractor*",
    "PBOM Code*",
    "SOW*",
    "Unit*",
    "Quantity*",
    "Remarks",
    "",
    "Contract Number",
]

FULL_DESCRIPTION = "2026-Detailed end to end transmission planning and design"
SINGLE_HOP_DESCRIPTION = "2026-Single-hop planning and design"
AA_DESCRIPTION = "Detailed end to end transmission planning and design (for AA modification & AA submisison sow only)"


def _write_input(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "data"
    worksheet.append(INPUT_HEADERS)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _row(
    site: str,
    *,
    region: str = "Central",
    source_subcon: str = "GCI",
    contract_subcon: str = "GCI",
    pbom: str = "350001143904",
    description: str = FULL_DESCRIPTION,
) -> list[object]:
    return [
        site,
        f"Site {site}",
        f"DU-{site}",
        region,
        source_subcon,
        contract_subcon,
        pbom,
        description,
        "Hop",
        1,
    ]


class PlanningRendererPresenceTest(unittest.TestCase):
    def test_planning_renderer_module_exists(self) -> None:
        self.assertTrue(RENDERER.exists(), "Issue #34 requires scripts/planning_ecc_renderer.py")


@unittest.skipUnless(RENDERER.exists(), "Planning renderer not implemented yet")
class PlanningRendererContractTest(unittest.TestCase):
    def _run(self, rows: list[list[object]], du_model: str):
        temporary = tempfile.TemporaryDirectory()
        root = Path(temporary.name)
        site_data = root / "planning-input.xlsx"
        output = root / "output"
        output.mkdir()
        _write_input(site_data, rows)
        result = subprocess.run(
            [
                sys.executable,
                str(RENDERER),
                "--site-data",
                str(site_data),
                "--mapping",
                str(MAPPING),
                "--output",
                str(output),
                "--scope",
                "Planning",
                "--all-sites",
                "--du-model-name",
                du_model,
            ],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=False,
        )
        return temporary, output, result

    def _read_details(self, path: Path) -> tuple[list[str], list[tuple[object, ...]], list[str]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["details"])
            worksheet = workbook["details"]
            header = ["" if value is None else str(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            return header, rows, workbook.sheetnames
        finally:
            workbook.close()

    def test_full_planning_and_aa_rows_share_base_gci_contract_group(self) -> None:
        temporary, output, result = self._run(
            [
                _row("A0001"),
                _row(
                    "A0002",
                    source_subcon="GCI_AA",
                    contract_subcon="GCI",
                    pbom="350001042321",
                    description=AA_DESCRIPTION,
                ),
            ],
            "2024 Celcomdigi BAU",
        )
        self.addCleanup(temporary.cleanup)
        self.assertEqual(result.returncode, 0, msg=f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}")
        files = sorted(output.glob("*.xlsx"))
        self.assertEqual(len(files), 1)
        self.assertIn("Central-GCI 2024 Celcomdigi BAU Planning PR", files[0].name)
        header, rows, _ = self._read_details(files[0])
        self.assertEqual(header, ECC_HEADERS)
        self.assertEqual(len(rows), 2)
        by_site = {str(row[3]): row for row in rows}

        standard = by_site["A0001"]
        self.assertEqual(str(standard[9]), "350001143904")
        self.assertEqual(standard[10], FULL_DESCRIPTION)
        self.assertEqual(standard[11], "Hop")
        self.assertEqual(standard[12], 1)
        self.assertEqual(standard[8], "GCI")
        self.assertEqual(standard[7], "S1MY2024071002WBF1")
        self.assertIn(standard[14], (None, ""))
        self.assertEqual(standard[15], "S1MY2024071002WBF1")

        aa = by_site["A0002"]
        self.assertEqual(str(aa[9]), "350001042321")
        self.assertEqual(aa[10], AA_DESCRIPTION)
        self.assertEqual(aa[8], "GCI")
        self.assertEqual(aa[7], "S1MY2024071002WBF1")
        self.assertNotIn(str(aa[9]), {"350001143904", "350001143905"})

    def test_single_hop_du_renders_350001143905(self) -> None:
        temporary, output, result = self._run(
            [_row("B0001", source_subcon="GTSB", contract_subcon="GTSB", pbom="350001143905", description=SINGLE_HOP_DESCRIPTION)],
            "MW EOS Swap",
        )
        self.addCleanup(temporary.cleanup)
        self.assertEqual(result.returncode, 0, msg=result.stderr)
        files = list(output.glob("*.xlsx"))
        self.assertEqual(len(files), 1)
        _, rows, _ = self._read_details(files[0])
        self.assertEqual(str(rows[0][9]), "350001143905")
        self.assertEqual(rows[0][10], SINGLE_HOP_DESCRIPTION)
        self.assertEqual(rows[0][8], "GTSB")
        self.assertEqual(rows[0][7], "S1MY2024071003WBF1")

    def test_renderer_splits_after_30_unique_sites(self) -> None:
        rows = [_row(f"C{index:04d}") for index in range(1, 32)]
        temporary, output, result = self._run(rows, "2023 TX Rollout")
        self.addCleanup(temporary.cleanup)
        self.assertEqual(result.returncode, 0, msg=result.stderr)
        files = sorted(output.glob("*.xlsx"))
        self.assertEqual(len(files), 2)
        counts = []
        for path in files:
            _, rendered_rows, _ = self._read_details(path)
            counts.append(len({str(row[3]) for row in rendered_rows}))
        self.assertEqual(sorted(counts), [1, 30])

    def test_unknown_region_fails_closed_without_partial_ecc(self) -> None:
        temporary, output, result = self._run([_row("D0001", region="UNKNOWN_REGION")], "2023 TX Rollout")
        self.addCleanup(temporary.cleanup)
        self.assertNotEqual(result.returncode, 0)
        self.assertEqual(list(output.glob("*.xlsx")), [])
        self.assertIn("PURCHASING_AREA_NOT_FOUND", result.stderr)


if __name__ == "__main__":
    unittest.main()
