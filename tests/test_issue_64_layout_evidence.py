import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from profile_du_export import build_header_inventory, resolve_approved_header_structure

FIXTURE = ROOT / "tests" / "fixtures" / "tx_mini_du_export_fixture.xlsx"
PROFILE_PATH = ROOT / "config" / "du_profiles" / "tx_mini_pr_v1.yaml"
MODEL_ID = "4188808420049567786"
NEW_VIEW_ID = "9999999999999999999"


class TestIssue64LayoutEvidence(unittest.TestCase):
    def test_structural_approval_uses_approved_site_mapping_not_identity_view_allowlist(self):
        profile = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
        profile["identity"]["accepted_view_ids"] = ["1111111111111111111"]

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "worker-upload.xlsx"
            shutil.copy2(FIXTURE, source)
            workbook = load_workbook(source)
            workbook["data"]["A1"] = (
                f"site|fix00012|{MODEL_ID}|{NEW_VIEW_ID}"
            )
            workbook.save(source)
            workbook.close()

            validation = resolve_approved_header_structure(
                build_header_inventory(source),
                profile,
            )

        self.assertTrue(validation["approved"])
        self.assertEqual(
            validation["approval_basis"],
            "VIEW_NORMALIZED_TO_APPROVED_LAYOUT",
        )

    def test_unapproved_site_mapping_cannot_supply_layout_reference(self):
        profile = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
        profile["identity"]["accepted_view_ids"] = ["2477626672974883536"]
        for candidate in profile["field_mapping"]["site_code"]["source_candidates"]:
            candidate["mapping_status"] = "UNVERIFIED"

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "worker-upload.xlsx"
            shutil.copy2(FIXTURE, source)
            workbook = load_workbook(source)
            workbook["data"]["A1"] = (
                f"site|fix00012|{MODEL_ID}|{NEW_VIEW_ID}"
            )
            workbook.save(source)
            workbook.close()

            validation = resolve_approved_header_structure(
                build_header_inventory(source),
                profile,
            )

        self.assertFalse(validation["approved"])
        self.assertEqual(validation["approval_basis"], "UNAPPROVED_STRUCTURE")


if __name__ == "__main__":
    unittest.main()
