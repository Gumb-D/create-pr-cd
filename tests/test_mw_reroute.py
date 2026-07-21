import os
import unittest
import sys
import subprocess
import tempfile
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import Workbook, load_workbook
import pandas as pd

# Add the scripts directory to path so we can import pr_helpers
scripts_dir = Path(__file__).parent.parent / 'scripts'
if str(scripts_dir) not in sys.path:
    sys.path.insert(0, str(scripts_dir))

from pr_helpers import (
    is_mw_reroute_row,
    parse_mw_new_link_reroute,
    filter_tss_mw_new_link_reroute_items,
    select_tss_items_for_site,
    has_duplicate_pbom
)


class TestTIRoutingControls(unittest.TestCase):
    """Test TI routing control logic."""

    def test_ipran_reroute_returns_false(self):
        """IPRAN Reroute should return False."""
        row = {'Tx SOW': 'IPRAN Reroute', 'TX Upgrade Scope': ''}
        self.assertFalse(is_mw_reroute_row(row))

    def test_mw_swap_with_dismantle_returns_false(self):
        """MW Swap with TX Upgrade Scope containing 'dismantle' should return False."""
        row = {'Tx SOW': 'MW Swap', 'TX Upgrade Scope': 'dismantle and remove'}
        self.assertFalse(is_mw_reroute_row(row))

    def test_mw_reroute_without_dismantle_returns_true(self):
        """MW Reroute SOW without 'dismantle' should return True."""
        row = {'Tx SOW': 'MW New Link / Reroute', 'TX Upgrade Scope': 'some other scope'}
        self.assertTrue(is_mw_reroute_row(row))

    def test_case_insensitive_matching(self):
        """Matching should be case insensitive."""
        row = {'Tx SOW': 'mw REROUTE work', 'TX Upgrade Scope': ''}
        self.assertTrue(is_mw_reroute_row(row))

    def test_empty_or_missing_sow_returns_false(self):
        """Empty or missing SOW should return False."""
        row1 = {'Tx SOW': '', 'TX Upgrade Scope': 'dismantle'}
        self.assertFalse(is_mw_reroute_row(row1))
        row2 = {'Tx SOW': None, 'TX Upgrade Scope': ''}
        self.assertFalse(is_mw_reroute_row(row2))

    def test_applicable_mw_new_link_reroute_ti_routing(self):
        """MW New Link / Reroute TI routing should work as expected."""
        # New Link case (no 'dismantle' in Upgrade Scope)
        row_new_link = {
            'Tx SOW': 'MW New Link / Reroute',
            'TX Upgrade Scope': 'install new equipment'
        }
        self.assertTrue(is_mw_reroute_row(row_new_link))

        # Reroute case (has 'dismantle' in Upgrade Scope)
        row_reroute = {
            'Tx SOW': 'MW New Link / Reroute',
            'TX Upgrade Scope': 'dismantle existing equipment'
        }
        self.assertTrue(is_mw_reroute_row(row_reroute))


class TestParseMWNewLinkReroute(unittest.TestCase):
    """Test parse_mw_new_link_reroute function."""

    def test_reroute_with_dismantle(self):
        """Should return True when SOW matches and upgrade_scope contains 'dismantle'."""
        self.assertTrue(parse_mw_new_link_reroute(
            'MW New Link / Reroute',
            'dismantle old equipment'
        ))

    def test_new_link_without_dismantle(self):
        """Should return False when SOW matches but no 'dismantle' in upgrade_scope."""
        self.assertFalse(parse_mw_new_link_reroute(
            'MW New Link / Reroute',
            'install new link'
        ))

    def test_non_matching_sow_returns_false(self):
        """Should return False when SOW doesn't match the pattern."""
        self.assertFalse(parse_mw_new_link_reroute(
            'MW New Link Only',
            'dismantle something'
        ))
        self.assertFalse(parse_mw_new_link_reroute(
            'Random SOW',
            'dismantle something'
        ))

    def test_case_insensitive_dismantle_check(self):
        """Should match 'dismantle' case-insensitively."""
        self.assertTrue(parse_mw_new_link_reroute(
            'MW New Link / Reroute',
            'DISMANTLE old gear'
        ))

    def test_empty_strings(self):
        """Should handle empty strings gracefully."""
        self.assertFalse(parse_mw_new_link_reroute('', ''))
        self.assertFalse(parse_mw_new_link_reroute('MW New Link / Reroute', ''))


class TestTSSFiltering(unittest.TestCase):
    """Test TSS model-driven filtering for MW New Link / Reroute."""

    def _create_tss_models(self):
        """Create the standard set of TSS model items for testing."""
        return [
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000062773', 'Description': 'LOS Survey', 'Unit': 'Each', 'Quantity': 1, 'Is_Mandatory': True, 'Remarks': ''},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000062776', 'Description': 'LOS Survey LOS', 'Unit': 'Each', 'Quantity': 1, 'Is_Mandatory': True, 'Remarks': ''},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000589343', 'Description': 'Item A', 'Unit': 'Each', 'Quantity': 1.0, 'Is_Mandatory': True, 'Remarks': 'New Link'},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000589343', 'Description': 'Item A', 'Unit': 'Each', 'Quantity': 1.5, 'Is_Mandatory': True, 'Remarks': 'Reroute'},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000589344', 'Description': 'Item B', 'Unit': 'Each', 'Quantity': 1.0, 'Is_Mandatory': True, 'Remarks': 'New Link'},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000589344', 'Description': 'Item B', 'Unit': 'Each', 'Quantity': 1.5, 'Is_Mandatory': True, 'Remarks': 'Reroute'},
        ]

    def test_scenario_1_new_link_non_los_no_dismantle(self):
        """Scenario 1: New Link, non-LOS site, no dismantle."""
        tss_models = self._create_tss_models()
        site_id = 'A01073_AD'
        sow = 'MW New Link / Reroute'
        upgrade_scope = 'some installation'

        result = select_tss_items_for_site(site_id, sow, upgrade_scope, tss_models)
        pboms = [item['PBOM_Code'] for item in result]

        self.assertEqual(pboms.count('350000062773'), 1)
        self.assertNotIn('350000062776', pboms)

        # Check PBOM 350000589343: should have quantity 1.0 (New Link)
        qty_343 = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589343']
        self.assertIn(1.0, qty_343)
        self.assertNotIn(1.5, qty_343)

        # Check PBOM 350000589344: should have quantity 1.0 (New Link)
        qty_344 = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589344']
        self.assertIn(1.0, qty_344)
        self.assertNotIn(1.5, qty_344)

        self.assertFalse(has_duplicate_pbom(result))

    def test_scenario_2_new_link_with_los_no_dismantle(self):
        """Scenario 2: New Link, LOS site, no dismantle."""
        tss_models = self._create_tss_models()
        site_id = 'SITE_LOS_001'
        sow = 'MW New Link / Reroute'
        upgrade_scope = 'some installation'

        result = select_tss_items_for_site(site_id, sow, upgrade_scope, tss_models)
        pboms = [item['PBOM_Code'] for item in result]

        self.assertIn('350000062773', pboms)
        self.assertNotIn('350000062776', pboms)

        # Check PBOM 350000589343: should have quantity 1.0 (New Link)
        qty_343 = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589343']
        self.assertIn(1.0, qty_343)
        self.assertNotIn(1.5, qty_343)

        # Check PBOM 350000589344: should have quantity 1.0 (New Link)
        qty_344 = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589344']
        self.assertIn(1.0, qty_344)
        self.assertNotIn(1.5, qty_344)

        self.assertFalse(has_duplicate_pbom(result))

    def test_scenario_3_reroute_non_los_with_dismantle(self):
        """Scenario 3: Reroute, non-LOS site, with dismantle."""
        tss_models = self._create_tss_models()
        site_id = 'B00256'
        sow = 'MW New Link / Reroute'
        upgrade_scope = 'dismantle existing equipment'

        result = select_tss_items_for_site(site_id, sow, upgrade_scope, tss_models)
        pboms = [item['PBOM_Code'] for item in result]

        self.assertIn('350000062773', pboms)
        self.assertNotIn('350000062776', pboms)

        # Check PBOM 350000589343: should have quantity 1.5 (Reroute)
        qty_343 = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589343']
        self.assertIn(1.5, qty_343)
        self.assertNotIn(1.0, qty_343)

        # Check PBOM 350000589344: should have quantity 1.5 (Reroute)
        qty_344 = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589344']
        self.assertIn(1.5, qty_344)
        self.assertNotIn(1.0, qty_344)

        self.assertFalse(has_duplicate_pbom(result))

    def test_scenario_4_reroute_with_los_with_dismantle(self):
        """Scenario 4: Reroute, LOS site, with dismantle."""
        tss_models = self._create_tss_models()
        site_id = 'SITE_LOS_002'
        sow = 'MW New Link / Reroute'
        upgrade_scope = 'dismantle existing equipment'

        result = select_tss_items_for_site(site_id, sow, upgrade_scope, tss_models)
        pboms = [item['PBOM_Code'] for item in result]

        self.assertNotIn('350000062773', pboms)
        self.assertIn('350000062776', pboms)

        # Check PBOM 350000589343: should have quantity 1.5 (Reroute)
        qty_343 = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589343']
        self.assertIn(1.5, qty_343)
        self.assertNotIn(1.0, qty_343)

        # Check PBOM 350000589344: should have quantity 1.5 (Reroute)
        qty_344 = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589344']
        self.assertIn(1.5, qty_344)
        self.assertNotIn(1.0, qty_344)

        self.assertFalse(has_duplicate_pbom(result))

    def test_remarks_exclusion_new_link_excludes_reroute_remark(self):
        """Remarks = 'Reroute' should be excluded for New Link."""
        tss_models = self._create_tss_models()
        site_id = 'A01073_AD'
        sow = 'MW New Link / Reroute'
        upgrade_scope = 'install new equipment'

        result = select_tss_items_for_site(site_id, sow, upgrade_scope, tss_models)

        # Should include the New Link version (qty 1.0) but exclude the Reroute version (qty 1.5)
        qty_343_items = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589343']
        self.assertIn(1.0, qty_343_items, "New Link version (qty 1.0) should be included")
        self.assertNotIn(1.5, qty_343_items, "Reroute version (qty 1.5) should be excluded")

        qty_344_items = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589344']
        self.assertIn(1.0, qty_344_items, "New Link version (qty 1.0) should be included")
        self.assertNotIn(1.5, qty_344_items, "Reroute version (qty 1.5) should be excluded")

    def test_remarks_exclusion_reroute_excludes_new_link_remark(self):
        """Remarks = 'New Link' should be excluded for Reroute."""
        tss_models = self._create_tss_models()
        site_id = 'B00256'
        sow = 'MW New Link / Reroute'
        upgrade_scope = 'dismantle existing equipment'

        result = select_tss_items_for_site(site_id, sow, upgrade_scope, tss_models)

        # Should include the Reroute version (qty 1.5) but exclude the New Link version (qty 1.0)
        qty_343_items = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589343']
        self.assertIn(1.5, qty_343_items, "Reroute version (qty 1.5) should be included")
        self.assertNotIn(1.0, qty_343_items, "New Link version (qty 1.0) should be excluded")

        qty_344_items = [item['Quantity'] for item in result if item['PBOM_Code'] == '350000589344']
        self.assertIn(1.5, qty_344_items, "Reroute version (qty 1.5) should be included")
        self.assertNotIn(1.0, qty_344_items, "New Link version (qty 1.0) should be excluded")

    def test_unrelated_empty_remarks_items_retained(self):
        """Unrelated mandatory items with empty Remarks should be retained unless handled by LOS exception."""
        tss_models = self._create_tss_models()
        # The LOS items (350000062773 and 350000062776) have empty Remarks
        site_id = 'A01073_AD'
        sow = 'MW New Link / Reroute'
        upgrade_scope = 'some installation'

        result = select_tss_items_for_site(site_id, sow, upgrade_scope, tss_models)
        pboms = [item['PBOM_Code'] for item in result]

        # Should include 350000062773 (LOS Survey with empty remarks)
        self.assertIn('350000062773', pboms)
        # Should not include 350000062776 unless site has _LOS
        self.assertNotIn('350000062776', pboms)


class TestDuplicatePrevention(unittest.TestCase):
    """Test duplicate PBOM prevention."""

    def test_no_duplicate_pbom(self):
        """has_duplicate_pbom should return False for unique PBOMs."""
        items = [
            {'PBOM_Code': 'A'},
            {'PBOM_Code': 'B'},
            {'PBOM_Code': 'C'}
        ]
        self.assertFalse(has_duplicate_pbom(items))

    def test_duplicate_pbom_detected(self):
        """has_duplicate_pbom should return True for duplicate PBOMs."""
        items = [
            {'PBOM_Code': 'A'},
            {'PBOM_Code': 'B'},
            {'PBOM_Code': 'A'}
        ]
        self.assertTrue(has_duplicate_pbom(items))

    def test_tss_scenario_no_duplicates(self):
        """Verify that the TSS scenarios produce no duplicate PBOMs."""
        tss_models = [
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000062773', 'Description': 'LOS Survey', 'Unit': 'Each', 'Quantity': 1, 'Is_Mandatory': True, 'Remarks': ''},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000062776', 'Description': 'LOS Survey LOS', 'Unit': 'Each', 'Quantity': 1, 'Is_Mandatory': True, 'Remarks': ''},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000589343', 'Description': 'Item A', 'Unit': 'Each', 'Quantity': 1.0, 'Is_Mandatory': True, 'Remarks': 'New Link'},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000589343', 'Description': 'Item A', 'Unit': 'Each', 'Quantity': 1.5, 'Is_Mandatory': True, 'Remarks': 'Reroute'},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000589344', 'Description': 'Item B', 'Unit': 'Each', 'Quantity': 1.0, 'Is_Mandatory': True, 'Remarks': 'New Link'},
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000589344', 'Description': 'Item B', 'Unit': 'Each', 'Quantity': 1.5, 'Is_Mandatory': True, 'Remarks': 'Reroute'},
        ]

        result = select_tss_items_for_site('A01073_AD', 'MW New Link / Reroute', 'install', tss_models)
        self.assertFalse(has_duplicate_pbom(result))


class TestLOSDetection(unittest.TestCase):
    """Test LOS site detection logic."""

    def test_los_site_detection(self):
        """Site IDs containing '_LOS' (case-insensitive) should be detected."""
        tss_models = [
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000062776', 'Description': 'LOS Survey LOS', 'Unit': 'Each', 'Quantity': 1, 'Is_Mandatory': True, 'Remarks': ''},
        ]

        # Non-LOS site
        result = select_tss_items_for_site('B00256', 'MW New Link / Reroute', 'dismantle something', tss_models)
        self.assertEqual(len(result), 0)  # Should exclude 350000062776

        # LOS site
        result = select_tss_items_for_site('SITE_LOS_002', 'MW New Link / Reroute', 'dismantle something', tss_models)
        self.assertEqual(len(result), 1)

    def test_case_insensitive_los_detection(self):
        """LOS detection should be case-insensitive."""
        tss_models = [
            {'SOW': 'MW New Link / Reroute', 'PBOM_Code': '350000062776', 'Description': 'LOS Survey LOS', 'Unit': 'Each', 'Quantity': 1, 'Is_Mandatory': True, 'Remarks': ''},
        ]

        result = select_tss_items_for_site('site_los_001', 'MW New Link / Reroute', 'dismantle', tss_models)
        self.assertEqual(len(result), 1)


class TestProductionExcelPBOMNormalization(unittest.TestCase):
    """Regression coverage for Excel-loaded PBOM values through the real generator path."""

    @classmethod
    def setUpClass(cls):
        cls.repo_root = Path(__file__).parent.parent
        cls.sample_site_path = cls.repo_root / 'Info' / 'input' / 'site_pr_po_view.xlsx'
        cls.mapping_path = cls.repo_root / 'Info' / 'input' / 'contract_info_reference.md'
        cls.template_path = cls.repo_root / 'Info' / 'input' / 'ecc_template.xls'
        cls.generator_path = cls.repo_root / 'scripts' / 'generate_tss_pr_ecc.py'
        cls.scenario_rows = {
            '4008B_AD': {'expected_survey_pbom': '350000062773', 'expected_qty': 1.0},
            '1679H_LOS': {'expected_survey_pbom': '350000062773', 'expected_qty': 1.0},
            '4982B': {'expected_survey_pbom': '350000062773', 'expected_qty': 1.5},
            '1258H_LOS': {'expected_survey_pbom': '350000062776', 'expected_qty': 1.5},
        }

    def _build_temp_pr_model(self, workbook_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = 'TX Line Item (After 21-Apr 26)'

        for _ in range(7):
            ws.append([None] * 7)

        # Force pandas to keep the PBOM column as floats when reading from Excel.
        ws.append(['Other SOW', 1234567890.5, 'Ignore float coercion row', 'Each', 1.0, 'Optional', ''])
        ws.append(['MW New Link / Reroute', 350000062773.0, 'LOS Survey', 'Each', 1.0, 'Mandatory', ''])
        ws.append(['MW New Link / Reroute', 350000062776.0, 'LOS Survey LOS', 'Each', 1.0, 'Mandatory', ''])
        ws.append(['MW New Link / Reroute', 350000589343.0, 'Item A', 'Each', 1.0, 'Mandatory', 'New Link'])
        ws.append(['MW New Link / Reroute', 350000589344.0, 'Item B', 'Each', 1.0, 'Mandatory', 'New Link'])
        ws.append(['MW New Link / Reroute', 350000589343.0, 'Item A', 'Each', 1.5, 'Mandatory', 'Reroute'])
        ws.append(['MW New Link / Reroute', 350000589344.0, 'Item B', 'Each', 1.5, 'Mandatory', 'Reroute'])
        ws.append([None] * 7)

        wb.save(workbook_path)

    def _build_temp_site_data(self, workbook_path: Path) -> None:
        df = pd.read_excel(self.sample_site_path, sheet_name='data', header=3)
        selected = df[df['customer site code'].isin(self.scenario_rows.keys())].copy()
        self.assertEqual(len(selected), 4, 'Expected four site rows for the production regression fixture')

        wb = Workbook()
        ws = wb.active
        ws.title = 'data'

        for _ in range(3):
            ws.append([None] * len(selected.columns))

        ws.append(list(selected.columns))
        for row in selected.itertuples(index=False, name=None):
            ws.append(list(row))

        wb.save(workbook_path)

    def _run_generator(self, site_path: Path, pr_model_path: Path, output_dir: Path) -> None:
        command = [
            sys.executable,
            str(self.generator_path),
            '--site-data',
            str(site_path),
            '--pr-model',
            str(pr_model_path),
            '--template',
            str(self.template_path),
            '--mapping',
            str(self.mapping_path),
            '--output',
            str(output_dir),
            '--scope',
            'TSS',
            '--all-sites',
        ]
        env = os.environ.copy()
        env["BYPASS_PR_MODEL_HASH_CHECK"] = "1"
        result = subprocess.run(
            command,
            cwd=self.repo_root,
            capture_output=True,
            text=True,
            check=False,
            env=env,
        )
        if result.returncode != 0:
            self.fail(
                f'Generator failed with exit code {result.returncode}\n'
                f'STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}'
            )

    def _load_output_rows(self, output_dir: Path) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for workbook_path in sorted(output_dir.glob('*.xlsx')):
            wb = load_workbook(workbook_path, data_only=True)
            ws = wb['details']
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(value is not None and str(value).strip() != '' for value in row):
                    continue
                rows.append(dict(zip(headers, row)))
        return rows

    def test_generator_normalizes_excel_loaded_pbom_codes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            temp_root = Path(tmpdir)
            pr_model_path = temp_root / 'pr_model.xlsx'
            site_path = temp_root / 'site_data.xlsx'
            output_dir = temp_root / 'output'
            output_dir.mkdir()

            self._build_temp_pr_model(pr_model_path)
            self._build_temp_site_data(site_path)
            self._run_generator(site_path, pr_model_path, output_dir)

            rows = self._load_output_rows(output_dir)
            self.assertTrue(rows, 'Expected the generator to produce ECC output rows')

            for site_id, expectation in self.scenario_rows.items():
                site_rows = [row for row in rows if str(row['Site ID*']).strip() == site_id]
                self.assertTrue(site_rows, f'Missing ECC rows for site {site_id}')

                pboms = [str(row['PBOM Code*']).strip() for row in site_rows]
                survey_pboms = [pbom for pbom in pboms if pbom.startswith('35000006277')]

                self.assertEqual(
                    survey_pboms.count(expectation['expected_survey_pbom']),
                    1,
                    f'{site_id} should include exactly one selected survey PBOM',
                )
                unwanted_pbom = '350000062776' if expectation['expected_survey_pbom'] == '350000062773' else '350000062773'
                self.assertNotIn(unwanted_pbom, survey_pboms, f'{site_id} should exclude the unwanted survey PBOM')
                self.assertNotIn('.0', ''.join(survey_pboms), f'{site_id} should emit canonical survey PBOM codes without .0')

                for controlled_pbom in ('350000589343', '350000589344'):
                    controlled_rows = [row for row in site_rows if str(row['PBOM Code*']).strip() == controlled_pbom]
                    self.assertEqual(len(controlled_rows), 1, f'{site_id} should include {controlled_pbom} exactly once')
                    self.assertEqual(
                        float(controlled_rows[0]['Quantity*']),
                        expectation['expected_qty'],
                        f'{site_id} should emit qty {expectation["expected_qty"]} for {controlled_pbom}',
                    )

                self.assertEqual(len(pboms), len(set(pboms)), f'{site_id} should not contain duplicate PBOM rows')


if __name__ == '__main__':
    unittest.main(verbosity=2)
